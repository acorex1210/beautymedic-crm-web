# -*- coding: utf-8 -*-
"""meta_ads.py
================
Importación de reportes de **Meta Ads** (Ads Manager) SIN API.

El usuario descarga el export (CSV o XLSX) desde Meta Ads Manager y lo sube a
la web. No se usa la Marketing API ni scraping, así que Meta no puede bloquear
la cuenta.

Cada carga se guarda con su fecha en ``META_DIR/cargas/{id}.json`` y un índice
``meta_cargas.json`` con el historial y los KPIs resumidos.
"""
import csv
import io
import json
import os
import re
from datetime import datetime

import openpyxl

# Cabecera normalizada (sin acentos, en mayúsculas) -> clave lógica
MAPA = {
    'CAMPANA': 'campania',
    'NOMBRE DE LA CAMPANA': 'campania',
    'CONJUNTO DE ANUNCIOS': 'conjunto',
    'ANUNCIO': 'anuncio',
    'PRESUPUESTO': 'presupuesto',
    'PRESUPUESTO DE LA CAMPANA': 'presupuesto',
    'PRESUPUESTO DEL CONJUNTO DE ANUNCIOS': 'presupuesto',
    'GASTO': 'gasto',
    'IMPORTE GASTADO': 'gasto',
    'IMPORTE INVERTIDO': 'gasto',
    'CANTIDAD GASTADA': 'gasto',
    'AMOUNT SPENT': 'gasto',
    'IMPRESIONES': 'impresiones',
    'ALCANCE': 'alcance',
    'CLICS EN EL ENLACE': 'clics',
    'CLICS (ENLACE)': 'clics',
    'CLICKS EN EL ENLACE': 'clics',
    'CLICS': 'clics',
    'CPC': 'cpc',
    'CPM': 'cpm',
    'RESULTADOS': 'resultados',
    'RESULTADO': 'resultados',
    'COSTE POR RESULTADO': 'costo_resultado',
    'COSTO POR RESULTADO': 'costo_resultado',
    'COSTE POR RESULTADOS': 'costo_resultado',
    'COSTO POR RESULTADOS': 'costo_resultado',
    'CONVERSIONES': 'conversiones',
    'COMPRAS': 'conversiones',
    'COSTE POR CONVERSION': 'costo_conversion',
    'COSTO POR CONVERSION': 'costo_conversion',
    'COSTE POR COMPRA': 'costo_conversion',
    'COSTO POR COMPRA': 'costo_conversion',
    'VALOR DE CONVERSION': 'valor_conversion',
    'VALOR DE LA CONVERSION': 'valor_conversion',
    'ROAS': 'roas',
    'FRECUENCIA': 'frecuencia',
    'ESTADO': 'estado',
    'ENTREGA': 'entrega',
    'ENTREGA DE LA CAMPANA': 'entrega',
    'CONTACTOS DE MENSAJES TOTALES': 'contactos',
    'NUEVOS CONTACTOS DE MENSAJES': 'nuevos_contactos',
    'MENSAJES INICIADOS': 'nuevos_contactos',
    'MENSAJES INICIADOS (TODO)': 'contactos',
}

MON = {'presupuesto', 'gasto', 'cpc', 'cpm', 'costo_resultado',
       'costo_conversion', 'valor_conversion'}
NUM = {'impresiones', 'alcance', 'clics', 'resultados', 'conversiones',
       'frecuencia', 'roas', 'contactos', 'nuevos_contactos'}
TXT = {'campania', 'conjunto', 'anuncio', 'estado', 'entrega'}
NUMERIC = MON | NUM


def _normalizar(v):
    s = str(v or '').strip().upper()
    for a, b in (('Á', 'A'), ('É', 'E'), ('Í', 'I'), ('Ó', 'O'),
                 ('Ú', 'U'), ('Ñ', 'N')):
        s = s.replace(a, b)
    # quita la moneda entre paréntesis de algunas cabeceras: "IMPORTE GASTADO (PEN)"
    for token in ('(PEN)', '(USD)', '(US$)', '(S/)', '($)'):
        s = s.replace(token, '')
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _num(v):
    """Convierte un valor de Meta (es-PE o en-US) a float, o None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(' ', '').replace('$', '').replace('%', '')
    s = s.replace('US$', '').replace('S/', '')
    if s in ('', '-', '—', 'N/A', 'NAN', 'NA', 'NULL', 'N/A'):
        return None
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        partes = s.split(',')
        if len(partes) == 2 and len(partes[1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    elif '.' in s:
        partes = s.split('.')
        if len(partes) == 2 and len(partes[1]) == 3 and len(partes[0]) <= 3:
            s = s.replace('.', '')
    try:
        return float(s)
    except ValueError:
        return None


def _detectar_fila_cabeceras(matriz):
    """Devuelve (fila_cabeceras, n_coincidencias)."""
    mejor = (0, 0)
    for i, fila in enumerate(matriz[:60]):
        n = sum(1 for c in fila if _normalizar(c) in MAPA)
        if n > mejor[1]:
            mejor = (i, n)
    return mejor


def _filas_a_dicts(cabeceras, filas_raw):
    """Mapea cada fila a {clave_lógica: valor} usando las cabeceras."""
    mapa_col = []
    for h in cabeceras:
        k = MAPA.get(_normalizar(h))
        if k is not None:
            mapa_col.append(k)
        else:
            mapa_col.append(None)
    filas = []
    for raw in filas_raw:
        d = {}
        n = 0
        for k, v in zip(mapa_col, raw):
            if k is None:
                continue
            s = str(v).strip() if v is not None else ''
            if k in NUMERIC:
                f = _num(s)
                if f is not None:
                    d[k] = f
                n += 1
            elif s:
                d[k] = s
        if d:
            filas.append(d)
    return filas


def _leer_csv(contenido):
    texto = contenido.decode('utf-8-sig', errors='replace')
    primera = texto.split('\n', 1)[0]
    if '\t' in primera and ',' not in primera:
        delim = '\t'
    elif ';' in primera and ',' not in primera:
        delim = ';'
    else:
        delim = ','
    reader = list(csv.reader(io.StringIO(texto), delimiter=delim))
    i, n = _detectar_fila_cabeceras(reader)
    if n == 0:
        raise ValueError('No se encontraron columnas de Meta Ads en el archivo')
    cabeceras = reader[i]
    datos = []
    for fila in reader[i + 1:]:
        if not any(str(c).strip() for c in fila):
            continue
        if _normalizar(fila[0]) in ('RESULTADOS DE LAS COLUMNAS', 'RESULTADOS'):
            continue
        datos.append(fila)
    return cabeceras, datos


def _leer_xlsx(contenido):
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True)
    ws = wb[wb.sheetnames[0]]
    matriz = []
    for row in ws.iter_rows(values_only=True):
        matriz.append(list(row))
    wb.close()
    i, n = _detectar_fila_cabeceras(matriz)
    if n == 0:
        raise ValueError('No se encontraron columnas de Meta Ads en el archivo')
    cabeceras = matriz[i]
    datos = []
    for fila in matriz[i + 1:]:
        if not any(str(c).strip() for c in fila):
            continue
        if _normalizar(fila[0]) in ('RESULTADOS DE LAS COLUMNAS', 'RESULTADOS'):
            continue
        datos.append(fila)
    return cabeceras, datos


def _columnas(cabeceras):
    out = []
    for h in cabeceras:
        k = MAPA.get(_normalizar(h))
        if k is None:
            continue
        out.append({'key': k, 'header': str(h).strip(),
                    'tipo': 'mon' if k in MON else ('num' if k in NUM else 'txt')})
    return out


def _kpis(filas):
    def tot(k):
        return sum((f.get(k) or 0) for f in filas
                   if isinstance(f.get(k), (int, float)))

    gasto = tot('gasto')
    imp = tot('impresiones')
    clics = tot('clics')
    conv = tot('conversiones')
    valor = tot('valor_conversion')
    return {
        'gasto': round(gasto, 2),
        'presupuesto': round(tot('presupuesto'), 2),
        'impresiones': int(imp),
        'alcance': int(tot('alcance')),
        'clics': int(clics),
        'cpc': round(clics and gasto / clics, 2),
        'cpm': round(imp and gasto / imp * 1000, 2),
        'resultados': int(tot('resultados')),
        'conversiones': int(conv),
        'costo_conversion': round(conv and gasto / conv, 2),
        'valor_conversion': round(valor, 2),
        'roas': round(gasto and valor / gasto, 2),
        'frecuencia': round(tot('frecuencia'), 2),
    }


def _por_campania(filas):
    agrup = {}
    for f in filas:
        c = f.get('campania') or '—'
        est = (f.get('entrega') or f.get('estado') or '').strip()
        g = agrup.setdefault(c, {'gasto': 0.0, 'presupuesto': 0.0,
                                 'impresiones': 0, 'alcance': 0, 'clics': 0,
                                 'conversiones': 0, 'valor': 0.0,
                                 'resultados': 0, 'contactos': 0,
                                 'nuevos_contactos': 0, 'estado': est})
        if est:
            g['estado'] = est
        for k, src in (('gasto', 'gasto'), ('presupuesto', 'presupuesto'),
                       ('impresiones', 'impresiones'), ('alcance', 'alcance'),
                       ('clics', 'clics'), ('conversiones', 'conversiones'),
                       ('valor', 'valor_conversion'),
                       ('resultados', 'resultados'),
                       ('contactos', 'contactos'),
                       ('nuevos_contactos', 'nuevos_contactos')):
            v = f.get(src)
            if isinstance(v, (int, float)):
                g[k] += v
    out = []
    for c, g in agrup.items():
        gasto = round(g['gasto'], 2)
        resultados = int(g['resultados'])
        contactos = int(g['contactos'])
        nuevos = int(g['nuevos_contactos'])
        imp = int(g['impresiones'])
        alc = int(g['alcance'])
        conv = int(g['conversiones'])
        valor = round(g['valor'], 2)
        clics = int(g['clics'])
        mensajes = nuevos or contactos or 0
        out.append({
            'campania': c,
            'estado': _estado(g['estado']),
            'gasto': gasto,
            'presupuesto': round(g['presupuesto'], 2),
            'impresiones': imp,
            'alcance': alc,
            'clics': clics,
            'conversiones': conv,
            'resultados': resultados,
            'contactos': contactos,
            'nuevos_contactos': nuevos,
            'valor': valor,
            'costo_resultado': round(resultados and gasto / resultados, 2),
            'costo_mensaje': round(mensajes and gasto / mensajes, 2),
            'costo_conversion': round(conv and gasto / conv, 2),
            'frecuencia': round(alc and imp / alc, 2),
            'cpm': round(imp and gasto / imp * 1000, 2),
            'cpc': round(clics and gasto / clics, 2),
            'roas': round(gasto and valor / gasto, 2),
        })
    out.sort(key=lambda x: x['gasto'], reverse=True)
    return out


def _estado(est):
    s = str(est or '').lower()
    if s in ('active', 'activa', 'activo', 'entregando', 'entregada', 'on', 'true', '1', 'en entrega'):
        return 'activa'
    if s in ('inactive', 'paused', 'en pausa', 'terminada', 'terminado',
             'archivada', 'archivado', 'off', 'false', '0', 'desactivada'):
        return 'inactiva'
    return 'inactiva' if not s else s


def _resumen(pc):
    res = {'activas': {'gasto': 0.0, 'resultados': 0, 'campanas': 0},
           'inactivas': {'gasto': 0.0, 'resultados': 0, 'campanas': 0}}
    for c in pc:
        d = res['activas'] if c['estado'] == 'activa' else res['inactivas']
        d['gasto'] += c['gasto']
        d['resultados'] += c['resultados']
        d['campanas'] += 1
    for d in res.values():
        d['gasto'] = round(d['gasto'], 2)
        d['costo_resultado'] = round(d['resultados'] and d['gasto'] / d['resultados'], 2)
    return res


def parsear(contenido, nombre):
    """Parsea un export de Meta Ads. Devuelve dict con columnas, filas y KPIs."""
    ext = (nombre or '').lower().rsplit('.', 1)[-1]
    if ext == 'csv':
        cabeceras, datos = _leer_csv(contenido)
    elif ext in ('xlsx', 'xls'):
        cabeceras, datos = _leer_xlsx(contenido)
    else:
        # intento automático: si parece texto -> csv, si no -> xlsx
        try:
            cabeceras, datos = _leer_csv(contenido)
        except Exception:
            try:
                cabeceras, datos = _leer_xlsx(contenido)
            except Exception:
                raise ValueError('El archivo no es un CSV ni un Excel de Meta Ads válido')
    filas = _filas_a_dicts(cabeceras, datos)
    por_campania = _por_campania(filas)
    return {
        'columnas': _columnas(cabeceras),
        'filas': filas,
        'kpis': _kpis(filas),
        'por_campania': por_campania,
        'resumen': _resumen(por_campania),
    }


def _ruta_index(dir):
    return os.path.join(dir, 'meta_cargas.json')


def _leer_index(dir):
    if not os.path.exists(_ruta_index(dir)):
        return []
    with open(_ruta_index(dir), encoding='utf-8') as f:
        return json.load(f)


def _escribir_index(dir, index):
    os.makedirs(dir, exist_ok=True)
    with open(_ruta_index(dir), 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)


def guardar(dir, contenido, nombre):
    """Guarda una carga nueva y devuelve su resumen. Lanza ValueError si es inválida."""
    parsed = parsear(contenido, nombre)
    if not parsed['filas']:
        raise ValueError('El archivo no tiene filas de datos de Meta Ads')
    ahora = datetime.now()
    carga_id = ahora.strftime('%Y%m%d_%H%M%S_%f')
    while os.path.exists(os.path.join(dir, 'cargas', f'{carga_id}.json')):
        ahora = datetime.now()
        carga_id = ahora.strftime('%Y%m%d_%H%M%S_%f')
    os.makedirs(os.path.join(dir, 'cargas'), exist_ok=True)
    detalle = {
        'id': carga_id,
        'archivo': nombre,
        'fecha': ahora.isoformat(timespec='seconds'),
        'columnas': parsed['columnas'],
        'filas': parsed['filas'],
        'kpis': parsed['kpis'],
        'por_campania': parsed['por_campania'],
        'resumen': parsed['resumen'],
    }
    with open(os.path.join(dir, 'cargas', f'{carga_id}.json'),
              'w', encoding='utf-8') as f:
        json.dump(detalle, f, ensure_ascii=False, indent=2)
    index = _leer_index(dir)
    index.insert(0, {
        'id': carga_id,
        'archivo': nombre,
        'fecha': detalle['fecha'],
        'filas': len(parsed['filas']),
        **parsed['kpis'],
    })
    _escribir_index(dir, index)
    return index[0]


def listar(dir):
    return _leer_index(dir)


def detalle(dir, carga_id):
    ruta = os.path.join(dir, 'cargas', f'{carga_id}.json')
    if not os.path.exists(ruta):
        raise ValueError('Carga no encontrada')
    with open(ruta, encoding='utf-8') as f:
        return json.load(f)


def borrar(dir, carga_id):
    ruta = os.path.join(dir, 'cargas', f'{carga_id}.json')
    if not os.path.exists(ruta):
        raise ValueError('Carga no encontrada')
    os.remove(ruta)
    index = [c for c in _leer_index(dir) if c['id'] != carga_id]
    _escribir_index(dir, index)
    return {'ok': True, 'borrados': len(index)}
