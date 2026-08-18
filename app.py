# -*- coding: utf-8 -*-
"""Web app de reportes Derma Essenza.

Genera los reportes PDF de ventas por campaña/CRM y sincroniza el maestro
BD DATA.xlsx desde Google Drive (AGENDADOS + VENTA DIARIA), igual que los
scripts de terminal reporte_ventas_pdf.py y alimentar_maestro.py.

Variables de entorno (además de las de alimentar_maestro.py):
  DATA_DIR   carpeta persistente para reportes y backups (default: ./data)
  REPORTES_DIR  carpeta para los PDF (default: DATA_DIR/reportes)
"""
import io
import json
import os
import shutil
import sys
import threading
import time
from datetime import datetime, timedelta
from typing import Dict, Optional

import openpyxl
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

import alimentar_maestro as am          # noqa: E402
import analitica as ana                 # noqa: E402
import crm_drive as crm                 # noqa: E402
import crm_plus as cp                   # noqa: E402
import meta_ads as mads                 # noqa: E402
import recibo_pdf as rp                 # noqa: E402
import reporte_ventas_pdf as rv         # noqa: E402
import whatsapp_leads as wa             # noqa: E402

DATA_DIR = os.environ.get('DATA_DIR', os.path.join(BASE_DIR, 'data'))
REPORTES_DIR = os.environ.get('REPORTES_DIR', os.path.join(DATA_DIR, 'reportes'))
BACKUP_DIR = os.environ.get('BACKUP_DIR', os.path.join(DATA_DIR, 'backups'))
RECIBOS_DIR = os.environ.get('RECIBOS_DIR', os.path.join(DATA_DIR, 'recibos'))
META_DIR = os.path.join(DATA_DIR, 'meta_ads')
os.makedirs(META_DIR, exist_ok=True)
MAX_BACKUPS = int(os.environ.get('MAX_BACKUPS', '12'))
os.makedirs(REPORTES_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(RECIBOS_DIR, exist_ok=True)


def _crear_backup():
    """Copia maestro y CRM.xlsx a la carpeta de backups, conservando los últimos N."""
    hecha = []
    hoy = datetime.now().strftime('%Y%m%d_%H%M%S')
    for origen, nombre in ((am.ruta_maestro_local(), 'BD DATA.xlsx'),
                           (os.path.join(am.TMP_DIR, 'CRM.xlsx'), 'CRM.xlsx')):
        if not os.path.exists(origen):
            continue
        destino = os.path.join(BACKUP_DIR, f'{hoy}_{nombre}')
        shutil.copy2(origen, destino)
        hecha.append(destino)
    for f in sorted(os.listdir(BACKUP_DIR)):
        if not f.lower().endswith('.xlsx'):
            continue
        versiones = sorted(g for g in os.listdir(BACKUP_DIR)
                           if g.endswith(f[f.find('_'):]))
        while len(versiones) > MAX_BACKUPS:
            os.remove(os.path.join(BACKUP_DIR, versiones.pop(0)))
    return hecha

_bloqueo = threading.Lock()

SYNC_HORA_UTC = int(os.environ.get('SYNC_HORA_UTC', '10'))  # 10:00 UTC = 05:00 Perú


def _sync_diario_loop():
    """Corre alimentar_maestro.ejecutar_sync(aplicar=True) una vez al día,
    para que el maestro (fuente de todo) siempre esté al día aunque nadie
    genere un reporte."""
    while True:
        ahora = datetime.utcnow()
        objetivo = ahora.replace(hour=SYNC_HORA_UTC, minute=0, second=0, microsecond=0)
        if objetivo <= ahora:
            objetivo += timedelta(days=1)
        time.sleep((objetivo - ahora).total_seconds())
        try:
            with _bloqueo:
                am.ejecutar_sync(aplicar=True)
        except Exception as e:  # noqa: BLE001
            print(f'[sync diario] error: {e}', file=sys.stderr)


threading.Thread(target=_sync_diario_loop, daemon=True).start()

app = FastAPI(title='Reportes Derma Essenza', version='1.0.0')
app.add_middleware(CORSMiddleware, allow_origins=['*'], allow_methods=['*'],
                   allow_headers=['*'])

BRAND = {
    'BRAND_NOMBRE': os.environ.get('BRAND_NOMBRE', 'Derma Essenza'),
    'BRAND_BADGE': os.environ.get('BRAND_BADGE', 'DE'),
}


def _html_index():
    with open(os.path.join(BASE_DIR, 'templates', 'index.html'),
              encoding='utf-8') as f:
        html = f.read()
    for k, v in BRAND.items():
        html = html.replace('{{' + k + '}}', v)
    return html


app.mount('/static', StaticFiles(directory=os.path.join(BASE_DIR, 'static')),
          name='static')


def _nombre_reporte(mes, anio, desde, hasta):
    nombre = f'Reporte_Ventas_{desde}-{hasta}_{rv.NOMBRES_MES.get(mes, mes)}_{anio}'
    for f in os.listdir(REPORTES_DIR):
        if f.lower().endswith('.pdf'):
            try:
                os.remove(os.path.join(REPORTES_DIR, f))
            except OSError:
                pass
    return os.path.join(REPORTES_DIR, f'{nombre}.pdf')


def _listar_reportes():
    out = []
    for f in sorted(os.listdir(REPORTES_DIR), reverse=True):
        if f.lower().endswith('.pdf'):
            p = os.path.join(REPORTES_DIR, f)
            out.append({'archivo': f, 'modificado': datetime.fromtimestamp(
                os.path.getmtime(p)).strftime('%d/%m/%Y %H:%M'),
                'tamano_kb': round(os.path.getsize(p) / 1024, 1)})
    return out


# ============================================================
# Schemas
# ============================================================
class ReporteReq(BaseModel):
    mes: str = 'AGO'
    anio: int = 2026
    desde: int = 1
    hasta: int = 10
    fuente: str = 'maestro'


class SyncReq(BaseModel):
    aplicar: bool = False


class AgendadoReq(BaseModel):
    crm: str = ''
    dia: Optional[int] = None
    mes: str = ''
    anio: Optional[int] = None
    nombre: str = ''
    red_social: str = ''
    telefono: str = ''
    correo: str = ''
    agendado_por: str = ''
    dia_cita: Optional[int] = None
    mes_cita: str = ''
    anio_cita: Optional[int] = None
    campana: str = ''
    hora: str = ''
    confirmado: str = ''
    observacion: str = ''
    reconfirmado: str = ''
    observacion2: str = ''


class VentaReq(BaseModel):
    dia: Optional[int] = None
    mes: str = ''
    anio: Optional[int] = None
    dni: str = ''
    cel: str = ''
    nombre: str = ''
    nuevo: str = ''
    distrito: str = ''
    edad: Optional[int] = None
    sexo: str = ''
    tratamiento: str = ''
    doctor: str = ''
    status: str = ''
    venta: Optional[float] = None
    pago: str = ''
    comisiona: Optional[float] = None
    observacion: str = ''
    producto_codigo: str = ''
    cantidad: Optional[float] = None


class ReprogramarReq(BaseModel):
    dia: Optional[int] = None
    mes: str = ''
    anio: Optional[int] = None


class LineaCompra(BaseModel):
    tratamiento: str = ''
    venta: Optional[float] = None
    producto_codigo: str = ''
    cantidad: Optional[float] = None


class ComproReq(BaseModel):
    dia: Optional[int] = None
    mes: str = ''
    anio: Optional[int] = None
    cel: str = ''
    nombre: str = ''
    doctor: str = ''
    status: str = 'SE REALIZO'
    pago: str = ''
    lineas: list[LineaCompra] = []


class MotivoReq(BaseModel):
    fila: int
    tipo: str = 'no_asistio'
    motivo: str = ''


class TarjetaReq(BaseModel):
    nombre: str = ''
    telefono: str = ''
    etapa: str = ''
    crm: str = ''
    campana: str = ''
    valor: Optional[float] = None
    prioridad: str = ''
    cita_dia: Optional[int] = None
    cita_mes: str = ''
    cita_hora: str = ''
    nota: str = ''


class TareaReq(BaseModel):
    titulo: str = ''
    tipo: str = ''
    fecha: str = ''
    hora: str = ''
    estado: str = ''
    prioridad: str = ''
    contacto: str = ''
    nota: str = ''


class NotaReq(BaseModel):
    telefono: str = ''
    contacto: str = ''
    tipo: str = ''
    texto: str = ''


class CuotaReq(BaseModel):
    paciente: str = ''
    telefono: str = ''
    tratamiento: str = ''
    monto_total: Optional[float] = None
    n_cuotas: Optional[int] = None
    pagadas: Optional[int] = None
    monto_cuota: Optional[float] = None
    prox_fecha: str = ''
    estado: str = ''
    nota: str = ''


class AperturaCajaReq(BaseModel):
    monto: float = 0
    monto_usd: float = 0
    responsable: str = ''
    nota: str = ''
    desglose: Optional[Dict[str, float]] = None
    desglose_usd: Optional[Dict[str, float]] = None


class MovimientoCajaReq(BaseModel):
    tipo: str
    monto: float = 0
    monto_usd: float = 0
    concepto: str = ''
    responsable: str = ''


class CierreCajaReq(BaseModel):
    monto: float = 0
    monto_usd: float = 0
    responsable: str = ''
    nota: str = ''
    desglose: Optional[Dict[str, float]] = None
    desglose_usd: Optional[Dict[str, float]] = None


class ProductoReq(BaseModel):
    codigo: str = ''
    producto: str = ''
    costo_bruto: Optional[float] = None
    stock: Optional[float] = None
    stock_minimo: Optional[float] = None
    unidad: str = ''


class MovimientoStockReq(BaseModel):
    tipo: str = 'ENTRADA'
    cantidad: Optional[float] = None
    referencia: str = ''
    nota: str = ''


class TrabajadorReq(BaseModel):
    nombre: str = ''
    cargo: str = ''
    sueldo_quincena: Optional[float] = None
    porcentaje_comision: Optional[float] = None
    metodo_pago: str = ''
    cuenta: str = ''
    estado: str = ''


class GenerarPlanillaReq(BaseModel):
    anio: int
    mes: str
    quincena: int


class PagoPlanillaReq(BaseModel):
    sueldo_base: Optional[float] = None
    comision: Optional[float] = None
    estado: str = ''
    metodo_pago: str = ''
    nota: str = ''


MESES_VALIDOS = {'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO',
                 'SET', 'SEP', 'OCT', 'NOV', 'DIC'}


def _validar_fecha(dia, mes, anio, requerido_mes=False):
    if dia is not None and not (1 <= dia <= 31):
        raise HTTPException(400, f'Día inválido: {dia}')
    if mes:
        m = mes.upper().replace('SEP', 'SET')
        if m not in MESES_VALIDOS:
            raise HTTPException(400, f'Mes inválido: {mes}')
    if anio is not None and not (2020 <= anio <= 2100):
        raise HTTPException(400, f'Año inválido: {anio}')


# ============================================================
# Página y estado
# ============================================================
@app.get('/', response_class=HTMLResponse)
async def index():
    return HTMLResponse(_html_index())


@app.get('/api/estado')
async def estado():
    maestro_ok = am.MAESTRO_FID or os.path.exists(am.ruta_maestro_local())
    return {
        'credenciales_ok': am.credenciales_disponibles(),
        'credenciales': am.CREDENCIALES,
        'maestro_ok': maestro_ok,
        'maestro': am.MAESTRO_FID or am.MAESTRO,
        'maestro_modificado': None,
        'agendados_ok': os.path.exists(os.path.join(am.TMP_DIR, 'AGENDADOS.xlsx')),
        'venta_ok': os.path.exists(os.path.join(am.TMP_DIR, 'VENTA_DIARIA.xlsx')),
        'meses': list(rv.NOMBRES_MES.keys()),
        'reportes': _listar_reportes(),
    }


# ============================================================
# Generar reporte PDF
# ============================================================
@app.post('/api/reporte')
async def generar_reporte(data: ReporteReq):
    mes = (data.mes or 'AGO').upper().replace('SEP', 'SET')
    if mes not in rv.NOMBRES_MES:
        raise HTTPException(400, f'Mes inválido: {mes}')
    if not (1 <= data.desde <= 31 and 1 <= data.hasta <= 31 and data.desde <= data.hasta):
        raise HTTPException(400, 'Rango de días inválido')
    if data.fuente not in ('maestro', 'auto'):
        raise HTTPException(400, 'Fuente inválida')
    if not (am.MAESTRO_FID or os.path.exists(am.ruta_maestro_local())):
        raise HTTPException(400, 'No hay maestro BD DATA.xlsx. Súbelo primero.')
    with _bloqueo:
        # El maestro es la fuente de todo: se corre la sincronización virtual
        # (AGENDADOS + VENTA DIARIA -> maestro) y se aplica siempre antes de
        # generar el reporte, sin importar la fuente elegida.
        am.ejecutar_sync(aplicar=True)
        ruta = _nombre_reporte(mes, data.anio, data.desde, data.hasta)
        try:
            res = rv.generar_reporte(mes=mes, anio=data.anio, desde=data.desde,
                                     hasta=data.hasta, fuente=data.fuente,
                                     salida=ruta)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(500, f'Error generando el reporte: {e}')
    verificacion = None
    if os.path.exists(os.path.join(am.TMP_DIR, 'VENTA_DIARIA.xlsx')):
        try:
            verificacion = am.verificar_venta_vs_td(
                os.path.join(am.TMP_DIR, 'VENTA_DIARIA.xlsx'),
                data.anio, mes, data.desde, data.hasta)
            if verificacion.get('ok'):
                reporte_monto = res['totales']['mon']
                monto_coincide = abs(reporte_monto - verificacion['venta']) < 0.01
                verificacion['reporte_monto'] = reporte_monto
                verificacion['monto_coincide'] = monto_coincide
                if not monto_coincide:
                    diffs = am.diferencias_maestro_venta(
                        am.ruta_maestro_local(),
                        os.path.join(am.TMP_DIR, 'VENTA_DIARIA.xlsx'),
                        data.anio, mes, data.desde, data.hasta)
                    verificacion['diferencias'] = diffs
                    if diffs:
                        detalle = '; '.join(
                            f"{d['nombre'] or d['telefono']} (fila {d['fila_maestro']}, "
                            f"{d['fecha']} {mes.lower()}): maestro S/ {d['monto_maestro']:,.2f} "
                            f"vs venta S/ {d['monto_venta']:,.2f}"
                            for d in diffs[:5])
                        verificacion['mensaje'] = (
                            f"El reporte maestro da S/ {reporte_monto:,.2f} pero VENTA 2026 / "
                            f"TD 2026 suman S/ {verificacion['venta']:,.2f} en el periodo. "
                            f"Revisa: {detalle}.")
                    else:
                        verificacion['mensaje'] = (
                            f"El reporte maestro da S/ {reporte_monto:,.2f} pero VENTA 2026 / "
                            f"TD 2026 suman S/ {verificacion['venta']:,.2f} en el periodo. "
                            'Algún dato está mal (revisa PAGO TOTAL acumulado en el maestro).')
                    verificacion['coincide'] = False
        except Exception:  # noqa: BLE001
            verificacion = None
    try:
        _crear_backup()
    except Exception:  # noqa: BLE001
        pass
    return {'ok': True, 'archivo': os.path.basename(res['archivo']),
            'url': f'/api/reporte/download/{os.path.basename(res["archivo"])}',
            'totales': res['totales'], 'por_crm': res['por_crm'],
            'detalle': res['detalle'], 'por_campana_meta': res.get('por_campana_meta', []),
            'verificacion': verificacion}


@app.get('/api/reportes')
async def listar_reportes():
    return _listar_reportes()


@app.get('/api/reporte/download/{archivo}')
async def descargar_reporte(archivo: str):
    p = os.path.join(REPORTES_DIR, os.path.basename(archivo))
    if not os.path.isfile(p):
        raise HTTPException(404, 'Reporte no encontrado')
    return FileResponse(p, filename=os.path.basename(p))


@app.delete('/api/reporte/{archivo}')
async def borrar_reporte(archivo: str):
    p = os.path.join(REPORTES_DIR, os.path.basename(archivo))
    if not os.path.isfile(p):
        raise HTTPException(404, 'Reporte no encontrado')
    os.remove(p)
    return {'ok': True}


# ============================================================
# Sincronizar maestro desde Drive
# ============================================================
@app.post('/api/sync')
async def sincronizar(data: SyncReq):
    with _bloqueo:
        resultado = am.ejecutar_sync(aplicar=data.aplicar)
    return resultado


# ============================================================
# Maestro BD DATA.xlsx (subir / descargar)
# ============================================================
@app.post('/api/maestro/upload')
async def subir_maestro(file: UploadFile = File(...)):
    contenido = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True)
        if 'BD DATA' not in wb.sheetnames:
            raise ValueError('El archivo no contiene la hoja "BD DATA"')
        wb.close()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f'Archivo inválido: {e}')
    ruta = os.path.join(am.TMP_DIR, 'BD DATA_upload.xlsx')
    with open(ruta, 'wb') as f:
        f.write(contenido)
    am.subir_maestro(ruta)
    return {'ok': True, 'maestro': am.MAESTRO_FID or am.MAESTRO}


@app.get('/api/maestro/download')
async def descargar_maestro():
    ruta = am.ruta_maestro_local()
    if not os.path.exists(ruta):
        raise HTTPException(404, 'No hay maestro subido')
    return FileResponse(ruta, filename='BD DATA.xlsx')


@app.put('/api/maestro/motivo')
async def registrar_motivo(req: MotivoReq):
    if req.tipo not in ('no_asistio', 'no_compra'):
        raise HTTPException(400, "tipo debe ser 'no_asistio' o 'no_compra'")
    if not req.motivo.strip():
        raise HTTPException(400, 'El motivo no puede estar vacío')
    try:
        col = am.actualizar_motivo_maestro(req.fila, req.tipo, req.motivo.strip().upper())
        return {'ok': True, 'col': col}
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo registrar el motivo: {e}')


# ============================================================
# Analítica interactiva (maestro BD DATA)
# ============================================================
def _filtros(mes: str = '', anio: str = '', desde: str = '', hasta: str = ''):
    m = (mes or 'AGO').upper().replace('SEP', 'SET')
    if m not in rv.NOMBRES_MES:
        raise HTTPException(400, f'Mes inválido: {m}')
    try:
        a = int(anio) if anio not in (None, '') else datetime.now().year
        d = int(desde) if desde not in (None, '') else 1
        h = int(hasta) if hasta not in (None, '') else 31
    except (TypeError, ValueError):
        raise HTTPException(400, 'Parámetros numéricos inválidos')
    if not (1 <= d <= 31 and 1 <= h <= 31 and d <= h):
        raise HTTPException(400, 'Rango de días inválido')
    return m, a, d, h


@app.get('/api/analitica/kpis')
async def analitica_kpis(mes: str = '', anio: str = '',
                         desde: str = '', hasta: str = ''):
    m, a, d, h = _filtros(mes, anio, desde, hasta)
    try:
        return {'ok': True, **ana.kpis(m, a, d, h)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular KPIs: {e}')


@app.get('/api/analitica/serie')
async def analitica_serie(mes: str = '', anio: str = '',
                          desde: str = '', hasta: str = ''):
    m, a, d, h = _filtros(mes, anio, desde, hasta)
    try:
        return {'ok': True, **ana.serie_diaria(m, a, d, h)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular la serie: {e}')


@app.get('/api/analitica/perfil')
async def analitica_perfil(mes: str = '', anio: str = '',
                           desde: str = '', hasta: str = ''):
    m, a, d, h = _filtros(mes, anio, desde, hasta)
    try:
        return {'ok': True, **ana.perfil(m, a, d, h)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular el perfil: {e}')


@app.get('/api/analitica/comparativo')
async def analitica_comparativo(mes: str = '', anio: str = '',
                                desde: str = '', hasta: str = ''):
    m, a, d, h = _filtros(mes, anio, desde, hasta)
    try:
        return {'ok': True, **ana.comparativo(m, a, d, h)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular el comparativo: {e}')


@app.get('/api/analitica/recurrentes')
async def analitica_recurrentes(mes: str = '', anio: str = '',
                                desde: str = '', hasta: str = ''):
    m, a, d, h = _filtros(mes, anio, desde, hasta)
    try:
        return {'ok': True, **ana.recurrentes(m, a, d, h)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudieron calcular recurrentes: {e}')


@app.get('/api/analitica/historico')
async def analitica_historico():
    try:
        return {'ok': True, 'meses': ana.ventas_por_mes()}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular el histórico: {e}')


@app.get('/api/analitica/paciente')
async def analitica_paciente(telefono: str = '', dni: str = ''):
    if not telefono and not dni:
        raise HTTPException(400, 'Indica teléfono o DNI del paciente')
    try:
        citas = ana.historial_paciente(telefono, dni or None)
        notas = cp.leer_notas(telefono=telefono) if telefono else []
        nombre = citas[0].get('nombre') if citas else ''
        return {'ok': True, 'telefono': telefono, 'nombre': nombre,
                'citas': citas, 'notas': notas}
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer el historial del paciente: {e}')


@app.get('/api/analitica/reactivar')
async def analitica_reactivar(meses: int = 3):
    try:
        return {'ok': True, 'pacientes': ana.pacientes_a_reactivar(meses)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular pacientes a reactivar: {e}')


@app.get('/api/analitica/ejecutivas')
async def analitica_ejecutivas(mes: str = '', anio: str = '',
                               desde: str = '', hasta: str = ''):
    m, a, d, h = _filtros(mes, anio, desde, hasta)
    try:
        return {'ok': True, **ana.ejecutivas(m, a, d, h)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular el ranking de ejecutivas: {e}')


@app.get('/api/analitica/motivos')
async def analitica_motivos(mes: str = '', anio: str = '',
                            desde: str = '', hasta: str = ''):
    m, a, d, h = _filtros(mes, anio, desde, hasta)
    try:
        return {'ok': True, **ana.motivos(m, a, d, h)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular los motivos: {e}')


# ============================================================
# Meta de ventas mensual (barra de progreso en el Panel)
# ============================================================
META_MENSUAL_PATH = os.path.join(DATA_DIR, 'meta_mensual.json')


def _leer_metas_mensuales():
    if not os.path.exists(META_MENSUAL_PATH):
        return {}
    try:
        with open(META_MENSUAL_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:  # noqa: BLE001
        return {}


class MetaMensualReq(BaseModel):
    mes: str
    anio: int
    meta: float


@app.get('/api/meta-mensual')
async def obtener_meta_mensual(mes: str = '', anio: str = ''):
    m, a, _, _ = _filtros(mes, anio, '1', '31')
    metas = _leer_metas_mensuales()
    return {'ok': True, 'mes': m, 'anio': a, 'meta': metas.get(f'{m}-{a}', 0)}


@app.put('/api/meta-mensual')
async def guardar_meta_mensual(data: MetaMensualReq):
    if data.meta < 0:
        raise HTTPException(400, 'La meta no puede ser negativa')
    m = data.mes.upper().replace('SEP', 'SET')
    if m not in rv.NOMBRES_MES:
        raise HTTPException(400, f'Mes inválido: {m}')
    metas = _leer_metas_mensuales()
    metas[f'{m}-{data.anio}'] = data.meta
    with open(META_MENSUAL_PATH, 'w', encoding='utf-8') as f:
        json.dump(metas, f, ensure_ascii=False, indent=2)
    return {'ok': True, 'mes': m, 'anio': data.anio, 'meta': data.meta}


# ============================================================
# Backups y exportaciones
# ============================================================
@app.post('/api/backup')
async def hacer_backup():
    try:
        hechas = _crear_backup()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo hacer el backup: {e}')
    return {'ok': True, 'backups': [os.path.basename(h) for h in hechas]}


@app.get('/api/backups')
async def listar_backups():
    out = []
    for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
        if f.lower().endswith('.xlsx'):
            p = os.path.join(BACKUP_DIR, f)
            out.append({'archivo': f, 'modificado': datetime.fromtimestamp(
                os.path.getmtime(p)).strftime('%d/%m/%Y %H:%M'),
                'tamano_kb': round(os.path.getsize(p) / 1024, 1)})
    return {'ok': True, 'backups': out}


@app.get('/api/backup/download/{archivo}')
async def descargar_backup(archivo: str):
    p = os.path.join(BACKUP_DIR, os.path.basename(archivo))
    if not os.path.isfile(p):
        raise HTTPException(404, 'Backup no encontrado')
    return FileResponse(p, filename=os.path.basename(p))


def _csv_response(columnas, filas, nombre):
    import csv
    import io as _io
    buf = _io.StringIO()
    w = csv.writer(buf)
    w.writerow(columnas)
    for f in filas:
        w.writerow([f.get(c) if not isinstance(f.get(c), (int, float)) or
                    isinstance(f.get(c), float) and f.get(c) == int(f.get(c))
                    else f.get(c) for c in columnas])
    return Response(buf.getvalue().encode('utf-8-sig'),
                    media_type='text/csv; charset=utf-8',
                    headers={'Content-Disposition':
                             f'attachment; filename="{nombre}.csv"'})


@app.get('/api/exportar/{tipo}')
async def exportar_csv(tipo: str):
    tipo = tipo.lower()
    try:
        if tipo == 'agendados':
            d = crm.leer_agendados()
            mapa = {'C': 'DIA', 'D': 'MES', 'E': 'AÑO', 'G': 'NOMBRE',
                    'H': 'RED SOCIAL', 'I': 'TELEFONO', 'K': 'AGENDADO POR',
                    'L': 'DIA2', 'M': 'MES3', 'N': 'AÑO4', 'O': 'CAMPAÑA',
                    'P': 'HORA', 'Q': 'ASISTENCIA'}
            filas = [{mapa.get(k, k): v for k, v in f.items()}
                     for f in d['filas']]
            cols = list(mapa.values())
            return _csv_response(cols, filas, 'agendados')
        if tipo == 'venta':
            d = crm.leer_venta()
            filas = [f for v in d['hojas'].values() for f in v['filas']]
            cols = sorted(filas[0].keys()) if filas else ['A']
            return _csv_response(cols, filas, 'venta_diaria')
        if tipo == 'cuotas':
            return _csv_response(
                ['ID', 'Paciente', 'Telefono', 'Tratamiento', 'Monto total',
                 'Cuotas', 'Pagadas', 'Saldo', 'Prox fecha', 'Estado'],
                [{'ID': c['id'], 'Paciente': c['paciente'], 'Telefono': c['telefono'],
                  'Tratamiento': c['tratamiento'], 'Monto total': c['monto_total'],
                  'Cuotas': c['n_cuotas'], 'Pagadas': c['pagadas'],
                  'Saldo': c['saldo'], 'Prox fecha': c['prox_fecha'],
                  'Estado': c['estado']} for c in cp.leer_cuotas()],
                'cuotas')
        if tipo == 'pacientes':
            filas = cp.leer_pacientes()
            cols = ['nombre', 'telefono', 'correo', 'crm', 'campana', 'citas',
                    'compras', 'total', 'proxima_cita', 'ultima_actividad', 'notas']
            return _csv_response(cols, filas, 'pacientes')
        if tipo == 'analitica':
            return _csv_response(
                ['Metrica', 'Actual', 'Mes anterior'],
                [{'Metrica': m, 'Actual': v, 'Mes anterior': ' '}
                 for m, v in ana.kpis('AGO', datetime.now().year, 1, 31).items()],
                'analitica')
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo exportar: {e}')
    raise HTTPException(404, f'Tipo no soportado: {tipo}')


# ============================================================
# CRM: Agendados y Venta diaria (lectura / alta en Drive)
# ============================================================
def _col_order(filas):
    cols = set()
    for f in filas:
        cols.update(f.keys())
    return sorted(cols, key=lambda c: openpyxl.utils.column_index_from_string(c))


def _normalizar_cabecera(v):
    return (str(v or '').strip().upper()
            .replace('Ñ', 'N').replace('Á', 'A').replace('É', 'E')
            .replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U'))


def _valores_data(ruta, mapa):
    """Valores únicos desde la pestaña 'DATA' de un archivo de Drive.

    ``mapa`` asocia cabecera normalizada -> clave. Detecta la fila de
    cabecera (la que contiene las columnas buscadas) y lee los valores
    de las filas siguientes, igual que los dropdowns de la hoja en Drive.
    """
    try:
        if not os.path.exists(ruta):
            return {}
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        ws = wb['DATA'] if 'DATA' in wb.sheetnames else None
        if ws is None:
            return {}
        filas_ws = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
        cab = None
        cols = {}
        for r, fila in enumerate(filas_ws, start=1):
            cabeceras = {_normalizar_cabecera(v): c
                         for c, v in enumerate(fila, start=1)}
            encontradas = {k: cabeceras[k] for k in mapa if k in cabeceras}
            if encontradas:
                cab = r
                cols = encontradas
                break
        if cab is None:
            return {}
        unicos = {k: set() for k in cols}
        for fila in ws.iter_rows(min_row=cab + 1, values_only=True):
            for k, c in cols.items():
                if c <= len(fila):
                    v = fila[c - 1]
                    if isinstance(v, str) and v.strip():
                        unicos[k].add(v.strip())
        return {mapa[k]: sorted(vals) for k, vals in unicos.items()}
    except Exception:  # noqa: BLE001
        return {}
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass


def _valores_agendados():
    """Valores de CAMPAÑA y RED SOCIAL desde la pestaña 'DATA' del archivo
    AGENDADOS (la fuente de los dropdowns de la hoja en Drive)."""
    return _valores_data(os.path.join(am.TMP_DIR, 'AGENDADOS.xlsx'),
                         {'CAMPAÑAS'.replace('Ñ', 'N'): 'campana',
                          'RED SOCIAL': 'red_social'})


def _valores_venta():
    """Valores de los selects de VENTA DIARIA desde la pestaña 'DATA' del
    archivo VENTA_DIARIA (la fuente de los dropdowns de la hoja en Drive)."""
    return _valores_data(os.path.join(am.TMP_DIR, 'VENTA_DIARIA.xlsx'),
                         {'NUEVO/RECURRENTE': 'nuevo',
                          'DISTRITO/DEPARTAMENTO': 'distrito',
                          'SEXO': 'sexo',
                          'TRATAMIENTO': 'tratamiento',
                          'DOCTOR': 'doctor',
                          'STATUS': 'status',
                          'PAGO': 'pago'})


@app.get('/api/crm/agendados')
async def crm_agendados():
    try:
        data = crm.leer_agendados()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer AGENDADOS de Drive: {e}')
    valores = crm.valores_unicos_agendados(data['filas'])
    extra = _valores_agendados()
    for campo in ('campana', 'red_social'):
        if extra.get(campo):
            valores[campo] = sorted(set(valores.get(campo, [])) | set(extra[campo]))
    return {
        'ok': True,
        'filas': data['filas'],
        'total': data['total'],
        'descargado': data['descargado'],
        'columnas': data['columnas'],
        'valores': valores,
    }


@app.post('/api/crm/agendados')
async def crm_agendados_nuevo(data: AgendadoReq):
    d = data.model_dump()
    _validar_fecha(d.get('dia'), d.get('mes'), d.get('anio'))
    _validar_fecha(d.get('dia_cita'), d.get('mes_cita'), d.get('anio_cita'),
                   requerido_mes=True)
    if not d.get('nombre') and not d.get('telefono'):
        raise HTTPException(400, 'Indica al menos el nombre o el teléfono del paciente')
    with _bloqueo:
        try:
            res = crm.agregar_agendado(d)
        except HTTPException:
            raise
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo guardar en AGENDADOS (Drive): {e}')
    return res


@app.delete('/api/crm/agendados/{fila}')
async def crm_agendados_borrar(fila: int):
    with _bloqueo:
        try:
            return crm.borrar_agendado(fila)
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo borrar el agendado en Drive: {e}')


@app.put('/api/crm/agendados/{fila}')
async def crm_agendados_editar(fila: int, data: AgendadoReq):
    d = data.model_dump()
    _validar_fecha(d.get('dia'), d.get('mes'), d.get('anio'))
    _validar_fecha(d.get('dia_cita'), d.get('mes_cita'), d.get('anio_cita'),
                   requerido_mes=True)
    if not d.get('nombre') and not d.get('telefono'):
        raise HTTPException(400, 'Indica al menos el nombre o el teléfono del paciente')
    with _bloqueo:
        try:
            res = crm.editar_agendado(fila, d)
        except HTTPException:
            raise
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo editar el agendado en Drive: {e}')
    return res


@app.post('/api/crm/agendados/{fila}/confirmar')
async def crm_agendados_confirmar(fila: int):
    with _bloqueo:
        try:
            return crm.actualizar_campos_agendado(fila, {'Q': 'CONFIRMADO'})
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo confirmar la cita en Drive: {e}')


@app.post('/api/crm/agendados/{fila}/cancelar')
async def crm_agendados_cancelar(fila: int):
    with _bloqueo:
        try:
            return crm.actualizar_campos_agendado(fila, {'Q': 'CANCELA'})
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo cancelar la cita en Drive: {e}')


@app.post('/api/crm/agendados/{fila}/reprogramar')
async def crm_agendados_reprogramar(fila: int, data: ReprogramarReq):
    d = data.model_dump()
    _validar_fecha(d.get('dia'), d.get('mes'), d.get('anio'))
    if d.get('dia') is None and not d.get('mes') and d.get('anio') is None:
        raise HTTPException(400, 'Indica la nueva fecha de la cita')
    with _bloqueo:
        try:
            return crm.actualizar_campos_agendado(
                fila, {'L': d.get('dia'), 'M': d.get('mes'), 'N': d.get('anio')})
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo reprogramar la cita en Drive: {e}')


@app.post('/api/crm/agendados/{fila}/compro')
async def crm_agendados_compro(fila: int, data: ComproReq):
    d = data.model_dump()
    _validar_fecha(d.get('dia'), d.get('mes'), d.get('anio'))
    if not d.get('nombre'):
        raise HTTPException(400, 'Indica el nombre del paciente')
    lineas = [ln for ln in d.get('lineas', []) if ln.get('tratamiento')]
    if not lineas:
        raise HTTPException(400, 'Indica al menos un tratamiento')
    avisos_inventario = []
    with _bloqueo:
        try:
            venta = crm.agregar_ventas_multi(
                {'dia': d.get('dia'), 'mes': d.get('mes'), 'anio': d.get('anio'),
                 'cel': d.get('cel'), 'nombre': d.get('nombre'),
                 'doctor': d.get('doctor'), 'status': d.get('status'),
                 'pago': d.get('pago')}, lineas)
            estado = crm.actualizar_campos_agendado(fila, {'Q': 'ASISTIO'})
        except HTTPException:
            raise
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo registrar la compra en Drive: {e}')
        for ln in lineas:
            codigo = str(ln.get('producto_codigo') or '').strip()
            if not codigo:
                continue
            try:
                cp.registrar_salida_stock(codigo, ln.get('cantidad') or 1,
                                          referencia=f"Venta: {d.get('nombre')}")
            except Exception as e:  # noqa: BLE001
                avisos_inventario.append(f'{codigo}: {e}')
    return {'ok': True, 'venta': venta, 'agendado': estado,
            'avisos_inventario': avisos_inventario}


@app.post('/api/crm/agendados/{fila}/nocompro')
async def crm_agendados_nocompro(fila: int):
    with _bloqueo:
        try:
            return crm.actualizar_campos_agendado(fila, {'Q': 'ASISTIO SIN COMPRA'})
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo registrar el no compró en Drive: {e}')


@app.post('/api/crm/agendados/{fila}/noasistio')
async def crm_agendados_noasistio(fila: int):
    with _bloqueo:
        try:
            return crm.actualizar_campos_agendado(fila, {'Q': 'NO ASISTIO'})
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo registrar el no asistió en Drive: {e}')


@app.post('/api/crm/agendados/{fila}/nocontesto')
async def crm_agendados_nocontesto(fila: int):
    with _bloqueo:
        try:
            return crm.actualizar_campos_agendado(fila, {'Q': 'NO CONTESTO'})
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo registrar el no contestó en Drive: {e}')


@app.get('/api/crm/venta')
async def crm_venta():
    try:
        data = crm.leer_venta()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer VENTA DIARIA de Drive: {e}')
    hojas = {h: {'filas': v['filas'], 'total': len(v['filas']),
                 'columnas': v['columnas']}
             for h, v in data['hojas'].items()}
    todas = [f for v in data['hojas'].values() for f in v['filas']]
    valores = crm.valores_unicos_venta(todas)
    extra = _valores_venta()
    for campo in extra:
        if extra[campo]:
            valores[campo] = sorted(set(valores.get(campo, [])) | set(extra[campo]))
    return {
        'ok': True,
        'hojas': hojas,
        'descargado': data['descargado'],
        'valores': valores,
    }


@app.post('/api/crm/venta')
async def crm_venta_nuevo(data: VentaReq):
    d = data.model_dump()
    _validar_fecha(d.get('dia'), d.get('mes'), d.get('anio'))
    if not d.get('nombre'):
        raise HTTPException(400, 'Indica el nombre del paciente')
    aviso_inventario = None
    with _bloqueo:
        try:
            res = crm.agregar_venta(d)
        except HTTPException:
            raise
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo guardar en VENTA DIARIA (Drive): {e}')
        codigo = str(d.get('producto_codigo') or '').strip()
        if codigo:
            try:
                cp.registrar_salida_stock(codigo, d.get('cantidad') or 1,
                                          referencia=f"Venta: {d.get('nombre')}")
            except Exception as e:  # noqa: BLE001
                aviso_inventario = f'{codigo}: {e}'
    if aviso_inventario:
        res['aviso_inventario'] = aviso_inventario
    return res


@app.delete('/api/crm/venta/{fila}')
async def crm_venta_borrar(fila: int, hoja: str):
    with _bloqueo:
        try:
            return crm.borrar_venta(hoja, fila)
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo borrar la venta en Drive: {e}')


@app.put('/api/crm/venta/{fila}')
async def crm_venta_editar(fila: int, hoja: str, data: VentaReq):
    d = data.model_dump()
    _validar_fecha(d.get('dia'), d.get('mes'), d.get('anio'))
    if not d.get('nombre'):
        raise HTTPException(400, 'Indica el nombre del paciente')
    with _bloqueo:
        try:
            return crm.editar_venta(hoja, fila, d)
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(502, f'No se pudo editar la venta en Drive: {e}')


@app.get('/api/crm/venta/recibo')
async def crm_venta_recibo(hoja: str, fila: int, n: int = 1):
    """Recibo interno en PDF de una venta (una o varias líneas consecutivas
    de VENTA DIARIA que comparten la misma compra). No es una boleta o
    factura electrónica válida ante SUNAT."""
    try:
        data = crm.leer_venta()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer VENTA DIARIA de Drive: {e}')
    v = data['hojas'].get(hoja)
    if not v:
        raise HTTPException(404, f'Hoja "{hoja}" no encontrada')
    objetivo = set(range(fila, fila + max(1, n)))
    seleccionadas = sorted((f for f in v['filas'] if f.get('_fila') in objetivo),
                           key=lambda f: f['_fila'])
    if not seleccionadas:
        raise HTTPException(404, 'No se encontraron filas de venta para ese recibo')
    base = seleccionadas[0]
    datos = {
        'hoja': hoja, 'fila': fila,
        'fecha': f"{base.get('B', '')}/{base.get('C', '')}/{base.get('D', '')}",
        'paciente': base.get('G', ''), 'telefono': base.get('F', ''),
        'doctor': base.get('M', ''), 'pago': base.get('P', ''),
        'lineas': [{'tratamiento': f.get('L', ''), 'venta': am.num(f.get('O'))}
                  for f in seleccionadas],
    }
    nombre_archivo = f"recibo_{hoja.replace(' ', '_')}_{fila}.pdf"
    ruta = os.path.join(RECIBOS_DIR, nombre_archivo)
    try:
        rp.generar_recibo(datos, ruta, brand_nombre=BRAND['BRAND_NOMBRE'])
    except Exception as e:  # noqa: BLE001
        raise HTTPException(500, f'No se pudo generar el recibo: {e}')
    return FileResponse(ruta, media_type='application/pdf', filename=nombre_archivo)


# ============================================================
# CRM Plus: Pipeline kanban
# ============================================================
@app.get('/api/crm/pipeline')
async def crm_pipeline():
    try:
        return {'ok': True, 'etapas': cp.ETAPAS, 'tarjetas': cp.leer_tarjetas()}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer el pipeline (Drive): {e}')


@app.post('/api/crm/pipeline')
async def crm_pipeline_nueva(data: TarjetaReq):
    try:
        res = cp.crear_tarjeta(data.model_dump(exclude_unset=True))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo guardar la tarjeta (Drive): {e}')
    return {'ok': True, 'tarjeta': res}


@app.patch('/api/crm/pipeline/{tid}')
async def crm_pipeline_actualizar(tid: int, data: TarjetaReq):
    try:
        res = cp.actualizar_tarjeta(tid, data.model_dump(exclude_unset=True))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo actualizar la tarjeta (Drive): {e}')
    if not res:
        raise HTTPException(404, 'Tarjeta no encontrada')
    return {'ok': True, 'tarjeta': res}


@app.delete('/api/crm/pipeline/{tid}')
async def crm_pipeline_borrar(tid: int):
    try:
        ok = cp.borrar_tarjeta(tid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo borrar la tarjeta (Drive): {e}')
    if not ok:
        raise HTTPException(404, 'Tarjeta no encontrada')
    return {'ok': True}


# ============================================================
# CRM Plus: Tareas y recordatorios
# ============================================================
@app.get('/api/crm/tareas')
async def crm_tareas(estado: str = ''):
    try:
        return {'ok': True, 'tareas': cp.leer_tareas(estado=estado or None)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudieron leer las tareas (Drive): {e}')


@app.post('/api/crm/tareas')
async def crm_tareas_nueva(data: TareaReq):
    if not data.titulo:
        raise HTTPException(400, 'Indica el título de la tarea')
    try:
        res = cp.crear_tarea(data.model_dump(exclude_unset=True))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo guardar la tarea (Drive): {e}')
    return {'ok': True, 'tarea': res}


@app.patch('/api/crm/tareas/{tid}')
async def crm_tareas_actualizar(tid: int, data: TareaReq):
    try:
        res = cp.actualizar_tarea(tid, data.model_dump(exclude_unset=True))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo actualizar la tarea (Drive): {e}')
    if not res:
        raise HTTPException(404, 'Tarea no encontrada')
    return {'ok': True, 'tarea': res}


@app.delete('/api/crm/tareas/{tid}')
async def crm_tareas_borrar(tid: int):
    try:
        ok = cp.borrar_tarea(tid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo borrar la tarea (Drive): {e}')
    if not ok:
        raise HTTPException(404, 'Tarea no encontrada')
    return {'ok': True}


# ============================================================
# CRM Plus: Notas de seguimiento
# ============================================================
@app.get('/api/crm/notas')
async def crm_notas(telefono: str = '', contacto: str = ''):
    try:
        return {'ok': True, 'notas': cp.leer_notas(
            telefono=telefono or None, contacto=contacto or None)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudieron leer las notas (Drive): {e}')


@app.post('/api/crm/notas')
async def crm_notas_nueva(data: NotaReq):
    if not data.texto:
        raise HTTPException(400, 'Escribe la nota')
    try:
        res = cp.agregar_nota(data.model_dump(exclude_unset=True))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo guardar la nota (Drive): {e}')
    return {'ok': True, 'nota': res}


# ============================================================
# CRM Plus: Cuotas / pagos a plazos
# ============================================================
@app.get('/api/crm/cuotas')
async def crm_cuotas(estado: str = '', telefono: str = ''):
    try:
        return {'ok': True, 'cuotas': cp.leer_cuotas(
            estado=estado or None, telefono=telefono or None)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudieron leer las cuotas (Drive): {e}')


@app.post('/api/crm/cuotas')
async def crm_cuotas_nueva(data: CuotaReq):
    if not data.paciente and not data.telefono:
        raise HTTPException(400, 'Indica el paciente o teléfono')
    if data.monto_total is not None and data.monto_total <= 0:
        raise HTTPException(400, 'El monto total debe ser mayor a 0')
    try:
        res = cp.crear_cuota(data.model_dump(exclude_unset=True))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo guardar la cuota (Drive): {e}')
    return {'ok': True, 'cuota': res}


@app.patch('/api/crm/cuotas/{cid}')
async def crm_cuotas_actualizar(cid: int, data: CuotaReq):
    try:
        res = cp.actualizar_cuota(cid, data.model_dump(exclude_unset=True))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo actualizar la cuota (Drive): {e}')
    if not res:
        raise HTTPException(404, 'Cuota no encontrada')
    return {'ok': True, 'cuota': res}


@app.post('/api/crm/cuotas/{cid}/pago')
async def crm_cuotas_pago(cid: int):
    try:
        res = cp.registrar_pago_cuota(cid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo registrar el pago (Drive): {e}')
    if not res:
        raise HTTPException(404, 'Cuota no encontrada')
    if res.get('error'):
        raise HTTPException(400, res['error'])
    return {'ok': True, 'cuota': res}


@app.delete('/api/crm/cuotas/{cid}')
async def crm_cuotas_borrar(cid: int):
    try:
        ok = cp.borrar_cuota(cid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo borrar la cuota (Drive): {e}')
    if not ok:
        raise HTTPException(404, 'Cuota no encontrada')
    return {'ok': True}


# ============================================================
# CRM Plus: Caja (apertura / movimientos / cierre del día)
# ============================================================
@app.get('/api/crm/caja')
async def crm_caja_estado(fecha: str = ''):
    try:
        return {'ok': True, **cp.estado_caja(fecha or None)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer el estado de caja (Drive): {e}')


@app.get('/api/crm/caja/historial')
async def crm_caja_historial(limite: int = 30):
    try:
        return {'ok': True, 'historial': cp.historial_caja(limite)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer el historial de caja (Drive): {e}')


@app.post('/api/crm/caja/apertura')
async def crm_caja_apertura(data: AperturaCajaReq):
    try:
        if cp.estado_caja()['abierta']:
            raise HTTPException(400, 'La caja de hoy ya está abierta')
        if not data.desglose and not data.monto and not data.desglose_usd and not data.monto_usd:
            raise HTTPException(400, 'Indica el desglose de billetes y monedas con el que abres caja')
        res = cp.abrir_caja({'monto': data.monto, 'monto_usd': data.monto_usd,
                             'concepto': data.nota, 'responsable': data.responsable,
                             'desglose': data.desglose, 'desglose_usd': data.desglose_usd})
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo abrir la caja (Drive): {e}')
    return {'ok': True, 'apertura': res}


@app.post('/api/crm/caja/movimiento')
async def crm_caja_movimiento(data: MovimientoCajaReq):
    try:
        estado = cp.estado_caja()
        if not estado['abierta']:
            raise HTTPException(400, 'Primero debes abrir la caja de hoy')
        if estado['cerrada']:
            raise HTTPException(400, 'La caja de hoy ya está cerrada')
        res = cp.registrar_movimiento_caja(data.model_dump())
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo registrar el movimiento (Drive): {e}')
    return {'ok': True, 'movimiento': res}


@app.delete('/api/crm/caja/{cid}')
async def crm_caja_borrar(cid: int):
    try:
        ok = cp.borrar_movimiento_caja(cid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo borrar el movimiento (Drive): {e}')
    if not ok:
        raise HTTPException(404, 'Movimiento no encontrado')
    return {'ok': True}


@app.post('/api/crm/caja/cierre')
async def crm_caja_cierre(data: CierreCajaReq):
    try:
        estado = cp.estado_caja()
        if not estado['abierta']:
            raise HTTPException(400, 'Primero debes abrir la caja de hoy')
        if estado['cerrada']:
            raise HTTPException(400, 'La caja de hoy ya está cerrada')
        res = cp.cerrar_caja({'monto': data.monto, 'monto_usd': data.monto_usd,
                              'concepto': data.nota, 'responsable': data.responsable,
                              'desglose': data.desglose, 'desglose_usd': data.desglose_usd})
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo cerrar la caja (Drive): {e}')
    return {'ok': True, 'cierre': res, 'estado': cp.estado_caja()}


# ============================================================
# CRM Plus: Inventario (catálogo de productos + kardex de stock)
# ============================================================
@app.get('/api/crm/inventario/productos')
async def inventario_productos():
    try:
        return {'ok': True, 'productos': cp.leer_productos()}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer el inventario (Drive): {e}')


@app.post('/api/crm/inventario/productos')
async def inventario_producto_nuevo(data: ProductoReq):
    try:
        p = cp.crear_producto(data.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo crear el producto (Drive): {e}')
    return {'ok': True, 'producto': p}


@app.patch('/api/crm/inventario/productos/{pid}')
async def inventario_producto_editar(pid: int, data: ProductoReq):
    try:
        p = cp.actualizar_producto(pid, data.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo editar el producto (Drive): {e}')
    if not p:
        raise HTTPException(404, 'Producto no encontrado')
    return {'ok': True, 'producto': p}


@app.delete('/api/crm/inventario/productos/{pid}')
async def inventario_producto_borrar(pid: int):
    try:
        ok = cp.borrar_producto(pid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo borrar el producto (Drive): {e}')
    if not ok:
        raise HTTPException(404, 'Producto no encontrado')
    return {'ok': True}


@app.get('/api/crm/inventario/movimientos')
async def inventario_movimientos(codigo: str = '', limite: int = 200):
    try:
        return {'ok': True, 'movimientos': cp.leer_movimientos_stock(codigo or None, limite)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer los movimientos (Drive): {e}')


@app.post('/api/crm/inventario/productos/{pid}/movimiento')
async def inventario_movimiento_nuevo(pid: int, data: MovimientoStockReq):
    try:
        productos = cp.leer_productos()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer el inventario (Drive): {e}')
    prod = next((p for p in productos if p['id'] == pid), None)
    if not prod:
        raise HTTPException(404, 'Producto no encontrado')
    if data.cantidad is None:
        raise HTTPException(400, 'Indica la cantidad')
    try:
        res = cp.registrar_movimiento_stock(prod['codigo'], data.tipo, data.cantidad,
                                            data.referencia, data.nota)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo registrar el movimiento (Drive): {e}')
    return {'ok': True, **res}


# ============================================================
# CRM Plus: Planilla (trabajadores + pagos quincenales)
# ============================================================
@app.get('/api/crm/planilla/trabajadores')
async def planilla_trabajadores(solo_activos: bool = False):
    try:
        return {'ok': True, 'trabajadores': cp.leer_trabajadores(solo_activos)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer los trabajadores (Drive): {e}')


@app.post('/api/crm/planilla/trabajadores')
async def planilla_trabajador_nuevo(data: TrabajadorReq):
    try:
        t = cp.crear_trabajador(data.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo crear el trabajador (Drive): {e}')
    return {'ok': True, 'trabajador': t}


@app.patch('/api/crm/planilla/trabajadores/{tid}')
async def planilla_trabajador_editar(tid: int, data: TrabajadorReq):
    try:
        t = cp.actualizar_trabajador(tid, data.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo editar el trabajador (Drive): {e}')
    if not t:
        raise HTTPException(404, 'Trabajador no encontrado')
    return {'ok': True, 'trabajador': t}


@app.delete('/api/crm/planilla/trabajadores/{tid}')
async def planilla_trabajador_borrar(tid: int):
    try:
        ok = cp.borrar_trabajador(tid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo borrar el trabajador (Drive): {e}')
    if not ok:
        raise HTTPException(404, 'Trabajador no encontrado')
    return {'ok': True}


@app.get('/api/crm/planilla')
async def planilla_listar(anio: int = 0, mes: str = '', quincena: int = 0):
    try:
        pagos = cp.leer_planilla(anio or None, mes or None, quincena or None)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer la planilla (Drive): {e}')
    tot = {'sueldo_base': 0.0, 'comision': 0.0, 'monto_total': 0.0,
           'pendiente': 0.0, 'pagado': 0.0}
    for p in pagos:
        tot['sueldo_base'] += p['sueldo_base']
        tot['comision'] += p['comision']
        tot['monto_total'] += p['monto_total']
        tot['pagado' if p['estado'] == 'PAGADO' else 'pendiente'] += p['monto_total']
    return {'ok': True, 'pagos': pagos, 'totales': {k: round(v, 2) for k, v in tot.items()}}


@app.post('/api/crm/planilla/generar')
async def planilla_generar(data: GenerarPlanillaReq):
    _validar_fecha(None, data.mes, data.anio, requerido_mes=True)
    if data.quincena not in (1, 2):
        raise HTTPException(400, 'La quincena debe ser 1 o 2')
    try:
        res = cp.generar_planilla_quincena(data.anio, data.mes, data.quincena)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo generar la planilla (Drive): {e}')
    return {'ok': True, **res}


@app.patch('/api/crm/planilla/{pid}')
async def planilla_pago_editar(pid: int, data: PagoPlanillaReq):
    cambios = data.model_dump()
    if cambios.get('estado') and cambios['estado'] not in cp.ESTADO_PAGO:
        raise HTTPException(400, f"Estado inválido: {cambios['estado']}")
    try:
        p = cp.actualizar_pago_planilla(pid, cambios)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo actualizar el pago (Drive): {e}')
    if not p:
        raise HTTPException(404, 'Pago de planilla no encontrado')
    return {'ok': True, 'pago': p}


@app.delete('/api/crm/planilla/{pid}')
async def planilla_pago_borrar(pid: int):
    try:
        ok = cp.borrar_pago_planilla(pid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo borrar el pago (Drive): {e}')
    if not ok:
        raise HTTPException(404, 'Pago de planilla no encontrado')
    return {'ok': True}


# ============================================================
# CRM Plus: Directorio, dashboard y actividades de hoy
# ============================================================
@app.get('/api/crm/pacientes')
async def crm_pacientes():
    try:
        return {'ok': True, 'pacientes': cp.leer_pacientes()}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo armar el directorio (Drive): {e}')


@app.get('/api/crm/dashboard')
async def crm_dashboard():
    try:
        return {'ok': True, **cp.leer_dashboard()}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo calcular el dashboard (Drive): {e}')


@app.get('/api/crm/hoy')
async def crm_hoy():
    try:
        return {'ok': True, **cp.leer_hoy()}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo armar el panel de hoy (Drive): {e}')


# ============================================================
# Meta Ads (import de reportes SIN API)
# ============================================================
@app.get('/api/meta')
async def meta_listar():
    try:
        return {'ok': True, 'cargas': mads.listar(META_DIR)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo listar el historial Meta Ads: {e}')


@app.post('/api/meta/upload')
async def meta_subir(file: UploadFile = File(...)):
    contenido = await file.read()
    try:
        res = mads.guardar(META_DIR, contenido, file.filename or 'reporte.csv')
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo guardar el reporte: {e}')
    return {'ok': True, 'carga': res}


@app.get('/api/meta/{carga_id}')
async def meta_detalle(carga_id: str):
    try:
        return {'ok': True, 'detalle': mads.detalle(META_DIR, carga_id)}
    except ValueError as e:
        raise HTTPException(404, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo leer el reporte: {e}')


@app.delete('/api/meta/{carga_id}')
async def meta_borrar(carga_id: str):
    try:
        return mads.borrar(META_DIR, carga_id)
    except ValueError as e:
        raise HTTPException(404, str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo borrar el reporte: {e}')


# ============================================================
# WhatsApp: webhook de solo-lectura (Cloud API) -> Pipeline
# ============================================================
@app.get('/api/whatsapp/webhook')
async def whatsapp_verificar(request: Request):
    """Handshake que pide Meta al configurar el webhook (hub.challenge)."""
    p = request.query_params
    if wa.VERIFY_TOKEN and p.get('hub.mode') == 'subscribe' and p.get('hub.verify_token') == wa.VERIFY_TOKEN:
        return PlainTextResponse(p.get('hub.challenge', ''))
    raise HTTPException(403, 'Token de verificación inválido o no configurado')


@app.post('/api/whatsapp/webhook')
async def whatsapp_recibir(request: Request):
    cuerpo = await request.body()
    firma = request.headers.get('x-hub-signature-256', '')
    if not wa.verificar_firma(cuerpo, firma):
        raise HTTPException(403, 'Firma inválida')
    try:
        payload = json.loads(cuerpo or b'{}')
        tarjetas = wa.procesar_webhook_payload(payload)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f'No se pudo procesar el mensaje: {e}')
    return {'ok': True, 'procesados': len(tarjetas)}


class CampanasWhatsappReq(BaseModel):
    campanas: dict


@app.get('/api/whatsapp/campanas')
async def whatsapp_campanas_get():
    return {'ok': True, 'campanas': wa.leer_campanas()}


@app.put('/api/whatsapp/campanas')
async def whatsapp_campanas_put(data: CampanasWhatsappReq):
    wa.guardar_campanas(data.campanas)
    return {'ok': True, 'campanas': data.campanas}
