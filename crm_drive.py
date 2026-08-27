# -*- coding: utf-8 -*-
"""crm_drive.py
============
Lectura y escritura de las hojas AGENDADOS y VENTA DIARIA (Google Drive)
para la página web tipo CRM de Derma Essenza.

- ``leer_*`` devuelve las filas actuales (con caché corta en TMP_DIR).
- ``agregar_*`` añade una fila nueva SIN crear archivos nuevos (el maestro
  sigue alimentándose de los .xlsx de siempre):
    * En .xlsx clásico usa cirugía XML (preserva autoFiltros, estilos,
      tablas dinámicas y fórmulas) y sube de vuelta conservando el mismo
      file id, con concurrencia best effort: un ``_lock`` serializa las
      escrituras de la app, un control preflight re-descarga y reintenta si
      la revisión del archivo cambió mientras se construía la fila (alguien
      editó directo en Drive), y un control postflight detecta con
      revisions.list si otra edición se coló y lo reporta en el resultado.
    * Si el archivo fuera una Google Sheet nativa, usa la Sheets API
      (values.append con INSERT_ROWS): append atómico del servidor.

Reutiliza la configuración de alimentar_maestro.py (CREDENCIALES,
AGENDADOS_FID, VENTA_FID, TMP_DIR) y sus normalizadores (num, esc_xml).
"""
import io
import os
import re
import shutil
import threading
import time
import unicodedata
import zipfile
from datetime import datetime

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

import alimentar_maestro as am  # noqa: E402

SCOPE = 'https://www.googleapis.com/auth/drive'
MIME_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
MIME_SHEETS = 'application/vnd.google-apps.spreadsheet'
CACHE_TTL = 600          # segundos que se conserva la copia local antes de re-descargar
# (cualquier escritura desde la propia app invalida el caché al toque via
# invalidar(); este TTL sólo cubre ediciones hechas fuera de la app
# directamente en Drive/Sheets, así que se puede estirar sin perder cambios
# propios. Antes en 120s: cada 2 minutos la primera pantalla que se abría
# pagaba el costo completo de descargar+exportar el xlsx desde Drive.)
MAX_REINTENTOS = 3       # reintentos ante conflicto 412 (alguien editó el xlsx a la vez)
_SELECTO_LIMIT = 200

_lock = threading.Lock()

# Columnas de cada hoja (letra -> nombre mostrable en el CRM)
AGENDADOS_COLS = {
    'B': 'CRM', 'C': 'DIA', 'D': 'MES', 'E': 'AÑO', 'G': 'NOMBRE',
    'H': 'RED SOCIAL', 'I': 'TELEFONO', 'J': 'CORREO', 'K': 'AGENDADO POR',
    'L': 'DIA CITA', 'M': 'MES CITA', 'N': 'AÑO CITA', 'O': 'CAMPAÑA',
    'P': 'HORA', 'Q': 'CONFIRMADO', 'R': 'OBSERVACION', 'S': 'RECONFIRMADO',
    'T': 'OBSERVACION2',
}
VENTA_COLS = {
    'B': 'DIA', 'C': 'MES', 'D': 'AÑO', 'E': 'DNI', 'F': 'CEL. PACIENTE',
    'G': 'NOMBRE Y APELLIDO', 'H': 'NUEVO/RECURRENTE', 'I': 'DISTRITO',
    'J': 'EDAD', 'K': 'SEXO', 'L': 'TRATAMIENTO', 'M': 'DOCTOR',
    'N': 'STATUS', 'O': 'VENTA', 'P': 'PAGO', 'Q': 'COMISIONA',
    'R': 'OBSERVACION', 'S': 'CAMPAÑA',
}

# Campo lógico -> letra canónica (layout fijo de AGENDADOS)
AGENDADOS_CANON = {
    'crm': 'B', 'dia': 'C', 'mes': 'D', 'anio': 'E',
    'dni': 'F', 'nombre': 'G', 'red_social': 'H', 'telefono': 'I',
    'asistencia': 'A', 'correo': 'J',
    'agendado_por': 'K', 'dia_cita': 'L', 'mes_cita': 'M', 'anio_cita': 'N',
    'campana': 'O', 'hora': 'P', 'confirmado': 'Q', 'observacion': 'R',
    'reconfirmado': 'S', 'observacion2': 'T',
}
# Campo lógico -> letra canónica para VENTA (layout fijo de VENTA DIARIA)
VENTA_CANON = {
    'dia': 'B', 'mes': 'C', 'anio': 'D', 'dni': 'E', 'cel': 'F',
    'nombre': 'G', 'nuevo': 'H', 'distrito': 'I', 'edad': 'J', 'sexo': 'K',
    'tratamiento': 'L', 'doctor': 'M', 'status': 'N', 'venta': 'O',
    'pago': 'P', 'comisiona': 'Q', 'campana': 'S', 'observacion': 'R',
}
# Nombres de encabezado normalizados (sin acentos ni Ñ) -> campo lógico
_HEADERS_AGENDADOS = {
    'CRM': 'crm', 'NOMBRE': 'nombre', 'RED SOCIAL': 'red_social',
    'TELEFONO': 'telefono', 'CORREO': 'correo', 'AGENDADO POR': 'agendado_por',
    'CAMPANA': 'campana', 'HORA': 'hora',
    'CONFIRMADO': 'confirmado', 'OBSERVACION': 'observacion',
    'RECONFIRMADO': 'reconfirmado', 'OBSERVACION2': 'observacion2',
    'DNI': 'dni', 'ASISTENCIA': 'asistencia',
}
_HEADERS_VENTA = {
    'DNI': 'dni', 'CEL': 'cel', 'CEL. PACIENTE': 'cel',
    'NOMBRE Y APELLIDO': 'nombre', 'NOMBRE': 'nombre',
    'NUEVO/RECURRENTE': 'nuevo', 'DISTRITO': 'distrito',
    'DISTRITO/DEPARTAMENTO': 'distrito', 'EDAD': 'edad', 'SEXO': 'sexo',
    'TRATAMIENTO': 'tratamiento', 'DOCTOR': 'doctor', 'STATUS': 'status',
    'VENTA': 'venta', 'PAGO': 'pago', 'COMISIONA': 'comisiona',
    'CAMPANA': 'campana', 'OBSERVACION': 'observacion',
}


def _normalizar_header(texto):
    s = str(texto or '').strip().upper()
    for a, b in (('Á', 'A'), ('É', 'E'), ('Í', 'I'), ('Ó', 'O'),
                 ('Ú', 'U'), ('Ñ', 'N')):
        s = s.replace(a, b)
    return re.sub(r'\s+', ' ', s)


def _campo_de_header(texto, mapas_header, contador):
    t = _normalizar_header(texto)
    if t in ('DIA', 'MES', 'ANO'):
        base = {'DIA': 'dia', 'MES': 'mes', 'ANO': 'anio'}[t]
        contador[base] = contador.get(base, 0) + 1
        return base if contador[base] == 1 else f'{base}_cita'
    if t in ('DIA 2', 'DIA CITA'):
        return 'dia_cita'
    if t in ('MES 2', 'MES CITA'):
        return 'mes_cita'
    if t in ('ANO 2', 'ANO CITA'):
        return 'anio_cita'
    return mapas_header.get(t)


_WB_CACHE = {}  # ruta -> (mtime, Workbook)


def _cargar_wb(ruta):
    """Parsea un .xlsx una sola vez por archivo (por mtime).

    leer_agendados/leer_venta llamaban a _detectar_columnas (que abre el
    libro de nuevo) y encima abrían su propia copia para leer las celdas:
    2-3 parseos completos del mismo archivo en una sola petición. Ese
    reparseo repetido era el cuello de botella real detrás de las
    pantallas lentas, no la descarga (que ya tenía su propio caché)."""
    mtime = os.path.getmtime(ruta)
    cache = _WB_CACHE.get(ruta)
    if cache and cache[0] == mtime:
        return cache[1]
    if cache:
        try:
            cache[1].close()
        except Exception:  # noqa: BLE001
            pass
    wb = openpyxl.load_workbook(ruta, data_only=True)
    _WB_CACHE[ruta] = (mtime, wb)
    return wb


def _detectar_columnas(ruta, hoja, mapas_header):
    """Detecta la fila de encabezados de la hoja.

    Devuelve (n_fila_encabezados, {letra_real: campo},
    {letra_real: encabezado_texto}).
    """
    wb = _cargar_wb(ruta)
    ws = wb[hoja]
    mejor = None
    for r in range(1, 13):
        contador = {}
        campos = {}
        textos = {}
        for c in range(1, 25):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            letra = openpyxl.utils.get_column_letter(c)
            campo = _campo_de_header(v, mapas_header, contador)
            if campo:
                campos[letra] = campo
                textos[letra] = str(v).strip()
        if len(campos) >= 3 and (mejor is None
                                 or len(campos) > len(mejor[1])):
            mejor = (r, campos, textos)
    if mejor is None:
        return 4, {}, {}
    return mejor


def _mapa_inverso(campos, canon):
    """{letra_canónica: letra_real} a partir de los campos detectados."""
    out = {}
    for real, campo in campos.items():
        cl = canon.get(campo)
        if cl:
            out[cl] = real
    return out


def _remapear(valores, inverso):
    """Re-mapea {letra_canónica: valor} a {letra_real: valor} (sólo campos presentes)."""
    return {inverso[cl]: v for cl, v in valores.items() if cl in inverso}


def _adaptador_agendados(ruta, valores):
    _, campos, _ = _detectar_columnas(ruta, 'AGENDADOS', _HEADERS_AGENDADOS)
    return _remapear(valores, _mapa_inverso(campos, AGENDADOS_CANON))


def _adaptador_venta(ruta, hoja, valores):
    _, campos, _ = _detectar_columnas(ruta, hoja, _HEADERS_VENTA)
    return _remapear(valores, _mapa_inverso(campos, VENTA_CANON))

# Campos para las listas de valores únicos (selects del formulario)
AGENDADOS_SELECTS = {'crm': 'B', 'red_social': 'H', 'agendado_por': 'K',
                     'campana': 'O', 'confirmado': 'Q', 'reconfirmado': 'S'}
VENTA_SELECTS = {'nuevo': 'H', 'distrito': 'I', 'sexo': 'K', 'tratamiento': 'L',
                 'doctor': 'M', 'status': 'N', 'pago': 'P', 'campana': 'S'}

# Estilo por defecto por hoja (sólo se usa si la última fila no define el estilo)
DEFAULT_STYLE = {'AGENDADOS': '115', 'VENTA 2026': '137', 'VENTA 2025': '99'}

# ============================================================
# DRIVE
# ============================================================
def _credenciales():
    am.garantizar_credenciales()
    if not os.path.exists(am.CREDENCIALES):
        raise FileNotFoundError(
            f'No hay credenciales en {am.CREDENCIALES}. Configura CREDENCIALES '
            'o GDRIVE_CREDENTIALS_JSON.')
    return service_account.Credentials.from_service_account_file(
        am.CREDENCIALES, scopes=[SCOPE])


def _drive():
    return build('drive', 'v3', credentials=_credenciales())


def _sheets():
    return build('sheets', 'v4', credentials=_credenciales())


_ES_SHEETS_CACHE = {}  # fid -> bool; el tipo de un archivo no cambia en la vida del proceso


def _es_sheets(fid):
    """True si el archivo de Drive es una Google Sheet nativa."""
    if fid in _ES_SHEETS_CACHE:
        return _ES_SHEETS_CACHE[fid]
    meta = _drive().files().get(fileId=fid, fields='mimeType').execute()
    es = meta.get('mimeType') == MIME_SHEETS
    _ES_SHEETS_CACHE[fid] = es
    return es


def subir_drive(fid, ruta):
    """Sube el contenido de ruta a Drive conservando el mismo file id."""
    drv = _drive()
    with open(ruta, 'rb') as fh:
        data = fh.read()
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=MIME_XLSX,
                              resumable=False)
    drv.files().update(fileId=fid, media_body=media).execute()


def _revision_actual(fid):
    """Revisión actual (headRevisionId) del archivo en Drive."""
    meta = _drive().files().get(fileId=fid, fields='headRevisionId').execute()
    return meta.get('headRevisionId')


def _conflicto_post_subida(fid, rev_padre):
    """Post-subida (best effort): True si entre la revisión que leímos y
    nuestra subida se coló otra edición (o nuestra subida no quedó como
    cabeza). La API v3 no ofrece actualización condicional, así que esto
    solo detecta y avisa; no puede evitarlo.
    """
    try:
        res = _drive().files().revisions().list(
            fileId=fid, pageSize=10, fields='revisions(id,modifiedTime)',
            orderBy='modifiedTime desc').execute()
        for i, r in enumerate(res.get('revisions', [])):
            if r['id'] == rev_padre:
                return i != 1
        return False
    except Exception:
        return False


def _ruta_cache(nombre):
    return os.path.join(am.TMP_DIR, f'{nombre}.xlsx')


def descargar(nombre, fid, forzar=False):
    """Descarga de Drive a TMP_DIR con caché de CACHE_TTL segundos.

    Si el archivo es un Google Sheets (Docs Editors), lo exporta a .xlsx.
    """
    ruta = _ruta_cache(nombre)
    if (not forzar and os.path.exists(ruta)
            and time.time() - os.path.getmtime(ruta) < CACHE_TTL):
        return ruta
    drv = _drive()
    meta = drv.files().get(fileId=fid, fields='mimeType').execute()
    if meta.get('mimeType') == MIME_XLSX:
        req = drv.files().get_media(fileId=fid)
    else:
        req = drv.files().export(fileId=fid, mimeType=MIME_XLSX)
    with open(ruta, 'wb') as fh:
        dl = MediaIoBaseDownload(fh, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
    return ruta


def invalidar(nombre):
    ruta = _ruta_cache(nombre)
    if os.path.exists(ruta):
        os.remove(ruta)


def fecha_descarga(nombre):
    ruta = _ruta_cache(nombre)
    if os.path.exists(ruta):
        return datetime.fromtimestamp(os.path.getmtime(ruta)).strftime('%d/%m/%Y %H:%M')
    return None


# ============================================================
# CIRUGÍA XML PARA AGREGAR FILAS
# ============================================================
def _mapa_hojas(ruta):
    """{nombre de hoja: ruta dentro del zip (xl/worksheets/sheetN.xml)}."""
    with zipfile.ZipFile(ruta) as z:
        wb = z.read('xl/workbook.xml').decode('utf-8')
        rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    rid_target = {}
    for rel in re.finditer(r'<Relationship\b[^>]*/?>', rels):
        tag = rel.group(0)
        rid = re.search(r'\bId="([^"]+)"', tag)
        target = re.search(r'\bTarget="([^"]+)"', tag)
        if not rid or not target:
            continue
        ruta_rel = target.group(1).lstrip('/')
        if not ruta_rel.startswith('xl/'):
            ruta_rel = 'xl/' + ruta_rel
        if ruta_rel.startswith('xl/worksheets/'):
            rid_target[rid.group(1)] = ruta_rel
    out = {}
    for m in re.finditer(r'<sheet\b[^>]*>', wb):
        tag = m.group(0)
        nm = re.search(r'name="([^"]+)"', tag)
        rid = re.search(r'r:id="(rId\d+)"', tag)
        if nm and rid and rid.group(1) in rid_target:
            out[nm.group(1)] = rid_target[rid.group(1)]
    return out


def _ultima_fila_datos(x):
    """Última fila (número) que contiene al menos una celda con valor."""
    ultima = 0
    for m in re.finditer(r'<row r="(\d+)"[^>]*>(?:(?!</row>).)*</row>', x, re.S):
        if re.search(r'<v>|<is>', m.group(0)):
            ultima = max(ultima, int(m.group(1)))
    return ultima


def _recortar_trailing(x, ultima):
    """Elimina las filas vacías de formato que quedan tras la última fila con datos."""
    def _keep(m):
        r = int(m.group(1))
        if r <= ultima or re.search(r'<v>|<is>', m.group(0)):
            return m.group(0)
        return ''
    return re.sub(r'<row r="(\d+)"[^>]*>(?:(?!</row>).)*</row>', _keep, x)


def _estilos_fila(x, fila):
    """Estilos (s) de cada columna de una fila."""
    m = re.search(r'<row r="%d"[^>]*>(?:(?!</row>).)*</row>' % fila, x, re.S)
    if not m:
        return {}
    estilos = {}
    for c in re.finditer(r'<c r="([A-Z]+)%d"([^>]*)>' % fila, m.group(0)):
        s = re.search(r's="(\d+)"', c.group(2))
        if s:
            estilos[c.group(1)] = s.group(1)
    return estilos


def _celda_xml(col, fila, valor, estilo):
    ref = f'{col}{fila}'
    s = f' s="{estilo}"' if estilo else ''
    if isinstance(valor, bool):
        return ''
    if isinstance(valor, (int, float)):
        v = int(valor) if float(valor) == int(valor) else valor
        return f'<c r="{ref}"{s}><v>{v}</v></c>'
    t = am.esc_xml(valor)
    return f'<c r="{ref}"{s} t="inlineStr"><is><t>{t}</t></is></c>'


def agregar_fila_xlsx(origen, destino, hoja, fila_num, valores, estilo_default):
    """Agrega una fila nueva (valores {col: valor}) a la hoja indicada."""
    build_dir = os.path.join(am.TMP_DIR, 'build_crm')
    if os.path.exists(build_dir):
        shutil.rmtree(build_dir)
    os.makedirs(build_dir)
    with zipfile.ZipFile(origen) as z:
        z.extractall(build_dir)

    hoja_file = _mapa_hojas(origen).get(hoja)
    if not hoja_file:
        raise ValueError(f'Hoja "{hoja}" no encontrada en el archivo')
    sheet_path = os.path.join(build_dir, hoja_file)
    with open(sheet_path, encoding='utf-8') as f:
        x = f.read()

    ultima = _ultima_fila_datos(x)
    fila_num = max(fila_num, ultima + 1)
    # quitar filas de formato vacías finales para que la fila nueva quede contigua
    x = _recortar_trailing(x, ultima)

    estilos = _estilos_fila(x, ultima)
    celdas = []
    for col in sorted(valores, key=lambda c: openpyxl.utils.column_index_from_string(c)):
        v = valores[col]
        if v is None or v == '':
            continue
        celdas.append(_celda_xml(col, fila_num, v, estilos.get(col, estilo_default)))
    fila_xml = f'<row r="{fila_num}" ht="14.25" customHeight="1">' + ''.join(celdas) + '</row>'
    x = x.replace('</sheetData>', fila_xml + '</sheetData>', 1)

    # extender los autoFilter si la fila nueva cae fuera de su rango
    def _extender(m):
        if int(m.group(4)) >= fila_num:
            return m.group(0)
        return f'{m.group(1)}{m.group(2)}{m.group(3)}{fila_num}{m.group(5)}'

    x = re.sub(r'(<autoFilter ref="\$[A-Z]+\$\d+:\$)([A-Z]+)(\$)(\d+)(")',
               _extender, x)

    with open(sheet_path, 'w', encoding='utf-8') as f:
        f.write(x)

    if os.path.exists(destino):
        os.remove(destino)
    with zipfile.ZipFile(destino, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(build_dir):
            for fn in files:
                full = os.path.join(root, fn)
                z.write(full, os.path.relpath(full, build_dir))
    shutil.rmtree(build_dir)


def agregar_filas_xlsx(origen, destino, hoja, fila_inicial, lista_valores, estilo_default):
    """Agrega varias filas nuevas consecutivas ({col: valor} por fila)."""
    if not lista_valores:
        raise ValueError('No hay filas para agregar')
    build_dir = os.path.join(am.TMP_DIR, 'build_crm_multi')
    if os.path.exists(build_dir):
        shutil.rmtree(build_dir)
    os.makedirs(build_dir)
    with zipfile.ZipFile(origen) as z:
        z.extractall(build_dir)

    hoja_file = _mapa_hojas(origen).get(hoja)
    if not hoja_file:
        raise ValueError(f'Hoja "{hoja}" no encontrada en el archivo')
    sheet_path = os.path.join(build_dir, hoja_file)
    with open(sheet_path, encoding='utf-8') as f:
        x = f.read()

    ultima = _ultima_fila_datos(x)
    fila_inicial = max(fila_inicial, ultima + 1)
    x = _recortar_trailing(x, ultima)

    estilos = _estilos_fila(x, ultima)
    filas_xml = []
    for i, valores in enumerate(lista_valores):
        fila_num = fila_inicial + i
        celdas = []
        for col in sorted(valores, key=lambda c: openpyxl.utils.column_index_from_string(c)):
            v = valores[col]
            if v is None or v == '':
                continue
            celdas.append(_celda_xml(col, fila_num, v, estilos.get(col, estilo_default)))
        filas_xml.append(f'<row r="{fila_num}" ht="14.25" customHeight="1">'
                         + ''.join(celdas) + '</row>')
    fila_fin = fila_inicial + len(lista_valores) - 1
    x = x.replace('</sheetData>', ''.join(filas_xml) + '</sheetData>', 1)

    def _extender(m):
        if int(m.group(4)) >= fila_fin:
            return m.group(0)
        return f'{m.group(1)}{m.group(2)}{m.group(3)}{fila_fin}{m.group(5)}'

    x = re.sub(r'(<autoFilter ref="\$[A-Z]+\$\d+:\$)([A-Z]+)(\$)(\d+)(")',
               _extender, x)

    with open(sheet_path, 'w', encoding='utf-8') as f:
        f.write(x)

    if os.path.exists(destino):
        os.remove(destino)
    with zipfile.ZipFile(destino, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(build_dir):
            for fn in files:
                full = os.path.join(root, fn)
                z.write(full, os.path.relpath(full, build_dir))
    shutil.rmtree(build_dir)


def vaciar_fila_xlsx(origen, destino, hoja, fila_num):
    """Vacía una fila de datos sin alterar el resto del libro.

    Se conserva la estructura, los estilos, fórmulas y tablas dinámicas del
    archivo; la fila queda en blanco y por eso deja de aparecer en el CRM.
    """
    build_dir = os.path.join(am.TMP_DIR, 'build_crm_delete')
    if os.path.exists(build_dir):
        shutil.rmtree(build_dir)
    os.makedirs(build_dir)
    with zipfile.ZipFile(origen) as z:
        z.extractall(build_dir)

    hoja_file = _mapa_hojas(origen).get(hoja)
    if not hoja_file:
        raise ValueError(f'Hoja "{hoja}" no encontrada en el archivo')
    sheet_path = os.path.join(build_dir, hoja_file)
    with open(sheet_path, encoding='utf-8') as f:
        x = f.read()

    patron = r'(<row r="%d"[^>]*>)(?:(?!</row>).)*</row>' % fila_num
    x_nuevo, cambios = re.subn(patron, r'\1</row>', x, count=1, flags=re.S)
    if not cambios:
        raise ValueError(f'Fila {fila_num} no encontrada en "{hoja}"')
    with open(sheet_path, 'w', encoding='utf-8') as f:
        f.write(x_nuevo)

    if os.path.exists(destino):
        os.remove(destino)
    with zipfile.ZipFile(destino, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(build_dir):
            for fn in files:
                full = os.path.join(root, fn)
                z.write(full, os.path.relpath(full, build_dir))
    shutil.rmtree(build_dir)


def reescribir_fila_xlsx(origen, destino, hoja, fila_num, valores, estilo_default):
    """Reemplaza el contenido de una fila existente (valores {col: valor}).

    Se conserva el tag de apertura de la fila (atributos, altura, estilos) y
    sólo se cambian las celdas; las celdas con valor None/'' quedan en blanco.
    """
    build_dir = os.path.join(am.TMP_DIR, 'build_crm_edit')
    if os.path.exists(build_dir):
        shutil.rmtree(build_dir)
    os.makedirs(build_dir)
    with zipfile.ZipFile(origen) as z:
        z.extractall(build_dir)

    hoja_file = _mapa_hojas(origen).get(hoja)
    if not hoja_file:
        raise ValueError(f'Hoja "{hoja}" no encontrada en el archivo')
    sheet_path = os.path.join(build_dir, hoja_file)
    with open(sheet_path, encoding='utf-8') as f:
        x = f.read()

    estilos = _estilos_fila(x, fila_num)
    celdas = []
    for col in sorted(valores, key=lambda c: openpyxl.utils.column_index_from_string(c)):
        v = valores[col]
        if v is None or v == '':
            continue
        celdas.append(_celda_xml(col, fila_num, v, estilos.get(col, estilo_default)))

    def _reemplazar(m):
        return m.group(1) + ''.join(celdas) + '</row>'

    patron = r'(<row r="%d"[^>]*>)(?:(?!</row>).)*</row>' % fila_num
    x, cambios = re.subn(patron, _reemplazar, x, count=1, flags=re.S)
    if not cambios:
        raise ValueError(f'Fila {fila_num} no encontrada en "{hoja}"')

    with open(sheet_path, 'w', encoding='utf-8') as f:
        f.write(x)

    if os.path.exists(destino):
        os.remove(destino)
    with zipfile.ZipFile(destino, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(build_dir):
            for fn in files:
                full = os.path.join(root, fn)
                z.write(full, os.path.relpath(full, build_dir))
    shutil.rmtree(build_dir)


def _siguiente_fila(ruta, hoja):
    hoja_file = _mapa_hojas(ruta).get(hoja)
    if not hoja_file:
        raise ValueError(f'Hoja "{hoja}" no encontrada')
    with zipfile.ZipFile(ruta) as z:
        x = z.read(hoja_file).decode('utf-8')
    return _ultima_fila_datos(x) + 1


# ============================================================
# NORMALIZACIÓN
# ============================================================
def _mes(m):
    s = str(m or '').strip().upper()
    if not s:
        return ''
    return 'SET' if s == 'SEP' else s


def _tel(t):
    if t is None:
        return None
    d = re.sub(r'\D', '', str(t))
    return int(d) if d else None


def _int(v):
    n = am.num(v)
    return int(n) if n is not None else None


def _txt(v):
    s = str(v or '').strip()
    return s if s else None


def normalizar_agendado(d):
    out = {}
    if _txt(d.get('crm')): out['B'] = _txt(d['crm'])
    if _int(d.get('dia')): out['C'] = _int(d['dia'])
    if _mes(d.get('mes')): out['D'] = _mes(d['mes'])
    if _int(d.get('anio')): out['E'] = _int(d['anio'])
    if _txt(d.get('nombre')): out['G'] = _txt(d['nombre'])
    if _txt(d.get('red_social')): out['H'] = _txt(d['red_social'])
    tel = _tel(d.get('telefono'))
    if tel is not None: out['I'] = tel
    if _txt(d.get('correo')): out['J'] = _txt(d['correo'])
    if _txt(d.get('agendado_por')): out['K'] = _txt(d['agendado_por'])
    if _int(d.get('dia_cita')): out['L'] = _int(d['dia_cita'])
    if _mes(d.get('mes_cita')): out['M'] = _mes(d['mes_cita'])
    if _int(d.get('anio_cita')): out['N'] = _int(d['anio_cita'])
    if _txt(d.get('campana')): out['O'] = _txt(d['campana'])
    if _txt(d.get('hora')): out['P'] = _txt(d['hora'])
    if _txt(d.get('confirmado')): out['Q'] = _txt(d['confirmado'])
    if _txt(d.get('observacion')): out['R'] = _txt(d['observacion'])
    if _txt(d.get('reconfirmado')): out['S'] = _txt(d['reconfirmado'])
    if _txt(d.get('observacion2')): out['T'] = _txt(d['observacion2'])
    return out


def normalizar_agendado_edicion(d):
    """Todos los campos editables (letras canónicas) con valor o None para limpiar."""
    return {
        'B': _txt(d.get('crm')),
        'C': _int(d.get('dia')),
        'D': _mes(d.get('mes')),
        'E': _int(d.get('anio')),
        'G': _txt(d.get('nombre')),
        'H': _txt(d.get('red_social')),
        'I': _tel(d.get('telefono')),
        'J': _txt(d.get('correo')),
        'K': _txt(d.get('agendado_por')),
        'L': _int(d.get('dia_cita')),
        'M': _mes(d.get('mes_cita')),
        'N': _int(d.get('anio_cita')),
        'O': _txt(d.get('campana')),
        'P': _txt(d.get('hora')),
        'Q': _txt(d.get('confirmado')),
        'R': _txt(d.get('observacion')),
        'S': _txt(d.get('reconfirmado')),
        'T': _txt(d.get('observacion2')),
    }


def normalizar_venta(d, hoja='VENTA 2026'):
    out = {}
    if _int(d.get('dia')): out['B'] = _int(d['dia'])
    if _mes(d.get('mes')): out['C'] = _mes(d['mes'])
    if _int(d.get('anio')): out['D'] = _int(d['anio'])
    if _int(d.get('dni')): out['E'] = _int(d['dni'])
    tel = _tel(d.get('cel'))
    if tel is not None: out['F'] = tel
    if _txt(d.get('nombre')): out['G'] = _txt(d['nombre'])
    if _txt(d.get('nuevo')): out['H'] = _txt(d['nuevo'])
    if _txt(d.get('distrito')): out['I'] = _txt(d['distrito'])
    if _int(d.get('edad')): out['J'] = _int(d['edad'])
    if _txt(d.get('sexo')): out['K'] = _txt(d['sexo'])
    if _txt(d.get('tratamiento')): out['L'] = _txt(d['tratamiento'])
    if _txt(d.get('doctor')): out['M'] = _txt(d['doctor'])
    if _txt(d.get('status')): out['N'] = _txt(d['status'])
    if am.num(d.get('venta')) is not None: out['O'] = am.num(d['venta'])
    if _txt(d.get('pago')): out['P'] = _txt(d['pago'])
    if hoja == 'VENTA 2025':
        # la hoja 2025 sólo llega hasta Q (OBSERVACION en Q)
        if _txt(d.get('observacion')): out['Q'] = _txt(d['observacion'])
    else:
        if _int(d.get('comisiona')): out['Q'] = _int(d['comisiona'])
        if _txt(d.get('observacion')): out['R'] = _txt(d['observacion'])
        if _txt(d.get('campana')): out['S'] = _txt(d['campana'])
    return out


def normalizar_venta_edicion(d, hoja='VENTA 2026'):
    """Todos los campos editables (letras canónicas) con valor o None para limpiar."""
    out = {
        'B': _int(d.get('dia')), 'C': _mes(d.get('mes')), 'D': _int(d.get('anio')),
        'E': _int(d.get('dni')), 'F': _tel(d.get('cel')), 'G': _txt(d.get('nombre')),
        'H': _txt(d.get('nuevo')), 'I': _txt(d.get('distrito')), 'J': _int(d.get('edad')),
        'K': _txt(d.get('sexo')), 'L': _txt(d.get('tratamiento')), 'M': _txt(d.get('doctor')),
        'N': _txt(d.get('status')), 'O': am.num(d.get('venta')), 'P': _txt(d.get('pago')),
    }
    if hoja == 'VENTA 2025':
        out['Q'] = _txt(d.get('observacion'))
    else:
        out['Q'] = _int(d.get('comisiona'))
        out['R'] = _txt(d.get('observacion'))
        out['S'] = _txt(d.get('campana'))
    return out


# ============================================================
# DETECCIÓN DE VENTAS DUPLICADAS
# ============================================================
class VentaDuplicada(Exception):
    """Ya hay una venta del mismo paciente en la misma fecha.

    ``existente`` describe la fila ya registrada para que la UI pueda
    mostrarla y dejar que el usuario decida (editarla o registrar igual).
    """

    def __init__(self, existente):
        self.existente = existente
        super().__init__('Ya existe una venta de este paciente en esa fecha')


def _clave_texto(v):
    """Texto comparable: sin acentos, en mayúsculas y con espacios colapsados."""
    s = unicodedata.normalize('NFKD', str(v if v is not None else ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.upper().split())


def buscar_venta_duplicada(ruta, hoja, valores):
    """Busca en ``hoja`` una venta ya registrada del mismo paciente y fecha.

    Compara la fecha (día/mes/año) y considera que es el mismo paciente si
    coincide el nombre normalizado o el teléfono. Sirve tanto para dobles
    envíos del formulario como para filas que alguien escribió a mano en el
    Drive antes de registrarlas por la web.

    Devuelve un dict con los datos de la fila encontrada, o None.
    """
    nombre = _clave_texto(valores.get('G'))
    tel = _tel(valores.get('F'))
    dia, mes, anio = valores.get('B'), valores.get('C'), valores.get('D')
    if not nombre or not dia or not mes:
        return None
    _, campos, _ = _detectar_columnas(ruta, hoja, _HEADERS_VENTA)
    inv = _mapa_inverso(campos, VENTA_CANON)
    if not {'B', 'C', 'G'} <= set(inv):
        return None
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[hoja]

    def celda(fila, cl):
        real = inv.get(cl)
        if not real:
            return None
        return ws.cell(row=fila,
                       column=openpyxl.utils.column_index_from_string(real)).value

    encontrada = None
    for r in range(1, ws.max_row + 1):
        if _int(celda(r, 'B')) != _int(dia):
            continue
        if _clave_texto(celda(r, 'C')) != _clave_texto(mes):
            continue
        if anio and _int(celda(r, 'D')) and _int(celda(r, 'D')) != _int(anio):
            continue
        mismo_nombre = _clave_texto(celda(r, 'G')) == nombre
        mismo_tel = bool(tel) and _tel(celda(r, 'F')) == tel
        if not (mismo_nombre or mismo_tel):
            continue
        encontrada = {
            'fila': r, 'hoja': hoja,
            'nombre': _txt(celda(r, 'G')),
            'dia': _int(celda(r, 'B')), 'mes': _txt(celda(r, 'C')),
            'anio': _int(celda(r, 'D')),
            'tratamiento': _txt(celda(r, 'L')), 'doctor': _txt(celda(r, 'M')),
            'status': _txt(celda(r, 'N')), 'venta': am.num(celda(r, 'O')),
            'coincide_por': 'nombre' if mismo_nombre else 'telefono',
        }
        break
    wb.close()
    return encontrada


# ============================================================
# AGREGAR FILAS (Google Sheet nativa: append atómico | xlsx: cirugía XML)
# ============================================================
def _append_sheets(fid, hoja, valores, ruta):
    """Escribe una fila nueva al final de los datos de una Google Sheet nativa.

    ``ruta`` es la copia exportada a .xlsx; de ella se calcula la fila destino
    (última fila con datos + 1) y se escribe con ``values.update``.

    No se usa ``values.append`` porque con hojas cuya columna A está vacía
    (como el AGENDADOS de Derma Essenza) Google no detecta la tabla e inserta
    la fila en la fila 1, corriendo todo el contenido hacia abajo.

    Devuelve el número de fila escrito.
    """
    srv = _sheets()
    fila = _siguiente_fila(ruta, hoja)
    letras = sorted(valores, key=openpyxl.utils.column_index_from_string)
    ncols = openpyxl.utils.column_index_from_string(letras[-1]) if letras else 1
    fila_row = [valores.get(openpyxl.utils.get_column_letter(c))
                for c in range(1, ncols + 1)]
    fila_row = ['' if v is None else v for v in fila_row]
    srv.spreadsheets().values().update(
        spreadsheetId=fid,
        range=f"'{hoja}'!A{fila}",
        valueInputOption='RAW',
        body={'values': [fila_row]}).execute()
    return fila


def _append_sheets_multi(fid, hoja, lista_valores, ruta):
    """Escribe varias filas nuevas al final de una Google Sheet nativa (un solo update)."""
    srv = _sheets()
    fila = _siguiente_fila(ruta, hoja)
    ncols = 1
    for valores in lista_valores:
        letras = sorted(valores, key=openpyxl.utils.column_index_from_string)
        if letras:
            ncols = max(ncols, openpyxl.utils.column_index_from_string(letras[-1]))
    filas = []
    for valores in lista_valores:
        fila_row = [valores.get(openpyxl.utils.get_column_letter(c))
                    for c in range(1, ncols + 1)]
        fila_row = ['' if v is None else v for v in fila_row]
        filas.append(fila_row)
    ultima_letra = openpyxl.utils.get_column_letter(ncols)
    srv.spreadsheets().values().update(
        spreadsheetId=fid,
        range=f"'{hoja}'!A{fila}:{ultima_letra}{fila + len(filas) - 1}",
        valueInputOption='RAW',
        body={'values': filas}).execute()
    return fila


def _agregar_xlsx_seguro_multi(nombre, fid, hoja, estilo, lista_valores, adaptador=None):
    """Append de varias filas sobre .xlsx de Drive con control de concurrencia."""
    destino = os.path.join(am.TMP_DIR, f'{nombre}_editado.xlsx')
    fila = None
    conflicto = False
    for intento in range(MAX_REINTENTOS):
        rev_antes = _revision_actual(fid)
        ruta = descargar(nombre, fid, forzar=True)
        valores_hoja = ([adaptador(ruta, v) for v in lista_valores]
                        if adaptador else lista_valores)
        fila = _siguiente_fila(ruta, hoja)
        agregar_filas_xlsx(ruta, destino, hoja, fila, valores_hoja, estilo)
        if _revision_actual(fid) != rev_antes:
            if intento < MAX_REINTENTOS - 1:
                continue
            raise RuntimeError(
                'El archivo de Drive cambió durante la escritura; '
                'reintenta la operación')
        subir_drive(fid, destino)
        conflicto = _conflicto_post_subida(fid, rev_antes)
        break
    invalidar(nombre)
    return fila, conflicto


def _agregar_xlsx_seguro(nombre, fid, hoja, estilo, valores, adaptador=None):
    """Append sobre .xlsx de Drive con control de concurrencia best effort.

    La API v3 de Drive no tiene actualización condicional (CAS), así que:
      1. ``_lock`` serializa las escrituras de la propia app en el proceso.
      2. Preflight: si la revisión del archivo cambió mientras descargábamos
         y construíamos la fila (alguien editó directo en Drive), se
         re-descarga y se reintenta con la versión nueva.
      3. Postflight: tras subir, se comprueba con revisions.list si otra
         edición se coló en el ínterin y se reporta en el resultado.

    ``adaptador`` (opcional) recibe (ruta_descargada, valores) y devuelve
    los valores ya re-mapeados a las letras reales de la hoja.

    Devuelve (fila, conflicto).
    """
    destino = os.path.join(am.TMP_DIR, f'{nombre}_editado.xlsx')
    fila = None
    conflicto = False
    for intento in range(MAX_REINTENTOS):
        rev_antes = _revision_actual(fid)
        ruta = descargar(nombre, fid, forzar=True)
        valores_hoja = adaptador(ruta, valores) if adaptador else valores
        fila = _siguiente_fila(ruta, hoja)
        agregar_fila_xlsx(ruta, destino, hoja, fila, valores_hoja, estilo)
        if _revision_actual(fid) != rev_antes:
            if intento < MAX_REINTENTOS - 1:
                continue
            raise RuntimeError(
                'El archivo de Drive cambió durante la escritura; '
                'reintenta la operación')
        subir_drive(fid, destino)
        conflicto = _conflicto_post_subida(fid, rev_antes)
        break
    invalidar(nombre)
    return fila, conflicto


def _borrar_xlsx_seguro(nombre, fid, hoja, fila_num, mapas_header):
    """Vacía una fila de un xlsx de Drive con la misma protección de concurrencia."""
    destino = os.path.join(am.TMP_DIR, f'{nombre}_borrado.xlsx')
    conflicto = False
    for intento in range(MAX_REINTENTOS):
        rev_antes = _revision_actual(fid)
        ruta = descargar(nombre, fid, forzar=True)
        encabezado, _, _ = _detectar_columnas(ruta, hoja, mapas_header)
        if fila_num <= encabezado:
            raise ValueError('No se puede borrar una fila de encabezados')
        vaciar_fila_xlsx(ruta, destino, hoja, fila_num)
        if _revision_actual(fid) != rev_antes:
            if intento < MAX_REINTENTOS - 1:
                continue
            raise RuntimeError(
                'El archivo de Drive cambió durante la eliminación; reintenta la operación')
        subir_drive(fid, destino)
        conflicto = _conflicto_post_subida(fid, rev_antes)
        break
    invalidar(nombre)
    return conflicto


def agregar_agendado(datos):
    valores = normalizar_agendado(datos)
    if not valores.get('G') and 'I' not in valores:
        raise ValueError('Indica al menos el nombre o el teléfono del paciente')
    with _lock:
        fid = am.AGENDADOS_FID
        if _es_sheets(fid):
            ruta = descargar('AGENDADOS', fid, forzar=True)
            valores_hoja = _adaptador_agendados(ruta, valores)
            fila = _append_sheets(fid, 'AGENDADOS', valores_hoja, ruta)
            invalidar('AGENDADOS')
            return {'ok': True, 'fila': fila, 'hoja': 'AGENDADOS',
                    'valores': valores}
        fila, conflicto = _agregar_xlsx_seguro('AGENDADOS', fid, 'AGENDADOS',
                                               DEFAULT_STYLE['AGENDADOS'],
                                               valores,
                                               adaptador=_adaptador_agendados)
    return {'ok': True, 'fila': fila, 'hoja': 'AGENDADOS',
            'valores': valores, 'conflicto': conflicto}


def _adaptador_edicion_agendados(ruta, fila_num, valores):
    """Construye {letra_real: valor} para reescribir la fila editada.

    Parte de los valores actuales de la fila (así se conservan columnas que el
    formulario no edita, p. ej. DNI o ASISTENCIA) y sobreescribe con los campos
    editados; los campos editados a None quedan en blanco.
    """
    _, campos, _ = _detectar_columnas(ruta, 'AGENDADOS', _HEADERS_AGENDADOS)
    inverso = _mapa_inverso(campos, AGENDADOS_CANON)
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb['AGENDADOS']
    actual = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=fila_num, column=c).value
        if v is not None and v != '':
            actual[openpyxl.utils.get_column_letter(c)] = v
    wb.close()
    for cl, v in valores.items():
        real = inverso.get(cl)
        if real:
            actual[real] = v
    return actual


def _adaptador_edicion_venta(ruta, hoja, fila_num, valores):
    """Construye {letra_real: valor} para reescribir una fila editada de VENTA
    DIARIA. Parte de los valores actuales de la fila (así se conservan
    columnas que el formulario no edita) y sobreescribe con los campos
    editados; los campos editados a None quedan en blanco."""
    _, campos, _ = _detectar_columnas(ruta, hoja, _HEADERS_VENTA)
    inverso = _mapa_inverso(campos, VENTA_CANON)
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[hoja]
    actual = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=fila_num, column=c).value
        if v is not None and v != '':
            actual[openpyxl.utils.get_column_letter(c)] = v
    wb.close()
    for cl, v in valores.items():
        real = inverso.get(cl)
        if real:
            actual[real] = v
    return actual


def _fila_actual_sheets(fid, hoja, fila_num, ultima_col='Z'):
    """Lee sólo la fila ``fila_num`` de una Google Sheet nativa vía la API de
    Sheets (values.get), sin exportar ni volver a parsear el archivo entero.

    Los ediciones de una sola fila (confirmar, reprogramar, editar venta...)
    necesitan el contenido actual de la fila para no borrar columnas que no
    tocan, pero no necesitan el archivo completo: exportarlo a .xlsx y
    parsearlo con openpyxl solo para leer una fila costaba ~2-3s por click,
    que era el cuello de botella real detrás de "se siente lento" en cada
    botón de acción."""
    res = _sheets().spreadsheets().values().get(
        spreadsheetId=fid, range=f"'{hoja}'!A{fila_num}:{ultima_col}{fila_num}",
        valueRenderOption='UNFORMATTED_VALUE').execute()
    fila = (res.get('values') or [[]])
    fila = fila[0] if fila else []
    out = {}
    for i, v in enumerate(fila):
        if v in (None, ''):
            continue
        if isinstance(v, float) and v == int(v):
            v = int(v)
        out[openpyxl.utils.get_column_letter(i + 1)] = v
    return out


def editar_venta(hoja, fila_num, datos):
    if hoja not in ('VENTA 2026', 'VENTA 2025'):
        raise ValueError('Hoja de venta no válida')
    if not isinstance(fila_num, int) or fila_num < 1:
        raise ValueError('Fila inválida')
    valores = normalizar_venta_edicion(datos, hoja)
    if not valores.get('G'):
        raise ValueError('Indica el nombre del paciente')
    with _lock:
        fid = am.VENTA_FID
        if _es_sheets(fid):
            ruta = descargar('VENTA_DIARIA', fid, forzar=True)
            valores_hoja = _adaptador_edicion_venta(ruta, hoja, fila_num, valores)
            letras = sorted(valores_hoja,
                            key=openpyxl.utils.column_index_from_string)
            ncols = openpyxl.utils.column_index_from_string(letras[-1]) if letras else 1
            fila_row = [valores_hoja.get(openpyxl.utils.get_column_letter(c))
                        for c in range(1, ncols + 1)]
            fila_row = ['' if v is None else v for v in fila_row]
            _sheets().spreadsheets().values().update(
                spreadsheetId=fid,
                range=f"'{hoja}'!A{fila_num}",
                valueInputOption='RAW',
                body={'values': [fila_row]}).execute()
            invalidar('VENTA_DIARIA')
            return {'ok': True, 'fila': fila_num, 'hoja': hoja, 'valores': valores}
        conflicto = _editar_xlsx_seguro('VENTA_DIARIA', fid, hoja, fila_num, valores,
                                        adaptador=lambda ruta, v:
                                        _adaptador_edicion_venta(ruta, hoja, fila_num, v))
    return {'ok': True, 'fila': fila_num, 'hoja': hoja, 'valores': valores,
            'conflicto': conflicto}


def _editar_xlsx_seguro(nombre, fid, hoja, fila_num, valores, adaptador=None):
    """Reescribe una fila de un xlsx de Drive con control de concurrencia."""
    destino = os.path.join(am.TMP_DIR, f'{nombre}_editado.xlsx')
    conflicto = False
    for intento in range(MAX_REINTENTOS):
        rev_antes = _revision_actual(fid)
        ruta = descargar(nombre, fid, forzar=True)
        valores_hoja = adaptador(ruta, valores) if adaptador else valores
        reescribir_fila_xlsx(ruta, destino, hoja, fila_num, valores_hoja,
                             DEFAULT_STYLE[hoja])
        if _revision_actual(fid) != rev_antes:
            if intento < MAX_REINTENTOS - 1:
                continue
            raise RuntimeError(
                'El archivo de Drive cambió durante la edición; reintenta la operación')
        subir_drive(fid, destino)
        conflicto = _conflicto_post_subida(fid, rev_antes)
        break
    invalidar(nombre)
    return conflicto


def editar_agendado(fila_num, datos):
    if not isinstance(fila_num, int) or fila_num < 1:
        raise ValueError('Fila inválida')
    valores = normalizar_agendado_edicion(datos)
    with _lock:
        fid = am.AGENDADOS_FID
        if _es_sheets(fid):
            ruta = descargar('AGENDADOS', fid, forzar=True)
            valores_hoja = _adaptador_edicion_agendados(ruta, fila_num, valores)
            letras = sorted(valores_hoja,
                            key=openpyxl.utils.column_index_from_string)
            ncols = openpyxl.utils.column_index_from_string(letras[-1]) if letras else 1
            fila_row = [valores_hoja.get(openpyxl.utils.get_column_letter(c))
                        for c in range(1, ncols + 1)]
            fila_row = ['' if v is None else v for v in fila_row]
            _sheets().spreadsheets().values().update(
                spreadsheetId=fid,
                range=f"'AGENDADOS'!A{fila_num}",
                valueInputOption='RAW',
                body={'values': [fila_row]}).execute()
            invalidar('AGENDADOS')
            return {'ok': True, 'fila': fila_num, 'hoja': 'AGENDADOS',
                    'valores': valores}
        conflicto = _editar_xlsx_seguro('AGENDADOS', fid, 'AGENDADOS', fila_num,
                                        valores,
                                        adaptador=lambda ruta, v:
                                        _adaptador_edicion_agendados(ruta, fila_num, v))
    return {'ok': True, 'fila': fila_num, 'hoja': 'AGENDADOS',
            'valores': valores, 'conflicto': conflicto}


def agregar_venta(datos, forzar=False):
    """Agrega una venta. Con ``forzar=False`` aborta si detecta un duplicado.

    Lanza ``VentaDuplicada`` si ya hay una venta del mismo paciente en la
    misma fecha, para que el usuario confirme antes de crear una fila doble.
    """
    anio = am.num(datos.get('anio')) or 2026
    hoja = 'VENTA 2026' if anio >= 2026 else 'VENTA 2025'
    valores = normalizar_venta(datos, hoja)
    if not valores.get('G'):
        raise ValueError('Indica el nombre del paciente')

    def adaptador(ruta, v):
        # Se valida aquí porque ``ruta`` es la copia recién descargada que
        # usará el append: así el chequeo ve el estado real del Drive.
        if not forzar:
            existente = buscar_venta_duplicada(ruta, hoja, v)
            if existente:
                raise VentaDuplicada(existente)
        return _adaptador_venta(ruta, hoja, v)

    with _lock:
        fid = am.VENTA_FID
        if _es_sheets(fid):
            ruta = descargar('VENTA_DIARIA', fid, forzar=True)
            valores_hoja = adaptador(ruta, valores)
            fila = _append_sheets(fid, hoja, valores_hoja, ruta)
            invalidar('VENTA_DIARIA')
            return {'ok': True, 'fila': fila, 'hoja': hoja, 'valores': valores}
        fila, conflicto = _agregar_xlsx_seguro('VENTA_DIARIA', fid, hoja,
                                               DEFAULT_STYLE[hoja], valores,
                                               adaptador=adaptador)
    return {'ok': True, 'fila': fila, 'hoja': hoja, 'valores': valores,
            'conflicto': conflicto}


def actualizar_campos_agendado(fila_num, campos):
    """Actualiza sólo los campos canónicos indicados de una fila de AGENDADOS.

    Conserva los valores actuales de la fila; sólo sobrescribe las letras
    presentes en ``campos`` (p. ej. {'Q': 'CONFIRMADO'} o
    {'L': 15, 'M': 'AGO', 'N': 2026} para reprogramar).
    """
    if not isinstance(fila_num, int) or fila_num < 1:
        raise ValueError('Fila inválida')
    campos = {cl: v for cl, v in (campos or {}).items()
              if v is not None and v != ''}
    if not campos:
        raise ValueError('No hay campos para actualizar')
    with _lock:
        fid = am.AGENDADOS_FID
        if _es_sheets(fid):
            ruta = descargar('AGENDADOS', fid)
            _, campos_hdr, _ = _detectar_columnas(ruta, 'AGENDADOS', _HEADERS_AGENDADOS)
            inverso = _mapa_inverso(campos_hdr, AGENDADOS_CANON)
            actual = _fila_actual_sheets(fid, 'AGENDADOS', fila_num)
            for cl, v in campos.items():
                real = inverso.get(cl)
                if real:
                    actual[real] = v
            letras = sorted(actual, key=openpyxl.utils.column_index_from_string)
            ncols = openpyxl.utils.column_index_from_string(letras[-1]) if letras else 1
            fila_row = [actual.get(openpyxl.utils.get_column_letter(c))
                        for c in range(1, ncols + 1)]
            fila_row = ['' if v is None else v for v in fila_row]
            _sheets().spreadsheets().values().update(
                spreadsheetId=fid,
                range=f"'AGENDADOS'!A{fila_num}",
                valueInputOption='RAW',
                body={'values': [fila_row]}).execute()
            invalidar('AGENDADOS')
            return {'ok': True, 'fila': fila_num, 'hoja': 'AGENDADOS',
                    'campos': campos}
        conflicto = _editar_xlsx_seguro('AGENDADOS', fid, 'AGENDADOS', fila_num,
                                        campos,
                                        adaptador=lambda ruta, v:
                                        _adaptador_edicion_agendados(ruta, fila_num, v))
    return {'ok': True, 'fila': fila_num, 'hoja': 'AGENDADOS',
            'campos': campos, 'conflicto': conflicto}


def reprogramar_agendado(fila_num, dia, mes, anio):
    """Cambia la fecha de la cita y deja huella del cambio anterior en
    OBSERVACION2 (columna T, sin uso hasta ahora), como entradas
    "D-MES-A>D-MES-A" separadas por ';'. Así el calendario puede mostrar
    que una cita se movió en vez de simplemente desaparecer del día
    original, y se puede ver cuántas veces se reagendó un paciente."""
    actual = None
    for f in leer_agendados()['filas']:
        if f.get('_fila') == fila_num:
            actual = f
            break
    campos = {'L': dia, 'M': mes, 'N': anio}
    if actual and actual.get('L') and actual.get('M') and actual.get('N'):
        entrada = f"{actual['L']}-{actual['M']}-{actual['N']}>{dia}-{mes}-{anio}"
        previo = [p for p in str(actual.get('T') or '').split(';') if p.strip()]
        campos['T'] = ';'.join((previo + [entrada])[-12:])
    return actualizar_campos_agendado(fila_num, campos)


def agregar_ventas_multi(datos, lineas):
    """Registra una venta con varios tratamientos: una fila por tratamiento.

    ``datos`` son los campos compartidos de la venta (dia/mes/anio, nombre,
    cel, doctor, status, pago, ...) y ``lineas`` una lista de
    {'tratamiento': ..., 'venta': ...}. Devuelve la fila inicial escrita.
    """
    lineas = [ln for ln in (lineas or []) if ln.get('tratamiento')]
    if not lineas:
        raise ValueError('Indica al menos un tratamiento')
    anio = am.num(datos.get('anio')) or 2026
    hoja = 'VENTA 2026' if anio >= 2026 else 'VENTA 2025'
    lista = []
    for ln in lineas:
        dd = dict(datos)
        dd['tratamiento'] = ln.get('tratamiento')
        dd['venta'] = ln.get('venta')
        v = normalizar_venta(dd, hoja)
        if not v.get('G'):
            raise ValueError('Indica el nombre del paciente')
        if not v.get('L'):
            raise ValueError('Indica el tratamiento')
        lista.append(v)
    if not lista:
        raise ValueError('No hay líneas de tratamiento válidas')

    def adaptador(ruta, v):
        return _adaptador_venta(ruta, hoja, v)

    with _lock:
        fid = am.VENTA_FID
        if _es_sheets(fid):
            ruta = descargar('VENTA_DIARIA', fid, forzar=True)
            valores_hoja = [adaptador(ruta, v) for v in lista]
            fila = _append_sheets_multi(fid, hoja, valores_hoja, ruta)
            invalidar('VENTA_DIARIA')
            return {'ok': True, 'fila': fila, 'hoja': hoja, 'n': len(lista)}
        fila, conflicto = _agregar_xlsx_seguro_multi('VENTA_DIARIA', fid, hoja,
                                                     DEFAULT_STYLE[hoja], lista,
                                                     adaptador=adaptador)
    return {'ok': True, 'fila': fila, 'hoja': hoja, 'n': len(lista),
            'conflicto': conflicto}


def _borrar_fila(fid, nombre, hoja, fila_num, mapas_header):
    if not isinstance(fila_num, int) or fila_num < 1:
        raise ValueError('Fila inválida')
    with _lock:
        if _es_sheets(fid):
            _sheets().spreadsheets().values().clear(
                spreadsheetId=fid,
                range=f"'{hoja}'!A{fila_num}:ZZ{fila_num}",
                body={}).execute()
            invalidar(nombre)
            return {'ok': True, 'conflicto': False}
        conflicto = _borrar_xlsx_seguro(nombre, fid, hoja, fila_num, mapas_header)
    return {'ok': True, 'conflicto': conflicto}


def borrar_agendado(fila_num):
    return _borrar_fila(am.AGENDADOS_FID, 'AGENDADOS', 'AGENDADOS', fila_num,
                         _HEADERS_AGENDADOS)


def borrar_venta(hoja, fila_num):
    if hoja not in ('VENTA 2026', 'VENTA 2025'):
        raise ValueError('Hoja de venta no válida')
    return _borrar_fila(am.VENTA_FID, 'VENTA_DIARIA', hoja, fila_num,
                         _HEADERS_VENTA)


# ============================================================
# LECTURA PARA MOSTRAR
# ============================================================
def leer_agendados():
    """Devuelve {'filas': [{letra_canónica: valor}...], 'total': n,
    'columnas': {letra_canónica: encabezado}}."""
    ruta = descargar('AGENDADOS', am.AGENDADOS_FID)
    n_fila, campos, textos = _detectar_columnas(ruta, 'AGENDADOS',
                                                _HEADERS_AGENDADOS)
    inverso = _mapa_inverso(campos, AGENDADOS_CANON)
    if not inverso:
        n_fila = 4
        inverso = {c: c for c in AGENDADOS_COLS}
        textos = AGENDADOS_COLS
    columnas = {}
    for cl, real in sorted(inverso.items(),
                           key=lambda kv: openpyxl.utils.column_index_from_string(kv[1])):
        columnas[cl] = textos[real]
    # Letra -> índice numérico calculado una sola vez: recalcularlo por cada
    # celda (antes, dentro del doble for) era el costo dominante de esta
    # lectura en hojas con cientos de filas.
    inverso_idx = {cl: openpyxl.utils.column_index_from_string(real)
                   for cl, real in inverso.items()}
    wb = _cargar_wb(ruta)
    ws = wb['AGENDADOS']
    filas = []
    for r in range(n_fila + 1, ws.max_row + 1):
        fila = {}
        for cl, idx in inverso_idx.items():
            v = ws.cell(row=r, column=idx).value
            if v is None:
                continue
            if isinstance(v, float) and v == int(v):
                v = int(v)
            fila[cl] = v
        if any(v is not None and v != '' for v in fila.values()):
            fila['_fila'] = r
            filas.append(fila)
    return {'filas': filas, 'total': len(filas),
            'descargado': fecha_descarga('AGENDADOS'), 'columnas': columnas}


def leer_venta():
    """Devuelve {'hojas': {hoja: {'filas': [...], 'columnas': {...}}}}."""
    ruta = descargar('VENTA_DIARIA', am.VENTA_FID)
    wb = _cargar_wb(ruta)
    out = {}
    for hoja in ('VENTA 2026', 'VENTA 2025'):
        if hoja not in wb.sheetnames:
            continue
        n_fila, campos, textos = _detectar_columnas(ruta, hoja, _HEADERS_VENTA)
        inverso = _mapa_inverso(campos, VENTA_CANON)
        if not inverso:
            n_fila = 5 if hoja == 'VENTA 2026' else 8
            inverso = {c: c for c in VENTA_COLS}
            textos = VENTA_COLS
        columnas = {}
        for cl, real in sorted(inverso.items(),
                               key=lambda kv: openpyxl.utils.column_index_from_string(kv[1])):
            columnas[cl] = textos[real]
        inverso_idx = {cl: openpyxl.utils.column_index_from_string(real)
                       for cl, real in inverso.items()}
        ws = wb[hoja]
        filas = []
        for r in range(n_fila + 1, ws.max_row + 1):
            fila = {}
            for cl, idx in inverso_idx.items():
                v = ws.cell(row=r, column=idx).value
                if v is None:
                    continue
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                fila[cl] = v
            if any(v is not None and v != '' for v in fila.values()):
                fila['_fila'] = r
                filas.append(fila)
        out[hoja] = {'filas': filas, 'columnas': columnas}
    return {'hojas': out, 'descargado': fecha_descarga('VENTA_DIARIA')}


def _unicos(filas, col):
    s = set()
    for f in filas:
        v = f.get(col)
        if v is not None and str(v).strip():
            s.add(str(v).strip())
    return sorted(s)[:_SELECTO_LIMIT]


def valores_unicos_agendados(filas):
    return {k: _unicos(filas, col) for k, col in AGENDADOS_SELECTS.items()}


def valores_unicos_venta(filas):
    return {k: _unicos(filas, col) for k, col in VENTA_SELECTS.items()}
