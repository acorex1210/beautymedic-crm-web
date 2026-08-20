# -*- coding: utf-8 -*-
"""
alimentar_maestro.py
====================
Sincroniza el archivo maestro 'BD DATA.xlsx' desde dos fuentes de Google Drive:

  1) AGENDADOS  -> copia columnas B..O (nuevas citas agendadas)
  2) VENTA DIARIA -> consulta consultiva: completa ASISTENCIA, DISTRITO, EDAD,
     SEXO, DNI y los TRAT/PAGO 1..4 + PAGO TOTAL de quien asistió.

La edición del maestro se hace por cirugía XML (se preservan tablas dinámicas,
slicers, fórmulas y formato). No se usa openpyxl para guardar.

Uso:
  python3 alimentar_maestro.py            # modo revisión (no toca el maestro)
  python3 alimentar_maestro.py --aplicar  # aplica cambios al maestro (con backup)
  python3 alimentar_maestro.py --sin-descarga   # reutiliza archivos ya descargados

Configuración vía variables de entorno (útiles para despliegue web):
  CREDENCIALES           ruta al JSON de la cuenta de servicio (default ~/credenciales-bm.json)
  GDRIVE_CREDENTIALS_JSON contenido del JSON (JSON crudo o base64; se escribe en CREDENCIALES)
  GOOGLE_APPLICATION_CREDENTIALS ruta a un archivo JSON, contenido crudo o base64 (alternativa)
   AGENDADOS_FID          id de archivo AGENDADOS en Drive
   VENTA_FID              id de archivo VENTA DIARIA en Drive
   MAESTRO_PATH           ruta local al maestro BD DATA.xlsx (sin MAESTRO_FID)
   MAESTRO_FID            id del maestro BD DATA.xlsx en Drive (lo hace portable)
   TMP_DIR                directorio temporal para descargas y simulaciones
"""
import argparse
import base64
import io
import json
import os
import re
import shutil
import sys
import zipfile
from collections import defaultdict
from datetime import datetime

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# ============================================================
# CONFIGURACIÓN
# ============================================================
_TMP_DEFAULT = os.path.expanduser(
    '/var/folders/fc/2jp7n0610cbfckv3jpbzt73h0000gn/T/opencode/maestro_auto')
CREDENCIALES = os.environ.get('CREDENCIALES',
                              os.path.expanduser('~/credenciales-derma-essenza.json'))
GDRIVE_CREDENTIALS_JSON = os.environ.get('GDRIVE_CREDENTIALS_JSON', '')
AGENDADOS_FID = os.environ.get('AGENDADOS_FID', '1So_1Fh744c3K9kss2oA1twjBLJpgrSxZCu2lqhWpqJM')
VENTA_FID = os.environ.get('VENTA_FID', '1TDM7ZFV6Jdsqc6i4CadNkwPQNdrIBhu7')
MAESTRO = os.environ.get('MAESTRO_PATH',
                         os.path.expanduser('~/Downloads/BD DATA.xlsx'))
MAESTRO_FID = os.environ.get('MAESTRO_FID', '').strip()
MIME_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
SCOPE_RW = 'https://www.googleapis.com/auth/drive'
TMP_DIR = os.environ.get('TMP_DIR', _TMP_DEFAULT)
os.makedirs(TMP_DIR, exist_ok=True)


def _normalizar_json_credenciales(v):
    """Devuelve el JSON crudo de la cuenta de servicio, aceptando JSON o base64."""
    if v is None:
        return ''
    v = str(v).strip()
    if not v:
        return ''
    try:
        json.loads(v)
        return v
    except (ValueError, TypeError):
        pass
    try:
        dec = base64.b64decode(v, validate=True).decode('utf-8')
        json.loads(dec)
        return dec
    except (ValueError, TypeError, Exception):  # noqa: BLE001
        return v


def _contenido_credenciales():
    """Busca el JSON de credenciales en el orden: GDRIVE_CREDENTIALS_JSON,
    GOOGLE_APPLICATION_CREDENTIALS (ruta o contenido), CREDENCIALES (archivo)."""
    for var in ('GDRIVE_CREDENTIALS_JSON', 'GOOGLE_APPLICATION_CREDENTIALS'):
        v = os.environ.get(var, '')
        if not v:
            continue
        if os.path.isfile(v):
            try:
                with open(v, encoding='utf-8') as f:
                    return f.read()
            except OSError:
                continue
        return _normalizar_json_credenciales(v)
    if os.path.exists(CREDENCIALES):
        try:
            with open(CREDENCIALES, encoding='utf-8') as f:
                return f.read()
        except OSError:
            return ''
    return ''


def garantizar_credenciales():
    """Escribe las credenciales en CREDENCIALES si no existen y hay fuente."""
    if not os.path.exists(CREDENCIALES):
        contenido = _contenido_credenciales()
        if contenido:
            os.makedirs(os.path.dirname(CREDENCIALES) or '.', exist_ok=True)
            with open(CREDENCIALES, 'w', encoding='utf-8') as f:
                f.write(contenido)


def credenciales_disponibles():
    return os.path.exists(CREDENCIALES) or bool(_contenido_credenciales())

MESES = {'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN',
         'JUL', 'AGO', 'SET', 'SEP', 'OCT', 'NOV', 'DIC'}

# STATUS de VENTA DIARIA que indican que el tratamiento SÍ se realizó
ASISTE_POR_TEXTO = ('SE REALIZO', 'COMPRO', 'COMPLETA', 'SESION', 'DEJO PAGADO')
REVISAR_STATUS = ()

# ============================================================
# UTILIDADES DE NORMALIZACIÓN
# ============================================================
_ACENTOS = {'Á': 'A', 'À': 'A', 'Ä': 'A', 'É': 'E', 'È': 'E', 'Ë': 'E',
            'Í': 'I', 'Ì': 'I', 'Ï': 'I', 'Ó': 'O', 'Ò': 'O', 'Ö': 'O',
            'Ú': 'U', 'Ù': 'U', 'Ü': 'U', 'Ñ': 'N'}


def norm_name(n):
    if n is None:
        return ''
    s = str(n).upper().strip()
    s = ''.join(_ACENTOS.get(ch, ch) for ch in s)
    return re.sub(r'\s+', ' ', s)


def norm_phone(p):
    if p is None or p == '':
        return ''
    if isinstance(p, float) and p.is_integer():
        p = int(p)
    d = re.sub(r'\D', '', str(p))
    return d[-9:] if len(d) >= 9 else d


def norm_fecha(dia, mes, anio):
    try:
        d = int(dia)
    except (TypeError, ValueError):
        d = None
    m = str(mes).strip().upper() if mes is not None else None
    if m in ('SEP',):
        m = 'SET'
    try:
        a = int(anio)
    except (TypeError, ValueError):
        a = None
    return (d, m, a)


def num(x):
    if x is None:
        return None
    try:
        v = float(x)
        return int(v) if v == int(v) else v
    except (TypeError, ValueError):
        return None


def txt(x):
    if x is None:
        return None
    s = str(x).strip()
    return s if s else None


# ============================================================
# DESCARGA DESDE DRIVE
# ============================================================
def descargar(fid, nombre, forzar=False):
    """Descarga de Drive a TMP_DIR (o reutiliza la copia local si existe).

    Si el archivo es un Google Sheets (Docs Editors), lo exporta a .xlsx.
    """
    ruta = os.path.join(TMP_DIR, f'{nombre}.xlsx')
    if os.path.exists(ruta) and not forzar:
        return ruta
    garantizar_credenciales()
    if not os.path.exists(CREDENCIALES):
        raise FileNotFoundError(
            f'No hay credenciales en {CREDENCIALES}. Configura CREDENCIALES o '
            'GDRIVE_CREDENTIALS_JSON.')
    creds = service_account.Credentials.from_service_account_file(
        CREDENCIALES, scopes=['https://www.googleapis.com/auth/drive.readonly'])
    drive = build('drive', 'v3', credentials=creds)
    meta = drive.files().get(fileId=fid, fields='mimeType').execute()
    if meta.get('mimeType') == MIME_XLSX:
        req = drive.files().get_media(fileId=fid)
    else:
        req = drive.files().export(fileId=fid, mimeType=MIME_XLSX)
    with open(ruta, 'wb') as fh:
        dl = MediaIoBaseDownload(fh, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
    print(f'  Descargado {nombre} -> {os.path.basename(ruta)}')
    return ruta


def _drive_rw():
    """Cliente Drive con scope de escritura (subidas)."""
    garantizar_credenciales()
    if not os.path.exists(CREDENCIALES):
        raise FileNotFoundError(
            f'No hay credenciales en {CREDENCIALES}. Configura CREDENCIALES o '
            'GDRIVE_CREDENTIALS_JSON.')
    creds = service_account.Credentials.from_service_account_file(
        CREDENCIALES, scopes=[SCOPE_RW])
    return build('drive', 'v3', credentials=creds)


def subir_archivo(fid, ruta):
    """Sube el contenido de ruta a Drive conservando el mismo file id."""
    drv = _drive_rw()
    with open(ruta, 'rb') as fh:
        data = fh.read()
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=MIME_XLSX,
                              resumable=False)
    drv.files().update(fileId=fid, media_body=media).execute()


# ============================================================
# MAESTRO PORTABLE (Drive o archivo local)
# ============================================================
def ruta_maestro_local(forzar=False):
    """Devuelve una ruta local válida al maestro BD DATA.xlsx.

    Si MAESTRO_FID está configurado, descarga de Drive a TMP_DIR
    (fuente de verdad en Drive; compatible con hosts sin disco
    persistente como Cloud Run). Si no, usa MAESTRO_PATH como siempre.
    """
    if MAESTRO_FID:
        return descargar(MAESTRO_FID, 'BD DATA', forzar=forzar)
    return MAESTRO


def subir_maestro(ruta):
    """Persiste el maestro: sube a Drive si hay MAESTRO_FID, si no lo copia
    localmente a MAESTRO_PATH."""
    if MAESTRO_FID:
        subir_archivo(MAESTRO_FID, ruta)
    else:
        if os.path.abspath(ruta) == os.path.abspath(MAESTRO):
            return
        os.makedirs(os.path.dirname(MAESTRO) or '.', exist_ok=True)
        shutil.copy2(ruta, MAESTRO)


# ============================================================
# LECTURA DE FUENTES (solo lectura)
# ============================================================
_SEM_AG = {
    'DIA': 'DIA', 'MES': 'MES', 'AÑO': 'ANIO', 'ANIO': 'ANIO',
    'DNI': 'DNI', 'NOMBRE': 'NOMBRE', 'RED SOCIAL': 'RED_SOCIAL',
    'TELEFONO': 'TELEFONO', 'CORREO': 'CORREO', 'ASISTENCIA': 'ASISTENCIA',
    'AGENDADO POR': 'AGENDADO',
    'DIA 2': 'DIA2', 'DIA2': 'DIA2',
    'MES 2': 'MES3', 'MES3': 'MES3',
    'AÑO 2': 'ANIO4', 'AÑO4': 'ANIO4', 'ANIO 2': 'ANIO4',
    'CAMPAÑA': 'CAMPANA', 'CAMPAÑA': 'CAMPANA',
}

# Columnas que se copian de AGENDADOS a las filas nuevas del maestro (el resto
# ASISTENCIA, DISTRITO, EDAD, SEXO, TRAT/PAGO se completan desde VENTA).
AG_COPY_SEMS = {'DIA', 'MES', 'ANIO', 'DNI', 'NOMBRE', 'RED_SOCIAL',
                'TELEFONO', 'CORREO', 'AGENDADO', 'DIA2', 'MES3', 'ANIO4',
                'CAMPANA'}


def detectar_agendados(ws):
    """Detecta las columnas de la hoja AGENDADOS por su cabecera (fila 4),
    soportando el formato Derma Essenza (sin CRM) y el BM (con CRM).
    Devuelve {semántico: letra} o None si no se reconoce."""
    hdr = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    col = {}
    for i, h in enumerate(hdr, 1):
        if h is None:
            continue
        hh = str(h).strip().upper().replace(' ', ' ')
        hh2 = ' '.join(hh.split())
        sem = _SEM_AG.get(hh2) or _SEM_AG.get(hh)
        if sem and sem not in col:
            col[sem] = openpyxl.utils.get_column_letter(i)
    if not (col.get('NOMBRE') and col.get('TELEFONO') and col.get('DIA2')):
        return None
    return col


def leer_agendados(path):
    """Devuelve (filas, ag_col): filas = [(fila_excel, {letra: valor})] de la hoja
    AGENDADOS y ag_col = {semántico: letra} detectado por cabecera."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['AGENDADOS']
    ag_col = detectar_agendados(ws)
    filas = []
    if not ag_col:
        return filas, None
    c_nom = openpyxl.utils.column_index_from_string(ag_col['NOMBRE'])
    c_dia = openpyxl.utils.column_index_from_string(ag_col['DIA'])
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=c_nom).value is None and ws.cell(row=r, column=c_dia).value is None:
            continue
        fila = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                fila[openpyxl.utils.get_column_letter(c)] = v
        filas.append((r, fila))
    return filas, ag_col


def leer_venta(path):
    """Devuelve lista de dicts de las hojas VENTA 2026 y VENTA 2025."""
    wb = openpyxl.load_workbook(path, data_only=True)
    filas = []
    for hoja, hr in [('VENTA 2026', 5), ('VENTA 2025', 8)]:
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        for r in range(hr + 1, ws.max_row + 1):
            if ws.cell(row=r, column=3).value is None and ws.cell(row=r, column=7).value is None:
                continue
            fila = {
                'hoja': hoja,
                'dia': ws.cell(row=r, column=2).value,
                'mes': ws.cell(row=r, column=3).value,
                'anio': ws.cell(row=r, column=4).value,
                'dni': ws.cell(row=r, column=5).value,
                'cel': ws.cell(row=r, column=6).value,
                'nombre': ws.cell(row=r, column=7).value,
                'distrito': ws.cell(row=r, column=9).value,
                'edad': ws.cell(row=r, column=10).value,
                'sexo': ws.cell(row=r, column=11).value,
                'tratamiento': ws.cell(row=r, column=12).value,
                'status': ws.cell(row=r, column=14).value,
                'venta': ws.cell(row=r, column=15).value,
                'fila': r,
            }
            filas.append(fila)
    return filas


def leer_maestro(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['BD DATA']
    return ws


_SEMANTICOS = {
    'DIA': 'DIA', 'MES': 'MES', 'AÑO': 'ANIO', 'DNI': 'DNI', 'NOMBRE': 'NOMBRE',
    'RED SOCIAL': 'RED_SOCIAL', 'TELEFONO': 'TELEFONO', 'CORREO': 'CORREO',
    'AGENDADO POR': 'AGENDADO', 'DIA2': 'DIA2', 'MES3': 'MES3', 'AÑO4': 'ANIO4',
    'CAMPAÑA': 'CAMPANA', 'ASISTENCIA': 'ASISTENCIA', 'DISTRITO': 'DISTRITO',
    'EDAD': 'EDAD', 'SEXO': 'SEXO', 'PAGO TOTAL': 'PAGO_TOTAL',
    'MOTIVO NO ASISTIO': 'MOTIVO_NO_ASISTIO',
    'MOTIVO NO COMPRA': 'MOTIVO_NO_COMPRA',
}

# Columnas (letras) de la fuente AGENDADOS (formato fijo BM) -> semántico.
# Con detección por cabecera (detectar_agendados) se usa el mapeo dinámico.
FUENTE_AG_COLS = {
    'C': 'DIA', 'D': 'MES', 'E': 'ANIO', 'G': 'NOMBRE', 'H': 'RED_SOCIAL',
    'I': 'TELEFONO', 'J': 'CORREO', 'K': 'AGENDADO', 'L': 'DIA2', 'M': 'MES3',
    'N': 'ANIO4', 'O': 'CAMPANA',
}


def detectar_maestro(ws):
    """Detecta las columnas del maestro por su cabecera (fila 4) y devuelve un
    dict {semántico: letra} compatible con los formatos BM (con columna CRM) y
    Derma Essenza (con RED SOCIAL, sin CRM). Devuelve None si el formato no se
    reconoce (faltan columnas clave como NOMBRE o TELEFONO)."""
    hdr = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    first = {}
    trat = []
    pago = []
    for i, h in enumerate(hdr, 1):
        if h is None:
            continue
        hh = str(h).strip().upper()
        if hh in _SEMANTICOS:
            first.setdefault(hh, i)
        elif hh.startswith('TRAT '):
            try:
                trat.append((int(hh[5:]), i))
            except ValueError:
                pass
        elif hh.startswith('PAGO '):
            try:
                pago.append((int(hh[5:]), i))
            except ValueError:
                pass
    L = openpyxl.utils.get_column_letter

    def g(h):
        return L(first[h]) if h in first else None

    if not (first.get('NOMBRE') and first.get('TELEFONO') and first.get('DIA2')):
        return None
    return {
        'DIA': g('DIA'), 'MES': g('MES'), 'ANIO': g('AÑO'), 'DNI': g('DNI'),
        'NOMBRE': g('NOMBRE'), 'RED_SOCIAL': g('RED SOCIAL'),
        'TELEFONO': g('TELEFONO'), 'CORREO': g('CORREO'), 'AGENDADO': g('AGENDADO POR'),
        'DIA2': g('DIA2'), 'MES3': g('MES3'), 'ANIO4': g('AÑO4'),
        'CAMPANA': g('CAMPAÑA'), 'ASISTENCIA': g('ASISTENCIA'),
        'DISTRITO': g('DISTRITO'), 'EDAD': g('EDAD'), 'SEXO': g('SEXO'),
        'TRAT': [L(c) for _, c in sorted(trat)],
        'PAGO': [L(c) for _, c in sorted(pago)],
        'PAGO_TOTAL': g('PAGO TOTAL'),
        'MOTIVO_NO_ASISTIO': g('MOTIVO NO ASISTIO'),
        'MOTIVO_NO_COMPRA': g('MOTIVO NO COMPRA'),
    }


def _rango_pivot(path, nombre_pivot):
    """Devuelve (hoja_origen, fila_inicio, fila_fin) del origen de una tabla
    dinámica (pivot) leyendo el XML del workbook, o None si no se encuentra."""
    try:
        with zipfile.ZipFile(path) as z:
            nombres = z.namelist()
            for pt in nombres:
                if not (pt.startswith('xl/pivotTables/') and pt.endswith('.xml')):
                    continue
                xml = z.read(pt).decode('utf-8', 'replace')
                m = re.search(r'<pivotTableDefinition[^>]*name="([^"]+)"', xml)
                if not m or m.group(1) != nombre_pivot:
                    continue
                target = None
                rels = pt.replace('.xml', '.xml.rels')
                if rels in nombres:
                    r = z.read(rels).decode('utf-8', 'replace')
                    m2 = re.search(r'Target="([^"]*pivotCache[^"]*)"', r)
                    if m2:
                        target = m2.group(1).replace('../', 'xl/')
                if target and target in nombres:
                    cdef = z.read(target).decode('utf-8', 'replace')
                else:
                    m2 = re.search(r'cacheId="(\d+)"', xml)
                    cid = int(m2.group(1)) if m2 else 0
                    cdef = z.read(
                        f'xl/pivotCache/pivotCacheDefinition{cid + 1}.xml'
                    ).decode('utf-8', 'replace')
                m3 = re.search(
                    r'<worksheetSource[^>]*ref="([A-Z]+)(\d+):[A-Z]+(\d+)"'
                    r'[^>]*sheet="([^"]+)"', cdef)
                if m3:
                    return m3.group(4), int(m3.group(2)), int(m3.group(3))
    except (KeyError, zipfile.BadZipFile):
        return None
    return None


def verificar_venta_vs_td(path, anio, mes, desde, hasta):
    """Verifica que la hoja VENTA {anio} coincida con su tabla dinámica TD {anio}
    para el periodo (mes, anio, desde..hasta).

    Devuelve dict con la suma de VENTA (col O) de la hoja completa, la suma del
    rango cubierto por la TD (pivot) y si coinciden. Si hay filas de venta fuera
    del rango de la tabla dinámica, algún dato está mal.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:  # noqa: BLE001
        return {'ok': False, 'error': 'No se pudo abrir VENTA_DIARIA.xlsx'}
    hoja = f'VENTA {anio}'
    if hoja not in wb.sheetnames:
        return {'ok': False, 'error': f'No hay hoja {hoja} en VENTA_DIARIA.xlsx'}
    ws = wb[hoja]
    rango = _rango_pivot(path, f'TD {anio}')
    fila_fin = ws.max_row
    if rango:
        fila_fin = min(ws.max_row, rango[2])
    venta_tot = 0.0
    td_tot = 0.0
    filas_fuera = 0
    for r in range(5, ws.max_row + 1):
        d = ws.cell(row=r, column=2).value
        m = ws.cell(row=r, column=3).value
        a = ws.cell(row=r, column=4).value
        v = ws.cell(row=r, column=15).value
        if (a == anio and m == mes and isinstance(d, (int, float))
                and desde <= int(d) <= hasta and isinstance(v, (int, float))):
            venta_tot += v
            if r <= fila_fin:
                td_tot += v
            else:
                filas_fuera += 1
    coincide = abs(venta_tot - td_tot) < 0.01
    mensaje = (f'VENTA {anio} y TD {anio} coinciden en el periodo: S/ {venta_tot:,.2f}'
               if coincide else
               f'VENTA {anio} (S/ {venta_tot:,.2f}) NO coincide con TD {anio} '
               f'(S/ {td_tot:,.2f}): {filas_fuera} fila(s) de venta quedan fuera del '
               'rango de la tabla dinámica.')
    return {'ok': True, 'venta': round(venta_tot, 2), 'td': round(td_tot, 2),
            'coincide': coincide, 'filas_fuera': filas_fuera, 'mensaje': mensaje}


def diferencias_maestro_venta(maestro_path, venta_path, anio, mes, desde, hasta):
    """Compara, paciente por paciente (teléfono+fecha), lo que el maestro tiene
    registrado como compra contra lo que VENTA DIARIA registra como venta
    confirmada en el periodo. Devuelve la lista de casos donde el monto no
    coincide, para señalar exactamente qué fila del maestro hay que revisar
    (en vez de solo avisar que "algo no cuadra")."""
    try:
        ws = leer_maestro(maestro_path)
    except Exception:  # noqa: BLE001
        return []
    col = detectar_maestro(ws)
    if not col:
        return []

    def idx(letra):
        return openpyxl.utils.column_index_from_string(letra) if letra else None

    c_tel, c_nom = idx(col.get('TELEFONO')), idx(col.get('NOMBRE'))
    c_d2, c_m3, c_a4 = idx(col.get('DIA2')), idx(col.get('MES3')), idx(col.get('ANIO4'))
    c_asist, c_ptot = idx(col.get('ASISTENCIA')), idx(col.get('PAGO_TOTAL'))
    c_pago = [idx(c) for c in col.get('PAGO', [])]
    if not (c_tel and c_d2 and c_m3 and c_a4 and c_asist):
        return []

    por_paciente = defaultdict(lambda: {'maestro': 0.0, 'venta': 0.0, 'nombre': None,
                                        'fila_maestro': None, 'fecha': None})
    for r in range(5, ws.max_row + 1):
        if not (ws.cell(row=r, column=c_a4).value == anio
                and ws.cell(row=r, column=c_m3).value == mes
                and isinstance(ws.cell(row=r, column=c_d2).value, (int, float))
                and desde <= int(ws.cell(row=r, column=c_d2).value) <= hasta
                and ws.cell(row=r, column=c_asist).value == 'ASISTIO'):
            continue
        ph = norm_phone(ws.cell(row=r, column=c_tel).value)
        if not ph:
            continue
        p_total = ws.cell(row=r, column=c_ptot).value if c_ptot else None
        if not (isinstance(p_total, (int, float)) and p_total > 0):
            p_total = sum(x for x in (ws.cell(row=r, column=c).value for c in c_pago)
                          if isinstance(x, (int, float)))
        d = por_paciente[ph]
        d['maestro'] += p_total
        d['nombre'] = d['nombre'] or (ws.cell(row=r, column=c_nom).value if c_nom else None)
        d['fila_maestro'] = r
        d['fecha'] = int(ws.cell(row=r, column=c_d2).value)

    for v in leer_venta(venta_path):
        st = (txt(v['status']) or '').upper()
        if not (st.startswith(ASISTE_POR_TEXTO) or any(x in st for x in ASISTE_POR_TEXTO)):
            continue
        if not (v['anio'] == anio and v['mes'] == mes
                and isinstance(v['dia'], (int, float)) and desde <= int(v['dia']) <= hasta
                and isinstance(v['venta'], (int, float))):
            continue
        ph = norm_phone(v['cel'])
        if not ph:
            continue
        d = por_paciente[ph]
        d['venta'] += v['venta']
        d['nombre'] = d['nombre'] or v['nombre']
        d['fecha'] = d['fecha'] or (int(v['dia']) if isinstance(v['dia'], (int, float)) else None)

    diffs = []
    for ph, d in por_paciente.items():
        if abs(d['maestro'] - d['venta']) > 0.01:
            diffs.append({
                'telefono': ph, 'nombre': d['nombre'], 'fecha': d['fecha'],
                'fila_maestro': d['fila_maestro'],
                'monto_maestro': round(d['maestro'], 2), 'monto_venta': round(d['venta'], 2),
                'diferencia': round(d['maestro'] - d['venta'], 2),
            })
    diffs.sort(key=lambda x: abs(x['diferencia']), reverse=True)
    return diffs


# ============================================================
# CÁLCULO DE CAMBIOS
# ============================================================
class Calculo:
    def __init__(self, maestro_ws, agendados, venta, col=None, ag_col=None):
        self.maestro = maestro_ws
        self.col = col or detectar_maestro(maestro_ws) or {}
        self.ag_col = ag_col or {}
        self.new_rows = []       # lista de (provisional, {col: valor})
        self.updates = {}        # {fila_maestro: {col: valor}} (existentes y provisionales)
        self.matches = []        # (venta_fila, maestro_fila, modo, hoja)
        self.pendientes = []     # ventas sin match
        self.revisar = []        # casos a revisar
        self.incompletas = []    # filas AGENDADOS incompletas/sin nombre
        self.sin_hueco = []      # (venta, maestro_fila) sin par TRAT/PAGO libre
        self._nueva_info = {}    # provisional -> (ph, nm, fecha)
        self._indexar()
        self._calcular_nuevos(agendados)
        self._indexar_nuevos()
        self._calcular_ventas(venta)

    def _idx(self, sem):
        c = self.col.get(sem)
        return openpyxl.utils.column_index_from_string(c) if c else None

    def _ag(self, sem):
        """Letra de la fuente AGENDADOS para el semántico (o fallback BM)."""
        c = self.ag_col.get(sem) or FUENTE_AG_COLS.get(sem)
        return c if c else None

    # ----- índice del maestro -----
    def _indexar(self):
        ws = self.maestro
        c_nom = self._idx('NOMBRE')
        c_tel = self._idx('TELEFONO')
        c_d2 = self._idx('DIA2')
        c_m3 = self._idx('MES3')
        c_a4 = self._idx('ANIO4')
        c_camp = self._idx('CAMPANA')
        self.m_rows = []
        self.by_phone = defaultdict(list)
        self.by_name = defaultdict(list)
        self.by_phone_date = defaultdict(set)  # (tel, fecha) -> set(campañas)
        self.keys_full = set()
        self.keys_loose = set()
        self.last_row = 4
        for r in range(5, ws.max_row + 1):
            if c_nom and c_tel and c_d2 and all(ws.cell(row=r, column=c).value is None
                                                for c in (c_nom, c_tel, c_d2)):
                continue
            self.last_row = r
            self.m_rows.append(r)
            cel = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    cel[openpyxl.utils.get_column_letter(c)] = v
            ph = norm_phone(cel.get(self.col.get('TELEFONO')))
            nm = norm_name(cel.get(self.col.get('NOMBRE')))
            fc = norm_fecha(cel.get(self.col.get('DIA2')),
                            cel.get(self.col.get('MES3')),
                            cel.get(self.col.get('ANIO4')))
            camp = norm_name(cel.get(self.col.get('CAMPANA')))
            if ph:
                self.by_phone[ph].append(r)
            if nm:
                self.by_name[nm].append(r)
            if ph and fc:
                self.by_phone_date[(ph, fc)].add(camp)
            self.keys_full.add((ph, nm, fc, camp))
            if nm:
                self.keys_loose.add((nm, fc, camp))

    # ----- 1) AGENDADOS -> filas nuevas (con fila provisional) -----
    def _calcular_nuevos(self, agendados):
        c_ph = self._ag('TELEFONO')
        c_nm = self._ag('NOMBRE')
        c_d2 = self._ag('DIA2')
        c_m3 = self._ag('MES3')
        c_a4 = self._ag('ANIO4')
        c_camp = self._ag('CAMPANA')
        vistos = set()
        for r, fila in agendados:
            ph = norm_phone(fila.get(c_ph))
            nm = norm_name(fila.get(c_nm))
            fc = norm_fecha(fila.get(c_d2), fila.get(c_m3), fila.get(c_a4))
            camp = norm_name(fila.get(c_camp))
            if not nm and not ph:
                self.incompletas.append((r, fila, 'sin nombre y sin teléfono'))
                continue
            if (ph, nm, fc, camp) in self.keys_full:
                continue
            if nm and (nm, fc, camp) in self.keys_loose:
                continue
            # Teléfono + fecha + campaña: misma persona, misma cita, misma campaña → duplicado
            if ph and camp and camp in self.by_phone_date.get((ph, fc), set()):
                self.incompletas.append((r, fila, 'teléfono+fecha+campaña ya existen en maestro'))
                continue
            if (ph, nm, fc, camp) in vistos:
                continue
            vistos.add((ph, nm, fc, camp))
            if not nm:
                self.incompletas.append((r, fila, 'sin nombre en AGENDADOS'))
            prov = self.last_row + len(self.new_rows) + 1
            self.new_rows.append((prov, fila))

    def _indexar_nuevos(self):
        c_ph = self._ag('TELEFONO')
        c_nm = self._ag('NOMBRE')
        c_d2 = self._ag('DIA2')
        c_m3 = self._ag('MES3')
        c_a4 = self._ag('ANIO4')
        for prov, fila in self.new_rows:
            ph = norm_phone(fila.get(c_ph))
            nm = norm_name(fila.get(c_nm))
            fc = norm_fecha(fila.get(c_d2), fila.get(c_m3), fila.get(c_a4))
            self._nueva_info[prov] = (ph, nm, fc)
            if ph:
                self.by_phone[ph].append(prov)
            if nm:
                self.by_name[nm].append(prov)

    # ----- helpers que funcionan para filas existentes y provisionales -----
    def _valor(self, fila, letra):
        if fila in self._nueva_info:
            return None
        col = openpyxl.utils.column_index_from_string(letra)
        return self.maestro.cell(row=fila, column=col).value

    def _fecha(self, fila):
        if fila in self._nueva_info:
            return self._nueva_info[fila][2]
        return norm_fecha(self._valor(fila, self.col.get('DIA2')),
                          self._valor(fila, self.col.get('MES3')),
                          self._valor(fila, self.col.get('ANIO4')))

    def _asistio(self, fila):
        return txt(self._valor(fila, self.col.get('ASISTENCIA'))) == 'ASISTIO'

    def _tiene_tratamiento(self, fila, v):
        """¿Alguno de los pares TRAT/PAGO de la fila ya trae esta venta?

        Se compara tratamiento + monto: sin esto, cada corrida diaria volvería
        a escribir la misma venta en el siguiente par libre."""
        trat_v = txt(v.get('tratamiento'))
        monto_v = num(v.get('venta'))
        for trat_col, pago_col in zip(self.col.get('TRAT', []), self.col.get('PAGO', [])):
            if txt(self._valor(fila, trat_col)) != trat_v:
                continue
            # Mismo tratamiento: sólo cuenta como la misma venta si el monto
            # coincide (o si no hay monto con qué distinguirlas).
            if monto_v is None or num(self._valor(fila, pago_col)) == monto_v:
                return True
        return False

    # ----- 2) VENTA DIARIA -> completar P..AB -----
    def _calcular_ventas(self, venta):
        ws = self.maestro
        para_llenar = defaultdict(list)  # maestro_row -> [venta_row]
        walkins = defaultdict(list)      # (tel, nombre, fecha) -> [venta_row] sin match
        for v in venta:
            status = txt(v['status']) or ''
            st = status.upper()
            ph = norm_phone(v['cel'])
            nm = norm_name(v['nombre'])
            fv = norm_fecha(v['dia'], v['mes'], v['anio'])

            if st in REVISAR_STATUS:
                self.revisar.append((v, 'STATUS por revisar'))
                continue
            if not (st.startswith(ASISTE_POR_TEXTO) or any(x in st for x in ASISTE_POR_TEXTO)):
                self.pendientes.append((v, f'STATUS "{status}" (no asistió o desconocido)'))
                continue

            def empty_p_rows(lista):
                return [r for r in lista if not self._asistio(r)]

            def ya_asistio(cand, etiqueta):
                """La fila ya está marcada ASISTIO. Si el tratamiento de esta
                venta todavía no figura en sus pares TRAT/PAGO, es un segundo
                tratamiento del mismo paciente: va al siguiente par libre de esa
                misma fila, no a una fila nueva. Si ya figura, está sincronizada
                y no se toca (esto es lo que evita duplicar en cada corrida)."""
                fila = next(r for r in cand if self._asistio(r))
                if self._tiene_tratamiento(fila, v):
                    self.pendientes.append((v, f'ya sincronizada ({etiqueta})'))
                    return None
                return fila

            modo = None
            candidatos = []
            # Etapa 1: mismo teléfono + misma fecha de cita
            if ph:
                cand = [r for r in self.by_phone[ph] if self._fecha(r) == fv]
                if any(self._asistio(r) for r in cand):
                    fila = ya_asistio(cand, 'teléfono+fecha')
                    if fila is None:
                        continue
                    candidatos, modo = [fila], 'telefono+fecha (2do tratamiento)'
                if not candidatos:
                    cand = empty_p_rows(cand)
                    if len(cand) == 1:
                        candidatos, modo = cand, 'telefono+fecha'
                    elif len(cand) > 1:
                        self.revisar.append((v, f'{len(cand)} filas con mismo teléfono+fecha'))
            # Etapa 2: mismo nombre + misma fecha
            if not candidatos and nm:
                cand = [r for r in self.by_name[nm] if self._fecha(r) == fv]
                if any(self._asistio(r) for r in cand):
                    fila = ya_asistio(cand, 'nombre+fecha')
                    if fila is None:
                        continue
                    candidatos, modo = [fila], 'nombre+fecha (2do tratamiento)'
                if not candidatos:
                    cand = empty_p_rows(cand)
                    if len(cand) == 1:
                        candidatos, modo = cand, 'nombre+fecha'
                    elif len(cand) > 1:
                        self.revisar.append((v, f'{len(cand)} filas con mismo nombre+fecha'))
            if not candidatos:
                if ph or nm:
                    walkins[(ph, nm, fv)].append(v)
                else:
                    self.pendientes.append((v, 'sin coincidencia en maestro'))
                continue

            fila_m = candidatos[0]
            para_llenar[fila_m].append(v)
            self.matches.append((v['fila'], fila_m, modo, v['hoja']))

        # ----- ventas sin agendado previo: crear fila nueva (walk-in) para que -----
        # ----- el dinero no se pierda del maestro y siempre cuadre con VENTA DIARIA -----
        c_nom = self.col.get('NOMBRE')
        c_tel = self.col.get('TELEFONO')
        c_dia = self.col.get('DIA')
        c_mes = self.col.get('MES')
        c_anio = self.col.get('ANIO')
        c_dia2 = self.col.get('DIA2')
        c_mes3 = self.col.get('MES3')
        c_anio4 = self.col.get('ANIO4')
        c_camp = self.col.get('CAMPANA')
        for (ph, nm, fv), ventas in walkins.items():
            ventas.sort(key=lambda x: (x['anio'] or 0, str(x['mes'] or ''), x['dia'] or 0, x['fila']))
            primer = ventas[0]
            prov = self.last_row + len(self.new_rows) + 1
            self.new_rows.append((prov, {}))
            self._nueva_info[prov] = (ph, nm, fv)
            if ph:
                self.by_phone[ph].append(prov)
            if nm:
                self.by_name[nm].append(prov)
            u = self.updates.setdefault(prov, {})
            if c_nom:
                u[c_nom] = primer['nombre']
            if c_tel:
                u[c_tel] = primer['cel']
            for c, val in ((c_dia, primer['dia']), (c_mes, primer['mes']), (c_anio, primer['anio']),
                           (c_dia2, primer['dia']), (c_mes3, primer['mes']), (c_anio4, primer['anio'])):
                if c and val is not None:
                    u[c] = val
            if c_camp:
                u[c_camp] = 'VENTA SIN AGENDAR'
            for v in ventas:
                para_llenar[prov].append(v)
                self.matches.append((v['fila'], prov, 'venta sin agendado (fila nueva)', v['hoja']))

        # ----- armar las celdas a escribir -----
        c_asist = self.col.get('ASISTENCIA')
        c_dni = self.col.get('DNI')
        c_dist = self.col.get('DISTRITO')
        c_edad = self.col.get('EDAD')
        c_sexo = self.col.get('SEXO')
        c_ptot = self.col.get('PAGO_TOTAL')
        for fila_m, ventas in para_llenar.items():
            ventas.sort(key=lambda x: (x['anio'] or 0, str(x['mes'] or ''), x['dia'] or 0, x['fila']))
            primer = ventas[0]
            u = self.updates.setdefault(fila_m, {})
            if not self._asistio(fila_m):
                u[c_asist] = 'ASISTIO'
            for col, v in ((c_dni, primer['dni']), (c_dist, primer['distrito']),
                           (c_edad, primer['edad']), (c_sexo, primer['sexo'])):
                if col and v is not None and txt(self._valor(fila_m, col)) is None:
                    u[col] = v
            # Cada venta va al PRIMER par TRAT/PAGO libre, no al par que le
            # tocaría por su orden: si TRAT 1 ya trae el tratamiento de una
            # sincronización anterior, el segundo tratamiento debe caer en
            # TRAT 2, no perderse.
            par = list(zip(self.col.get('TRAT', []), self.col.get('PAGO', [])))
            for v in ventas:
                libre = next(((t, p) for t, p in par
                              if txt(self._valor(fila_m, t)) is None and t not in u), None)
                if libre is None:
                    self.sin_hueco.append((v, fila_m))
                    continue
                trat_col, pago_col = libre
                u[trat_col] = v['tratamiento']
                if num(self._valor(fila_m, pago_col)) is None and v['venta'] is not None:
                    u[pago_col] = num(v['venta'])
            # Recalcular PAGO TOTAL como suma de todos los pares TRAT/PAGO
            # (incluye valores que acabamos de escribir + los que ya existían)
            if c_ptot:
                s = 0.0
                for c_pago in self.col.get('PAGO', []):
                    val = u.get(c_pago) or self._valor(fila_m, c_pago)
                    if isinstance(val, (int, float)):
                        s += val
                u[c_ptot] = s

    def resumen(self):
        n_match_exacto = sum(1 for _, _, m, _ in self.matches if m in ('telefono+fecha', 'nombre+fecha'))
        n_sin_fecha = sum(1 for _, _, m, _ in self.matches if m == 'sin fecha' or 'sin fecha' in m)
        return {
            'filas_nuevas': len(self.new_rows),
            'ventas_totales': len(self.pendientes) + len(self.matches) + len(self.revisar),
            'matches': len(self.matches),
            'matches_exactos': n_match_exacto,
            'matches_sin_fecha': n_sin_fecha,
            'pendientes': len(self.pendientes),
            'revisar': len(self.revisar),
            'incompletas': len(self.incompletas),
            'sin_hueco': len(self.sin_hueco),
            'filas_a_actualizar': len(self.updates),
            'celdas_a_escribir': sum(len(v) for v in self.updates.values()),
        }


# ============================================================
# APLICACIÓN VÍA XML (preserva tablas dinámicas, slicers, fórmulas)
# ============================================================
AB_FORMULA = ('+Tabla1[[#This Row],[PAGO 1]]+Tabla1[[#This Row],[PAGO 2]]'
              '+Tabla1[[#This Row],[PAGO 3]]+Tabla1[[#This Row],[PAGO 4]]')


def esc_xml(s):
    return (str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


def celda_xml(col, fila, valor, formula=False):
    ref = f'{col}{fila}'
    if formula:
        return f'<c r="{ref}" s="1"><f>{esc_xml(AB_FORMULA)}</f><v>0</v></c>'
    if valor is None:
        return f'<c r="{ref}" s="1"/>'
    if isinstance(valor, (int, float)):
        return f'<c r="{ref}" s="1"><v>{num(valor)}</v></c>'
    t = esc_xml(valor)
    return f'<c r="{ref}" s="1" t="inlineStr"><is><t>{t}</t></is></c>'


_CELL_RE = re.compile(
    r'<c r="([A-Z]{1,3})\d+"[^>]*?/>(?=<c|</row>|$)'
    r'|<c r="([A-Z]{1,3})\d+"[^>]*?>(?:(?!</c>).)*?</c>', re.S)
_CELL_OPEN_RE = re.compile(r'<c r="([A-Z]{1,3})\d+"[^>]*>')


def parse_row(row_xml):
    """Extrae celdas de una fila -> dict {col: xml de celda}."""
    celdas = {}
    pos = 0
    for m in _CELL_RE.finditer(row_xml):
        xml = m.group(0)
        col = _CELL_OPEN_RE.search(xml).group(1)
        celdas[col] = xml
    return celdas


_COL_ORDER = {openpyxl.utils.get_column_letter(i): i for i in range(1, 64)}


def build_row(fila, celdas):
    items = sorted(celdas.items(), key=lambda kv: _COL_ORDER[kv[0]])
    ult = max((_COL_ORDER.get(k, 0) for k, _ in items), default=28)
    spans = f'2:{max(ult, 28)}'
    return f'<row r="{fila}" spans="{spans}" x14ac:dyDescent="0.3">' + ''.join(x for _, x in items) + '</row>'


def aplicar_xml(origen, destino, new_rows, updates, col=None, ag_col=None):
    col = col or {}
    ag_col = ag_col or {}
    p_total = col.get('PAGO_TOTAL') or 'AB'
    primera = col.get('DIA') or 'B'
    ref_inicio = f'{primera}4'
    ref = f'{ref_inicio}:{p_total}'
    build_dir = os.path.join(TMP_DIR, 'build')
    if os.path.exists(build_dir):
        shutil.rmtree(build_dir)
    os.makedirs(build_dir)
    with zipfile.ZipFile(origen) as z:
        z.extractall(build_dir)

    sheet_path = os.path.join(build_dir, 'xl', 'worksheets', 'sheet1.xml')
    with open(sheet_path, encoding='utf-8') as f:
        x = f.read()

    # ---- actualizar celdas de filas existentes ----
    row_re = re.compile(r'<row r="(\d+)"[^>]*>.*?</row>', re.S)

    def reemplazar_row(m):
        fila = int(m.group(1))
        if fila not in updates:
            return m.group(0)
        celdas = parse_row(m.group(0))
        for colc, val in updates[fila].items():
            celdas[colc] = celda_xml(colc, fila, val)
        return build_row(fila, celdas)

    x = row_re.sub(reemplazar_row, x)

    # letra de AGENDADOS -> semántico (detección por cabecera, fallback BM)
    letra_a_sem = {letra: sem for sem, letra in ag_col.items()}

    # ---- agregar filas nuevas ----
    ultima = max(int(m.group(1)) for m in row_re.finditer(x)) or 4
    for i, (fila_num, fila) in enumerate(new_rows):
        celdas = {}
        for col_src, val in fila.items():
            sem = letra_a_sem.get(col_src) or FUENTE_AG_COLS.get(col_src)
            if not sem or sem not in AG_COPY_SEMS:
                continue
            mcol = col.get(sem)
            if mcol:
                celdas[mcol] = celda_xml(mcol, fila_num, val)
        for colc, val in updates.get(fila_num, {}).items():
            celdas[colc] = celda_xml(colc, fila_num, val)
        celdas[p_total] = celda_xml(p_total, fila_num, 0, formula=True)
        x = x.replace('</sheetData>', build_row(fila_num, celdas) + '</sheetData>', 1)
        if fila_num > ultima:
            ultima = fila_num

    # ---- actualizar tabla y dimension ----
    x = re.sub(r'<table ref="[^"]+"', f'<table ref="{ref}{ultima}"', x)
    x = re.sub(r'<autoFilter ref="[^"]+"', f'<autoFilter ref="{ref}{ultima}"', x)
    x = re.sub(r'<dimension ref="[^"]+"', f'<dimension ref="A1:{p_total}{ultima}"', x)

    with open(sheet_path, 'w', encoding='utf-8') as f:
        f.write(x)

    # ---- actualizar tabla1.xml ----
    tab_path = os.path.join(build_dir, 'xl', 'tables', 'table1.xml')
    if os.path.exists(tab_path):
        with open(tab_path, encoding='utf-8') as f:
            t = f.read()
        t = re.sub(r'ref="[A-Z]+\d+:[A-Z]+\d+"', f'ref="{ref}{ultima}"', t)
        with open(tab_path, 'w', encoding='utf-8') as f:
            f.write(t)

    # ---- re-empacar ----
    if os.path.exists(destino):
        os.remove(destino)
    with zipfile.ZipFile(destino, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(build_dir):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, build_dir)
                z.write(full, rel)


# ============================================================
# MIGRACIÓN DE COLUMNAS DE MOTIVO (preserva tablas dinámicas)
# ============================================================
MOTIVO_NO_ASISTIO_HEADER = 'MOTIVO NO ASISTIO'
MOTIVO_NO_COMPRA_HEADER = 'MOTIVO NO COMPRA'
MOTIVOS_NO_ASISTIO = ['REAGENDÓ', 'NO CONTESTÓ', 'CANCELÓ', 'OTRO']
MOTIVOS_NO_COMPRA = ['SOLO CONSULTA', 'NO TENÍA PRESUPUESTO', 'REAGENDÓ', 'OTRO']


def detectar_columnas_motivo(ws):
    """Devuelve {semántico: letra} de las columnas de motivo del maestro (o {} si no existen)."""
    col = detectar_maestro(ws) or {}
    return {k: col[k] for k in ('MOTIVO_NO_ASISTIO', 'MOTIVO_NO_COMPRA') if col.get(k)}


def migrar_columnas_motivo(maestro, nuevo_path=None):
    """Agrega las columnas MOTIVO NO ASISTIO / MOTIVO NO COMPRA al maestro si no
    existen, vía XML surgery (no rompe tablas dinámicas ni fórmulas).

    Devuelve el dict {semántico: letra} de las columnas de motivo tras la
    migración. Idempotente: si ya existen ambas, no modifica nada.
    """
    destino = nuevo_path or maestro
    build_dir = os.path.join(TMP_DIR, 'build_motivo')
    if os.path.exists(build_dir):
        shutil.rmtree(build_dir)
    os.makedirs(build_dir)

    with zipfile.ZipFile(maestro) as z:
        z.extractall(build_dir)

    sheet_path = os.path.join(build_dir, 'xl', 'worksheets', 'sheet1.xml')
    with open(sheet_path, encoding='utf-8') as f:
        x = f.read()

    # --- detectar cabeceras existentes en la fila 4 (via openpyxl) ---
    salida = detectar_columnas_motivo(leer_maestro(maestro))
    if len(salida) == 2:
        return salida

    # --- calcular letras libres (después de la última columna usada) ---
    ws = leer_maestro(maestro)
    L = openpyxl.utils.get_column_letter
    ultima = max((ws.max_column, 27))
    if 'MOTIVO_NO_ASISTIO' not in salida:
        ultima += 1
        salida['MOTIVO_NO_ASISTIO'] = L(ultima)
    if 'MOTIVO_NO_COMPRA' not in salida:
        ultima += 1
        salida['MOTIVO_NO_COMPRA'] = L(ultima)

    # --- reescribir la fila 4 con las nuevas cabeceras ---
    fila4 = re.search(r'<row r="4"[^>]*>.*?</row>', x, re.S)
    if not fila4:
        raise ValueError('No se encontró la fila 4 (cabecera) del maestro')
    celdas = parse_row(fila4.group(0))
    for sem, texto in (('MOTIVO_NO_ASISTIO', MOTIVO_NO_ASISTIO_HEADER),
                       ('MOTIVO_NO_COMPRA', MOTIVO_NO_COMPRA_HEADER)):
        letra = salida[sem]
        celdas[letra] = celda_xml(letra, 4, texto)

    def reemplazar_row4(m):
        return build_row(4, celdas)

    x = row_re_sub4(x, reemplazar_row4)

    # --- actualizar dimension ---
    x = re.sub(r'<dimension ref="A1:[A-Z]+', f'<dimension ref="A1:{salida["MOTIVO_NO_COMPRA"]}', x)

    with open(sheet_path, 'w', encoding='utf-8') as f:
        f.write(x)

    if os.path.exists(destino):
        os.remove(destino)
    with zipfile.ZipFile(destino, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(build_dir):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, build_dir)
                z.write(full, rel)
    return salida


def row_re_sub4(x, fn):
    """Reemplaza la fila 4 de sheet1.xml (revisada antes con `<row r="4"`)."""
    return re.sub(r'<row r="4"[^>]*>.*?</row>', fn, x, count=1, flags=re.S)


def actualizar_motivo_maestro(fila, tipo, motivo):
    """Registra un motivo (MOTIVO NO ASISTIO / MOTIVO NO COMPRA) en la fila del
    maestro. tipo en ('no_asistio', 'no_compra'). Aplica y sube el maestro.

    Devuelve el dict col detectado tras la escritura."""
    maestro = ruta_maestro_local()
    sem = 'MOTIVO_NO_ASISTIO' if tipo == 'no_asistio' else 'MOTIVO_NO_COMPRA'
    col = detectar_maestro(leer_maestro(maestro)) or {}
    letra = col.get(sem)
    if not letra:
        salida = migrar_columnas_motivo(maestro)
        letra = salida.get(sem)
        col = detectar_maestro(leer_maestro(maestro)) or {}
    if not letra:
        raise ValueError(f'No se encontró la columna {sem} en el maestro')
    aplicar_xml(maestro, maestro, [], {int(fila): {letra: motivo}}, col=col)
    migrar_columnas_motivo(maestro)
    subir_maestro(maestro)
    return col


# ============================================================
# REPORTE
# ============================================================
def escribir_reporte(calc, ruta):
    s = []
    s.append('=' * 60)
    s.append(f'REPORTE DE ACTUALIZACION DEL MAESTRO - {datetime.now():%d/%m/%Y %H:%M}')
    s.append('=' * 60)
    r = calc.resumen()
    s.append(f'Filas nuevas (AGENDADOS):        {r["filas_nuevas"]}')
    s.append(f'Ventas procesadas:               {r["ventas_totales"]}')
    s.append(f'  con match en maestro:          {r["matches"]} (exactas {r["matches_exactos"]}, sin fecha {r["matches_sin_fecha"]})')
    s.append(f'  pendientes (sin match):        {r["pendientes"]}')
    s.append(f'  a revisar:                     {r["revisar"]}')
    s.append(f'Filas del maestro a actualizar:  {r["filas_a_actualizar"]}')
    s.append(f'Celdas a escribir:               {r["celdas_a_escribir"]}')
    s.append(f'Incompletos en AGENDADOS (no cop.): {r["incompletas"]}')
    s.append('')

    c_ag = {'d': calc._ag('DIA'), 'm': calc._ag('MES'), 'a': calc._ag('ANIO'),
            'd2': calc._ag('DIA2'), 'm2': calc._ag('MES3'), 'a2': calc._ag('ANIO4'),
            'nom': calc._ag('NOMBRE'), 'tel': calc._ag('TELEFONO'),
            'camp': calc._ag('CAMPANA')}

    def fmt_ag(fila):
        return (f'{fila.get(c_ag["d"])}/{fila.get(c_ag["m"])}/{fila.get(c_ag["a"])} -> '
                f'cita {fila.get(c_ag["d2"])}/{fila.get(c_ag["m2"])}/{fila.get(c_ag["a2"])} | '
                f'{fila.get(c_ag["nom"]) or "(sin nombre)"} | {fila.get(c_ag["tel"])} | '
                f'{fila.get(c_ag["camp"])}')

    if calc.incompletas:
        s.append('--- AGENDADOS INCOMPLETOS (no copiados, revisar fuente) ---')
        for r_ag, fila, motivo in calc.incompletas[:60]:
            s.append(f'  [{motivo}] fila {r_ag}: {fmt_ag(fila)}')
        s.append('')

    if calc.new_rows:
        s.append('--- AGENDADOS NUEVOS ---')
        for r_ag, fila in calc.new_rows[:200]:
            s.append(f'  {fmt_ag(fila)}')

    s.append('')
    s.append('--- VENTAS PENDIENTES (sin coincidencia) ---')
    for v, motivo in calc.pendientes:
        s.append(f'  [{motivo}] {v["fila"]} {v["nombre"]} | {v["cel"]} | '
                 f'{v["dia"]}/{v["mes"]}/{v["anio"]} | {v["tratamiento"]} | S/{v["venta"]}')
    s.append('')
    s.append('--- MATCHES SIN FECHA EXACTA (revisar) ---')
    ws = calc.maestro
    c_nom = calc._idx('NOMBRE')
    c_fecha = [calc._idx(x) for x in ('DIA2', 'MES3', 'ANIO4')]
    for vf, mf, modo, hoja in calc.matches:
        if 'sin fecha' in modo:
            nombre_m = ws.cell(row=mf, column=c_nom).value if mf not in calc._nueva_info else '(fila nueva)'
            fecha_m = [ws.cell(row=mf, column=c).value for c in c_fecha] if mf not in calc._nueva_info else '(nueva)'
            s.append(f'  venta {vf} -> maestro fila {mf} | {nombre_m} | cita {fecha_m}')
    s.append('')
    s.append('--- CASOS A REVISAR ---')
    for v, motivo in calc.revisar:
        s.append(f'  [{motivo}] {v["fila"]} {v["nombre"]} | {v["cel"]} | '
                 f'{v["dia"]}/{v["mes"]}/{v["anio"]} | {v["tratamiento"]} | {v["status"]}')
    with open(ruta, 'w', encoding='utf-8') as f:
        f.write('\n'.join(s))
    return '\n'.join(s)


# ============================================================
# FLUJO PRINCIPAL (reutilizable desde CLI o web)
# ============================================================
def agendados_por_periodo(ag_path, anio, mes, desde, hasta):
    """Cuenta agendados directamente desde AGENDADOS para un periodo.

    Filtra por fecha de creación (DIA/MES/ANIO), no fecha de cita.
    Devuelve dict[(campana, crm)] -> ag, con crm='SIN CRM' (AGENDADOS no tiene CRM).
    """
    agendados, ag_col = leer_agendados(ag_path)
    if not ag_col:
        return {}
    c_dia = ag_col.get('DIA')
    c_mes = ag_col.get('MES')
    c_anio = ag_col.get('ANIO')
    c_camp = ag_col.get('CAMPANA')
    c_tel = ag_col.get('TELEFONO')
    c_nm = ag_col.get('NOMBRE')
    agg = defaultdict(lambda: {'ag': 0})
    for _r, fila in agendados:
        if (fila.get(c_anio) == anio
                and fila.get(c_mes) == mes
                and isinstance(fila.get(c_dia), (int, float))
                and desde <= int(fila.get(c_dia)) <= hasta):
            if fila.get(c_tel) or fila.get(c_nm):
                camp = str(fila.get(c_camp) or '').strip() or '(SIN CAMPANA)'
                agg[(camp, 'SIN CRM')]['ag'] += 1
    return dict(agg)


def ejecutar_sync(aplicar=True, sin_descarga=False):
    """Ejecuta el flujo completo de sincronización.

    Devuelve un dict con el resumen, el texto del reporte y los archivos usados.
    Si aplicar=True, respalda y escribe los cambios en el maestro.
    """
    resultado = {'ok': True, 'aplicar': aplicar, 'errores': [],
                 'backup': None, 'maestro': None,
                 'agendados': None, 'venta_diaria': None}
    try:
        if sin_descarga:
            ag = os.path.join(TMP_DIR, 'AGENDADOS.xlsx')
            ve = os.path.join(TMP_DIR, 'VENTA_DIARIA.xlsx')
        else:
            ag = descargar(AGENDADOS_FID, 'AGENDADOS', forzar=True)
            ve = descargar(VENTA_FID, 'VENTA_DIARIA', forzar=True)

        maestro = ruta_maestro_local(forzar=True)
        resultado['maestro'] = maestro
        maestro_ws = leer_maestro(maestro)
        col = detectar_maestro(maestro_ws)
        resultado['formato'] = ('Derma Essenza' if col and col.get('CAMPANA') and not col.get('CRM')
                                else 'BM' if col else 'desconocido')
        agendados, ag_col = leer_agendados(ag)
        resultado['formato_agendados'] = ('Derma Essenza' if ag_col and 'CAMPANA' in ag_col
                                          else 'BM' if ag_col else 'desconocido')
        venta = leer_venta(ve)

        calc = Calculo(maestro_ws, agendados, venta, col=col, ag_col=ag_col)
        texto = escribir_reporte(calc, os.path.join(TMP_DIR, 'reporte_actualizacion.txt'))
        resultado['reporte'] = texto
        resultado['resumen'] = calc.resumen()
        resultado['agendados'] = ag
        resultado['venta_diaria'] = ve

        if aplicar:
            if not col or not ag_col:
                resultado['aviso'] = (
                    'No se reconocieron las columnas del maestro o de AGENDADOS '
                    '(fila 4 sin NOMBRE/TELEFONO/DIA2). NO se aplicaron cambios.')
            else:
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup = maestro.replace('.xlsx', f'_backup_{ts}.xlsx')
                shutil.copy2(maestro, backup)
                resultado['backup'] = backup
                aplicar_xml(maestro, maestro, calc.new_rows, calc.updates,
                            col=col, ag_col=ag_col)
                migrar_columnas_motivo(maestro)
                subir_maestro(maestro)
    except Exception as e:  # noqa: BLE001
        resultado['ok'] = False
        resultado['errores'].append(str(e))
    return resultado


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--aplicar', action='store_true', help='Guardar los cambios en el maestro')
    ap.add_argument('--sin-descarga', action='store_true', help='Reutilizar archivos ya descargados')
    args = ap.parse_args()

    print('1) Descargando fuentes...')
    print('2) Leyendo maestro y fuentes...')
    print('3) Calculando cambios...')

    resultado = ejecutar_sync(aplicar=args.aplicar, sin_descarga=args.sin_descarga)

    if not resultado['ok']:
        print('\nERROR:', resultado['errores'])
        sys.exit(1)

    print(resultado['reporte'])

    if not args.aplicar:
        print('\n[DRY RUN] No se modificó el maestro. Usa --aplicar para guardar.')
    else:
        print('4) Aplicando cambios (con backup)...')
        print(f'  Backup: {resultado["backup"]}')
        print(f'  Maestro actualizado: {resultado.get("maestro") or MAESTRO}')


if __name__ == '__main__':
    main()
