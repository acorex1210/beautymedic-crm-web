#!/usr/bin/env python3
"""Reporte de ventas por campana de agendados en PDF (multi-pagina).

Mismos criterios que los pivotes del maestro BD DATA.xlsx:
  - Hoja AGENDADO : Agendados = filas con FECHA DE AGENDADO (DIA/MES/ANO, C/D/E)
                    dentro del periodo y con telefono.
  - Hoja ASISTIDO : Asistidos = filas con FECHA DE CITA (DIA2/MES3/ANO4, L/M/N)
                    dentro del periodo y ASISTENCIA = ASISTIO.
                    Compraron = de los asistidos, los que tienen PAGO TOTAL > 0.
                    Monto = suma de PAGO TOTAL.

Paginas del PDF:
  1. Resumen comercial   (KPIs, tasas, resumen por CRM)
  2. Flujo operativo     (embudo agendados -> venta y ticket promedio)
  3. Resumen de campanas (tablas por campana y por CRM)
  4. Campanas por canal  (graficas apiladas y detalle Kommo/WhatsApp/Organico)
  5. Metricas            (tratamientos, distritos, perfil del paciente)
  6. Hallazgos y sugerencias

Uso:
  python3 reporte_ventas_pdf.py                       # AGO 1-10, maestro directo
  python3 reporte_ventas_pdf.py --desde 11 --hasta 11 --fuente auto \
      --salida Reporte_Ventas_11_Agosto_2026.pdf
"""
import argparse
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import FancyBboxPatch
import numpy as np
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import alimentar_maestro as am
import meta_ads as mads

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# Tipografía: Playfair Display (titulares, cifras grandes) + Manrope
# (texto/tablas) — el mismo par editorial-premium que se usa en la web
# (fuentes propias, licencia OFL de Google Fonts, en fonts/). Se registran
# una vez al importar el módulo; si por algún motivo faltan los archivos
# (entorno nuevo sin el folder fonts/), se cae de vuelta a la fuente por
# defecto de matplotlib sin romper el reporte.
FONT_DIR = os.path.join(BASE_DIR, 'fonts')
FUENTE_SERIF = 'Playfair Display'
FUENTE_SANS = 'Manrope'
try:
    for _f in os.listdir(FONT_DIR):
        if _f.endswith('.ttf'):
            fm.fontManager.addfont(os.path.join(FONT_DIR, _f))
    # DejaVu Sans (la que trae matplotlib) como respaldo: Manrope no incluye
    # los glifos ▲▼▶ que se usan para las variaciones — matplotlib cae
    # automáticamente a la siguiente fuente de la lista si falta un glifo.
    plt.rcParams['font.family'] = [FUENTE_SANS, 'DejaVu Sans']
except (FileNotFoundError, OSError):
    FUENTE_SERIF = FUENTE_SANS = plt.rcParams['font.family'][0]
# Fuentes TrueType embebidas de verdad en el PDF (no Type 3): texto nítido,
# seleccionable y buscable en cualquier lector.
matplotlib.rcParams['pdf.fonttype'] = 42
SALIDA = os.path.join(BASE_DIR, 'Reporte_Ventas_1-10_Agosto_2026.pdf')
MES = 'AGO'
ANIO = 2026
D1, D2 = 1, 10
FUENTE = 'maestro'   # maestro = pivotes guardados | auto = integra AGENDADOS+VENTA (Drive)

# Paleta alineada a la identidad de marca de la web (navy + dorado + crema,
# ver templates/index.html :root) para que el PDF se sienta parte del mismo
# producto premium, no un reporte de Excel aparte.
FONDO_TITULO = '#16264d'    # navy del sidebar de la web
AZUL = '#2c5aa0'            # --azul
CELESTE = '#5aa0e8'         # --azul3
VERDE = '#38a543'           # --verde
NARANJA = '#c9a847'         # --dorado — acento premium (antes era naranja puro)
NARANJA_VIVO = '#e58e2d'    # --naranja original, para gráficas donde 3+ series
                            # necesitan distinguirse de VERDE y del dorado
GRIS = '#5b6a7e'            # --gris
ROJO = '#d44436'            # --rojo
FONDO_CLARO = '#faf3e0'     # tinte crema (--crema), reemplaza el celeste frío
GRIS_LINEA = '#e4e8f0'      # --borde
GRIS_PIE = '#93a0b5'
DORADO_CLARO = '#f3e6c2'    # fondo suave para acentos dorados (barras, chips)

# Alto (fracción de página) de una fila de tabla a fontsize=9 cuando se usa
# estilo_tabla(..., max_row_h=1.0), es decir cuando la tabla llena exactamente
# el axes que se le da (sin el auto-ajuste proporcional de estilo_tabla).
ALTO_FILA_TABLA = 0.032

CRM_ORDEN = ['KOMMO', 'WHATSAPP', 'ORGANICO', 'SIN CRM']
CANALES = ['KOMMO', 'WHATSAPP', 'ORGANICO']
COLOR_CANAL = {'KOMMO': AZUL, 'WHATSAPP': VERDE, 'ORGANICO': NARANJA_VIVO}
PALETA = [AZUL, VERDE, NARANJA_VIVO, CELESTE]

# ============================================================
# Sistema de diseño: encabezado/pie consistentes en todas las páginas
# ============================================================
# Márgenes horizontales comunes a todas las páginas (misma franja útil que
# usan las tablas y gráficas, para que todo quede alineado verticalmente).
MARGEN_IZQ, MARGEN_DER = 0.07, 0.93
ANCHO_UTIL = MARGEN_DER - MARGEN_IZQ
Y_LINEA_TITULO = 0.895   # línea divisoria bajo el título; el contenido de
                         # cada página empieza en o por debajo de esta cota
Y_LINEA_PIE = 0.045      # línea divisoria sobre el pie; el contenido de cada
                         # página debe terminar por encima de esta cota


def _marca_nombre():
    return os.environ.get('BRAND_NOMBRE', 'Derma Essenza')


def nueva_pagina(titulo, subtitulo=None, num=None, total=None):
    """Crea una página A4 en blanco con el encabezado (franja de marca +
    título + línea) y el pie (línea + numeración + confidencialidad) que
    comparten todas las páginas del reporte. Devuelve (fig, ax) — ``ax``
    cubre toda la página (transAxes = fracción de página) para que cada
    función de página siga agregando texto/formas sobre ella, y las demás
    gráficas/tablas se agregan con ``fig.add_axes`` como hasta ahora."""
    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor('white')
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.axis('off')

    # Franja superior de marca — versalitas con tracking (se simula con
    # espacios entre letras; matplotlib no soporta letter-spacing real) para
    # que se lea como una cabecera editorial, no una barra de título de Excel.
    ax.add_patch(plt.Rectangle((0, 0.978), 1, 0.022, facecolor=FONDO_TITULO,
                               edgecolor='none', transform=ax.transAxes, zorder=3))
    ax.text(MARGEN_IZQ, 0.989, ' '.join(_marca_nombre().upper()), ha='left', va='center',
            fontsize=7.5, color=NARANJA, fontweight='bold', family=FUENTE_SANS,
            transform=ax.transAxes, zorder=4)
    ax.text(MARGEN_DER, 0.989, f'{D1} al {D2} de {NOMBRES_MES.get(MES, MES)} {ANIO}',
            ha='right', va='center', fontsize=8, color='white', family=FUENTE_SANS,
            transform=ax.transAxes, zorder=4)

    # Título editorial: serif grande + filete dorado fino debajo (en vez del
    # tick de color a la izquierda) — más elegante que un bloque de color.
    ax.text(MARGEN_IZQ, 0.940, titulo, ha='left', va='center', fontsize=19,
            fontweight='bold', color=FONDO_TITULO, family=FUENTE_SERIF, transform=ax.transAxes)
    ax.plot([MARGEN_IZQ, MARGEN_IZQ + 0.052], [0.920, 0.920], color=NARANJA, linewidth=2,
            solid_capstyle='round', transform=ax.transAxes, zorder=4)
    if subtitulo:
        ax.text(MARGEN_IZQ, 0.906, subtitulo, ha='left', va='center', fontsize=9.5,
                color=GRIS, family=FUENTE_SANS, transform=ax.transAxes)
    ax.plot([MARGEN_IZQ, MARGEN_DER], [Y_LINEA_TITULO, Y_LINEA_TITULO], color=GRIS_LINEA,
            linewidth=0.8, transform=ax.transAxes)

    # Pie: línea + confidencialidad + numeración
    ax.plot([MARGEN_IZQ, MARGEN_DER], [Y_LINEA_PIE, Y_LINEA_PIE], color=GRIS_LINEA,
            linewidth=0.8, transform=ax.transAxes)
    ax.text(MARGEN_IZQ, 0.03, f'{_marca_nombre()} · documento interno, uso confidencial',
            ha='left', va='center', fontsize=7, color=GRIS_PIE, family=FUENTE_SANS,
            transform=ax.transAxes)
    if num is not None and total is not None:
        ax.text(MARGEN_DER, 0.03, f'Página {num} de {total}', ha='right', va='center',
                fontsize=7, color=GRIS_PIE, family=FUENTE_SANS, transform=ax.transAxes)
    return fig, ax


def canales_desde_datos(agg):
    """Canales reales presentes en los datos (top 3 por agendados), sin SIN CRM."""
    totales = Counter()
    for (_, crm), d in agg.items():
        totales[crm] += d['ag']
    top = [c for c, _ in totales.most_common(3) if str(c).strip().upper() != 'SIN CRM']
    return top or ['WHATSAPP']


# ============================================================
# Campañas reales vs "Otros" (no son campañas de anuncios)
# ============================================================
# Valores que puede tener la columna CAMPAÑA del maestro sin ser una campaña
# de verdad: tipos de venta que se registran por costumbre en esa celda, no
# nombres de campaña de Meta Ads.
CATEGORIAS_OTROS = {
    'EVALUACION', 'RETOQUE', 'RECURRENTE', 'RECOMENDADO', 'SESION',
    'ORGANICO REDES', 'VENTA SIN AGENDAR',
}
_RE_HORA = re.compile(r'^\d{1,2}[:.]\d{2}\s*(AM|PM)?$', re.I)


def _normalizar_categoria(s):
    s = unicodedata.normalize('NFKD', str(s or '').upper())
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^A-Z0-9]+', ' ', s).strip()


def campana_valida(nombre):
    """False si ``nombre`` no es una campaña real: vacío, un tipo de venta
    (evaluación, retoque, recurrente...) o un dato corrupto (p.ej. una hora
    tipeada por error en la columna CAMPAÑA)."""
    s = str(nombre or '').strip()
    if not s or s == '(SIN CAMPANA)':
        return False
    if _RE_HORA.match(s):
        return False
    return _normalizar_categoria(s) not in CATEGORIAS_OTROS


def categoria_otros(nombre):
    """Nombre de categoría a mostrar en la tabla 'Otros' para un valor de
    CAMPAÑA que no pasó ``campana_valida``."""
    s = str(nombre or '').strip()
    if not s or s == '(SIN CAMPANA)':
        return 'SIN CAMPAÑA'
    if _RE_HORA.match(s):
        return 'SIN CAMPAÑA (dato inválido)'
    return s


def separar_campanas_otros(agg):
    """Divide un agg {(campana, crm): datos} en (campanas_reales, otros),
    agrupando en 'otros' lo que no es una campaña de anuncios (ver
    ``campana_valida``) para que no ensucie las tablas de campañas."""
    campanas, otros = {}, {}
    for (camp, crm), d in agg.items():
        if campana_valida(camp):
            campanas[(camp, crm)] = d
        else:
            cat = categoria_otros(camp)
            dest = otros.setdefault((cat, crm), dict(ag=0, as_=0, co=0, mon=0.0,
                                                      no_fueron=0, fueron_sin_compra=0))
            for k in dest:
                dest[k] += d.get(k, 0)
    return campanas, otros


COL_BM = {
    'CRM': 2, 'RED_SOCIAL': 7, 'DIA': 3, 'MES': 4, 'ANIO': 5, 'TELEFONO': 9,
    'DIA2': 12, 'MES3': 13, 'ANIO4': 14, 'CAMPANA': 15, 'ASISTENCIA': 16,
    'DISTRITO': 17, 'EDAD': 18, 'SEXO': 19,
    'TRAT': [20, 22, 24, 26], 'PAGO': [21, 23, 25, 27], 'PAGO_TOTAL': 28,
}
_CABECERA = {
    'DIA': 'DIA', 'MES': 'MES', 'AÑO': 'ANIO', 'TELEFONO': 'TELEFONO',
    'DIA2': 'DIA2', 'MES3': 'MES3', 'AÑO4': 'ANIO4', 'CAMPAÑA': 'CAMPANA',
    'ASISTENCIA': 'ASISTENCIA', 'DISTRITO': 'DISTRITO', 'EDAD': 'EDAD',
    'SEXO': 'SEXO', 'CRM': 'CRM', 'RED SOCIAL': 'RED_SOCIAL',
    'PAGO TOTAL': 'PAGO_TOTAL', 'DNI': 'DNI', 'NOMBRE': 'NOMBRE',
    'CORREO': 'CORREO',
    'AGENDADO POR': 'AGENDADO',
    'MOTIVO NO ASISTIO': 'MOTIVO_NO_ASISTIO',
    'MOTIVO NO COMPRA': 'MOTIVO_NO_COMPRA',
}


def detectar_columnas(ws):
    """Ubica las columnas del maestro por su cabecera (fila 4), tolerando el
    formato BM (con columna CRM) y el Derma Essenza (sin CRM). Si una columna
    no se detecta, usa la posición del formato BM."""
    col = dict(COL_BM)
    hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    pos = {}
    for i, h in enumerate(hdr, 1):
        if h is None:
            continue
        hh = str(h).strip().upper()
        if hh in _CABECERA:
            pos[_CABECERA[hh]] = i
        elif hh.startswith('TRAT '):
            pos.setdefault('TRAT', []).append((int(hh[5:]), i))
        elif hh.startswith('PAGO '):
            pos.setdefault('PAGO', []).append((int(hh[5:]), i))
    for k, v in pos.items():
        col[k] = [i for _, i in sorted(v)] if isinstance(v, list) else v
    col['ES_BM'] = 'CRM' in pos
    col['CANAL'] = pos.get('CRM') or pos.get('RED_SOCIAL') or COL_BM['CRM']
    return col


COL = dict(COL_BM)

NOMBRES_MES = {'ENE': 'Enero', 'FEB': 'Febrero', 'MAR': 'Marzo', 'ABR': 'Abril',
               'MAY': 'Mayo', 'JUN': 'Junio', 'JUL': 'Julio', 'AGO': 'Agosto',
               'SET': 'Septiembre', 'OCT': 'Octubre', 'NOV': 'Noviembre', 'DIC': 'Diciembre'}

# Derma Essenza opera desde julio 2026: el maestro BD DATA.xlsx es una hoja
# reutilizada de un negocio anterior (Beauty Medic) que arrastra filas suyas
# con fecha anterior — no son datos reales de Derma Essenza. generar_reporte()
# rechaza pedir un reporte para un periodo anterior a esto (ver también
# analitica._INICIO_OPERACION, que filtra lo mismo para las páginas de
# histórico/proyección/recurrencia).
#
# Además, dentro del periodo real de Derma Essenza, cualquier fila con canal
# CRM (KOMMO/WHATSAPP/ORGANICO) tampoco es de Derma Essenza — confirmado con
# Andre: ninguna fila con ese campo tipificado corresponde a un paciente
# real (ninguna tiene ASISTENCIA = ASISTIO), y la columna CRM en sí es un
# resto del formato de Beauty Medic. Derma Essenza siempre deja ese campo
# vacío ("SIN CRM"). _fila_real_derma() descarta ambos casos a la vez.
INICIO_OPERACION = (2026, 7)  # (año, mes)


def _antes_de_inicio(anio, mes_txt):
    """True si (anio, mes_txt) es anterior al inicio real de operación de
    Derma Essenza — mismo criterio que analitica._antes_de_inicio."""
    if not anio or not mes_txt:
        return False
    mes_txt = str(mes_txt).strip().upper()
    if mes_txt == 'SEP':
        mes_txt = 'SET'
    orden_mes = list(NOMBRES_MES)
    if mes_txt not in orden_mes:
        return False
    try:
        return (int(anio), orden_mes.index(mes_txt) + 1) < INICIO_OPERACION
    except (TypeError, ValueError):
        return False


def _fila_real_derma(ws, r):
    """True si la fila ``r`` del maestro es un dato real de Derma Essenza.

    Importante: el filtro de canal CRM (KOMMO/WHATSAPP/ORGANICO) sólo aplica
    cuando el maestro está en formato "BM" (``COL['ES_BM']`` — tiene una
    columna CRM literal, herencia de Beauty Medic, el negocio anterior que
    usó esa hoja). El maestro real de Derma Essenza (BD DATA DERMA ESSENZA)
    NO tiene columna CRM — ahí ``COL['CANAL']`` cae a RED SOCIAL, que es un
    campo propio y legítimo (no debe excluirse)."""
    if COL.get('ES_BM') and ws.cell(row=r, column=COL['CANAL']).value:
        return False
    if (_antes_de_inicio(ws.cell(row=r, column=COL['ANIO']).value,
                         ws.cell(row=r, column=COL['MES']).value)
            or _antes_de_inicio(ws.cell(row=r, column=COL['ANIO4']).value,
                               ws.cell(row=r, column=COL['MES3']).value)):
        return False
    return True


def en_periodo(dia):
    return isinstance(dia, (int, float)) and D1 <= int(dia) <= D2


def pago_total(ws, r):
    p = ws.cell(row=r, column=COL['PAGO_TOTAL']).value
    s = sum(x for x in (ws.cell(row=r, column=c).value for c in COL['PAGO'])
            if isinstance(x, (int, float)))
    if isinstance(p, (int, float)) and p > 0:
        return p
    return s


def _cita_pasada(ws, r):
    """True si la fecha de cita de la fila ya ocurrió (comparada con hoy)."""
    try:
        anio = int(ws.cell(row=r, column=COL['ANIO4']).value)
        mes = ws.cell(row=r, column=COL['MES3']).value
        dia = int(ws.cell(row=r, column=COL['DIA2']).value)
        if not mes:
            return False
        mes = str(mes).strip().upper()
        if mes == 'SEP':
            mes = 'SET'
        if mes not in NOMBRES_MES:
            return False
        hoy = datetime.now()
        mi = list(NOMBRES_MES).index(mes) + 1
        return (anio, mi, dia) < (hoy.year, hoy.month, hoy.day)
    except (TypeError, ValueError):
        return False


def monto(sol):
    return f"S/ {sol:,.0f}" if sol else "S/ 0"


def _campana_origen_por_telefono(ag_path):
    """Teléfono -> campaña de su PRIMERA cita en todo el historial de
    AGENDADOS (no solo el periodo del reporte). Un cliente reagendado para
    retoque/evaluación bajo otra "campaña" sigue contando en la que lo trajo,
    igual que en el cuadre manual."""
    agendados, ag_col = am.leer_agendados(ag_path)
    if not ag_col:
        return {}
    orden_mes = list(NOMBRES_MES)

    def clave(fc):
        dia, mes, anio = fc
        return (anio, orden_mes.index(mes) if mes in orden_mes else 99, dia)

    por_tel = defaultdict(list)
    for _r, f in agendados:
        tel = am.norm_phone(f.get(ag_col['TELEFONO']))
        camp = str(f.get(ag_col['CAMPANA']) or '').strip()
        fc = am.norm_fecha(f.get(ag_col['DIA2']), f.get(ag_col['MES3']), f.get(ag_col['ANIO4']))
        if tel and camp and fc:
            por_tel[tel].append((fc, camp))
    return {tel: min(opts, key=lambda x: clave(x[0]))[1] for tel, opts in por_tel.items()}


def build_data():
    global COL
    ag_path = os.path.join(am.TMP_DIR, 'AGENDADOS.xlsx')
    if FUENTE == 'auto':
        maestro_ws = am.leer_maestro(am.ruta_maestro_local())
        COL = detectar_columnas(maestro_ws)
        agendados, ag_col = am.leer_agendados(ag_path)
        venta = am.leer_venta(os.path.join(am.TMP_DIR, 'VENTA_DIARIA.xlsx'))
        col = am.detectar_maestro(maestro_ws)
        calc = am.Calculo(maestro_ws, agendados, venta, col=col, ag_col=ag_col)
        tmp = os.path.join(am.TMP_DIR, 'simulado_reporte.xlsx')
        am.aplicar_xml(am.ruta_maestro_local(), tmp, calc.new_rows, calc.updates,
                       col=col, ag_col=ag_col)
        ws = openpyxl.load_workbook(tmp, data_only=True)['BD DATA']
    else:
        ws = am.leer_maestro(am.ruta_maestro_local())
        COL = detectar_columnas(ws)

    agg = defaultdict(lambda: dict(ag=0, as_=0, co=0, mon=0.0,
                                   no_fueron=0, fueron_sin_compra=0))

    # Contar agendados desde AGENDADOS actual (no del maestro acumulativo)
    ag_counts = am.agendados_por_periodo(ag_path, ANIO, MES, D1, D2)
    for (camp, crm), counts in ag_counts.items():
        agg[(camp, crm)]['ag'] = counts['ag']

    # Asistidos y monto: desde el maestro (fuente de verdad para VENTA), con
    # la campaña de origen (primera cita) del cliente por teléfono, para que
    # una venta de un reagendado no se pierda en la campaña del día de venta.
    origen_por_telefono = _campana_origen_por_telefono(ag_path)
    for r in range(5, ws.max_row + 1):
        if not _fila_real_derma(ws, r):
            continue
        crm = ws.cell(row=r, column=COL['CANAL']).value or 'SIN CRM'
        tel = am.norm_phone(ws.cell(row=r, column=COL['TELEFONO']).value)
        camp_maestro = str(ws.cell(row=r, column=COL['CAMPANA']).value or '').strip()
        camp_ag_key = origen_por_telefono.get(tel) or (camp_maestro if camp_maestro else '(SIN CAMPANA)')
        d = agg[(camp_ag_key, crm)]
        if (ws.cell(row=r, column=COL['ANIO4']).value == ANIO
                and ws.cell(row=r, column=COL['MES3']).value == MES
                and en_periodo(ws.cell(row=r, column=COL['DIA2']).value)):
            asist = str(ws.cell(row=r, column=COL['ASISTENCIA']).value or '').strip()
            if asist == 'ASISTIO':
                d['as_'] += 1
                p = pago_total(ws, r)
                d['mon'] += p
                if p > 0:
                    d['co'] += 1
                else:
                    d['fueron_sin_compra'] += 1
            elif _cita_pasada(ws, r):
                d['no_fueron'] += 1
    return agg


def datos_analiticos():
    """Metricas de perfil de los asistentes del periodo."""
    global COL
    ws = am.leer_maestro(am.ruta_maestro_local())
    COL = detectar_columnas(ws)
    trat = Counter(); dist = Counter(); sexo = Counter()
    edades = []; montos = []; crm = Counter(); camp = Counter()
    for r in range(5, ws.max_row + 1):
        if (ws.cell(row=r, column=COL['ANIO4']).value == ANIO
                and ws.cell(row=r, column=COL['MES3']).value == MES
                and en_periodo(ws.cell(row=r, column=COL['DIA2']).value)
                and ws.cell(row=r, column=COL['ASISTENCIA']).value == 'ASISTIO'
                and _fila_real_derma(ws, r)):
            crm[str(ws.cell(row=r, column=COL['CANAL']).value or 'SIN CRM')] += 1
            camp[str(ws.cell(row=r, column=COL['CAMPANA']).value or '').strip() or '(SIN)'] += 1
            for cc in COL['TRAT']:
                t = ws.cell(row=r, column=cc).value
                if t:
                    trat[str(t).strip().upper()] += 1
            d = ws.cell(row=r, column=COL['DISTRITO']).value
            if d:
                dist[str(d).strip().title()] += 1
            e = ws.cell(row=r, column=COL['EDAD']).value
            if isinstance(e, (int, float)):
                edades.append(e)
            s = ws.cell(row=r, column=COL['SEXO']).value
            if s:
                sexo[str(s).strip().upper()] += 1
            p = pago_total(ws, r)
            if p > 0:
                montos.append(p)
    return {'trat': trat, 'dist': dist, 'sexo': sexo, 'edades': edades,
            'montos': montos, 'crm': crm, 'camp': camp}


def datos_ejecutivas():
    """Ranking por ejecutiva (AGENDADO POR) del periodo, desde el maestro."""
    global COL
    ag_path = os.path.join(am.TMP_DIR, 'AGENDADOS.xlsx')
    ws = am.leer_maestro(am.ruta_maestro_local())
    COL = detectar_columnas(ws)
    c_ag = COL.get('AGENDADO')
    if not c_ag:
        return []

    # Claves de AGENDADOS para filtrar agendados y no_fueron
    ag_keys = set()
    agendados_raw, ag_col = am.leer_agendados(ag_path)
    if agendados_raw and ag_col:
        c_ph = ag_col.get('TELEFONO')
        c_nm = ag_col.get('NOMBRE')
        c_d2 = ag_col.get('DIA2')
        c_m3 = ag_col.get('MES3')
        c_a4 = ag_col.get('ANIO4')
        for _r, fila in agendados_raw:
            ph = am.norm_phone(fila.get(c_ph))
            nm = am.norm_name(fila.get(c_nm))
            fc = am.norm_fecha(fila.get(c_d2), fila.get(c_m3), fila.get(c_a4))
            ag_keys.add((ph, fc))
            if nm:
                ag_keys.add((nm, fc))

    agrup = {}
    for r in range(5, ws.max_row + 1):
        if (ws.cell(row=r, column=COL['ANIO4']).value == ANIO
                and ws.cell(row=r, column=COL['MES3']).value == MES
                and en_periodo(ws.cell(row=r, column=COL['DIA2']).value)
                and _fila_real_derma(ws, r)):
            ej = str(ws.cell(row=r, column=c_ag).value or '').strip() or 'SIN EJECUTIVA'
            d = agrup.setdefault(ej, {'ag': 0, 'as_': 0, 'co': 0, 'mon': 0.0,
                                      'no_fueron': 0, 'fueron_sin_compra': 0})
            asist = str(ws.cell(row=r, column=COL['ASISTENCIA']).value or '').strip()
            if asist == 'ASISTIO':
                d['as_'] += 1
                p = pago_total(ws, r)
                d['mon'] += p
                if p > 0:
                    d['co'] += 1
                else:
                    d['fueron_sin_compra'] += 1
            elif _cita_pasada(ws, r):
                d['no_fueron'] += 1
    out = []
    for ej, d in agrup.items():
        out.append({'ejecutiva': ej, 'ag': d['ag'], 'as_': d['as_'], 'co': d['co'],
                    'mon': round(d['mon'], 2), 'no_fueron': d['no_fueron'],
                    'fueron_sin_compra': d['fueron_sin_compra'],
                    'asistencia_pct': pct(d['as_'], d['ag']),
                    'conversion_pct': pct(d['co'], d['as_']),
                    'ticket': d['mon'] / d['co'] if d['co'] else 0})
    out.sort(key=lambda x: (x['mon'], x['ag']), reverse=True)
    return out


def datos_motivos():
    """Distribución de motivos de no asistencia y no compra del periodo."""
    global COL
    ag_path = os.path.join(am.TMP_DIR, 'AGENDADOS.xlsx')
    ws = am.leer_maestro(am.ruta_maestro_local())
    COL = detectar_columnas(ws)
    c_m_as = COL.get('MOTIVO_NO_ASISTIO')
    c_m_co = COL.get('MOTIVO_NO_COMPRA')

    # Claves de AGENDADOS para filtrar no_asistio
    ag_keys = set()
    agendados_raw, ag_col = am.leer_agendados(ag_path)
    if agendados_raw and ag_col:
        c_ph = ag_col.get('TELEFONO')
        c_nm = ag_col.get('NOMBRE')
        c_d2 = ag_col.get('DIA2')
        c_m3 = ag_col.get('MES3')
        c_a4 = ag_col.get('ANIO4')
        for _r, fila in agendados_raw:
            ph = am.norm_phone(fila.get(c_ph))
            nm = am.norm_name(fila.get(c_nm))
            fc = am.norm_fecha(fila.get(c_d2), fila.get(c_m3), fila.get(c_a4))
            ag_keys.add((ph, fc))
            if nm:
                ag_keys.add((nm, fc))

    no_asistio = Counter()
    no_compra = Counter()
    for r in range(5, ws.max_row + 1):
        if not _fila_real_derma(ws, r):
            continue
        asist = str(ws.cell(row=r, column=COL['ASISTENCIA']).value or '').strip()
        if (ag_keys and asist != 'ASISTIO' and _cita_pasada(ws, r) and c_m_as):
            ph = am.norm_phone(ws.cell(row=r, column=COL['TELEFONO']).value)
            nm = am.norm_name(ws.cell(row=r, column=COL['NOMBRE']).value)
            fc = am.norm_fecha(ws.cell(row=r, column=COL['DIA2']).value,
                               ws.cell(row=r, column=COL['MES3']).value,
                               ws.cell(row=r, column=COL['ANIO4']).value)
            if (ph, fc) in ag_keys or (nm, fc) in ag_keys:
                m = str(ws.cell(row=r, column=c_m_as).value or '').strip()
                if m:
                    no_asistio[m.upper()] += 1
        if (asist == 'ASISTIO' and pago_total(ws, r) == 0 and c_m_co
                and ws.cell(row=r, column=COL['ANIO4']).value == ANIO
                and ws.cell(row=r, column=COL['MES3']).value == MES
                and en_periodo(ws.cell(row=r, column=COL['DIA2']).value)):
            m = str(ws.cell(row=r, column=c_m_co).value or '').strip()
            if m:
                no_compra[m.upper()] += 1
    return {'no_asistio': dict(no_asistio.most_common()),
            'no_compra': dict(no_compra.most_common())}


def datos_dia_semana():
    """Monto vendido y N° de ventas por día de la semana dentro del periodo
    del reporte — un patrón que no se ve a simple vista revisando el Excel
    fila por fila."""
    global COL
    ws = am.leer_maestro(am.ruta_maestro_local())
    COL = detectar_columnas(ws)
    dias_es = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    orden_mes = list(NOMBRES_MES)
    monto_dia = [0.0] * 7
    ventas_dia = [0] * 7
    for r in range(5, ws.max_row + 1):
        if not (ws.cell(row=r, column=COL['ANIO4']).value == ANIO
                and ws.cell(row=r, column=COL['MES3']).value == MES
                and en_periodo(ws.cell(row=r, column=COL['DIA2']).value)
                and str(ws.cell(row=r, column=COL['ASISTENCIA']).value or '').strip() == 'ASISTIO'
                and _fila_real_derma(ws, r)):
            continue
        p = pago_total(ws, r)
        if p <= 0:
            continue
        dia = ws.cell(row=r, column=COL['DIA2']).value
        mes_txt = str(ws.cell(row=r, column=COL['MES3']).value or '').strip().upper()
        if mes_txt not in orden_mes or not isinstance(dia, (int, float)):
            continue
        try:
            wd = datetime(ANIO, orden_mes.index(mes_txt) + 1, int(dia)).weekday()
        except ValueError:
            continue
        monto_dia[wd] += p
        ventas_dia[wd] += 1
    return {'dias': dias_es, 'monto': monto_dia, 'ventas': ventas_dia}


def datos_pareto():
    """Concentración de ingresos por paciente dentro del periodo: qué % de
    los pacientes que compraron generó qué % del monto total — para ver si
    el mes depende de pocos clientes grandes o está bien repartido."""
    global COL
    ws = am.leer_maestro(am.ruta_maestro_local())
    COL = detectar_columnas(ws)
    por_paciente = defaultdict(float)
    for r in range(5, ws.max_row + 1):
        if not (ws.cell(row=r, column=COL['ANIO4']).value == ANIO
                and ws.cell(row=r, column=COL['MES3']).value == MES
                and en_periodo(ws.cell(row=r, column=COL['DIA2']).value)
                and str(ws.cell(row=r, column=COL['ASISTENCIA']).value or '').strip() == 'ASISTIO'
                and _fila_real_derma(ws, r)):
            continue
        p = pago_total(ws, r)
        if p <= 0:
            continue
        tel = am.norm_phone(ws.cell(row=r, column=COL['TELEFONO']).value)
        clave = tel or str(ws.cell(row=r, column=COL['NOMBRE']).value or '').strip()
        if not clave:
            continue
        por_paciente[clave] += p
    montos = sorted(por_paciente.values(), reverse=True)
    n = len(montos)
    total = sum(montos)
    if not n or total <= 0:
        return {'n_pacientes': 0, 'total': 0.0, 'top20_n': 0, 'top20_pct_monto': 0.0}
    corte = max(1, round(n * 0.2))
    return {'n_pacientes': n, 'total': round(total, 2), 'top20_n': corte,
            'top20_pct_monto': round(100 * sum(montos[:corte]) / total, 1)}


def _fecha_valida(dia, mes_txt, anio):
    """Construye una fecha a partir de columnas del maestro (DIA/MES/AÑO o
    DIA2/MES3/AÑO4). None si algún dato falta o es inválido."""
    try:
        if not (dia and mes_txt and anio):
            return None
        mes_txt = str(mes_txt).strip().upper()
        if mes_txt == 'SEP':
            mes_txt = 'SET'
        if mes_txt not in NOMBRES_MES:
            return None
        mi = list(NOMBRES_MES).index(mes_txt) + 1
        return datetime(int(anio), mi, int(dia)).date()
    except (TypeError, ValueError):
        return None


BUCKETS_ANTICIPACION = [(0, 2, '0-2 días'), (3, 6, '3-6 días'), (7, 14, '7-14 días'),
                        (15, 10**9, '15+ días')]


def datos_anticipacion():
    """Relación entre la anticipación con la que se agenda una cita (días
    entre FECHA DE AGENDADO y FECHA DE CITA) y la tasa de asistencia real,
    para las citas del periodo del reporte (mismo periodo que el resto del
    informe, para que las cifras sean consistentes entre páginas).

    Nota: el histórico completo del maestro sólo tiene la columna ASISTENCIA
    tipificada de forma confiable desde hace pocos meses (meses anteriores
    aparecen casi enteramente en blanco), así que un análisis "de todo el
    histórico" daría porcentajes falsos por datos faltantes, no por que la
    gente no haya asistido. Por eso esto usa el periodo del reporte, igual
    que el resto de páginas — y con periodos muy cortos la muestra por
    bucket puede ser chica (se informa el total, no sólo el %)."""
    global COL
    ws = am.leer_maestro(am.ruta_maestro_local())
    COL = detectar_columnas(ws)
    conteo = {nombre: {'total': 0, 'asistio': 0} for _, _, nombre in BUCKETS_ANTICIPACION}
    for r in range(5, ws.max_row + 1):
        if not (ws.cell(row=r, column=COL['ANIO4']).value == ANIO
                and ws.cell(row=r, column=COL['MES3']).value == MES
                and en_periodo(ws.cell(row=r, column=COL['DIA2']).value)
                and _fila_real_derma(ws, r)):
            continue
        if not _cita_pasada(ws, r):
            continue  # todavía no ocurre: no se sabe si va a venir
        fecha_cita = _fecha_valida(ws.cell(row=r, column=COL['DIA2']).value,
                                   ws.cell(row=r, column=COL['MES3']).value,
                                   ws.cell(row=r, column=COL['ANIO4']).value)
        fecha_ag = _fecha_valida(ws.cell(row=r, column=COL['DIA']).value,
                                 ws.cell(row=r, column=COL['MES']).value,
                                 ws.cell(row=r, column=COL['ANIO']).value)
        if not fecha_cita or not fecha_ag:
            continue
        lead = (fecha_cita - fecha_ag).days
        if lead < 0:
            continue
        asist = str(ws.cell(row=r, column=COL['ASISTENCIA']).value or '').strip() == 'ASISTIO'
        for dmin, dmax, nombre in BUCKETS_ANTICIPACION:
            if dmin <= lead <= dmax:
                conteo[nombre]['total'] += 1
                if asist:
                    conteo[nombre]['asistio'] += 1
                break
    out = []
    for _, _, nombre in BUCKETS_ANTICIPACION:
        d = conteo[nombre]
        out.append({'bucket': nombre, 'total': d['total'], 'asistio': d['asistio'],
                    'pct': round(pct(d['asistio'], d['total']), 1) if d['total'] else None})
    return out


def _periodo_anterior(mes, anio, desde, hasta):
    """Mes previo con el mismo rango de días (variación comparable)."""
    orden = list(NOMBRES_MES)
    i = orden.index(mes)
    if i == 0:
        return orden[-1], anio - 1, desde, hasta
    return orden[i - 1], anio, desde, hasta


def _variacion(actual, anterior):
    if anterior is None or anterior == 0:
        return None
    return round((actual - anterior) / anterior * 100.0, 1)


def totales_periodo(mes, anio, desde, hasta):
    """Totales del periodo sobre el maestro actual (sin tocar globales)."""
    ag_path = os.path.join(am.TMP_DIR, 'AGENDADOS.xlsx')
    ws = am.leer_maestro(am.ruta_maestro_local())
    col = detectar_columnas(ws)
    tot = {'ag': 0, 'as_': 0, 'co': 0, 'mon': 0.0,
           'no_fueron': 0, 'fueron_sin_compra': 0}

    # Claves de AGENDADOS para filtrar agendados y no_fueron
    ag_keys = set()
    agendados_raw, ag_col = am.leer_agendados(ag_path)
    if agendados_raw and ag_col:
        c_ph = ag_col.get('TELEFONO')
        c_nm = ag_col.get('NOMBRE')
        c_d2 = ag_col.get('DIA2')
        c_m3 = ag_col.get('MES3')
        c_a4 = ag_col.get('ANIO4')
        for _r, fila in agendados_raw:
            ph = am.norm_phone(fila.get(c_ph))
            nm = am.norm_name(fila.get(c_nm))
            fc = am.norm_fecha(fila.get(c_d2), fila.get(c_m3), fila.get(c_a4))
            ag_keys.add((ph, fc))
            if nm:
                ag_keys.add((nm, fc))

    for r in range(5, ws.max_row + 1):
        if (ws.cell(row=r, column=col['ANIO4']).value == anio
                and ws.cell(row=r, column=col['MES3']).value == mes
                and isinstance(ws.cell(row=r, column=col['DIA2']).value, (int, float))
                and desde <= int(ws.cell(row=r, column=col['DIA2']).value) <= hasta
                and not (col.get('ES_BM') and ws.cell(row=r, column=col['CANAL']).value)
                and not _antes_de_inicio(ws.cell(row=r, column=col['ANIO']).value,
                                        ws.cell(row=r, column=col['MES']).value)
                and not _antes_de_inicio(ws.cell(row=r, column=col['ANIO4']).value,
                                        ws.cell(row=r, column=col['MES3']).value)):
            asist = str(ws.cell(row=r, column=col['ASISTENCIA']).value or '').strip()
            if asist == 'ASISTIO':
                tot['as_'] += 1
                p = pago_total(ws, r)
                tot['mon'] += p
                if p > 0:
                    tot['co'] += 1
                else:
                    tot['fueron_sin_compra'] += 1
            elif _cita_pasada(ws, r):
                tot['no_fueron'] += 1
    return tot


def pct(a, b):
    return 100.0 * a / b if b else 0.0


# ============================================================
# Integración Meta Ads + maestro (tabla por campaña)
# ============================================================
META_A_MAESTRO = {
    'TOXINA 2026': 'TOXINA FULL FACE',
    'CONSULTA GRATIS 2026': 'CONSULTA GRATUITA',
    'ACIDO HIALURONICO': 'ACIDO HIALURONICO',
}


def _ruta_meta_dir():
    base = os.environ.get('DATA_DIR', os.path.join(BASE_DIR, 'data'))
    return os.path.join(base, 'meta_ads')


def _norm_tokens(s):
    s = unicodedata.normalize('NFKD', str(s).upper())
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return set(t for t in s.split() if len(t) >= 3)


def _fuzzy_maestro(meta_nombre, camp_maestros):
    """Encuentra la campaña del maestro más parecida por tokens compartidos."""
    mt = _norm_tokens(meta_nombre) - {'2026'}
    if not mt:
        return None
    best, best_n = None, 0
    for cm in camp_maestros:
        n = len(mt & _norm_tokens(cm))
        if n > best_n:
            best, best_n = cm, n
    return best if best_n > 0 else None


def _cargar_meta_campanas():
    """Por campaña de la carga de Meta Ads más reciente. None si no hay cargas."""
    try:
        idx = mads.listar(_ruta_meta_dir())
    except Exception:
        return None
    if not idx:
        return None
    try:
        det = mads.detalle(_ruta_meta_dir(), idx[0]['id'])
    except Exception:
        return None
    return det.get('por_campania') or []


def _fila_meta(nombre, mc, d, maes=None):
    gasto = mc.get('gasto') or 0
    leads = mc.get('resultados') or 0
    costo_res = mc.get('costo_resultado')
    if costo_res is None:
        costo_res = round(gasto / leads, 2) if leads else 0
    ag, as_, co, mon = d['ag'], d['as_'], d['co'], d['mon']
    return dict(campania=nombre, campania_maestro=maes, gasto=gasto, leads=leads,
                costo_res=costo_res, ag=ag, pct_ag=pct(ag, leads), as_=as_,
                pct_as=pct(as_, ag), co=co, pct_co=pct(co, as_), mon=mon,
                ticket=mon / co if co else 0, organica=False)


def _gasto_ads_total():
    por_meta = _cargar_meta_campanas()
    if not por_meta:
        return None
    return round(sum((mc.get('gasto') or 0) for mc in por_meta), 2)


def build_campana_meta(agg):
    """Une cada campaña de Meta Ads con su embudo del maestro (ag→as→co→monto)."""
    camps = defaultdict(lambda: dict(ag=0, as_=0, co=0, mon=0.0))
    for (camp, crm), d in agg.items():
        for k in ('ag', 'as_', 'co', 'mon'):
            camps[camp][k] += d[k]
    cero = dict(ag=0, as_=0, co=0, mon=0.0)
    filas = []
    usadas = set()
    por_meta = _cargar_meta_campanas()
    if por_meta is not None:
        for mc in por_meta:
            nombre = mc.get('campania') or '—'
            maes = META_A_MAESTRO.get(nombre) or _fuzzy_maestro(nombre, list(camps))
            d = camps.get(maes, cero) if maes else cero
            if maes:
                usadas.add(maes)
            filas.append(_fila_meta(nombre, mc, d, maes=maes))
        filas.sort(key=lambda x: -x['gasto'])
    for camp, d in sorted(camps.items(), key=lambda kv: -kv[1]['ag']):
        if camp in usadas or (not d['ag'] and not d['as_']):
            continue
        filas.append(dict(campania=camp + ' (organica)', campania_maestro=camp,
                          gasto=0, leads=0,
                          costo_res=0, ag=d['ag'], pct_ag=0,
                          as_=d['as_'], pct_as=pct(d['as_'], d['ag']),
                          co=d['co'], pct_co=pct(d['co'], d['as_']), mon=d['mon'],
                          ticket=d['mon'] / d['co'] if d['co'] else 0,
                          organica=True))
    return filas


def estilo_tabla(ax, datos, col_w, header=None, color_titulo=FONDO_TITULO,
                 fontsize=8.5, scale=1.7, alinear_izq_col0=True, altura_fila=None,
                 max_row_h=None):
    """scale ahora limita la altura de fila (fraccion de la altura del eje);
    la tabla siempre llena el eje en horizontal, y en vertical hasta ese
    tope, quedando pegada arriba en vez de centrada con un hueco grande."""
    ax.axis('off')
    n_filas = len(datos) + (1 if header else 0)
    tope = max_row_h if max_row_h is not None else 0.11 * (scale / 1.7)
    alto_tabla = min(1.0, n_filas * tope)
    tabla = ax.table(cellText=datos, colLabels=header, cellLoc='center',
                     bbox=[0, 1 - alto_tabla, 1, alto_tabla], colWidths=col_w)
    tabla.auto_set_font_size(False)
    tabla.set_fontsize(fontsize)
    for (i, j), cel in tabla.get_celld().items():
        cel.set_edgecolor(GRIS_LINEA)
        cel.set_linewidth(0.7)
        if altura_fila:
            cel.set_height(altura_fila)
        if i == 0:
            cel.set_facecolor(color_titulo)
            cel.set_text_props(color='white', fontweight='bold')
        elif alinear_izq_col0 and j == 0:
            cel.set_text_props(fontweight='bold', ha='left')
        elif i % 2 == 0:
            cel.set_facecolor('#f8f5ee')
    return tabla


def caja_kpi(fig, x, y, w, h, titulo, valor, sub, color, var_pct=None):
    ax = fig.add_axes([x, y, w, h]); ax.axis('off')
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.add_patch(FancyBboxPatch((0.015, 0.03), 0.97, 0.94, facecolor=color, edgecolor='none',
                                boxstyle='round,pad=0,rounding_size=0.05',
                                mutation_aspect=w / h if h else 1, transform=ax.transAxes))
    fs = 25 if len(valor) <= 3 else 23
    ax.text(0.5, 0.72, valor, ha='center', va='center', fontsize=fs, family=FUENTE_SERIF,
            fontweight='bold', color='white', transform=ax.transAxes)
    ax.text(0.5, 0.44, titulo, ha='center', va='center', fontsize=9.5,
            color='white', fontweight='bold', transform=ax.transAxes)
    sub_txt = sub
    if var_pct is not None:
        signo = '▲' if var_pct > 0 else ('▼' if var_pct < 0 else '▶')
        sub_txt = f"{signo} {var_pct:+.1f}% vs periodo ant.\n{sub}"
    ax.text(0.5, 0.16, sub_txt, ha='center', va='center', fontsize=7.5,
            color='white', alpha=0.9, transform=ax.transAxes)


def pagina_resumen(pdf, tot, agg, num, total, variacion=None, gasto_ads=None):
    fig, ax = nueva_pagina('Reporte de ventas', 'Resumen ejecutivo del periodo',
                           num=num, total=total)

    v = variacion or {}
    kpis = [
        ('AGENDADOS', str(tot['ag']), 'citas agendadas', AZUL, v.get('ag')),
        ('ASISTIERON', str(tot['as_']), f"{pct(tot['as_'], tot['ag']):.0f}% de asistencia", VERDE, v.get('as_')),
        ('COMPRARON', str(tot['co']), f"{pct(tot['co'], tot['as_']):.0f}% de conversion", NARANJA, v.get('co')),
        ('MONTO TOTAL', monto(tot['mon']), 'ventas del periodo', FONDO_TITULO, v.get('mon')),
    ]
    n = len(kpis)
    w = 0.20
    gap = (ANCHO_UTIL - n * w) / (n - 1)
    for i, (t, val, s, c, vp) in enumerate(kpis):
        x = MARGEN_IZQ + i * (w + gap)
        caja_kpi(fig, x, 0.775, w, 0.085, t, val, s, c, var_pct=vp)

    ax.add_patch(FancyBboxPatch((MARGEN_IZQ, 0.555), ANCHO_UTIL, 0.175, facecolor=FONDO_CLARO,
                                edgecolor=NARANJA, linewidth=1.1, boxstyle='round,pad=0,rounding_size=0.012',
                                mutation_aspect=8.27 / 11.69, transform=fig.transFigure))
    ax.text(MARGEN_IZQ + 0.03, 0.702, 'Indicadores clave', fontsize=11.5, fontweight='bold',
            color=FONDO_TITULO, transform=fig.transFigure)
    if tot['co']:
        ticket = tot['mon'] / tot['co']
        lineas = [
            f"Ticket promedio: S/ {ticket:,.0f} por compra   ·   Asistencia: {pct(tot['as_'], tot['ag']):.0f}%   ·   Conversión a compra: {pct(tot['co'], tot['as_']):.0f}%",
            f"Agendaron pero no fueron: {tot['no_fueron']}   ·   Fueron pero no compraron: {tot['fueron_sin_compra']}",
            f"Cada {tot['ag']} agendados generaron {tot['co']} ventas por S/ {tot['mon']:,.0f}",
        ]
        if gasto_ads and gasto_ads > 0:
            roas = tot['mon'] / gasto_ads
            cac = gasto_ads / tot['co']
            lineas.append(f"Inversión Meta Ads: S/ {gasto_ads:,.0f}   ·   ROAS: {roas:.2f}x   ·   CAC: S/ {cac:,.0f} por compra")
    else:
        lineas = ['No se registraron ventas en el periodo.',
                  f"Agendaron pero no fueron: {tot['no_fueron']}   ·   Fueron pero no compraron: {tot['fueron_sin_compra']}"]
        if gasto_ads and gasto_ads > 0:
            lineas.append(f"Inversión Meta Ads: S/ {gasto_ads:,.0f} (sin ventas registradas aún)")
    y = 0.660
    for i, linea in enumerate(lineas):
        color = AZUL if i == len(lineas) - 1 and gasto_ads else GRIS
        peso = 'bold' if color == AZUL else 'normal'
        ax.text(MARGEN_IZQ + 0.03, y, linea, fontsize=9.5, color=color,
                fontweight=peso, transform=fig.transFigure)
        y -= 0.033

    ax.text(MARGEN_IZQ, 0.505, 'Notas del periodo', fontsize=12.5, fontweight='bold',
            color=FONDO_TITULO, transform=fig.transFigure)
    notas = []
    if v.get('mon') is not None:
        signo = 'crecieron' if v['mon'] >= 0 else 'cayeron'
        notas.append(f"Las ventas {signo} {abs(v['mon']):.1f}% frente al periodo comparable anterior.")
    if tot['ag'] and pct(tot['as_'], tot['ag']) < 50:
        notas.append('La asistencia está por debajo del 50%: conviene reforzar la confirmación previa a la cita.')
    if not notas:
        notas.append('Sin observaciones adicionales para este periodo.')
    y = 0.465
    for n_txt in notas[:3]:
        ax.text(MARGEN_IZQ + 0.012, y, '•  ' + n_txt, fontsize=9.3, color=GRIS,
                transform=fig.transFigure)
        y -= 0.032

    ax.text(MARGEN_IZQ, 0.065, 'Fuente: maestro BD DATA.xlsx (hojas AGENDADO y ASISTIDO, '
                                'mismas cifras de los pivotes).', fontsize=8, color=GRIS,
            transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


def pagina_anticipacion(pdf, buckets, gasto_ads, tot, num, total):
    """Relación entre cuánto falta para la cita al momento de agendarla y la
    probabilidad real de que el paciente venga, con una propuesta de cambios
    concreta basada en esa relación."""
    fig, ax = nueva_pagina('Anticipación de la cita', 'Entre más lejos agendas, menos '
                           'vienen: cuánto importa la anticipación en la asistencia real',
                           num=num, total=total)

    con_datos = [b for b in buckets if b['total']]
    if not con_datos:
        ax.text(0.5, 0.55, 'No hay suficientes citas con fecha de agendado y de cita '
                            'válidas en este periodo para medir este patrón.',
                ha='center', fontsize=10.5, color=GRIS, transform=fig.transFigure)
        pdf.savefig(fig); plt.close(fig)
        return

    ax.text(MARGEN_IZQ, 0.855, '% de asistencia según anticipación', fontsize=12.5,
            fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
    axg = fig.add_axes([0.13, 0.60, 0.80, 0.22])
    nombres = [b['bucket'] for b in buckets]
    vals = [b['pct'] or 0 for b in buckets]
    colores = [VERDE if v == max(vals) else (ROJO if v == min(vals) and v < max(vals) else AZUL)
              for v in vals]
    axg.bar(nombres, vals, color=colores)
    axg.set_ylim(0, max(vals + [10]) * 1.28)
    axg.tick_params(axis='both', labelsize=9)
    axg.spines[['top', 'right']].set_visible(False)
    for i, b in enumerate(buckets):
        etiqueta = f"{b['pct']:.0f}%\n(n={b['total']})" if b['total'] else 'sin datos'
        axg.text(i, (b['pct'] or 0), etiqueta, ha='center', va='bottom', fontsize=8, color=GRIS)

    mejor = max(con_datos, key=lambda b: b['pct'])
    peor = min(con_datos, key=lambda b: b['pct'])
    if mejor['bucket'] != peor['bucket'] and mejor['pct'] > 0:
        veces = mejor['pct'] / peor['pct'] if peor['pct'] else None
        linea = (f"Agendar con {mejor['bucket'].lower()} de anticipación asiste el "
                f"{mejor['pct']:.0f}% de las veces; con {peor['bucket'].lower()}, sólo el "
                f"{peor['pct']:.0f}%")
        if veces and veces >= 1.3:
            linea += f" — {veces:.1f} veces más probable que venga si la cita es próxima."
        else:
            linea += '.'
    else:
        linea = 'La asistencia se mantiene relativamente pareja según la anticipación en este periodo.'
    ax.text(MARGEN_IZQ, 0.555, linea, fontsize=9, fontweight='bold', color=FONDO_TITULO,
            transform=fig.transFigure, wrap=True)
    ax.text(MARGEN_IZQ, 0.525, 'Muestra del periodo de este reporte; con pocos días el N por '
                               'grupo puede ser chico.', fontsize=7.8, color=GRIS_PIE,
            style='italic', transform=fig.transFigure)
    ax.text(MARGEN_IZQ, 0.503, 'Sigue esta tendencia mes a mes para confirmarla con más datos.',
            fontsize=7.8, color=GRIS_PIE, style='italic', transform=fig.transFigure)

    ax.plot([MARGEN_IZQ, MARGEN_DER], [0.475, 0.475], color=GRIS_LINEA, linewidth=0.7,
            transform=fig.transFigure)

    # ---- Por qué conviene: retener sale más barato que conseguir más ----
    ax.text(MARGEN_IZQ, 0.445, 'Por qué conviene resolver esto primero', fontsize=12.5,
            fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
    if gasto_ads and gasto_ads > 0 and tot.get('co'):
        cac = gasto_ads / tot['co']
        texto_costo = (f"Conseguir que una cita YA agendada se presente no cuesta nada en "
                       f"publicidad. Conseguir un comprador nuevo cuesta en promedio "
                       f"{monto(cac)} en este periodo (gasto en Meta Ads / compras). "
                       'Antes de pedir más presupuesto de anuncios, cerrar esta brecha de '
                       'asistencia es la forma más barata de vender más.')
    else:
        texto_costo = ('Recuperar una cita que ya está agendada (con una llamada o un '
                       'recordatorio) no cuesta prácticamente nada. Conseguir un comprador '
                       'nuevo sí tiene costo (anuncios, tiempo comercial). Mejorar la '
                       'asistencia de las citas lejanas suele ser la forma más barata de '
                       'vender más, antes de invertir en más leads.')
    ax.text(MARGEN_IZQ, 0.405, texto_costo, fontsize=8.7, color=GRIS, transform=fig.transFigure,
            wrap=True)

    ax.plot([MARGEN_IZQ, MARGEN_DER], [0.335, 0.335], color=GRIS_LINEA, linewidth=0.7,
            transform=fig.transFigure)

    # ---- Propuesta de cambios ----
    ax.text(MARGEN_IZQ, 0.305, 'Propuesta de cambios', fontsize=12.5, fontweight='bold',
            color=VERDE, transform=fig.transFigure)
    propuestas = [
        'Al agendar, ofrecer primero los cupos más próximos disponibles; usar una fecha '
        'lejana sólo cuando el paciente no tenga otra opción.',
        'Usar el panel "Confirmar (48h)" del CRM de forma sistemática para TODAS las citas, '
        'y no sólo como recordatorio de último momento: es la herramienta pensada para '
        'esto y ya está disponible.',
        'Sumar un recordatorio intermedio (a mitad de camino entre agendar y la cita) para '
        f'las citas con {peor["bucket"].lower()} de anticipación, que son las que más fallan.',
        'Medir este % de asistencia por anticipación cada mes en este mismo reporte, para '
        'confirmar si las mejoras en confirmación realmente elevan la asistencia de las citas lejanas.',
    ]
    y = 0.268
    for p in propuestas:
        ax.add_patch(plt.Rectangle((MARGEN_IZQ + 0.003, y - 0.006), 0.009, 0.009,
                                   facecolor=VERDE, edgecolor='none', transform=fig.transFigure))
        ax.text(MARGEN_IZQ + 0.026, y, p, fontsize=8.6, color='#333333', va='top',
                wrap=True, transform=fig.transFigure)
        y -= 0.052
    pdf.savefig(fig); plt.close(fig)


def pagina_flujo(pdf, tot, num, total):
    fig, ax = nueva_pagina('Flujo operativo', 'Del agendado a la venta, etapa por etapa',
                           num=num, total=total)

    etapas = [('AGENDADOS', tot['ag'], AZUL),
              ('ASISTIERON', tot['as_'], VERDE),
              ('COMPRARON', tot['co'], NARANJA)]
    maxv = max(tot['ag'], 1)
    ancho_max = 0.46
    bar_h = 0.095
    paso = 0.155
    y_tops = [0.855 - i * paso for i in range(len(etapas))]
    for i, (nom, val, c) in enumerate(etapas):
        y_top = y_tops[i]
        y_bot = y_top - bar_h
        w = max(0.14, ancho_max * val / maxv)
        x0 = 0.5 - w / 2
        x1 = x0 + w
        ax.add_patch(FancyBboxPatch((x0, y_bot), w, bar_h, facecolor=c, edgecolor='none',
                                    boxstyle='round,pad=0,rounding_size=0.012',
                                    mutation_aspect=8.27 / 11.69, transform=fig.transFigure,
                                    clip_on=False))
        ax.text(x1 + 0.025, y_bot + bar_h / 2, f'{nom}   {val}', ha='left', va='center',
                fontsize=13.5, fontweight='bold', color=c, family=FUENTE_SERIF, transform=fig.transFigure)
        if i < len(etapas) - 1:
            conv = pct(etapas[i + 1][1], val)
            y_next_top = y_tops[i + 1]
            ax.annotate('', xy=(0.5, y_next_top + 0.012), xytext=(0.5, y_bot - 0.004),
                        arrowprops=dict(arrowstyle='-|>', color=GRIS, lw=1.4,
                                        mutation_scale=14),
                        transform=fig.transFigure)
            ax.text(0.5 + max(w, ancho_max) / 2 + 0.025, (y_bot + y_next_top) / 2,
                    f'{conv:.0f}% convierte', ha='left', va='center',
                    fontsize=9, fontweight='bold', color=GRIS, transform=fig.transFigure)

    y_ultimo_bot = y_tops[-1] - bar_h
    banner_top = y_ultimo_bot - 0.055
    banner_h = 0.125
    ax.add_patch(FancyBboxPatch((0.20, banner_top - banner_h), 0.60, banner_h, facecolor=FONDO_TITULO,
                                edgecolor='none', boxstyle='round,pad=0,rounding_size=0.015',
                                mutation_aspect=8.27 / 11.69, transform=fig.transFigure, clip_on=False))
    ax.text(0.5, banner_top - 0.042, monto(tot['mon']), ha='center', va='center', fontsize=25,
            fontweight='bold', color='white', family=FUENTE_SERIF, transform=fig.transFigure)
    ax.text(0.5, banner_top - 0.088, 'MONTO VENDIDO EN EL PERIODO', ha='center', va='center',
            fontsize=9.5, color='white', transform=fig.transFigure)

    y_stats = banner_top - banner_h - 0.075
    if tot['co']:
        ticket = tot['mon'] / tot['co']
        stats = [
            (monto(ticket), 'Ticket promedio por compra'),
            (monto(tot['mon'] / max(tot['as_'], 1)), 'Venta por cada asistente'),
            (f"{pct(tot['co'], tot['ag']):.0f}%", 'De los agendados terminó comprando'),
        ]
        n = len(stats)
        w = 0.22
        gap = (ANCHO_UTIL - n * w) / (n - 1)
        for i, (val, lbl) in enumerate(stats):
            x = MARGEN_IZQ + i * (w + gap) + w / 2
            ax.text(x, y_stats, val, ha='center', va='center', fontsize=15,
                    fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
            ax.text(x, y_stats - 0.032, lbl, ha='center', va='center', fontsize=8.3,
                    color=GRIS, transform=fig.transFigure, wrap=True)
    else:
        ax.text(0.5, y_stats, 'No hubo ventas en el periodo.', ha='center', fontsize=10.5,
                color=GRIS, transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


def pagina_campanas(pdf, agg, num, total):
    """``agg`` debe ser solo campañas reales (ver ``separar_campanas_otros``);
    los tipos de venta que no son campaña van en ``pagina_otros``."""
    fig, ax = nueva_pagina('Resumen de campañas', num=num, total=total)

    camps = defaultdict(lambda: dict(ag=0, as_=0, co=0, mon=0.0))
    for (camp, _crm), d in agg.items():
        for k in ('ag', 'as_', 'co', 'mon'):
            camps[camp][k] += d[k]
    camp_ag = sum(d['ag'] for d in camps.values())
    camp_as = sum(d['as_'] for d in camps.values())
    camp_co = sum(d['co'] for d in camps.values())
    camp_mon = sum(d['mon'] for d in camps.values())

    if not camps:
        ax.text(0.5, 0.55, 'No hubo campañas de anuncios en este periodo.\n'
                            'Ver "Otros" para evaluaciones, retoques, recurrentes, etc.',
                ha='center', fontsize=10.5, color=GRIS, transform=fig.transFigure)
        pdf.savefig(fig); plt.close(fig)
        return

    # Sólo campañas con alguna actividad en el periodo; las que no tuvieron
    # ni un agendado se resumen en una nota al pie en vez de ensuciar la tabla.
    camp_list_todas = sorted(camps.items(), key=lambda kv: -kv[1]['ag'])
    camp_list = [(c, d) for c, d in camp_list_todas if any(d[k] for k in ('ag', 'as_', 'co', 'mon'))]
    n_sin_actividad = len(camp_list_todas) - len(camp_list)
    filas = [[c, d['ag'], d['as_'], d['co'], f"{pct(d['co'], d['as_']):.0f}%",
              monto(d['mon'])] for c, d in camp_list[:8]]
    if len(camp_list) > 8:
        resto = [d for _, d in camp_list[8:]]
        filas.append(['Otras campañas', sum(d['ag'] for d in resto),
                      sum(d['as_'] for d in resto),
                      sum(d['co'] for d in resto),
                      f"{pct(sum(d['co'] for d in resto), sum(d['as_'] for d in resto)):.0f}%",
                      monto(sum(d['mon'] for d in resto))])
    filas.append(['TOTAL', camp_ag, camp_as, camp_co,
                  f"{pct(camp_co, camp_as):.0f}%", monto(camp_mon)])

    heading_y = 0.855
    axc_top = 0.825
    alto_camp = min(0.42, max(0.10, ALTO_FILA_TABLA * (len(filas) + 1)))
    ax.text(MARGEN_IZQ, heading_y, 'Por campaña (con actividad en el periodo)', fontsize=12,
            fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
    axc = fig.add_axes([MARGEN_IZQ, axc_top - alto_camp, ANCHO_UTIL, alto_camp])
    estilo_tabla(axc, filas, [0.36, 0.13, 0.13, 0.13, 0.11, 0.17],
                 header=['Campaña', 'Agend.', 'Asist.', 'Compr.', 'Conv.', 'Monto'],
                 fontsize=9, max_row_h=1.0)
    if n_sin_actividad:
        ax.text(MARGEN_IZQ, axc_top - alto_camp - 0.022,
                f'+ {n_sin_actividad} campañas más sin ningún agendado en el periodo.',
                fontsize=7.8, color=GRIS_PIE, style='italic', transform=fig.transFigure)

    pdf.savefig(fig); plt.close(fig)


def pagina_otros(pdf, agg_otros, num, total):
    """Tipos de venta que no son campañas de anuncios: evaluaciones, retoques,
    recurrentes, recomendados, orgánico/redes, sesiones y filas con la
    columna CAMPAÑA vacía o con un dato inválido. No tienen presupuesto ni
    leads de Meta Ads porque no vienen de un anuncio pagado."""
    fig, ax = nueva_pagina('Otros movimientos', 'Ventas que no vienen de una campaña de anuncios',
                           num=num, total=total)
    ax.text(MARGEN_IZQ, 0.865, 'Evaluaciones, retoques, recurrentes, recomendados, orgánico/redes, '
                               'sesiones y datos sin campaña válida.', fontsize=8.3, color=GRIS,
            transform=fig.transFigure)

    cats = defaultdict(lambda: dict(ag=0, as_=0, co=0, mon=0.0))
    for (cat, crm), d in agg_otros.items():
        for k in ('ag', 'as_', 'co', 'mon'):
            cats[cat][k] += d[k]

    if not cats:
        ax.text(0.5, 0.55, 'No hubo ventas sin campaña en este periodo: todo lo\n'
                            'registrado corresponde a una campaña de anuncios real.',
                ha='center', fontsize=10.5, color=GRIS, transform=fig.transFigure)
        pdf.savefig(fig); plt.close(fig)
        return

    # Sólo categorías con alguna actividad; el resto se resume en una nota.
    cat_list_todas = sorted(cats.items(), key=lambda kv: -kv[1]['ag'])
    cat_list = [(c, d) for c, d in cat_list_todas if any(d[k] for k in ('ag', 'as_', 'co', 'mon'))]
    n_sin_actividad = len(cat_list_todas) - len(cat_list)
    filas = [[c, d['ag'], d['as_'], f"{pct(d['as_'], d['ag']):.0f}%", d['co'],
              f"{pct(d['co'], d['as_']):.0f}%", monto(d['mon'])] for c, d in cat_list]
    tot = dict(ag=sum(d['ag'] for d in cats.values()), as_=sum(d['as_'] for d in cats.values()),
               co=sum(d['co'] for d in cats.values()), mon=sum(d['mon'] for d in cats.values()))
    filas.append(['TOTAL', tot['ag'], tot['as_'], f"{pct(tot['as_'], tot['ag']):.0f}%",
                  tot['co'], f"{pct(tot['co'], tot['as_']):.0f}%", monto(tot['mon'])])

    alto = min(0.55, max(0.09, ALTO_FILA_TABLA * (len(filas) + 1)))
    axc = fig.add_axes([MARGEN_IZQ, 0.825 - alto, ANCHO_UTIL, alto])
    estilo_tabla(axc, filas, [0.34, 0.11, 0.10, 0.10, 0.10, 0.10, 0.15],
                 header=['Categoría', 'Agend.', 'Asist.', '% Asist.', 'Compr.',
                         '% Conv.', 'Monto'],
                 fontsize=9, max_row_h=1.0)
    if n_sin_actividad:
        ax.text(MARGEN_IZQ, 0.825 - alto - 0.022,
                f'+ {n_sin_actividad} categorías más sin ningún agendado en el periodo.',
                fontsize=7.8, color=GRIS_PIE, style='italic', transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


def pagina_campanas_meta(pdf, filas, num, total_pag):
    fig, ax = nueva_pagina('Rendimiento por campaña',
                           'Meta Ads + embudo de conversión (gasto y leads de la carga de Meta Ads)',
                           num=num, total=total_pag)

    total = dict(gasto=0, leads=0, ag=0, as_=0, co=0, mon=0.0)
    datos = []
    for f in filas:
        fila = [f['campania'][:18],
                monto(f['gasto']) if f['gasto'] else '—',
                str(f['leads']) if f['leads'] else '—',
                f"S/ {f['costo_res']:.2f}" if f['leads'] else '—',
                str(f['ag']),
                f"{f['pct_ag']:.0f}%" if f['leads'] else '—',
                str(f['as_']),
                f"{f['pct_as']:.0f}%",
                str(f['co']),
                f"{f['pct_co']:.0f}%",
                monto(f['mon']),
                monto(f['ticket']) if f['co'] else '—']
        datos.append(fila)
        if not f['organica']:
            total['gasto'] += f['gasto']; total['leads'] += f['leads']
        total['ag'] += f['ag']; total['as_'] += f['as_']
        total['co'] += f['co']; total['mon'] += f['mon']
    datos.append(['TOTAL',
                  monto(total['gasto']), str(total['leads']),
                  f"S/ {total['gasto'] / total['leads']:.2f}" if total['leads'] else '—',
                  str(total['ag']),
                  f"{pct(total['ag'], total['leads']):.0f}%" if total['leads'] else '—',
                  str(total['as_']), f"{pct(total['as_'], total['ag']):.0f}%",
                  str(total['co']), f"{pct(total['co'], total['as_']):.0f}%",
                  monto(total['mon']),
                  monto(total['mon'] / total['co']) if total['co'] else '—'])

    alto = min(0.68, max(0.14, ALTO_FILA_TABLA * 1.35 * (len(datos) + 1)))
    axc = fig.add_axes([0.035, 0.855 - alto, 0.93, alto])
    estilo_tabla(axc, datos,
                 [0.19, 0.065, 0.04, 0.065, 0.06, 0.065, 0.06, 0.065, 0.06, 0.065, 0.08, 0.09],
                 header=['Campaña', 'Gasto\nMeta', 'Leads', 'Costo\n/lead', 'Agend.',
                         'Agend.\n/lead', 'Asist.', '%\nAsist.', 'Compr.', '%\nCompr.',
                         'Monto', 'Ticket\n/venta'],
                 fontsize=7.3, max_row_h=1.0)
    y_nota = 0.855 - alto - 0.03
    ax.text(MARGEN_IZQ, y_nota, 'Agend/lead = % de agendados logrados por cada lead de Meta.  '
                                 'Ticket/venta = ventas entre los que compraron.',
            fontsize=7.5, color=GRIS, transform=fig.transFigure)
    ax.text(MARGEN_IZQ, y_nota - 0.022, 'El gasto y los leads de Meta son del rango del reporte cargado.',
            fontsize=7.5, color=GRIS, transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


def pagina_metricas(pdf, analitica, num, total):
    fig, ax = nueva_pagina('Métricas del periodo', 'Tratamientos, distritos y perfil de '
                           'los pacientes que asistieron', num=num, total=total)

    trat = analitica['trat'].most_common(8)
    if trat:
        ax1 = fig.add_axes([0.30, 0.56, 0.63, 0.27])
        names = [str(t[0])[:20] for t in trat]
        vals = [t[1] for t in trat]
        ypos = np.arange(len(names))[::-1]
        ax1.barh(ypos, vals, color=AZUL)
        ax1.set_yticks(ypos); ax1.set_yticklabels(names, fontsize=7.5)
        ax1.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax1.set_title('Tratamientos más frecuentes', fontsize=12, fontweight='bold',
                      color=FONDO_TITULO, pad=10)
        for yi, va in zip(ypos, vals):
            ax1.text(va + 0.05, yi, str(va), va='center', fontsize=9, color=GRIS)
        ax1.tick_params(axis='x', labelsize=8)
        ax1.spines[['top', 'right']].set_visible(False)
    else:
        ax.text(0.5, 0.70, 'Sin tratamientos registrados en el periodo.', ha='center',
                fontsize=10, color=GRIS, transform=fig.transFigure)

    dist = analitica['dist'].most_common(8)
    if dist:
        ax2 = fig.add_axes([MARGEN_IZQ + 0.03, 0.09, 0.32, 0.40])
        names2 = [str(d[0])[:13] for d in dist]
        vals2 = [d[1] for d in dist]
        ypos2 = np.arange(len(names2))[::-1]
        ax2.barh(ypos2, vals2, color=VERDE)
        ax2.set_yticks(ypos2); ax2.set_yticklabels(names2, fontsize=6.8)
        ax2.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax2.set_title('Distritos', fontsize=11, fontweight='bold', color=FONDO_TITULO, pad=8)
        ax2.tick_params(axis='x', labelsize=7)
        ax2.spines[['top', 'right']].set_visible(False)

    ax3 = fig.add_axes([0.52, 0.09, 0.41, 0.40]); ax3.axis('off')
    ax3.set_title('Perfil del paciente', fontsize=11, fontweight='bold',
                  color=FONDO_TITULO, pad=8)
    edades = analitica['edades']
    sexo = analitica['sexo']
    ed_med = round(sum(edades) / len(edades), 1) if edades else 'n/d'
    ed_min = min(edades) if edades else '-'
    ed_max = max(edades) if edades else '-'
    sf = sexo.get('F', 0)
    sm = sexo.get('M', 0)
    stot = sf + sm or 1
    ax3.text(0.02, 0.95, 'Edad media', fontsize=9, fontweight='bold', color=GRIS)
    ax3.text(0.02, 0.82, f'{ed_med} años (rango {ed_min}-{ed_max})', fontsize=12,
             fontweight='bold', color=FONDO_TITULO)

    ax3.text(0.02, 0.68, 'Bandas de edad', fontsize=9, fontweight='bold', color=GRIS)
    bandas = {'18-25': 0, '26-35': 0, '36-45': 0, '46-55': 0, '56+': 0}
    for e in edades:
        if e < 26:
            bandas['18-25'] += 1
        elif e < 36:
            bandas['26-35'] += 1
        elif e < 46:
            bandas['36-45'] += 1
        elif e < 56:
            bandas['46-55'] += 1
        else:
            bandas['56+'] += 1
    ed_tot = len(edades) or 1
    colores_banda = [AZUL, VERDE, NARANJA, CELESTE, GRIS]
    x0, y0, alto = 0.02, 0.58, 0.055
    for (nom, cnt), col in zip(bandas.items(), colores_banda):
        w = 0.96 * cnt / ed_tot
        if w > 0:
            ax3.add_patch(plt.Rectangle((x0, y0), w, alto, facecolor=col,
                                        edgecolor='white', linewidth=0.6,
                                        transform=ax3.transAxes))
        x0 += w
    for i, ((nom, cnt), col) in enumerate(zip(bandas.items(), colores_banda)):
        ax3.text(0.02 + (i % 2) * 0.50, 0.46 - (i // 2) * 0.10,
                 f'{nom}: {pct(cnt, ed_tot):.0f}%', fontsize=7.8, color=col, fontweight='bold')

    ax3.text(0.02, 0.12, 'Sexo', fontsize=9, fontweight='bold', color=GRIS)
    ax3.text(0.02, 0.00, f'F: {sf} ({pct(sf, stot):.0f}%)   |   M: {sm} ({pct(sm, stot):.0f}%)',
             fontsize=10, fontweight='bold', color=FONDO_TITULO)
    pdf.savefig(fig); plt.close(fig)


def pagina_ejecutivas(pdf, ejecutivas, num, total):
    fig, ax = nueva_pagina('Performance por ejecutiva', 'Según el campo "Agendado por"',
                           num=num, total=total)

    if not ejecutivas:
        ax.text(0.5, 0.5, 'No hay datos de ejecutivas en el periodo',
                ha='center', fontsize=12, color=GRIS, transform=fig.transFigure)
        pdf.savefig(fig); plt.close(fig)
        return

    filas = []
    tot_ag_e = sum(e['ag'] for e in ejecutivas)
    tot_as_e = sum(e['as_'] for e in ejecutivas)
    tot_co_e = sum(e['co'] for e in ejecutivas)
    tot_mon_e = sum(e['mon'] for e in ejecutivas)
    tot_nf_e = sum(e['no_fueron'] for e in ejecutivas)
    for e in ejecutivas:
        filas.append([
            e['ejecutiva'][:18],
            e['ag'],
            e['as_'],
            f"{e['asistencia_pct']:.0f}%",
            e['co'],
            f"{e['conversion_pct']:.0f}%",
            e['no_fueron'],
            monto(e['mon']),
            monto(e['ticket']) if e['co'] else '—',
        ])
    filas.append([
        'TOTAL',
        tot_ag_e,
        tot_as_e,
        f"{pct(tot_as_e, tot_ag_e):.0f}%",
        tot_co_e,
        f"{pct(tot_co_e, tot_as_e):.0f}%",
        tot_nf_e,
        monto(tot_mon_e),
        monto(tot_mon_e / tot_co_e) if tot_co_e else '—',
    ])

    header = ['Ejecutiva', 'Agend.', 'Asist.', '%\nAsist.', 'Compr.',
              '%\nConv.', 'No\nAsist.', 'Monto', 'Ticket']
    col_w = [0.20, 0.09, 0.09, 0.09, 0.09, 0.09, 0.09, 0.13, 0.13]
    sw = sum(col_w)
    col_w = [w / sw for w in col_w]

    alto_ej = min(0.30, max(0.09, ALTO_FILA_TABLA * 1.5 * (len(filas) + 1)))
    ax_t = fig.add_axes([MARGEN_IZQ, 0.83 - alto_ej, ANCHO_UTIL, alto_ej])
    estilo_tabla(ax_t, filas, col_w, header=header, fontsize=7.8, max_row_h=1.0)

    valid_ej = [e for e in ejecutivas if e['ejecutiva'] != 'SIN EJECUTIVA'] or ejecutivas
    if valid_ej:
        ax_g = fig.add_axes([0.14, 0.10, 0.76, 0.34])
        names = [e['ejecutiva'][:16] for e in valid_ej][:8]
        montos_ej = [e['mon'] for e in valid_ej][:8]
        ypos = np.arange(len(names))[::-1]
        ax_g.barh(ypos, montos_ej, color=AZUL, height=0.55)
        ax_g.set_yticks(ypos)
        ax_g.set_yticklabels(names, fontsize=8.5)
        ax_g.set_title('Ventas generadas por ejecutiva (S/)', fontsize=11,
                       fontweight='bold', color=FONDO_TITULO, pad=8)
        max_m = max(montos_ej) if montos_ej else 100
        for yi, va in zip(ypos, montos_ej):
            ax_g.text(va + max(max_m * 0.02, 1), yi, monto(va),
                      va='center', fontsize=8.5, color=FONDO_TITULO, fontweight='bold')
        ax_g.tick_params(axis='x', labelsize=8)
        ax_g.spines[['top', 'right']].set_visible(False)
        ax_g.set_xlim(0, max_m * 1.25 if max_m > 0 else 100)

    pdf.savefig(fig); plt.close(fig)


def pagina_motivos(pdf, motivos, num, total):
    fig, ax = nueva_pagina('Motivos de pérdida', num=num, total=total)

    m_as = motivos.get('no_asistio', {})
    m_co = motivos.get('no_compra', {})

    # Bloque 1: No Asistencia (arriba)
    ax.text(MARGEN_IZQ, 0.855, '1. Motivos de NO Asistencia (agendaron pero no acudieron)',
            fontsize=12, fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
    tot_as_loss = sum(m_as.values())
    if m_as:
        filas_as = []
        for mot, cnt in m_as.items():
            filas_as.append([mot, cnt, f"{cnt / tot_as_loss * 100.0:.1f}%"])
        filas_as.append(['TOTAL', tot_as_loss, '100.0%'])
        ax_t1 = fig.add_axes([MARGEN_IZQ, 0.58, 0.40, 0.25])
        estilo_tabla(ax_t1, filas_as, [0.55, 0.22, 0.23],
                     header=['Motivo', 'Cant.', '%'], fontsize=8, max_row_h=1.0)

        ax_g1 = fig.add_axes([0.55, 0.58, 0.38, 0.25])
        names = list(m_as.keys())[:6]
        vals = list(m_as.values())[:6]
        ypos = np.arange(len(names))[::-1]
        ax_g1.barh(ypos, vals, color=ROJO, alpha=0.85, height=0.55)
        ax_g1.set_yticks(ypos)
        ax_g1.set_yticklabels(names, fontsize=7.5)
        ax_g1.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax_g1.spines[['top', 'right']].set_visible(False)
        for yi, va in zip(ypos, vals):
            ax_g1.text(va + 0.05, yi, str(va), va='center', fontsize=8, color=GRIS)
    else:
        ax.add_patch(FancyBboxPatch((MARGEN_IZQ, 0.60), ANCHO_UTIL, 0.23, facecolor=FONDO_CLARO,
                                    edgecolor=CELESTE, linewidth=1, boxstyle='round,pad=0,rounding_size=0.012',
                                    mutation_aspect=8.27 / 11.69, transform=fig.transFigure))
        ax.text(0.5, 0.72, 'Sin motivos de no asistencia registrados en el periodo.',
                ha='center', va='center', fontsize=10, color=GRIS, transform=fig.transFigure)
        ax.text(0.5, 0.67, 'Tipificación en la columna "MOTIVO NO ASISTIO" (col AB) del maestro BD DATA.',
                ha='center', va='center', fontsize=8.5, color=AZUL, transform=fig.transFigure)

    # Bloque 2: No Compra (abajo)
    ax.text(MARGEN_IZQ, 0.475, '2. Motivos de NO Compra (asistieron pero no compraron)',
            fontsize=12, fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
    tot_co_loss = sum(m_co.values())
    if m_co:
        filas_co = []
        for mot, cnt in m_co.items():
            filas_co.append([mot, cnt, f"{cnt / tot_co_loss * 100.0:.1f}%"])
        filas_co.append(['TOTAL', tot_co_loss, '100.0%'])
        ax_t2 = fig.add_axes([MARGEN_IZQ, 0.19, 0.40, 0.25])
        estilo_tabla(ax_t2, filas_co, [0.55, 0.22, 0.23],
                     header=['Motivo', 'Cant.', '%'], fontsize=8, max_row_h=1.0)

        ax_g2 = fig.add_axes([0.55, 0.19, 0.38, 0.25])
        names2 = list(m_co.keys())[:6]
        vals2 = list(m_co.values())[:6]
        ypos2 = np.arange(len(names2))[::-1]
        ax_g2.barh(ypos2, vals2, color=NARANJA, alpha=0.85, height=0.55)
        ax_g2.set_yticks(ypos2)
        ax_g2.set_yticklabels(names2, fontsize=7.5)
        ax_g2.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax_g2.spines[['top', 'right']].set_visible(False)
        for yi, va in zip(ypos2, vals2):
            ax_g2.text(va + 0.05, yi, str(va), va='center', fontsize=8, color=GRIS)
    else:
        ax.add_patch(FancyBboxPatch((MARGEN_IZQ, 0.21), ANCHO_UTIL, 0.23, facecolor=FONDO_CLARO,
                                    edgecolor=CELESTE, linewidth=1, boxstyle='round,pad=0,rounding_size=0.012',
                                    mutation_aspect=8.27 / 11.69, transform=fig.transFigure))
        ax.text(0.5, 0.33, 'Sin motivos de no compra registrados en el periodo.',
                ha='center', va='center', fontsize=10, color=GRIS, transform=fig.transFigure)
        ax.text(0.5, 0.28, 'Tipificación en la columna "MOTIVO NO COMPRA" (col AC) del maestro BD DATA.',
                ha='center', va='center', fontsize=8.5, color=AZUL, transform=fig.transFigure)

    ax.text(MARGEN_IZQ, 0.10, 'Fuente: maestro BD DATA.xlsx (columnas AB y AC).',
            fontsize=8, color=GRIS, transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


def pagina_comparativo_historico(pdf, comp, historico, num, total):
    fig, ax = nueva_pagina('Comparativo y tendencia', 'Frente a periodos comparables',
                           num=num, total=total)

    actual = comp['actual']
    ant = comp['anterior_mes']
    anio_ant = comp['mismo_mes_anio_anterior']

    ax1 = fig.add_axes([0.12, 0.61, 0.76, 0.25])
    metricas = ['agendados', 'asistidos', 'compraron']
    labels_m = ['Agendados', 'Asistieron', 'Compraron']
    periodos = [('Periodo actual', actual, FONDO_TITULO),
                ('Mes anterior\n(mismos dias)', ant, CELESTE),
                ('Mismo periodo\nano anterior', anio_ant, GRIS)]
    x = np.arange(len(metricas))
    width = 0.25
    for i, (nombre, d, color) in enumerate(periodos):
        vals = [d.get(m, 0) for m in metricas]
        ax1.bar(x + (i - 1) * width, vals, width, label=nombre, color=color)
    ax1.set_xticks(x); ax1.set_xticklabels(labels_m, fontsize=9)
    ax1.legend(fontsize=6.5, loc='upper right', ncol=1)
    ax1.set_title('Embudo: periodo actual vs comparables', fontsize=11,
                  fontweight='bold', color=FONDO_TITULO, pad=10)
    ax1.tick_params(axis='y', labelsize=8)
    ax1.spines[['top', 'right']].set_visible(False)

    def var(a, b):
        return f"{((a - b) / b * 100):+.1f}%" if b else 'n/d'

    filas = [
        ['Agendados', actual['agendados'], ant.get('agendados', 0),
         anio_ant.get('agendados', 0), var(actual['agendados'], ant.get('agendados', 0))],
        ['Asistieron', actual['asistidos'], ant.get('asistidos', 0),
         anio_ant.get('asistidos', 0), var(actual['asistidos'], ant.get('asistidos', 0))],
        ['Compraron', actual['compraron'], ant.get('compraron', 0),
         anio_ant.get('compraron', 0), var(actual['compraron'], ant.get('compraron', 0))],
        ['Monto (S/)', f"{actual['monto']:,.0f}", f"{ant.get('monto', 0):,.0f}",
         f"{anio_ant.get('monto', 0):,.0f}", var(actual['monto'], ant.get('monto', 0))],
    ]
    axt = fig.add_axes([MARGEN_IZQ + 0.02, 0.43, ANCHO_UTIL - 0.04, 0.145])
    estilo_tabla(axt, filas, [0.24, 0.19, 0.21, 0.21, 0.15],
                 header=['Métrica', 'Actual', 'Mes ant.', 'Año ant.', 'Var. vs mes ant.'],
                 fontsize=8, max_row_h=1.0)

    if historico:
        ax2 = fig.add_axes([0.10, 0.09, 0.82, 0.30])
        etiquetas = [f"{h['mes']} {str(h['anio'])[2:]}" for h in historico]
        montos_h = [h['monto'] for h in historico]
        ax2.plot(range(len(etiquetas)), montos_h, marker='o', color=AZUL,
                 linewidth=2, markersize=4)
        ax2.fill_between(range(len(etiquetas)), montos_h, alpha=0.12, color=AZUL)
        ax2.set_xticks(range(len(etiquetas)))
        ax2.set_xticklabels(etiquetas, fontsize=6.5, rotation=45, ha='right')
        ax2.set_title(f'Tendencia de ventas ({len(historico)} meses)', fontsize=11,
                      fontweight='bold', color=FONDO_TITULO, pad=8)
        ax2.spines[['top', 'right']].set_visible(False)
        ax2.tick_params(axis='y', labelsize=7)
    pdf.savefig(fig); plt.close(fig)


def pagina_proyeccion(pdf, proy, patrones, pareto, num, total):
    """Página de análisis que no se ve a simple vista en el Excel: proyección
    de cierre de mes (pipeline real + ritmo de agendamiento), mejor día de la
    semana para vender, y qué tan concentrados están los ingresos en pocos
    pacientes."""
    fig, ax = nueva_pagina('Proyección y patrones', 'Lo que no se ve a simple vista '
                           'revisando el Excel', num=num, total=total)

    # ---- Bloque 1: proyección de cierre de mes ----
    ax.text(MARGEN_IZQ, 0.855, f'Proyección de cierre de {NOMBRES_MES.get(MES, MES)}',
            fontsize=12.5, fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
    if proy:
        kpis = [
            ('YA VENDIDO', monto(proy['ya_vendido']), f"días {proy['dias_transcurridos']} de {proy['dias_mes']}", AZUL),
            ('PIPELINE AGENDADO', monto(proy['pipeline_esperado']), f"{proy['citas_pendientes']} citas pendientes", VERDE),
            ('NUEVOS ESPERADOS', monto(proy['venta_nuevos_esperada']), f"~{proy['nuevos_esperados']:.0f} leads más", NARANJA),
            ('PROYECCIÓN TOTAL', monto(proy['proyeccion']), 'al cierre del mes', FONDO_TITULO),
        ]
        n = len(kpis)
        w = 0.20
        gap = (ANCHO_UTIL - n * w) / (n - 1)
        for i, (t, val, s, c) in enumerate(kpis):
            x = MARGEN_IZQ + i * (w + gap)
            caja_kpi(fig, x, 0.735, w, 0.10, t, val, s, c)
        ax.text(MARGEN_IZQ, 0.685, f"Tasa de conversión reciente (asiste→compra): "
                f"{proy['tasa_conversion_pct']:.0f}%   ·   Ticket promedio: "
                f"{monto(proy['ticket_promedio'])}   ·   Ritmo de agendamiento: "
                f"{proy['ritmo_agendados_dia']:.1f} leads/día", fontsize=8.7, color=GRIS,
                transform=fig.transFigure)
        ax.text(MARGEN_IZQ, 0.658, f"Si sólo se mantuviera el ritmo de venta ya facturado "
                f"(sin sumar pipeline ni leads nuevos), el cierre lineal simple daría "
                f"{monto(proy['ritmo_lineal'])} — la proyección combinada de arriba es "
                'más realista porque usa las citas que ya están en agenda.',
                fontsize=8, color=GRIS_PIE, style='italic', wrap=True, transform=fig.transFigure)
    else:
        ax.text(0.5, 0.72, 'No se pudo calcular la proyección para este periodo.',
                ha='center', fontsize=10, color=GRIS, transform=fig.transFigure)

    ax.plot([MARGEN_IZQ, MARGEN_DER], [0.615, 0.615], color=GRIS_LINEA, linewidth=0.7,
            transform=fig.transFigure)

    # ---- Bloque 2: mejor día de la semana para vender ----
    ax.text(MARGEN_IZQ, 0.585, 'Mejor día de la semana para vender', fontsize=12.5,
            fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
    montos_sem = patrones['monto']
    if any(montos_sem):
        axg = fig.add_axes([0.13, 0.375, 0.80, 0.17])
        colores = [NARANJA if v == max(montos_sem) else AZUL for v in montos_sem]
        axg.bar(patrones['dias'], montos_sem, color=colores)
        axg.tick_params(axis='both', labelsize=8)
        axg.spines[['top', 'right']].set_visible(False)
        for i, v in enumerate(montos_sem):
            if v:
                axg.text(i, v, f'{v:,.0f}', ha='center', va='bottom', fontsize=7.5, color=GRIS)
        mejor_i = max(range(7), key=lambda i: montos_sem[i])
        dias_con_venta = sum(1 for v in montos_sem if v)
        dia_plural = {'Sábado': 'sábados', 'Domingo': 'domingos'}.get(
            patrones['dias'][mejor_i], patrones['dias'][mejor_i].lower())
        linea = f"Los {dia_plural} concentran más ventas: {monto(montos_sem[mejor_i])} en el periodo"
        # Con muy pocos días de venta en el periodo, un % "vs. el resto" puede
        # dispararse a cifras absurdas (dividiendo entre días que en realidad
        # nunca tuvieron actividad) — sólo se muestra con una base razonable.
        if dias_con_venta >= 3:
            otros = [v for i, v in enumerate(montos_sem) if i != mejor_i]
            prom_otros = sum(otros) / len(otros) if otros else 0
            extra = pct(montos_sem[mejor_i] - prom_otros, prom_otros) if prom_otros else None
            if extra and 0 < extra <= 300:
                linea += f", {extra:.0f}% más que el promedio del resto de días."
            else:
                linea += '.'
        else:
            linea += '.'
        ax.text(MARGEN_IZQ, 0.345, linea, fontsize=8.7, fontweight='bold', color=AZUL,
                transform=fig.transFigure, wrap=True)
        if dias_con_venta < 3:
            ax.text(MARGEN_IZQ, 0.318, f'Sólo {dias_con_venta} día(s) con ventas en el periodo — '
                    'tómalo como tendencia inicial, no como patrón confirmado.',
                    fontsize=7.8, color=GRIS_PIE, style='italic', transform=fig.transFigure)
    else:
        ax.text(0.5, 0.44, 'Sin ventas suficientes en el periodo para ver un patrón semanal.',
                ha='center', fontsize=9.5, color=GRIS, transform=fig.transFigure)

    ax.plot([MARGEN_IZQ, MARGEN_DER], [0.27, 0.27], color=GRIS_LINEA, linewidth=0.7,
            transform=fig.transFigure)

    # ---- Bloque 3: concentración de ingresos (pareto) ----
    ax.text(MARGEN_IZQ, 0.24, 'Concentración de ingresos', fontsize=12.5, fontweight='bold',
            color=FONDO_TITULO, transform=fig.transFigure)
    if pareto['n_pacientes']:
        alto_riesgo = pareto['top20_pct_monto'] >= 50
        color_caja = '#fbe9e7' if alto_riesgo else FONDO_CLARO
        color_borde = ROJO if alto_riesgo else CELESTE
        ax.add_patch(FancyBboxPatch((MARGEN_IZQ, 0.12), ANCHO_UTIL, 0.095, facecolor=color_caja,
                                    edgecolor=color_borde, linewidth=1, boxstyle='round,pad=0,rounding_size=0.012',
                                    mutation_aspect=8.27 / 11.69, transform=fig.transFigure))
        ax.text(MARGEN_IZQ + 0.02, 0.192, f"El {round(100 * pareto['top20_n'] / pareto['n_pacientes'])}% de los "
                f"pacientes que compraron ({pareto['top20_n']} de {pareto['n_pacientes']}) generó el "
                f"{pareto['top20_pct_monto']:.0f}% de los S/ {pareto['total']:,.0f} vendidos en el periodo.",
                fontsize=9, fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure, wrap=True)
        interpretacion = ('Alta dependencia de pocos clientes grandes: conviene cuidar '
                          'especialmente a ese grupo y diversificar la base.' if alto_riesgo else
                          'Los ingresos del periodo están razonablemente bien repartidos entre pacientes.')
        ax.text(MARGEN_IZQ + 0.02, 0.152, interpretacion, fontsize=8.3, color=GRIS,
                transform=fig.transFigure, wrap=True)
    else:
        ax.text(0.5, 0.17, 'Sin compras registradas en el periodo para medir concentración.',
                ha='center', fontsize=9.5, color=GRIS, transform=fig.transFigure)

    pdf.savefig(fig); plt.close(fig)


def pagina_evolucion_diaria(pdf, serie, num, total):
    fig, ax = nueva_pagina('Evolución diaria', 'Del periodo del reporte', num=num, total=total)

    dias = serie['dias']
    montos_d = serie['monto']
    if not dias or (not any(montos_d) and not any(serie['agendados'])):
        ax.text(0.5, 0.5, 'Sin actividad diaria registrada en el periodo.',
                ha='center', fontsize=11, color=GRIS, transform=fig.transFigure)
        pdf.savefig(fig); plt.close(fig)
        return

    ax1 = fig.add_axes([0.10, 0.55, 0.82, 0.31])
    ax1.bar(dias, montos_d, color=FONDO_TITULO, width=0.6)
    ax1.set_title('Monto vendido por dia (S/)', fontsize=12, fontweight='bold',
                  color=FONDO_TITULO, pad=8)
    ax1.set_xticks(dias)
    ax1.tick_params(axis='both', labelsize=7.5)
    ax1.spines[['top', 'right']].set_visible(False)
    for d, v in zip(dias, montos_d):
        if v:
            ax1.text(d, v, f'{v:,.0f}', ha='center', va='bottom', fontsize=6.5, color=GRIS)

    ax2 = fig.add_axes([0.10, 0.17, 0.82, 0.29])
    ax2.plot(dias, serie['agendados'], marker='o', color=AZUL, label='Agendados',
             linewidth=1.6, markersize=4)
    ax2.plot(dias, serie['asistidos'], marker='o', color=VERDE, label='Asistieron',
             linewidth=1.6, markersize=4)
    ax2.plot(dias, serie['compraron'], marker='o', color=NARANJA, label='Compraron',
             linewidth=1.6, markersize=4)
    ax2.set_xticks(dias)
    ax2.tick_params(axis='both', labelsize=7.5)
    ax2.legend(fontsize=8, loc='upper right')
    ax2.set_title('Agendados, asistencia y compras por dia', fontsize=12,
                  fontweight='bold', color=FONDO_TITULO, pad=8)
    ax2.spines[['top', 'right']].set_visible(False)

    if any(montos_d):
        mejor_i = max(range(len(dias)), key=lambda i: montos_d[i])
        ax.text(0.5, 0.095, f'Mejor día del periodo: {dias[mejor_i]} de {NOMBRES_MES[MES]} '
                            f'con S/ {montos_d[mejor_i]:,.0f} en ventas.',
                ha='center', fontsize=9.5, color=AZUL, fontweight='bold',
                transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


def pagina_recurrentes(pdf, rec, num, total):
    fig, ax = nueva_pagina('Pacientes recurrentes y LTV', 'De los pacientes que asistieron '
                           'en este periodo (no del histórico completo)', num=num, total=total)

    total_p = rec['total_pacientes']
    total_r = rec['total_recurrentes']
    kpis = [
        ('PACIENTES ÚNICOS', str(total_p), 'con al menos 1 compra', AZUL),
        ('RECURRENTES', str(total_r), f'{pct(total_r, total_p):.1f}% del total', VERDE),
        ('LTV PROMEDIO', monto(rec['ltv_promedio']), 'valor de vida por paciente', FONDO_TITULO),
    ]
    n = len(kpis)
    w = 0.26
    gap = (ANCHO_UTIL - n * w) / (n - 1)
    for i, (t, val, s, c) in enumerate(kpis):
        x = MARGEN_IZQ + i * (w + gap)
        caja_kpi(fig, x, 0.72, w, 0.11, t, val, s, c)

    ax.text(MARGEN_IZQ, 0.615, 'Top pacientes recurrentes (por número de compras)', fontsize=12,
            fontweight='bold', color=FONDO_TITULO, transform=fig.transFigure)
    pacientes = rec['pacientes'][:15]
    if pacientes:
        filas = [[p['nombre'][:26], p['compras'], monto(p['monto']),
                  monto(p['monto'] / p['compras']), p['ultima']] for p in pacientes]
        alto = min(0.49, max(0.10, ALTO_FILA_TABLA * 1.3 * (len(filas) + 1)))
        axt = fig.add_axes([MARGEN_IZQ, 0.59 - alto, ANCHO_UTIL, alto])
        estilo_tabla(axt, filas, [0.30, 0.13, 0.19, 0.19, 0.19],
                     header=['Paciente', 'Compras', 'Monto total', 'Ticket prom.', 'Última visita'],
                     fontsize=8, max_row_h=1.0)
    else:
        ax.text(0.5, 0.4, 'Aún no hay pacientes con 2 o más compras registradas.',
                ha='center', fontsize=11, color=GRIS, transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


def pagina_reactivacion(pdf, react, num, total, meses_umbral=3):
    fig, ax = nueva_pagina('Oportunidad de reactivación', f'Pacientes que asistieron alguna '
                           f'vez pero no vuelven hace más de {meses_umbral} meses (a hoy)',
                           num=num, total=total)

    ax.add_patch(FancyBboxPatch((MARGEN_IZQ, 0.795), ANCHO_UTIL, 0.075, facecolor=FONDO_TITULO,
                                edgecolor='none', boxstyle='round,pad=0,rounding_size=0.015',
                                mutation_aspect=8.27 / 11.69, transform=fig.transFigure))
    ax.text(0.5, 0.8325, f'{len(react)} pacientes para contactar', ha='center', va='center',
            fontsize=17, fontweight='bold', color='white', family=FUENTE_SERIF, transform=fig.transFigure)

    if react:
        filas = [[p['nombre'][:26], p['telefono'], p['ultima_cita'], p['meses_sin_volver']]
                 for p in react[:25]]
        alto = min(0.575, max(0.10, ALTO_FILA_TABLA * 1.3 * (len(filas) + 1)))
        axt = fig.add_axes([MARGEN_IZQ, 0.765 - alto, ANCHO_UTIL, alto])
        estilo_tabla(axt, filas, [0.34, 0.24, 0.22, 0.20],
                     header=['Paciente', 'Teléfono', 'Última cita', 'Meses sin volver'],
                     fontsize=8, max_row_h=1.0)
        if len(react) > 25:
            ax.text(0.5, 0.765 - alto - 0.025, f'Mostrando los 25 casos más antiguos de '
                    f'{len(react)} totales.', ha='center', fontsize=8.5, color=GRIS,
                    transform=fig.transFigure)
    else:
        ax.text(0.5, 0.5, 'No hay pacientes pendientes de reactivar en este momento.',
                ha='center', fontsize=11, color=GRIS, transform=fig.transFigure)
    if react:
        ax.text(0.5, 0.105, 'Contactar a esta lista es una oportunidad directa de venta:',
                ha='center', fontsize=9, color=AZUL, fontweight='bold', transform=fig.transFigure)
        ax.text(0.5, 0.078, 'son pacientes que ya confiaron en el consultorio.',
                ha='center', fontsize=9, color=AZUL, fontweight='bold', transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


def pagina_hallazgos(pdf, tot, analitica, num, total, rec=None, react=None, serie=None,
                     proy=None, patrones=None):
    fig, ax = nueva_pagina('Hallazgos y sugerencias', 'Cierre del reporte', num=num, total=total)

    hallazgos = []
    asis = pct(tot['as_'], tot['ag'])
    conv = pct(tot['co'], tot['as_'])
    hallazgos.append(f"La asistencia fue del {asis:.0f}% ({tot['as_']} de {tot['ag']} "
                     f"agendados) y la conversion a compra del {conv:.0f}%.")
    if proy and proy.get('dias_restantes'):
        hallazgos.append(f"Con {proy['dias_restantes']} días restantes del mes, la "
                         f"proyección de cierre es {monto(proy['proyeccion'])} (pipeline "
                         f"agendado: {monto(proy['pipeline_esperado'])}).")
    if patrones and any(patrones.get('monto', [])):
        montos_s = patrones['monto']
        mejor_i = max(range(7), key=lambda i: montos_s[i])
        hallazgos.append(f"Los {patrones['dias'][mejor_i].lower()} son el día de la "
                         f"semana con más ventas del periodo ({monto(montos_s[mejor_i])}).")
    if tot['co']:
        ticket = tot['mon'] / tot['co']
        hallazgos.append(f"Ticket promedio de {monto(ticket)} por compra.")
    mejor_crm = max(((c, v) for c, v in agg_tot_crm.items() if v['co']),
                    key=lambda kv: kv[1]['co'], default=None)
    if mejor_crm:
        hallazgos.append(f"{mejor_crm[0]} genera la mayor parte de las ventas "
                         f"({mejor_crm[1]['co']} compras, S/ {mejor_crm[1]['mon']:,.0f}).")
    top_trat = analitica['trat'].most_common(1)
    if top_trat:
        t, n = top_trat[0]
        hallazgos.append(f"El servicio mas solicitado fue {t} ({n} veces).")
    if rec and rec.get('total_pacientes'):
        pct_rec = pct(rec['total_recurrentes'], rec['total_pacientes'])
        hallazgos.append(f"{pct_rec:.0f}% de los pacientes de la base son recurrentes "
                         f"(2+ compras), con un LTV promedio de {monto(rec['ltv_promedio'])}.")
    if react:
        hallazgos.append(f"{len(react)} pacientes no vuelven hace mas de 3 meses: "
                         'oportunidad directa de reactivacion.')
    if serie and any(serie.get('monto', [])):
        montos_s = serie['monto']
        mejor_i = max(range(len(montos_s)), key=lambda i: montos_s[i])
        hallazgos.append(f"El mejor dia del periodo fue el {serie['dias'][mejor_i]} "
                         f"con S/ {montos_s[mejor_i]:,.0f} en ventas.")
    if tot['no_fueron']:
        # Nota: no_fueron cuenta citas por FECHA DE CITA en el periodo (incluye
        # reagendados que agendaron en un periodo anterior), así que no es
        # comparable como % directo de tot['ag'] (que cuenta por FECHA DE
        # AGENDADO) — se muestra como conteo, no como porcentaje del agendado.
        hallazgos.append(f"{tot['no_fueron']} citas agendadas para este periodo no se "
                         'presentaron (no-show).')
    if tot['fueron_sin_compra']:
        hallazgos.append(f"{tot['fueron_sin_compra']} pacientes fueron a su cita pero "
                         f"no compraron ({pct(tot['fueron_sin_compra'], tot['as_']):.0f}"
                         '% de los asistentes).')
    top_camp = analitica['camp'].most_common(1)
    if top_camp:
        c, n = top_camp[0]
        hallazgos.append(f"La campana con mas asistentes fue {c} ({n}).")
    top_dist = analitica['dist'].most_common(1)
    if top_dist:
        d, n = top_dist[0]
        hallazgos.append(f"Mayor concentracion de pacientes: {d} ({n} asistentes).")
    hallazgos = hallazgos[:8]

    sugerencias = []
    if asis < 65:
        sugerencias.append('Reforzar confirmacion y recordatorios (dia previo) para '
                           'subir la asistencia por encima del 65%.')
    else:
        sugerencias.append('Mantener el protocolo de recordatorios; la asistencia '
                           'ya supera el 65%.')
    if conv < 50:
        sugerencias.append('Capacitar en upselling durante la consulta para elevar '
                           'la conversion a compra.')
    else:
        sugerencias.append('La conversion es solida; sube el ticket con paquetes y '
                           'promociones.')
    if react:
        sugerencias.append(f'Contactar a los {len(react)} pacientes inactivos con una '
                           'promocion de regreso puede generar ventas rapidas.')
    if top_trat:
        t, n = top_trat[0]
        sugerencias.append(f'Reserva mas stock y agenda para {t}, el mas demandado.')
    sugerencias = sugerencias[:4]

    def bloque(titulo, items, y0, color, paso):
        ax.text(MARGEN_IZQ, y0, titulo, fontsize=13, fontweight='bold', color=color,
                transform=fig.transFigure)
        y = y0 - 0.05
        for it in items:
            ax.add_patch(plt.Rectangle((MARGEN_IZQ + 0.003, y - 0.006), 0.009, 0.009,
                                       facecolor=color, edgecolor='none',
                                       transform=fig.transFigure))
            ax.text(MARGEN_IZQ + 0.026, y, it, fontsize=9.8, color='#333333', va='top',
                    wrap=True, transform=fig.transFigure)
            y -= paso
        return y

    paso_h = min(0.062, max(0.040, 0.30 / max(len(hallazgos), 1)))
    y_fin = bloque('Hallazgos del periodo', hallazgos, 0.845, FONDO_TITULO, paso_h)
    paso_s = min(0.075, max(0.050, 0.22 / max(len(sugerencias), 1)))
    bloque('Sugerencias', sugerencias, y_fin - 0.05, VERDE, paso_s)

    ax.text(MARGEN_IZQ, 0.09, 'Este reporte se genera autom\u00e1ticamente a partir del maestro '
                              'BD DATA.xlsx. Ante cualquier diferencia frente al cuadre manual, '
                              'revisar primero la tipificaci\u00f3n de columnas (CRM, CAMPA\u00d1A, '
                              'ASISTENCIA) en AGENDADOS y VENTA DIARIA.',
            fontsize=7.8, color=GRIS_PIE, style='italic', wrap=True, transform=fig.transFigure)
    pdf.savefig(fig); plt.close(fig)


agg_tot_crm = {}
tot_ag = tot_as = tot_co = 0
tot_mon = 0.0


def generar_reporte(mes='AGO', anio=2026, desde=1, hasta=10, fuente='maestro',
                    salida=None):
    """Genera el PDF del reporte y devuelve dict con resumen (reutilizable CLI/web)."""
    global MES, ANIO, D1, D2, FUENTE, SALIDA, tot_ag, tot_as, tot_co, tot_mon
    global CRM_ORDEN, CANALES, COLOR_CANAL
    MES, ANIO, D1, D2, FUENTE = mes, int(anio), int(desde), int(hasta), fuente
    mes_ord = list(NOMBRES_MES).index(MES) + 1 if MES in NOMBRES_MES else None
    if mes_ord and (ANIO, mes_ord) < INICIO_OPERACION:
        raise ValueError(
            f'{NOMBRES_MES.get(MES, MES)} {ANIO} es anterior al inicio real de operación '
            'de Derma Essenza (julio 2026). Ese periodo del maestro BD DATA.xlsx es de un '
            'negocio anterior (Beauty Medic) y no corresponde a Derma Essenza.')
    if salida:
        SALIDA = salida if os.path.isabs(salida) else os.path.join(BASE_DIR, salida)
    agg_tot_crm.clear()
    agg = build_data()
    if COL.get('ES_BM'):
        CRM_ORDEN = ['KOMMO', 'WHATSAPP', 'ORGANICO', 'SIN CRM']
        CANALES = ['KOMMO', 'WHATSAPP', 'ORGANICO']
    else:
        CANALES = canales_desde_datos(agg)
        CRM_ORDEN = CANALES + ['SIN CRM']
    COLOR_CANAL = {c: PALETA[i % len(PALETA)] for i, c in enumerate(CANALES)}
    tot = dict(ag=0, as_=0, co=0, mon=0.0, no_fueron=0, fueron_sin_compra=0)
    for d in agg.values():
        tot['ag'] += d['ag']; tot['as_'] += d['as_']
        tot['co'] += d['co']; tot['mon'] += d['mon']
        tot['no_fueron'] += d['no_fueron']
        tot['fueron_sin_compra'] += d['fueron_sin_compra']
    for c in CRM_ORDEN:
        sel = {k: v for k, v in agg.items() if k[1] == c}
        if sel:
            agg_tot_crm[c] = dict(ag=sum(v['ag'] for v in sel.values()),
                                  as_=sum(v['as_'] for v in sel.values()),
                                  co=sum(v['co'] for v in sel.values()),
                                  mon=sum(v['mon'] for v in sel.values()),
                                  no_fueron=sum(v['no_fueron'] for v in sel.values()),
                                  fueron_sin_compra=sum(v['fueron_sin_compra'] for v in sel.values()))
    tot_ag, tot_as, tot_co, tot_mon = tot['ag'], tot['as_'], tot['co'], tot['mon']
    agg_campanas, agg_otros = separar_campanas_otros(agg)
    analitica = datos_analiticos()
    filas_meta = build_campana_meta(agg_campanas)
    ejecutivas = datos_ejecutivas()
    motivos = datos_motivos()
    gasto_ads = _gasto_ads_total()

    # Variación vs periodo anterior equivalente (mismo rango de días)
    pm, pa, pd, ph = _periodo_anterior(MES, ANIO, D1, D2)
    tot_ant = totales_periodo(pm, pa, pd, ph)
    variacion = {
        'ag': _variacion(tot['ag'], tot_ant['ag']),
        'as_': _variacion(tot['as_'], tot_ant['as_']),
        'co': _variacion(tot['co'], tot_ant['co']),
        'mon': _variacion(tot['mon'], tot_ant['mon']),
    }

    # Analítica adicional (comparativos, tendencia, recurrencia, reactivación).
    # Import diferido para evitar el ciclo de imports con analitica.py, que a
    # su vez importa este módulo.
    import analitica as ana
    comp = ana.comparativo(MES, ANIO, D1, D2)
    historico = ana.ventas_por_mes()
    serie = ana.serie_diaria(MES, ANIO, D1, D2)
    rec = ana.recurrentes(MES, ANIO, D1, D2)
    react = ana.pacientes_a_reactivar(meses=3)
    try:
        proy = ana.proyeccion_mes(MES, ANIO)
    except Exception:  # noqa: BLE001 — la proyección es informativa, no debe romper el reporte
        proy = None
    patrones_semana = datos_dia_semana()
    pareto = datos_pareto()
    anticipacion = datos_anticipacion()

    total_paginas = 13 + (1 if filas_meta else 0) + (1 if ejecutivas else 0)
    pag = [0]  # contador mutable (Python 2/3-safe sin nonlocal en closures)

    def sig():
        pag[0] += 1
        return pag[0]

    with PdfPages(SALIDA) as pdf:
        pagina_resumen(pdf, tot, agg, sig(), total_paginas, variacion=variacion, gasto_ads=gasto_ads)
        pagina_comparativo_historico(pdf, comp, historico, sig(), total_paginas)
        pagina_proyeccion(pdf, proy, patrones_semana, pareto, sig(), total_paginas)
        pagina_anticipacion(pdf, anticipacion, gasto_ads, tot, sig(), total_paginas)
        pagina_flujo(pdf, tot, sig(), total_paginas)
        pagina_evolucion_diaria(pdf, serie, sig(), total_paginas)
        pagina_campanas(pdf, agg_campanas, sig(), total_paginas)
        if filas_meta:
            pagina_campanas_meta(pdf, filas_meta, sig(), total_paginas)
        # pagina_campana_canal ya no se genera: Derma Essenza no tipifica CRM
        # (todo queda "SIN CRM"), así que esa página siempre saldría vacía.
        pagina_otros(pdf, agg_otros, sig(), total_paginas)
        if ejecutivas:
            pagina_ejecutivas(pdf, ejecutivas, sig(), total_paginas)
        pagina_metricas(pdf, analitica, sig(), total_paginas)
        pagina_recurrentes(pdf, rec, sig(), total_paginas)
        pagina_reactivacion(pdf, react, sig(), total_paginas)
        pagina_motivos(pdf, motivos, sig(), total_paginas)
        pagina_hallazgos(pdf, tot, analitica, sig(), total_paginas, rec=rec, react=react,
                        serie=serie, proy=proy, patrones=patrones_semana)
    return {'archivo': SALIDA, 'totales': tot,
            'por_crm': {c: dict(v) for c, v in agg_tot_crm.items()},
            'detalle': {f'{k[0]} | {k[1]}': dict(v) for k, v in sorted(
                agg.items(), key=lambda kv: -kv[1]['ag'])},
            'por_campana_meta': filas_meta,
            'ejecutivas': ejecutivas,
            'motivos': motivos,
            'variacion': variacion,
            'recurrentes': {'total_pacientes': rec['total_pacientes'],
                            'total_recurrentes': rec['total_recurrentes'],
                            'ltv_promedio': rec['ltv_promedio']},
            'reactivacion': {'pendientes': len(react)},
            'proyeccion': proy,
            'pareto': pareto,
            'anticipacion': anticipacion}


def generar_reporte_breve(mes='AGO', anio=2026, desde=1, hasta=10, fuente='maestro',
                          salida=None):
    """PDF de una sola página: la tabla de campañas (PPTO/CPL/LEAD/AGENDADO/
    ASISTIDO/REALIZADO/FACTURADO/TICKET), igual al cuadre manual en Excel."""
    global MES, ANIO, D1, D2, FUENTE, SALIDA
    MES, ANIO, D1, D2, FUENTE = mes, int(anio), int(desde), int(hasta), fuente
    if salida:
        SALIDA = salida if os.path.isabs(salida) else os.path.join(BASE_DIR, salida)
    agg = build_data()
    tot = dict(ag=0, as_=0, co=0, mon=0.0)
    for d in agg.values():
        tot['ag'] += d['ag']; tot['as_'] += d['as_']
        tot['co'] += d['co']; tot['mon'] += d['mon']
    agg_campanas, _agg_otros = separar_campanas_otros(agg)
    filas_meta = build_campana_meta(agg_campanas)
    with PdfPages(SALIDA) as pdf:
        pagina_campanas_meta(pdf, filas_meta)
    return {'archivo': SALIDA, 'totales': tot,
            'detalle': {f'{k[0]} | {k[1]}': dict(v) for k, v in sorted(
                agg.items(), key=lambda kv: -kv[1]['ag'])},
            'por_campana_meta': filas_meta}


def datos_reporte_breve_web(mes='AGO', anio=2026, desde=1, hasta=10):
    """Igual que generar_reporte_breve pero para la web: junto a la tabla de
    campañas (PPTO/CPL/LEAD/agendado/asistido/realizado/facturado/ticket) arma
    el detalle de pacientes por campaña (quién agendó/asistió/compró), sin
    generar PDF. No vuelve a sincronizar Drive: usa el maestro y AGENDADOS ya
    descargados por /api/reporte."""
    global MES, ANIO, D1, D2, FUENTE, COL
    MES, ANIO, D1, D2, FUENTE = mes, int(anio), int(desde), int(hasta), 'maestro'
    ag_path = os.path.join(am.TMP_DIR, 'AGENDADOS.xlsx')
    ws = am.leer_maestro(am.ruta_maestro_local())
    COL = detectar_columnas(ws)

    agg = defaultdict(lambda: dict(ag=0, as_=0, co=0, mon=0.0,
                                   no_fueron=0, fueron_sin_compra=0))
    pac = defaultdict(lambda: dict(agendaron=[], asistieron=[], compraron=[]))

    ag_counts = am.agendados_por_periodo(ag_path, ANIO, MES, D1, D2)
    for key, counts in ag_counts.items():
        agg[key]['ag'] = counts['ag']
        pac[key]['agendaron'] = counts.get('pacientes', [])

    origen_por_telefono = _campana_origen_por_telefono(ag_path)
    for r in range(5, ws.max_row + 1):
        if not _fila_real_derma(ws, r):
            continue
        crm = ws.cell(row=r, column=COL['CANAL']).value or 'SIN CRM'
        tel = am.norm_phone(ws.cell(row=r, column=COL['TELEFONO']).value)
        nombre = ws.cell(row=r, column=COL['NOMBRE']).value or ''
        camp_maestro = str(ws.cell(row=r, column=COL['CAMPANA']).value or '').strip()
        camp_ag_key = origen_por_telefono.get(tel) or (camp_maestro if camp_maestro else '(SIN CAMPANA)')
        key = (camp_ag_key, crm)
        d = agg[key]
        if (ws.cell(row=r, column=COL['ANIO4']).value == ANIO
                and ws.cell(row=r, column=COL['MES3']).value == MES
                and en_periodo(ws.cell(row=r, column=COL['DIA2']).value)):
            fecha = f"{ws.cell(row=r, column=COL['DIA2']).value}/{MES}/{ANIO}"
            asist = str(ws.cell(row=r, column=COL['ASISTENCIA']).value or '').strip()
            if asist == 'ASISTIO':
                d['as_'] += 1
                pac[key]['asistieron'].append({'nombre': nombre, 'telefono': tel, 'fecha': fecha})
                p = pago_total(ws, r)
                d['mon'] += p
                if p > 0:
                    d['co'] += 1
                    pac[key]['compraron'].append({'nombre': nombre, 'telefono': tel,
                                                  'fecha': fecha, 'monto': p})
                else:
                    d['fueron_sin_compra'] += 1
            elif _cita_pasada(ws, r):
                d['no_fueron'] += 1

    agg_campanas, _agg_otros = separar_campanas_otros(dict(agg))
    filas_meta = build_campana_meta(agg_campanas)

    # El detalle de pacientes se junta por campaña (sumando entre canales/CRM,
    # igual que build_campana_meta hace con los conteos).
    detalle_por_campana = defaultdict(lambda: dict(agendaron=[], asistieron=[], compraron=[]))
    for (camp, _crm), d in pac.items():
        for k in ('agendaron', 'asistieron', 'compraron'):
            detalle_por_campana[camp][k].extend(d[k])
    for fila in filas_meta:
        nombre_camp = fila.get('campania_maestro') or fila['campania'].replace(' (organica)', '')
        fila['detalle'] = detalle_por_campana.get(
            nombre_camp, dict(agendaron=[], asistieron=[], compraron=[]))

    tot = dict(ag=0, as_=0, co=0, mon=0.0)
    for d in agg.values():
        tot['ag'] += d['ag']; tot['as_'] += d['as_']
        tot['co'] += d['co']; tot['mon'] += d['mon']

    return {'totales': tot, 'campanas': filas_meta}


def main(argv=None):
    ap = argparse.ArgumentParser(description='Reporte de ventas por campana de agendados (PDF)')
    ap.add_argument('--mes', default='AGO')
    ap.add_argument('--anio', type=int, default=2026)
    ap.add_argument('--desde', type=int, default=1)
    ap.add_argument('--hasta', type=int, default=10)
    ap.add_argument('--fuente', choices=['maestro', 'auto'], default='maestro',
                    help='maestro = pivotes guardados; auto = integra AGENDADOS+VENTA de Drive')
    ap.add_argument('--salida', default=None)
    a = ap.parse_args(argv)
    res = generar_reporte(mes=a.mes, anio=a.anio, desde=a.desde, hasta=a.hasta,
                          fuente=a.fuente, salida=a.salida)
    tot = res['totales']
    print(f"PDF generado: {res['archivo']}")
    print(f"Agendados: {tot['ag']} | Asistieron: {tot['as_']} | Compraron: {tot['co']} | "
          f"Monto: S/ {tot['mon']:,.0f}".replace(',', ' '))


if __name__ == '__main__':
    main()
