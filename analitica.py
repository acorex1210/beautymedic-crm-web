# -*- coding: utf-8 -*-
"""analitica.py
===============
Analítica interactiva sobre el maestro BD DATA (formato Derma Essenza o BM).
Reutiliza la detección de columnas de reporte_ventas_pdf para leer el maestro
y devuelve diccionarios listos para graficar en el frontend (Chart.js).

Funciones principales:
  kpis(mes, anio, desde, hasta)      -> indicadores del periodo
  serie_diaria(mes, anio, desde, hasta) -> evolución día a día
  perfil(mes, anio, desde, hasta)    -> edad, sexo, distrito, tratamiento, canal
  comparativo(mes, anio, desde, hasta)  -> periodo actual vs anterior vs año pasado
  recurrentes(mes, anio, desde, hasta)  -> pacientes que repiten y LTV
"""
import calendar
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timedelta

import openpyxl  # noqa: F401

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import alimentar_maestro as am  # noqa: E402
import reporte_ventas_pdf as rv  # noqa: E402

_MM = ['', 'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO',
       'SET', 'OCT', 'NOV', 'DIC']

_MM_ORD = {m: i for i, m in enumerate(_MM)}

# Derma Essenza opera desde julio 2026. El maestro BD DATA.xlsx es una hoja
# reutilizada de un negocio anterior (Beauty Medic) y arrastra filas suyas
# con fecha anterior a esa (p.ej. la ejecutiva "STEFANY", ago-set 2025) — no
# son datos reales de Derma Essenza y no deben entrar en ningún cálculo
# histórico (tendencia, comparativo año anterior, recurrencia, reactivación,
# proyección). _filas_maestro() las descarta en el origen para que todas las
# funciones de este módulo queden protegidas automáticamente.
_INICIO_OPERACION = (2026, 7)  # (año, mes) del primer mes real de Derma Essenza


def _antes_de_inicio(anio, mes_txt):
    """True si (anio, mes_txt) es anterior al inicio real de operación de
    Derma Essenza — o sea, no puede ser un dato genuino de este negocio."""
    if not anio or not mes_txt:
        return False
    mi = _MM_ORD.get(str(mes_txt).strip().upper())
    if not mi:
        return False
    try:
        return (int(anio), mi) < _INICIO_OPERACION
    except (TypeError, ValueError):
        return False


def _cita_pasada(f):
    """True si la fecha de cita de la fila ya ocurrió (comparada con hoy)."""
    if not (f['anio_cita'] and f['mes_cita'] and f['dia_cita']):
        return False
    try:
        ci = (int(f['anio_cita']), _MM_ORD.get(str(f['mes_cita']).strip().upper(), 0),
              int(f['dia_cita']))
    except (TypeError, ValueError):
        return False
    ho = (datetime.now().year, _MM_ORD.get(_MM[datetime.now().month], 0),
          datetime.now().day)
    return ci < ho

TOTAL_COLS = 12  # años a mirar para el histórico de ventas por mes


def _meses_previos(mes, anio, n):
    """Devuelve los n meses anteriores como lista de (mes, anio)."""
    ord_mes = list(rv.NOMBRES_MES.keys())
    idx = ord_mes.index(mes)
    out = []
    a, i = anio, idx
    for _ in range(n):
        i -= 1
        if i < 0:
            i = len(ord_mes) - 1
            a -= 1
        out.append((ord_mes[i], a))
    return out


def _periodo_anterior(mes, anio, desde, hasta):
    """Periodo anterior equivalente (mismo rango de días, mes previo)."""
    pm, pa = _meses_previos(mes, anio, 1)[0]
    return pm, pa, desde, hasta


def _variacion(actual, anterior):
    """Variación % de actual vs anterior. None si anterior es 0/no hay dato."""
    if actual is None or anterior is None:
        return None
    if not isinstance(actual, (int, float)) or not isinstance(anterior, (int, float)):
        return None
    if anterior == 0:
        return None
    return round((actual - anterior) / abs(anterior) * 100, 1)


def _gasto_ads():
    """Gasto total de la carga Meta Ads más reciente (o None si no hay)."""
    try:
        import meta_ads as mads
        META_DIR = os.path.join(os.environ.get('DATA_DIR', 'data'), 'meta_ads')
        cargas = mads.listar(META_DIR)
        if not cargas:
            return None
        return float(cargas[0].get('gasto') or 0) or None
    except Exception:  # noqa: BLE001
        return None


def _ventas_venta_diaria(mes, anio, desde, hasta):
    """Nº de ventas registradas en VENTA DIARIA en el periodo (para CAC)."""
    ruta = os.path.join(am.TMP_DIR, 'VENTA_DIARIA.xlsx')
    if not os.path.exists(ruta):
        return None
    try:
        ventas = am.leer_venta(ruta)
    except Exception:  # noqa: BLE001
        return None
    n = 0
    for v in ventas:
        if not isinstance(v.get('venta'), (int, float)) or v['venta'] <= 0:
            continue
        if (v['anio'] == anio
                and str(v.get('mes') or '').strip().upper() == str(mes).strip().upper()
                and isinstance(v.get('dia'), (int, float))
                and desde <= int(v['dia']) <= hasta):
            n += 1
    return n


def _leer_col(ws):
    return rv.detectar_columnas(ws)


def _en_rango(dia, desde, hasta):
    return isinstance(dia, (int, float)) and desde <= int(dia) <= hasta


def _valores(ws, col, r, claves):
    return {k: ws.cell(row=r, column=col[k]).value for k in claves}


def _filas_maestro():
    ws = am.leer_maestro(am.ruta_maestro_local())
    col = _leer_col(ws)
    filas = []
    c_mot_as = col.get('MOTIVO_NO_ASISTIO')
    c_mot_co = col.get('MOTIVO_NO_COMPRA')
    for r in range(5, ws.max_row + 1):
        f = {'r': r,
             'dia_ag': ws.cell(row=r, column=col['DIA']).value,
             'mes_ag': ws.cell(row=r, column=col['MES']).value,
             'anio_ag': ws.cell(row=r, column=col['ANIO']).value,
             'dia_cita': ws.cell(row=r, column=col['DIA2']).value,
             'mes_cita': ws.cell(row=r, column=col['MES3']).value,
             'anio_cita': ws.cell(row=r, column=col['ANIO4']).value,
             'telefono': ws.cell(row=r, column=col['TELEFONO']).value,
             'dni': ws.cell(row=r, column=col['DNI']).value if 'DNI' in col else None,
             'nombre': ws.cell(row=r, column=col['NOMBRE']).value if 'NOMBRE' in col else None,
             'canal': ws.cell(row=r, column=col['CANAL']).value,
             'campana': ws.cell(row=r, column=col['CAMPANA']).value,
             'ejecutiva': ws.cell(row=r, column=col['AGENDADO']).value if col.get('AGENDADO') else None,
             'asistencia': ws.cell(row=r, column=col['ASISTENCIA']).value,
             'distrito': ws.cell(row=r, column=col['DISTRITO']).value,
             'edad': ws.cell(row=r, column=col['EDAD']).value,
             'sexo': ws.cell(row=r, column=col['SEXO']).value,
             'motivo_no_asistio': ws.cell(row=r, column=c_mot_as).value if c_mot_as else None,
             'motivo_no_compra': ws.cell(row=r, column=c_mot_co).value if c_mot_co else None,
             'trats': [ws.cell(row=r, column=c).value for c in col['TRAT']],
             'montos': [ws.cell(row=r, column=c).value for c in col['PAGO']],
             'pago_total': ws.cell(row=r, column=col['PAGO_TOTAL']).value,
             }
        # Descarta filas anteriores al inicio real de operación de Derma
        # Essenza (arrastradas del negocio anterior que usaba esta hoja) —
        # ver _INICIO_OPERACION. Si cualquiera de las dos fechas de la fila
        # (agendado o cita) cae antes, la fila entera no es un dato real.
        if (_antes_de_inicio(f['anio_ag'], f['mes_ag'])
                or _antes_de_inicio(f['anio_cita'], f['mes_cita'])):
            continue
        filas.append(f)
    return filas, col


def _agendado(f, mes, anio, desde, hasta):
    return (f['anio_ag'] == anio and f['mes_ag'] == mes
            and _en_rango(f['dia_ag'], desde, hasta) and f['telefono'])


def _asistido(f, mes, anio, desde, hasta):
    return (f['anio_cita'] == anio and f['mes_cita'] == mes
            and _en_rango(f['dia_cita'], desde, hasta)
            and str(f['asistencia'] or '').strip() == 'ASISTIO')


def _monto(f):
    p = f['pago_total']
    if isinstance(p, (int, float)) and p > 0:
        return p
    return sum(x for x in f['montos'] if isinstance(x, (int, float)))


def kpis(mes, anio, desde, hasta):
    base = _kpis_core(mes, anio, desde, hasta)

    # ---- variación vs. periodo anterior equivalente ----
    pm, pa, pd, ph = _periodo_anterior(mes, anio, desde, hasta)
    anterior = _kpis_core(pm, pa, pd, ph)
    base['variacion'] = {
        'monto': _variacion(base['monto'], anterior['monto']),
        'agendados': _variacion(base['agendados'], anterior['agendados']),
        'asistidos': _variacion(base['asistidos'], anterior['asistidos']),
        'compraron': _variacion(base['compraron'], anterior['compraron']),
        'asistencia_pct': _variacion(base['asistencia_pct'], anterior['asistencia_pct']),
        'conversion_pct': _variacion(base['conversion_pct'], anterior['conversion_pct']),
        'ticket_promedio': _variacion(base['ticket_promedio'], anterior['ticket_promedio']),
        'roas': _variacion(base['roas'], anterior['roas']) if base['roas'] is not None else None,
    }
    return base


def _kpis_core(mes, anio, desde, hasta):
    filas, col = _filas_maestro()
    ag = as_ = co = 0
    mon = 0.0
    no_fueron = 0
    fueron_sin_compra = 0
    for f in filas:
        if _agendado(f, mes, anio, desde, hasta):
            ag += 1
            if (str(f['asistencia'] or '').strip() != 'ASISTIO'
                    and _cita_pasada(f)):
                no_fueron += 1
        if _asistido(f, mes, anio, desde, hasta):
            as_ += 1
            p = _monto(f)
            mon += p
            if p > 0:
                co += 1
            else:
                fueron_sin_compra += 1
    recompra = 0
    if co:
        compradores = _compradores_frecuentes(filas, mes, anio, desde, hasta,
                                              min_compras=2)
        recompra = len(compradores)

    base = {
        'agendados': ag,
        'asistidos': as_,
        'compraron': co,
        'monto': round(mon, 2),
        'no_fueron': no_fueron,
        'fueron_sin_compra': fueron_sin_compra,
        'asistencia_pct': round(100.0 * as_ / ag, 1) if ag else 0,
        'conversion_pct': round(100.0 * co / as_, 1) if as_ else 0,
        'ticket_promedio': round(mon / co, 2) if co else 0,
        'monto_por_asistido': round(mon / as_, 2) if as_ else 0,
        'recompra': recompra,
    }

    # ---- ROAS / CAC (gasto de la carga Meta Ads más reciente) ----
    gasto = _gasto_ads()
    ventas_venta = _ventas_venta_diaria(mes, anio, desde, hasta)
    if gasto and mon:
        base['gasto_ads'] = round(gasto, 2)
        base['roas'] = round(mon / gasto, 2)
        base['cac'] = round(gasto / ventas_venta, 2) if ventas_venta else None
        base['ventas_venta_diaria'] = ventas_venta
    else:
        base['gasto_ads'] = None
        base['roas'] = None
        base['cac'] = None
        base['ventas_venta_diaria'] = ventas_venta
    return base


def serie_diaria(mes, anio, desde, hasta):
    filas, _ = _filas_maestro()
    dias = list(range(desde, hasta + 1))
    ag = {d: 0 for d in dias}
    as_ = {d: 0 for d in dias}
    mon = {d: 0.0 for d in dias}
    co = {d: 0 for d in dias}
    for f in filas:
        if _asistido(f, mes, anio, desde, hasta):
            d = int(f['dia_cita'])
            as_[d] += 1
            p = _monto(f)
            mon[d] += p
            if p > 0:
                co[d] += 1
        if _agendado(f, mes, anio, desde, hasta):
            ag[int(f['dia_ag'])] += 1
    return {'dias': dias,
            'agendados': [ag[d] for d in dias],
            'asistidos': [as_[d] for d in dias],
            'compraron': [co[d] for d in dias],
            'monto': [round(mon[d], 2) for d in dias]}


def _fecha(dia, mes, anio):
    """(dia, 'AGO', anio) -> date, o None si algo no calza."""
    try:
        mi = _MM_ORD.get(str(mes).strip().upper())
        if not mi:
            return None
        return datetime(int(anio), mi, int(dia)).date()
    except (TypeError, ValueError):
        return None


def proyeccion_mes(mes, anio, dia_referencia=None):
    """Proyección de venta al cierre del mes, sin asumir un ritmo diario
    plano. Combina tres factores:

      1. Ya vendido: lo que ya se facturó este mes (citas hasta hoy).
      2. Pipeline conocido: citas YA agendadas para lo que resta del mes
         (aún no ocurren) × tasa de conversión y ticket promedio de los
         últimos 3 meses — a diferencia del monto/día, esto usa compromisos
         reales ya en la agenda, no un promedio.
      3. Leads todavía sin agendar: para los días que faltan y que aún no
         tienen ninguna cita en agenda, se estima cuántos agendados nuevos
         aparecerán según el ritmo de agendamiento de los últimos 30 días
         (no del mes en curso, que puede estar a mitad de camino), y se les
         aplica la misma tasa de conversión y ticket promedio.

    ``dia_referencia``, si se da, fuerza qué día del mes se usa como "hoy"
    para el cálculo (por ejemplo 1, para obtener el pronóstico de cierre de
    mes calculado en frío, antes de que avance ninguna venta real -- así se
    puede congelar una única meta fija para todo el mes en vez de recalcular
    un número que sube y baja según cuánto haya avanzado el mes). Por
    defecto usa la fecha real de hoy.

    Devuelve el desglose completo para que la pantalla explique el número,
    no sólo lo muestre.
    """
    mes = str(mes).strip().upper()
    anio = int(anio)
    mi = _MM_ORD.get(mes)
    if not mi:
        return None
    dias_mes = calendar.monthrange(anio, mi)[1]
    if dia_referencia is not None:
        hoy = datetime(anio, mi, min(max(1, int(dia_referencia)), dias_mes)).date()
    else:
        hoy = datetime.now().date()
    es_mes_actual = (hoy.year, hoy.month) == (anio, mi)
    dia_hoy = hoy.day if es_mes_actual else dias_mes
    dias_transcurridos = min(dia_hoy, dias_mes)
    dias_restantes = max(0, dias_mes - dias_transcurridos)

    # ---- tasa de conversión, tasa de efectividad y ticket promedio: últimos
    # 3 meses (incluye el actual) para no depender de un solo mes con pocos
    # datos ----
    meses_hist = [(mes, anio)] + _meses_previos(mes, anio, 2)
    ag_hist = as_hist = co_hist = 0
    mon_hist = 0.0
    for m, a in meses_hist:
        k = _kpis_core(m, a, 1, 31)
        ag_hist += k['agendados']; as_hist += k['asistidos']
        co_hist += k['compraron']; mon_hist += k['monto']
    tasa_conversion = co_hist / as_hist if as_hist else 0.0
    ticket_promedio = mon_hist / co_hist if co_hist else 0.0
    # tasa_efectividad = agendado -> compra (agendado -> asiste -> compra),
    # a diferencia de tasa_conversion que solo mide asiste -> compra. Una
    # cita agendada que todavía no ocurre puede no llegar a asistir nunca
    # (no contesta, cancela...), así que aplicarle sólo tasa_conversion
    # infla el pipeline: asume que TODO agendado se presenta.
    tasa_efectividad = co_hist / ag_hist if ag_hist else 0.0

    filas, _ = _filas_maestro()
    ya_vendido = 0.0
    citas_pendientes = 0
    for f in filas:
        if f['anio_cita'] != anio or str(f['mes_cita'] or '').strip().upper() != mes:
            continue
        fc = _fecha(f['dia_cita'], f['mes_cita'], f['anio_cita'])
        if fc is None:
            continue
        if fc <= hoy:
            if str(f['asistencia'] or '').strip() == 'ASISTIO':
                ya_vendido += _monto(f)
        else:
            citas_pendientes += 1
    pipeline_esperado = round(citas_pendientes * tasa_efectividad * ticket_promedio, 2)

    # ---- ritmo de agendamiento reciente (últimos 30 días naturales, no el
    # mes en curso) para estimar cuántos leads NUEVOS (sin cita todavía)
    # aparecerán en lo que falta del mes ----
    desde30 = hoy - timedelta(days=30)
    nuevos_30d = 0
    for f in filas:
        fa = _fecha(f['dia_ag'], f['mes_ag'], f['anio_ag'])
        if fa and desde30 <= fa <= hoy and f['telefono']:
            nuevos_30d += 1
    ritmo_agendados_dia = nuevos_30d / 30.0
    nuevos_esperados = round(ritmo_agendados_dia * dias_restantes, 1)
    venta_nuevos_esperada = round(nuevos_esperados * tasa_efectividad * ticket_promedio, 2)

    total = round(ya_vendido + pipeline_esperado + venta_nuevos_esperada, 2)
    # Piso de seguridad: la proyección lineal simple (ritmo actual llevado a
    # fin de mes) nunca debería superar al modelo combinado si el mes venía
    # fuerte y el pipeline es flojo — se informa como referencia, no se usa
    # para recortar el resultado.
    ritmo_lineal = round((ya_vendido / dias_transcurridos) * dias_mes, 2) if dias_transcurridos else 0.0

    return {
        'mes': mes, 'anio': anio, 'dias_mes': dias_mes,
        'dias_transcurridos': dias_transcurridos, 'dias_restantes': dias_restantes,
        'ya_vendido': round(ya_vendido, 2),
        'citas_pendientes': citas_pendientes,
        'pipeline_esperado': pipeline_esperado,
        'nuevos_esperados': nuevos_esperados,
        'venta_nuevos_esperada': venta_nuevos_esperada,
        'tasa_conversion_pct': round(tasa_conversion * 100, 1),
        'tasa_efectividad_pct': round(tasa_efectividad * 100, 1),
        'ticket_promedio': round(ticket_promedio, 2),
        'ritmo_agendados_dia': round(ritmo_agendados_dia, 2),
        'ritmo_lineal': ritmo_lineal,
        'proyeccion': total,
    }


# Objetivo de eficiencia publicitaria: por cada sol vendido, como máximo
# ese porcentaje de sol invertido en anuncios. 0.1 = invertir 1 para vender
# 10 (ROAS 10x). Es el número que la clínica se puso como meta.
OBJETIVO_INVERSION_VENTA = 0.1


def _norm_campana(v):
    s = unicodedata.normalize('NFD', str(v or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.upper().split())


def _carga_meta(carga_id=None):
    """Carga de Meta Ads a usar: la indicada, o la más reciente subida.

    Las cargas no guardan qué periodo cubre el reporte (Meta no lo trae en el
    export), así que se devuelve también la fecha de subida para que la
    pantalla diga con qué archivo está calculando en vez de dar por hecho
    que corresponde al mes que se está mirando.
    """
    try:
        import meta_ads as mads
        d = os.path.join(os.environ.get('DATA_DIR', 'data'), 'meta_ads')
        cargas = mads.listar(d)
        if not cargas:
            return None
        elegida = next((c for c in cargas if c['id'] == carga_id), cargas[0])
        return mads.detalle(d, elegida['id'])
    except Exception:  # noqa: BLE001
        return None


def _meta_venta_mes(mes, anio):
    """Meta de VENTA del mes (meta_mensual.json), la del Panel.

    OJO: "meta" acá es objetivo de venta, no Meta/Facebook. Este archivo no
    tiene nada que ver con la inversión en publicidad — leerlo como gasto da
    un ratio inversión/venta inventado. La inversión sale del export de Meta
    Ads o de lo que se escriba a mano por campaña (ver _inversion_manual).
    """
    try:
        ruta = os.path.join(os.environ.get('DATA_DIR', 'data'), 'meta_mensual.json')
        with open(ruta, encoding='utf-8') as f:
            return float(json.load(f).get(f'{mes}-{anio}') or 0) or None
    except Exception:  # noqa: BLE001
        return None


def _inversion_manual(mes, anio):
    """Inversión por campaña escrita a mano para un mes (inversion_campanas.json).

    Sin API de Meta y sin export subido no hay forma de saber cuánto costó
    cada campaña, pero el equipo sí lo sabe: es el presupuesto que le puso.
    Lo escrito a mano manda sobre el export cuando existen los dos, porque
    el export puede cubrir sólo parte del mes.
    """
    try:
        ruta = os.path.join(os.environ.get('DATA_DIR', 'data'),
                            'inversion_campanas.json')
        with open(ruta, encoding='utf-8') as f:
            datos = json.load(f).get(f'{mes}-{anio}') or {}
        return {_norm_campana(k): float(v) for k, v in datos.items()
                if isinstance(v, (int, float))}
    except Exception:  # noqa: BLE001
        return {}


# Palabras que no distinguen una campaña de otra y sólo generan cruces falsos.
_RUIDO_CAMPANA = {'CAMPANA', 'LEADS', 'MENSAJES', 'WHATSAPP', 'ADS', 'META',
                  'NUEVO', 'NUEVA', 'PROMO', 'CLINICA', 'DERMA', 'ESSENZA'}


def _tokens_campana(nombre):
    """Palabras que de verdad identifican a una campaña.

    Se van los años y números sueltos ("TOXINA 2026" y "TOXINA FULL FACE" son
    la misma campaña con el año pegado) y las palabras de relleno.
    """
    return {p for p in re.split(r'[^A-Z0-9]+', _norm_campana(nombre))
            if len(p) >= 4 and not p.isdigit() and p not in _RUIDO_CAMPANA}


def _cruce_campana(nombre, disponibles):
    """Campaña de Meta que corresponde a una campaña de las hojas (o None).

    Los nombres nunca calzan literal: en Meta la campaña se llama
    "TOXINA 2026" o "CONSULTA GRATIS 2026" y en AGENDADOS se escribe
    "TOXINA FULL FACE" y "CONSULTA GRATUITA". Se prueba en tres niveles, del
    más seguro al más flexible: nombre idéntico, uno contenido en el otro, y
    por último palabras clave en común. Sin ninguna coincidencia se devuelve
    None y el gasto aparece en "sin cruce", que es lo correcto: mejor que se
    vea el gasto suelto a inventarle un embudo que no le corresponde.
    """
    n = _norm_campana(nombre)
    if not n:
        return None
    if n in disponibles:
        return disponibles[n]
    mejor = None
    for otro, dato in disponibles.items():
        if len(otro) >= 4 and (otro in n or n in otro):
            largo = min(len(otro), len(n))
            if mejor is None or largo > mejor[0]:
                mejor = (largo, dato)
    if mejor:
        return mejor[1]
    mios = _tokens_campana(n)
    if not mios:
        return None
    mejor = None
    for otro, dato in disponibles.items():
        comunes = mios & _tokens_campana(otro)
        if not comunes:
            continue
        # Más palabras en común gana; a igualdad, la palabra más larga (es
        # la más específica: "HIALURONICO" distingue más que "TOXINA").
        peso = (len(comunes), sum(len(x) for x in comunes))
        if mejor is None or peso > mejor[0]:
            mejor = (peso, dato)
    return mejor[1] if mejor else None


def _pct(parte, total):
    return round(100.0 * parte / total, 1) if total else 0.0


def plan_campanas(mes, anio, objetivo=OBJETIVO_INVERSION_VENTA, carga_id=None):
    """Embudo, costo e inversión/venta de cada campaña, para decidir presupuesto.

    Junta las dos mitades que hoy viven separadas:

      - Lo que costó traer al paciente: gasto y resultados (leads) por campaña
        del export de Meta Ads.
      - Lo que ese paciente dejó: agendados -> asistió -> realizó -> ticket
        promedio -> venta, sacado del maestro por su columna CAMPAÑA.

    Con las dos, cada campaña queda medida por ``inversion / venta`` (lo que
    la clínica llama "0.2" o "0.1": soles de publicidad por cada sol vendido)
    y por su ROAS.

    OJO con la lectura: mientras el costo por lead y las tasas no cambien,
    ese ratio NO mejora invirtiendo más — el doble de presupuesto da el doble
    de leads y el doble de venta, con el mismo ratio. Por eso se devuelve
    ``costo_lead_objetivo``: el costo por lead máximo con el que esa campaña
    llegaría al objetivo con su conversión y su ticket actuales. El
    presupuesto decide el VOLUMEN; el ratio lo deciden costo por lead,
    conversión y ticket.
    """
    mes = str(mes).strip().upper()
    anio = int(anio)
    if mes not in _MM_ORD or not _MM_ORD[mes]:
        return None
    objetivo = float(objetivo or 0) or OBJETIVO_INVERSION_VENTA

    # ---- embudo por campaña, desde el maestro ----
    filas, _ = _filas_maestro()
    emb = defaultdict(lambda: {'agendados': 0, 'asistidos': 0,
                               'realizados': 0, 'monto': 0.0})
    for f in filas:
        camp = str(f.get('campana') or '').strip() or '(sin campaña)'
        if _agendado(f, mes, anio, 1, 31):
            emb[camp]['agendados'] += 1
        if _asistido(f, mes, anio, 1, 31):
            emb[camp]['asistidos'] += 1
            m = _monto(f)
            emb[camp]['monto'] += m
            if m > 0:
                emb[camp]['realizados'] += 1

    # ---- costo por campaña, desde el export de Meta ----
    carga = _carga_meta(carga_id)
    por_meta = {}
    if carga:
        for c in carga.get('por_campania') or []:
            por_meta[_norm_campana(c.get('campania'))] = c
    usados = set()
    manual = _inversion_manual(mes, anio)

    campanas = []
    for camp, e in emb.items():
        m = _cruce_campana(camp, por_meta)
        if m:
            usados.add(_norm_campana(m.get('campania')))
        gasto = round(float(m['gasto']), 2) if m and m.get('gasto') else None
        a_mano = manual.get(_norm_campana(camp))
        if a_mano:
            gasto = round(a_mano, 2)
        # "Resultados" en Meta es lo que la campaña optimiza (mensajes,
        # formularios...): es el lead, el paso previo a que alguien lo agende.
        leads = int(m['resultados'] or 0) or None if m else None
        monto = round(e['monto'], 2)
        ticket = round(monto / e['realizados'], 2) if e['realizados'] else 0.0
        venta_lead = round(monto / leads, 2) if leads else None
        venta_agendado = round(monto / e['agendados'], 2) if e['agendados'] else 0.0
        ratio = round(gasto / monto, 3) if gasto and monto else None
        campanas.append({
            'campana': camp,
            'campana_meta': m.get('campania') if m else None,
            'gasto': gasto,
            'gasto_origen': 'manual' if a_mano else ('meta' if gasto else None),
            'leads': leads,
            'costo_lead': round(gasto / leads, 2) if gasto and leads else None,
            # Sin export de Meta no hay leads, pero el agendado sí es dato
            # propio: sirve de ancla para el mismo cálculo.
            'costo_agendado': round(gasto / e['agendados'], 2) if gasto and e['agendados'] else None,
            'costo_realizado': round(gasto / e['realizados'], 2) if gasto and e['realizados'] else None,
            'agendados': e['agendados'],
            'asistidos': e['asistidos'],
            'realizados': e['realizados'],
            'monto': monto,
            'ticket': ticket,
            'pct_lead_agenda': _pct(e['agendados'], leads) if leads else None,
            'pct_agenda_asiste': _pct(e['asistidos'], e['agendados']),
            'pct_asiste_realiza': _pct(e['realizados'], e['asistidos']),
            'pct_agenda_realiza': _pct(e['realizados'], e['agendados']),
            'venta_por_lead': venta_lead,
            'venta_por_agendado': venta_agendado,
            'ratio': ratio,
            'roas': round(monto / gasto, 2) if gasto and monto else None,
            # Techo de costo por lead para llegar al objetivo con la
            # conversión y el ticket que esta campaña tiene hoy.
            'costo_lead_objetivo': round(venta_lead * objetivo, 2) if venta_lead else None,
            'costo_agendado_objetivo': round(venta_agendado * objetivo, 2) if venta_agendado else None,
            'cumple': (ratio is not None and ratio <= objetivo),
        })
    campanas.sort(key=lambda c: (-(c['monto'] or 0), -(c['agendados'] or 0)))

    # Campañas que Meta cobró y que no aparecen en ninguna ficha: o el nombre
    # no calza con lo que se escribe en AGENDADOS, o no trajeron a nadie.
    # En los dos casos es gasto sin retorno visible y hay que mostrarlo.
    sin_cruce = [{'campana': c.get('campania'),
                  'gasto': round(float(c.get('gasto') or 0), 2),
                  'leads': int(c.get('resultados') or 0),
                  'costo_lead': c.get('costo_resultado')}
                 for k, c in por_meta.items()
                 if k not in usados and (c.get('gasto') or 0) > 0]
    sin_cruce.sort(key=lambda c: -c['gasto'])

    monto_total = round(sum(c['monto'] for c in campanas), 2)
    gasto_cruzado = round(sum(c['gasto'] or 0 for c in campanas), 2)
    gasto_total = round(gasto_cruzado + sum(c['gasto'] for c in sin_cruce), 2)
    # Sólo cuenta como inversión lo que de verdad se invirtió: el export de
    # Meta o lo escrito a mano por campaña. Si no hay ninguno de los dos, el
    # ratio queda en None y la pantalla lo pide, en vez de inventar un número.
    gasto_global = gasto_total or None
    meta_venta = _meta_venta_mes(mes, anio)

    return {
        'mes': mes, 'anio': anio, 'objetivo': objetivo,
        'carga': ({'id': carga['id'], 'archivo': carga['archivo'],
                   'fecha': carga['fecha']} if carga else None),
        'campanas': campanas,
        'sin_cruce': sin_cruce,
        'totales': {
            'gasto': gasto_global,
            'meta_venta': meta_venta,
            # Techo de inversión que esa meta de venta soporta al objetivo:
            # si la meta son 15 000 y el objetivo 0.1, más de 1 500 en
            # anuncios ya rompe el objetivo aunque la meta se cumpla.
            'inversion_para_meta': round(meta_venta * objetivo, 2) if meta_venta else None,
            'gasto_en_campanas': gasto_total,
            'leads': sum(c['leads'] or 0 for c in campanas) or None,
            'agendados': sum(c['agendados'] for c in campanas),
            'asistidos': sum(c['asistidos'] for c in campanas),
            'realizados': sum(c['realizados'] for c in campanas),
            'monto': monto_total,
            'ticket': round(monto_total / sum(c['realizados'] for c in campanas), 2)
                      if sum(c['realizados'] for c in campanas) else 0.0,
            'ratio': round(gasto_global / monto_total, 3) if gasto_global and monto_total else None,
            'roas': round(monto_total / gasto_global, 2) if gasto_global and monto_total else None,
        },
    }


def perfil(mes, anio, desde, hasta):
    filas, _ = _filas_maestro()
    trat = Counter(); dist = Counter(); sexo = Counter(); canal = Counter()
    camp = Counter(); edades = []
    for f in filas:
        if not _asistido(f, mes, anio, desde, hasta):
            continue
        for t in f['trats']:
            if t:
                trat[str(t).strip().upper()] += 1
        if f['distrito']:
            dist[str(f['distrito']).strip().title()] += 1
        if f['sexo']:
            sexo[str(f['sexo']).strip().upper()] += 1
        if f['canal']:
            canal[str(f['canal']).strip().title()] += 1
        if f['campana']:
            camp[str(f['campana']).strip()] += 1
        if isinstance(f['edad'], (int, float)):
            edades.append(int(f['edad']))

    def buckets(lst):
        bandas = {'18-25': 0, '26-35': 0, '36-45': 0, '46-55': 0, '56+': 0}
        for e in lst:
            if e < 26:
                bandas['18-25'] += 1
            elif e < 36:
                bandas['26-35'] += 1
            elif e < 46:
                bandas['36-45'] += 1
            elif e < 56:
                bandas['46-55'] += 1
            else:
                bandas['56+'] += 1
        return bandas

    return {
        'tratamientos': dict(trat.most_common(12)),
        'distritos': dict(dist.most_common(10)),
        'sexo': dict(sexo.most_common()),
        'canal': dict(canal.most_common()),
        'campanas': dict(camp.most_common(10)),
        'edad_bandas': buckets(edades),
        'edad_promedio': round(sum(edades) / len(edades), 1) if edades else 0,
        'total_asistidos': len([f for f in filas if _asistido(f, mes, anio, desde, hasta)]),
    }


def _totales_periodo(mes, anio, desde=1, hasta=31):
    """Totales (ag, as_, co, mon) de un periodo con rango de días dado."""
    filas, _ = _filas_maestro()
    ag = as_ = co = 0
    mon = 0.0
    for f in filas:
        if _agendado(f, mes, anio, desde, hasta):
            ag += 1
        if _asistido(f, mes, anio, desde, hasta):
            as_ += 1
            p = _monto(f)
            mon += p
            if p > 0:
                co += 1
    return {'agendados': ag, 'asistidos': as_, 'compraron': co,
            'monto': round(mon, 2)}


def comparativo(mes, anio, desde, hasta):
    actual = kpis(mes, anio, desde, hasta)
    pm, pa = _meses_previos(mes, anio, 1)[0]
    anterior = _totales_periodo(pm, pa, desde, hasta)
    anio_mes = _totales_periodo(mes, anio - 1, desde, hasta)
    return {
        'actual': actual,
        'anterior_mes': anterior,
        'mismo_mes_anio_anterior': anio_mes,
        'labels': [f'{_MM[i+1]}' for i in range(12)],
    }


def ejecutivas(mes, anio, desde, hasta):
    """Ranking por ejecutiva (AGENDADO POR): agendados, asistencia, conversión,
    monto y ticket promedio del periodo."""
    filas, _ = _filas_maestro()
    agrup = {}
    for f in filas:
        ej = str(f.get('ejecutiva') or '').strip() or 'SIN EJECUTIVA'
        d = agrup.setdefault(ej, {'agendados': 0, 'asistidos': 0, 'compraron': 0,
                                  'monto': 0.0, 'no_fueron': 0,
                                  'fueron_sin_compra': 0})
        if _agendado(f, mes, anio, desde, hasta):
            d['agendados'] += 1
            if (str(f['asistencia'] or '').strip() != 'ASISTIO'
                    and _cita_pasada(f)):
                d['no_fueron'] += 1
        if _asistido(f, mes, anio, desde, hasta):
            d['asistidos'] += 1
            p = _monto(f)
            d['monto'] += p
            if p > 0:
                d['compraron'] += 1
            else:
                d['fueron_sin_compra'] += 1
    out = []
    for ej, d in agrup.items():
        d['ejecutiva'] = ej
        d['monto'] = round(d['monto'], 2)
        d['asistencia_pct'] = round(100.0 * d['asistidos'] / d['agendados'], 1) \
            if d['agendados'] else 0
        d['conversion_pct'] = round(100.0 * d['compraron'] / d['asistidos'], 1) \
            if d['asistidos'] else 0
        d['ticket_promedio'] = round(d['monto'] / d['compraron'], 2) \
            if d['compraron'] else 0
        out.append(d)
    out.sort(key=lambda x: (x['monto'], x['agendados']), reverse=True)
    total = {'agendados': sum(x['agendados'] for x in out),
             'asistidos': sum(x['asistidos'] for x in out),
             'compraron': sum(x['compraron'] for x in out),
             'monto': round(sum(x['monto'] for x in out), 2)}
    return {'ejecutivas': out, 'total': total}


def motivos(mes, anio, desde, hasta):
    """Distribución de motivos de no asistencia y de no compra en el periodo."""
    filas, _ = _filas_maestro()
    no_asistio = Counter()
    no_compra = Counter()
    for f in filas:
        if _agendado(f, mes, anio, desde, hasta):
            if (str(f['asistencia'] or '').strip() != 'ASISTIO'
                    and _cita_pasada(f)):
                m = str(f.get('motivo_no_asistio') or '').strip()
                if m:
                    no_asistio[m.upper()] += 1
        if _asistido(f, mes, anio, desde, hasta):
            if _monto(f) == 0:
                m = str(f.get('motivo_no_compra') or '').strip()
                if m:
                    no_compra[m.upper()] += 1

    def lista(counter):
        tot = sum(counter.values())
        return [{'motivo': k, 'cantidad': v,
                 'pct': round(100.0 * v / tot, 1) if tot else 0}
                for k, v in counter.most_common()]

    return {
        'no_asistio': lista(no_asistio),
        'no_compra': lista(no_compra),
        'total_no_asistio': sum(no_asistio.values()),
        'total_no_compra': sum(no_compra.values()),
    }


def _compradores_frecuentes(filas, mes, anio, desde, hasta, min_compras=2):
    por_paciente = defaultdict(list)
    for f in filas:
        if not _asistido(f, mes, anio, desde, hasta):
            continue
        p = _monto(f)
        if p <= 0:
            continue
        clave = str(f['telefono'] or '').strip() or str(f['dni'] or '').strip()
        if not clave:
            continue
        por_paciente[clave].append(p)
    return {k: v for k, v in por_paciente.items() if len(v) >= min_compras}


def recurrentes(mes, anio, desde, hasta):
    filas, _ = _filas_maestro()
    por_paciente = defaultdict(lambda: {'compras': 0, 'monto': 0.0,
                                        'nombre': '', 'trats': [], 'ultima': ''})
    for f in filas:
        # Antes no se aplicaba ningún filtro de periodo aquí (a diferencia
        # de _compradores_frecuentes, que sí usa _asistido): con un mes/año
        # elegido en el dashboard, esta tarjeta seguía mostrando recurrentes
        # y LTV de TODO el histórico, sin relación con el periodo filtrado.
        if not _asistido(f, mes, anio, desde, hasta):
            continue
        clave = str(f['telefono'] or '').strip() or str(f['dni'] or '').strip()
        if not clave:
            continue
        p = _monto(f)
        if p > 0:
            d = por_paciente[clave]
            d['compras'] += 1
            d['monto'] += p
            if not d['nombre']:
                d['nombre'] = str(f['nombre'] or '').strip()
            for t in f['trats']:
                if t:
                    d['trats'].append(str(t).strip().upper())
            d['ultima'] = f'{f["dia_cita"]}/{f["mes_cita"]}/{f["anio_cita"]}'
    frecuentes = []
    for clave, d in por_paciente.items():
        if d['compras'] >= 2:
            frecuentes.append({
                'nombre': d['nombre'] or clave,
                'telefono': clave,
                'compras': d['compras'],
                'monto': round(d['monto'], 2),
                'ultima': d['ultima'],
            })
    frecuentes.sort(key=lambda x: x['compras'], reverse=True)
    ltv = round(sum(d['monto'] for d in por_paciente.values()) /
                len(por_paciente), 2) if por_paciente else 0
    return {'pacientes': frecuentes[:40],
            'total_pacientes': len(por_paciente),
            'total_recurrentes': len(frecuentes),
            'ltv_promedio': ltv}


def ventas_por_mes():
    """Histórico de ventas por mes (para gráfico de tendencia)."""
    filas, _ = _filas_maestro()
    por_mes = defaultdict(float)
    por_mes_as = defaultdict(int)
    por_mes_co = defaultdict(int)
    for f in filas:
        if f['anio_cita'] and f['mes_cita']:
            p = _monto(f)
            k = (int(f['anio_cita']), str(f['mes_cita']))
            por_mes[k] += p
            if p > 0:
                por_mes_co[k] += 1
    out = []
    claves = sorted(por_mes.items(),
                    key=lambda kv: (kv[0][0], _MM_ORD.get(str(kv[0][1]).strip().upper(), 0)))
    for (a, m), tot in claves:
        out.append({'anio': a, 'mes': m, 'monto': round(tot, 2),
                    'ventas': por_mes_co.get((a, m), 0)})
    return out[-24:]


def _norm_tel(v):
    import re
    return re.sub(r'\D', '', str(v or ''))


def historial_paciente(telefono, dni=None):
    """Historial completo de un paciente desde el maestro BD DATA."""
    filas, _ = _filas_maestro()
    tel = _norm_tel(telefono)
    dni_s = _norm_tel(dni) if dni else None
    citas = []
    for f in filas:
        ok = False
        if tel and _norm_tel(f['telefono']) == tel:
            ok = True
        elif dni_s and _norm_tel(f['dni']) == dni_s:
            ok = True
        if not ok:
            continue
        citas.append({
            'dia_cita': f['dia_cita'], 'mes_cita': f['mes_cita'],
            'anio_cita': f['anio_cita'],
            'fecha_cita': f'{f["dia_cita"]}/{f["mes_cita"]}/{f["anio_cita"]}'
                          if f['dia_cita'] and f['mes_cita'] and f['anio_cita'] else '',
            'dia_ag': f['dia_ag'], 'mes_ag': f['mes_ag'], 'anio_ag': f['anio_ag'],
            'fecha_ag': f'{f["dia_ag"]}/{f["mes_ag"]}/{f["anio_ag"]}'
                        if f['dia_ag'] and f['mes_ag'] and f['anio_ag'] else '',
            'campana': f['campana'], 'canal': f['canal'],
            'asistencia': f['asistencia'],
            'tratamientos': [str(t).strip() for t in f['trats'] if t],
            'montos': [round(m, 2) for m in f['montos'] if isinstance(m, (int, float))],
            'pago_total': round(_monto(f), 2),
            'distrito': f['distrito'], 'edad': f['edad'], 'sexo': f['sexo'],
            'nombre': f['nombre'],
        })
    citas.sort(key=lambda c: (str(c['anio_cita']), str(c['mes_cita']), c['dia_cita'] or 0))
    return citas


def pacientes_a_reactivar(meses=3):
    """Pacientes cuya última cita asistida fue hace más de N meses."""
    filas, _ = _filas_maestro()
    hoy = datetime.now()
    por_pac = {}
    for f in filas:
        if not (f['anio_cita'] and f['mes_cita'] and f['dia_cita']):
            continue
        if str(f['asistencia'] or '').strip() != 'ASISTIO':
            continue
        clave = _norm_tel(f['telefono']) or _norm_tel(f['dni'])
        if not clave:
            continue
        try:
            fecha = datetime(int(f['anio_cita']), _MM.index(str(f['mes_cita'])),
                             min(int(f['dia_cita']), 28))
        except Exception:  # noqa: BLE001
            continue
        p = por_pac.get(clave)
        if not p or fecha > p['ultima_fecha']:
            por_pac[clave] = {'ultima_fecha': fecha, 'nombre': f['nombre'],
                              'telefono': f['telefono'],
                              'monto_total': (por_pac[clave]['monto_total']
                                              if p else 0)}
    out = []
    for clave, p in por_pac.items():
        diff_meses = (hoy.year - p['ultima_fecha'].year) * 12 + \
                     hoy.month - p['ultima_fecha'].month
        if diff_meses > meses:
            out.append({
                'nombre': p['nombre'] or clave, 'telefono': p['telefono'],
                'ultima_cita': p['ultima_fecha'].strftime('%d/%m/%Y'),
                'meses_sin_volver': diff_meses,
            })
    out.sort(key=lambda x: x['meses_sin_volver'], reverse=True)
    return out[:100]
