# -*- coding: utf-8 -*-
"""crm_plus.py
============
Funciones estilo CRM (tipo KOMMO) sobre Google Drive: pipeline kanban,
tareas/recordatorios y notas de seguimiento por paciente.

Guarda todo en un workbook propio ``CRM.xlsx`` (hojas TARJETAS, TAREAS,
NOTAS, CUOTAS, CAJA, PRODUCTOS, MOVIMIENTOS_STOCK, TRABAJADORES y PLANILLA)
que se busca por nombre en Drive. Además agrega vistas derivadas: directorio
de pacientes, dashboard de métricas, panel de actividades de hoy (combinando
AGENDADOS y VENTA DIARIA), el catálogo/kardex de inventario y la planilla
quincenal de sueldos.
"""
import calendar
import json
import os
import re
import time
import unicodedata
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl

import alimentar_maestro as am  # noqa: E402
import crm_drive as cd          # noqa: E402

TZ = ZoneInfo('America/Lima')
NOMBRE_ARCHIVO = os.environ.get('CRM_ARCHIVO', 'CRM.xlsx')

ETAPAS = ['NUEVO', 'AGENDADO', 'CONFIRMADO', 'ATENDIDO', 'GANADO', 'PERDIDO']
TIPO_TAREA = ['LLAMADA', 'SEGUIMIENTO', 'RECORDATORIO', 'OTRO']
ESTADOS_TAREA = ['PENDIENTE', 'COMPLETADA', 'CANCELADA']
PRIORIDADES = ['ALTA', 'MEDIA', 'BAJA']
TIPO_NOTA = ['LLAMADA', 'WHATSAPP', 'INBOX', 'OTRO']

_MM = ['', 'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO',
       'SET', 'OCT', 'NOV', 'DIC']
_MM_IDX = {m: i for i, m in enumerate(_MM) if m}

HOJAS = {
    'TARJETAS': ['ID', 'NOMBRE', 'TELEFONO', 'ETAPA', 'CRM', 'CAMPANA', 'VALOR',
                 'PRIORIDAD', 'FECHA_ALTA', 'FECHA_ULT', 'CITA_DIA', 'CITA_MES',
                 'CITA_HORA', 'NOTA'],
    'TAREAS': ['ID', 'TITULO', 'TIPO', 'FECHA', 'HORA', 'ESTADO', 'PRIORIDAD',
               'CONTACTO', 'NOTA'],
    'NOTAS': ['ID', 'TELEFONO', 'CONTACTO', 'FECHA', 'TIPO', 'TEXTO'],
    'CUOTAS': ['ID', 'PACIENTE', 'TELEFONO', 'TRATAMIENTO', 'MONTO_TOTAL',
               'N_CUOTAS', 'PAGADAS', 'MONTO_CUOTA', 'PROX_FECHA', 'ESTADO',
               'NOTA'],
    'CAJA': ['ID', 'FECHA', 'HORA', 'TIPO', 'MONTO', 'CONCEPTO', 'RESPONSABLE',
             'DESGLOSE', 'MONTO_USD', 'DESGLOSE_USD'],
    'PRODUCTOS': ['ID', 'CODIGO', 'PRODUCTO', 'COSTO_BRUTO', 'STOCK',
                  'STOCK_MINIMO', 'UNIDAD', 'FECHA_ALTA', 'ACTUALIZADO'],
    'MOVIMIENTOS_STOCK': ['ID', 'FECHA', 'CODIGO', 'PRODUCTO', 'TIPO', 'CANTIDAD',
                          'STOCK_RESULTANTE', 'REFERENCIA', 'NOTA'],
    'TRABAJADORES': ['ID', 'NOMBRE', 'CARGO', 'SUELDO_QUINCENA', 'PORCENTAJE_COMISION',
                     'METODO_PAGO', 'CUENTA', 'ESTADO', 'FECHA_ALTA', 'NOMBRE_VENTAS'],
    'PLANILLA': ['ID', 'TRABAJADOR_ID', 'NOMBRE', 'ANIO', 'MES', 'QUINCENA',
                'SUELDO_BASE', 'COMISION', 'MONTO_TOTAL', 'ESTADO', 'FECHA_PAGO',
                'METODO_PAGO', 'NOTA', 'EXTRA', 'MOTIVO_EXTRA'],
    'HISTORIAS': ['ID', 'FECHA', 'HORA', 'PACIENTE', 'TELEFONO', 'DNI', 'EDAD',
                  'DOCTOR', 'MOTIVO', 'ANTECEDENTES', 'ALERGIAS', 'DIAGNOSTICO',
                  'TRATAMIENTO', 'INDICACIONES', 'PROXIMO_CONTROL',
                  'OBSERVACION', 'AGENDADO_FILA', 'DIRECCION', 'PATOLOGICO',
                  'PIEL', 'ANATOMIA', 'CITAS'],
}

TARJETA_COLS = {'id': 'A', 'nombre': 'B', 'telefono': 'C', 'etapa': 'D',
                'crm': 'E', 'campana': 'F', 'valor': 'G', 'prioridad': 'H',
                'fecha_alta': 'I', 'fecha_ult': 'J', 'cita_dia': 'K',
                'cita_mes': 'L', 'cita_hora': 'M', 'nota': 'N'}
TAREA_COLS = {'id': 'A', 'titulo': 'B', 'tipo': 'C', 'fecha': 'D', 'hora': 'E',
              'estado': 'F', 'prioridad': 'G', 'contacto': 'H', 'nota': 'I'}
NOTA_COLS = {'id': 'A', 'telefono': 'B', 'contacto': 'C', 'fecha': 'D',
             'tipo': 'E', 'texto': 'F'}
CUOTA_COLS = {'id': 'A', 'paciente': 'B', 'telefono': 'C', 'tratamiento': 'D',
              'monto_total': 'E', 'n_cuotas': 'F', 'pagadas': 'G',
              'monto_cuota': 'H', 'prox_fecha': 'I', 'estado': 'J', 'nota': 'K'}
CAJA_COLS = {'id': 'A', 'fecha': 'B', 'hora': 'C', 'tipo': 'D', 'monto': 'E',
             'concepto': 'F', 'responsable': 'G', 'desglose': 'H',
             'monto_usd': 'I', 'desglose_usd': 'J'}
TIPO_CAJA = ['APERTURA', 'INGRESO', 'EGRESO', 'CIERRE']
PRODUCTO_COLS = {'id': 'A', 'codigo': 'B', 'producto': 'C', 'costo_bruto': 'D',
                 'stock': 'E', 'stock_minimo': 'F', 'unidad': 'G',
                 'fecha_alta': 'H', 'actualizado': 'I'}
MOVIMIENTO_COLS = {'id': 'A', 'fecha': 'B', 'codigo': 'C', 'producto': 'D',
                   'tipo': 'E', 'cantidad': 'F', 'stock_resultante': 'G',
                   'referencia': 'H', 'nota': 'I'}
TIPO_MOVIMIENTO = ['ENTRADA', 'SALIDA', 'AJUSTE']
TRABAJADOR_COLS = {'id': 'A', 'nombre': 'B', 'cargo': 'C', 'sueldo_quincena': 'D',
                   'porcentaje_comision': 'E', 'metodo_pago': 'F', 'cuenta': 'G',
                   'estado': 'H', 'fecha_alta': 'I', 'nombre_ventas': 'J'}
ESTADO_TRABAJADOR = ['ACTIVO', 'INACTIVO']
PLANILLA_COLS = {'id': 'A', 'trabajador_id': 'B', 'nombre': 'C', 'anio': 'D',
                 'mes': 'E', 'quincena': 'F', 'sueldo_base': 'G', 'comision': 'H',
                 'monto_total': 'I', 'estado': 'J', 'fecha_pago': 'K',
                 'metodo_pago': 'L', 'nota': 'M', 'extra': 'N', 'motivo_extra': 'O'}
ESTADO_PAGO = ['PENDIENTE', 'PAGADO']
# La comisión se paga una sola vez al mes, en la segunda quincena, y se calcula
# sobre las ventas de todo el mes (así lo maneja el consultorio con el médico).
QUINCENA_COMISION = 2
# Denominaciones del arqueo (formato de recepción/caja)
BILLETES_SOLES = [200, 100, 50, 20, 10]
MONEDAS_SOLES = [5, 2, 1, 0.5, 0.2, 0.1]
DENOMINACIONES_CAJA = BILLETES_SOLES + MONEDAS_SOLES
DENOMINACIONES_USD = [100, 50, 20, 10]

_lock = am.LockConTiempos('CRM.xlsx', reentrante=True)


def _ahora():
    return datetime.now(TZ).strftime('%d/%m/%Y %H:%M')


def _hoy():
    hoy = datetime.now(TZ)
    return hoy.day, _MM[hoy.month], hoy.year


def _fecha_hoy_str():
    dia, mes, anio = _hoy()
    return f'{dia}/{mes}/{anio}'


def _fecha_orden(fecha):
    try:
        d, m, a = str(fecha).split('/')
        return (int(a), _MM_IDX.get(m, 0), int(d))
    except (ValueError, AttributeError):
        return (0, 0, 0)


# ============================================================
# WORKBOOK CRM.xlsx (Drive)
# ============================================================
_FID_CACHE = None  # el fid de CRM.xlsx no cambia en la vida del proceso (medido:
                   # ~500ms por búsqueda sin caché, y se llamaba 2 veces por escritura)


def _fid():
    global _FID_CACHE
    if _FID_CACHE:
        return _FID_CACHE
    with am.cronometro('crm_plus _fid() [drive, busca por nombre, 1 vez por proceso]', umbral_ms=200):
        drv = cd._drive()
        res = drv.files().list(
            q=f"name='{NOMBRE_ARCHIVO}' and trashed=false",
            fields='files(id)', pageSize=10).execute()
    archivos = res.get('files', [])
    if archivos:
        _FID_CACHE = archivos[0]['id']
        return _FID_CACHE
    # La cuenta de servicio no puede crear archivos (no tiene cuota de
    # almacenamiento): pedimos un archivo vacío compartido con el bot.
    try:
        correo = cd._credenciales().service_account_email
    except Exception:  # noqa: BLE001
        correo = 'la cuenta de servicio'
    raise RuntimeError(
        f'No se encontró el archivo "{NOMBRE_ARCHIVO}" en Google Drive. '
        f'Crea un archivo Excel vacío con ese nombre y compártelo con '
        f'{correo} (edición) para habilitar el CRM.')


def _bajar():
    """Descarga CRM.xlsx de Drive. Si es un Google Sheets, lo exporta a xlsx."""
    ruta = os.path.join(am.TMP_DIR, NOMBRE_ARCHIVO)
    if os.path.exists(ruta) and time.time() - os.path.getmtime(ruta) < cd.CACHE_TTL:
        return ruta
    with am.lock_de_descarga(ruta):
        # Otra request pudo haber refrescado el archivo mientras esperábamos
        # el candado (evita una descarga redundante y, sobre todo, evita que
        # dos descargas escriban el mismo archivo a la vez).
        if os.path.exists(ruta) and time.time() - os.path.getmtime(ruta) < cd.CACHE_TTL:
            return ruta
        with am.cronometro(f'crm_plus _bajar({NOMBRE_ARCHIVO}) [drive, cache vencido]'):
            fid = _fid()
            drv = cd._drive()
            meta = drv.files().get(fileId=fid, fields='mimeType').execute()
            if meta.get('mimeType') == cd.MIME_XLSX:
                req = drv.files().get_media(fileId=fid)
            else:
                req = drv.files().export(fileId=fid, mimeType=cd.MIME_XLSX)
            # Se escribe a un .tmp y se renombra al final (rename es atómico
            # en POSIX): así ruta nunca existe a medio escribir. El candado
            # de arriba evita descargas simultáneas redundantes, pero el
            # chequeo rápido de "¿está fresco?" de otras llamadas (más
            # arriba, sin candado) igual podía leer ruta justo mientras
            # open(ruta, 'wb') la había truncado pero no terminado de
            # llenar — eso es lo que producía "File is not a zip file".
            tmp = ruta + '.tmp'
            with open(tmp, 'wb') as fh:
                dl = cd.MediaIoBaseDownload(fh, req)
                done = False
                while not done:
                    _, done = dl.next_chunk()
            os.replace(tmp, ruta)
    return ruta


_WB_CACHE = {}  # ruta -> (mtime, Workbook)


def _cargar_wb(ruta):
    """Parsea CRM.xlsx una sola vez por archivo (por mtime): una sola vista
    de HOJA/AS puede pedir 10+ hojas del mismo libro (p.ej. generar_planilla_
    quincena lee todo HOJAS), y reabrir+parsear el .xlsx en cada una era el
    cuello de botella real de las pantallas lentas."""
    mtime = os.path.getmtime(ruta)
    cache = _WB_CACHE.get(ruta)
    if cache and cache[0] == mtime:
        return cache[1]
    if cache:
        try:
            cache[1].close()
        except Exception:  # noqa: BLE001
            pass
    with am.cronometro(f'crm_plus parsear xlsx ({os.path.basename(ruta)}) [cache miss]'):
        wb = openpyxl.load_workbook(ruta, data_only=True)
    _WB_CACHE[ruta] = (mtime, wb)
    return wb


def _leer_hoja(nombre):
    ruta = _bajar()
    wb = _cargar_wb(ruta)
    if nombre not in wb.sheetnames:
        return []
    ws = wb[nombre]
    cols = [c for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value is not None]
    filas = []
    for r in range(2, ws.max_row + 1):
        f = {}
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if v is None or v == '':
                continue
            f[openpyxl.utils.get_column_letter(c)] = v
        if f:
            filas.append(f)
    return filas


def _filas_a_valores(nombre, filas):
    """[{letra: valor}] -> filas en el formato de rango 2D de la API de
    Sheets, en el orden de columnas de HOJAS[nombre]."""
    letras = [openpyxl.utils.get_column_letter(i) for i in range(1, len(HOJAS[nombre]) + 1)]
    return [[f.get(letra, '') for letra in letras] for f in filas]


def _guardar(filas_por_hoja):
    """Reemplaza el contenido (desde la fila 2) de cada hoja indicada, vía la
    API de Sheets. CRM.xlsx es una Google Sheet nativa (no un .xlsx binario
    subido): escribir por la API de Sheets evita la exportación/reimportación
    de formato que hacía lenta cada escritura (medido: 4-6s por acción, la
    mayor parte en subir el archivo completo). Sólo se tocan las hojas que
    vienen en filas_por_hoja; las demás quedan intactas sin necesidad de
    "preservarlas" — nunca se reconstruye el archivo entero."""
    with _lock, am.cronometro('crm_plus _guardar() [Sheets API, solo hojas tocadas]'):
        fid = _fid()
        sheets = cd._sheets()
        for nombre, filas in filas_por_hoja.items():
            ultima_col = openpyxl.utils.get_column_letter(len(HOJAS[nombre]))
            with am.cronometro(f'crm_plus _guardar(): {nombre} [drive]'):
                sheets.spreadsheets().values().clear(
                    spreadsheetId=fid, range=f"'{nombre}'!A2:{ultima_col}", body={}).execute()
                if filas:
                    sheets.spreadsheets().values().update(
                        spreadsheetId=fid, range=f"'{nombre}'!A2",
                        valueInputOption='RAW',
                        body={'values': _filas_a_valores(nombre, filas)}).execute()
        cd.invalidar(os.path.splitext(NOMBRE_ARCHIVO)[0])


def _reescribir(hoja, filas):
    """Reemplaza el contenido completo de una hoja (fila 2 en adelante)."""
    _guardar({hoja: filas})


def _siguiente_id(filas):
    return max([am.num(f.get('A')) or 0 for f in filas] + [0]) + 1


# ============================================================
# TARJETAS (pipeline)
# ============================================================
def _aplicar_tarjeta(f, d):
    def pon(campo, letra):
        if campo in d and d[campo] not in (None, ''):
            f[letra] = str(d[campo]).strip()
    if 'telefono' in d:
        tel = re.sub(r'\D', '', str(d.get('telefono') or ''))
        if tel:
            f['C'] = tel
    if 'etapa' in d:
        etapa = str(d.get('etapa') or '').upper().strip()
        if etapa in ETAPAS:
            f['D'] = etapa
    if 'prioridad' in d:
        p = str(d.get('prioridad') or '').upper().strip()
        if p in PRIORIDADES:
            f['H'] = p
    if 'valor' in d and am.num(d.get('valor')) is not None:
        f['G'] = am.num(d['valor'])
    if 'cita_dia' in d and am.num(d.get('cita_dia')) is not None:
        f['K'] = am.num(d['cita_dia'])
    if 'cita_mes' in d:
        m = cd._mes(d.get('cita_mes'))
        if m:
            f['L'] = m
    for campo, letra in (('nombre', 'B'), ('crm', 'E'), ('campana', 'F'),
                         ('cita_hora', 'M'), ('nota', 'N')):
        pon(campo, letra)


def _tarjeta_named(f):
    return {
        'id': am.num(f.get('A')), 'nombre': f.get('B'), 'telefono': f.get('C'),
        'etapa': f.get('D') or 'NUEVO', 'crm': f.get('E'), 'campana': f.get('F'),
        'valor': am.num(f.get('G')), 'prioridad': f.get('H') or 'MEDIA',
        'fecha_alta': f.get('I'), 'fecha_ult': f.get('J'),
        'cita_dia': am.num(f.get('K')), 'cita_mes': f.get('L'),
        'cita_hora': f.get('M'), 'nota': f.get('N'),
    }


def leer_tarjetas():
    return [_tarjeta_named(f) for f in _leer_hoja('TARJETAS')]


def crear_tarjeta(datos):
    with _lock:
        filas = _leer_hoja('TARJETAS')
        f = {'A': _siguiente_id(filas), 'I': _ahora(), 'J': _ahora()}
        _aplicar_tarjeta(f, datos)
        f['D'] = f.get('D') or 'NUEVO'
        f['H'] = f.get('H') or 'MEDIA'
        filas.append(f)
        _reescribir('TARJETAS', filas)
        return _tarjeta_named(f)


def actualizar_tarjeta(tid, cambios):
    with _lock:
        filas = _leer_hoja('TARJETAS')
        for f in filas:
            if am.num(f.get('A')) == tid:
                _aplicar_tarjeta(f, cambios)
                f['J'] = _ahora()
                _reescribir('TARJETAS', filas)
                return _tarjeta_named(f)
    return None


def borrar_tarjeta(tid):
    with _lock:
        filas = _leer_hoja('TARJETAS')
        nuevas = [f for f in filas if am.num(f.get('A')) != tid]
        if len(nuevas) == len(filas):
            return False
        _reescribir('TARJETAS', nuevas)
        return True


# ============================================================
# TAREAS
# ============================================================
def _aplicar_tarea(f, d):
    def pon(campo, letra):
        if campo in d and d[campo] not in (None, ''):
            f[letra] = str(d[campo]).strip()
    if 'tipo' in d:
        t = str(d.get('tipo') or '').upper().strip()
        if t in TIPO_TAREA:
            f['C'] = t
    if 'estado' in d:
        e = str(d.get('estado') or '').upper().strip()
        if e in ESTADOS_TAREA:
            f['F'] = e
    if 'prioridad' in d:
        p = str(d.get('prioridad') or '').upper().strip()
        if p in PRIORIDADES:
            f['G'] = p
    for campo, letra in (('titulo', 'B'), ('fecha', 'D'), ('hora', 'E'),
                         ('contacto', 'H'), ('nota', 'I')):
        pon(campo, letra)


def _tarea_named(f):
    return {
        'id': am.num(f.get('A')), 'titulo': f.get('B'), 'tipo': f.get('C'),
        'fecha': f.get('D'), 'hora': f.get('E'), 'estado': f.get('F') or 'PENDIENTE',
        'prioridad': f.get('G') or 'MEDIA', 'contacto': f.get('H'),
        'nota': f.get('I'),
    }


def leer_tareas(estado=None):
    filas = _leer_hoja('TAREAS')
    out = [_tarea_named(f) for f in filas]
    if estado:
        out = [t for t in out if t['estado'] == estado]
    out.sort(key=lambda t: (t['fecha'] or '9999', t['hora'] or '99'))
    return out


def crear_tarea(datos):
    with _lock:
        filas = _leer_hoja('TAREAS')
        f = {'A': _siguiente_id(filas)}
        _aplicar_tarea(f, datos)
        f['C'] = f.get('C') or 'OTRO'
        f['F'] = f.get('F') or 'PENDIENTE'
        f['G'] = f.get('G') or 'MEDIA'
        filas.append(f)
        _reescribir('TAREAS', filas)
        return _tarea_named(f)


def actualizar_tarea(tid, cambios):
    with _lock:
        filas = _leer_hoja('TAREAS')
        for f in filas:
            if am.num(f.get('A')) == tid:
                _aplicar_tarea(f, cambios)
                _reescribir('TAREAS', filas)
                return _tarea_named(f)
    return None


def borrar_tarea(tid):
    with _lock:
        filas = _leer_hoja('TAREAS')
        nuevas = [f for f in filas if am.num(f.get('A')) != tid]
        if len(nuevas) == len(filas):
            return False
        _reescribir('TAREAS', nuevas)
        return True


# ============================================================
# NOTAS de seguimiento
# ============================================================
def _aplicar_nota(f, d):
    if d.get('telefono'):
        tel = re.sub(r'\D', '', str(d['telefono']))
        if tel:
            f['B'] = tel
    if d.get('contacto') not in (None, ''):
        f['C'] = str(d['contacto']).strip()
    if 'tipo' in d:
        t = str(d.get('tipo') or '').upper().strip()
        if t in TIPO_NOTA:
            f['E'] = t
    if d.get('texto') not in (None, ''):
        f['F'] = str(d['texto']).strip()


def _nota_named(f):
    return {
        'id': am.num(f.get('A')), 'telefono': f.get('B'), 'contacto': f.get('C'),
        'fecha': f.get('D'), 'tipo': f.get('E'), 'texto': f.get('F'),
    }


def leer_notas(telefono=None, contacto=None):
    filas = _leer_hoja('NOTAS')
    out = [_nota_named(f) for f in filas]
    if telefono:
        tel = re.sub(r'\D', '', str(telefono))
        out = [n for n in out if n['telefono'] and tel in str(n['telefono'])]
    elif contacto:
        c = str(contacto).strip().lower()
        out = [n for n in out if c and c in str(n.get('contacto') or '').lower()]
    out.sort(key=lambda n: n['fecha'] or '', reverse=True)
    return out


def agregar_nota(datos):
    with _lock:
        filas = _leer_hoja('NOTAS')
        f = {'A': _siguiente_id(filas), 'D': _ahora()}
        _aplicar_nota(f, datos)
        f['E'] = f.get('E') or 'OTRO'
        filas.append(f)
        _reescribir('NOTAS', filas)
        return _nota_named(f)


# ============================================================
# DIRECTORIO DE PACIENTES (agregado)
# ============================================================
def _clave(tel, nombre):
    if tel:
        return 't' + tel
    return 'n' + str(nombre or '').strip().lower()


# Orden de meses tal como se escriben en las hojas (SET, no SEP)
_MESES_NUM = {'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
              'JUL': 7, 'AGO': 8, 'SET': 9, 'SEP': 9, 'OCT': 10, 'NOV': 11,
              'DIC': 12}


def _fecha_iso(dia, mes, anio):
    """'YYYY-MM-DD' a partir de día/mes-texto/año de las hojas. None si falta algo."""
    d = am.num(dia)
    a = am.num(anio)
    m = _MESES_NUM.get(_status_normalizado(mes))
    if not d or not m or not a:
        return None
    return f'{int(a):04d}-{m:02d}-{int(d):02d}'


def _asistio_agendado(v):
    """True si el agendado quedó marcado como que el paciente fue al consultorio."""
    s = _status_normalizado(v)
    return 'ASISTIO' in s and 'NO ASISTIO' not in s


def _vino_a_consulta(v):
    """True si una fila de VENTA DIARIA implica que el paciente fue al consultorio.

    Una fila en VENTA DIARIA significa que el paciente vino; sea que compró
    (SE REALIZO) o que sólo fue la consulta (NO SE REALIZO). Sólo se descarta
    si el status dice explícitamente que no asistió.
    """
    return 'NO ASISTIO' not in _status_normalizado(v)


def _claves_persona(tel, nombre):
    """Claves para cruzar a una misma persona: por teléfono y por nombre."""
    claves = []
    t = _tel_digitos(tel)
    if t:
        claves.append('t' + t)
    n = _status_normalizado(nombre)
    if n:
        claves.append('n' + n)
    return claves


def asistencia_agendados(filas=None):
    """Cruza AGENDADOS con VENTA DIARIA para saber quién fue y quién compró.

    No basta con la columna CONFIRMADO: en la práctica casi nunca se marca, así
    que la asistencia se deduce también de tener una venta el mismo día de la
    cita. Devuelve {fila: {'asistio', 'compro', 'monto', 'fecha_cita'}}.
    """
    if filas is None:
        filas = cd.leer_agendados()['filas']

    ventas = {}
    for v in cd.leer_venta()['hojas'].values():
        for f in v['filas']:
            iso = _fecha_iso(f.get('B'), f.get('C'), f.get('D'))
            if not iso:
                continue
            for k in _claves_persona(f.get('F'), f.get('G')):
                d = ventas.setdefault((k, iso),
                                      {'vino': False, 'compro': False, 'monto': 0.0})
                if _vino_a_consulta(f.get('N')):
                    d['vino'] = True
                if _es_venta_registrada(f.get('N')):
                    d['compro'] = True
                    d['monto'] += am.num(f.get('O')) or 0

    out = {}
    for f in filas:
        fila = f.get('_fila')
        if not isinstance(fila, int):
            continue
        iso = _fecha_iso(f.get('L'), f.get('M'), f.get('N'))
        asistio = _asistio_agendado(f.get('Q'))
        compro, monto = False, 0.0
        if iso:
            for k in _claves_persona(f.get('I'), f.get('G')):
                d = ventas.get((k, iso))
                if not d:
                    continue
                asistio = asistio or d['vino']
                compro = compro or d['compro']
                monto = max(monto, d['monto'])
        out[fila] = {'asistio': asistio, 'compro': compro, 'monto': monto,
                     'fecha_cita': iso}
    return out


def leer_pacientes(desde=None, hasta=None, solo_atendidos=True):
    """Directorio de pacientes.

    Con ``solo_atendidos`` (por defecto) sólo devuelve a quienes fueron al
    consultorio —compraron o no—, no a todos los agendados. ``desde``/``hasta``
    ('YYYY-MM-DD') acotan por fecha de atención.
    """
    ag = cd.leer_agendados()['filas']
    ve = [dict(f, _hoja=hoja)
          for hoja, v in cd.leer_venta()['hojas'].items() for f in v['filas']]
    tar = leer_tarjetas()
    notas = _leer_hoja('NOTAS')

    pacientes = {}

    def base(nombre, tel):
        k = _clave(tel, nombre)
        p = pacientes.get(k)
        if not p:
            p = {'clave': k, 'nombre': nombre, 'telefono': tel, 'correo': '',
                 'crm': '', 'campana': '', 'etapa': None, 'citas': 0,
                 'compras': 0, 'total': 0.0, 'proxima_cita': None,
                 'ultima_actividad': None, 'notas': 0, 'doctores': '',
                 'atendido': False, 'atenciones': 0, 'ultima_atencion': None,
                 'ultima_atencion_hora': None, 'tratamientos': []}
            pacientes[k] = p
        return p

    def marcar_atencion(p, fecha, hora=None):
        p['atendido'] = True
        if fecha and (p['ultima_atencion'] is None or fecha > p['ultima_atencion']):
            p['ultima_atencion'] = fecha
            p['ultima_atencion_hora'] = str(hora).strip() if hora else None

    for f in ag:
        tel = f.get('I')
        tel = re.sub(r'\D', '', str(tel)) if tel is not None else ''
        p = base(f.get('G'), tel)
        if not p['crm'] and f.get('B'):
            p['crm'] = str(f['B']).strip()
        if not p['campana'] and f.get('O'):
            p['campana'] = str(f['O']).strip()
        if not p['correo'] and f.get('J'):
            p['correo'] = str(f['J']).strip()
        p['citas'] += 1
        if _asistio_agendado(f.get('Q')):
            marcar_atencion(p, _fecha_iso(f.get('L'), f.get('M'), f.get('N')), f.get('P'))
        if f.get('L') and f.get('M') and f.get('N'):
            fecha = f'{f["L"]}/{f["M"]}/{f["N"]}'
            hora = f.get('P') or ''
            if p['proxima_cita'] is None or fecha > p['proxima_cita'].split(' ')[0]:
                p['proxima_cita'] = f'{fecha} {hora}'.strip()
            if p['ultima_actividad'] is None or fecha > p['ultima_actividad'].split(' ')[0]:
                p['ultima_actividad'] = f'{fecha} {hora}'.strip()

    for f in ve:
        tel = f.get('F')
        tel = re.sub(r'\D', '', str(tel)) if tel is not None else ''
        p = base(f.get('G'), tel)
        if not p['crm'] and f.get('H'):
            p['crm'] = str(f['H']).strip()
        if not p['campana'] and f.get('L'):
            p['campana'] = str(f['L']).strip()
        if f.get('O'):
            p['compras'] += 1
            p['total'] = (p['total'] or 0) + (am.num(f['O']) or 0)
        if f.get('M') and str(f['M']).strip() not in p['doctores']:
            p['doctores'] = (p['doctores'] + ' · ' if p['doctores'] else '') + str(f['M']).strip()
        if f.get('B') and f.get('C') and f.get('D'):
            fecha = f'{f["B"]}/{f["C"]}/{f["D"]}'
            if p['ultima_actividad'] is None or fecha > p['ultima_actividad'].split(' ')[0]:
                p['ultima_actividad'] = fecha
        if _vino_a_consulta(f.get('N')):
            iso = _fecha_iso(f.get('B'), f.get('C'), f.get('D'))
            marcar_atencion(p, iso)
            p['atenciones'] += 1
            p['tratamientos'].append({
                'fecha': iso,
                'fecha_texto': f'{f.get("B") or ""} {f.get("C") or ""} {f.get("D") or ""}'.strip(),
                'tratamiento': str(f.get('L') or '').strip(),
                'doctor': str(f.get('M') or '').strip(),
                'status': str(f.get('N') or '').strip(),
                'precio': am.num(f.get('O')) or 0,
                'pago': str(f.get('P') or '').strip(),
                'hoja': f.get('_hoja', ''), 'fila': f.get('_fila'),
            })

    for t in tar:
        tel = re.sub(r'\D', '', str(t.get('telefono') or ''))
        p = base(t.get('nombre'), tel)
        if not p['crm'] and t.get('crm'):
            p['crm'] = str(t['crm']).strip()
        if not p['campana'] and t.get('campana'):
            p['campana'] = str(t['campana']).strip()
        if t.get('etapa'):
            p['etapa'] = t['etapa']
        if t.get('fecha_ult') and (p['ultima_actividad'] is None
                                   or t['fecha_ult'] > p['ultima_actividad']):
            p['ultima_actividad'] = t['fecha_ult']

    for f in notas:
        tel = f.get('B')
        tel = re.sub(r'\D', '', str(tel)) if tel is not None else ''
        p = base(f.get('C'), tel)
        p['notas'] += 1
        if f.get('D') and (p['ultima_actividad'] is None
                           or f['D'] > p['ultima_actividad']):
            p['ultima_actividad'] = f['D']

    out = list(pacientes.values())
    if solo_atendidos:
        out = [p for p in out if p['atendido']]
    if desde or hasta:
        def en_rango(p):
            fechas = [t['fecha'] for t in p['tratamientos'] if t['fecha']]
            if p['ultima_atencion']:
                fechas.append(p['ultima_atencion'])
            return any((not desde or fe >= desde) and (not hasta or fe <= hasta)
                       for fe in fechas)
        out = [p for p in out if en_rango(p)]
    for p in out:
        p['tratamientos'].sort(key=lambda t: (t['fecha'] or ''), reverse=True)
    out.sort(key=lambda p: (p['ultima_atencion'] or p['ultima_actividad'] or ''),
             reverse=True)
    return out


def leer_pacientes_inactivos(dias=45):
    """Pacientes atendidos alguna vez pero sin visita en más de ``dias`` días.

    Ordenados por gasto histórico (``total``) descendente, para priorizar a
    quién llamar primero para reactivar.
    """
    hoy_iso = datetime.now(TZ).strftime('%Y-%m-%d')
    out = []
    for p in leer_pacientes(solo_atendidos=True):
        if not p['ultima_atencion']:
            continue
        dias_inactivo = (datetime.strptime(hoy_iso, '%Y-%m-%d')
                          - datetime.strptime(p['ultima_atencion'], '%Y-%m-%d')).days
        if dias_inactivo > dias:
            q = dict(p)
            q['dias_inactivo'] = dias_inactivo
            out.append(q)
    out.sort(key=lambda p: (p['total'] or 0), reverse=True)
    return out


# ============================================================
# DASHBOARD
# ============================================================
def _status_normalizado(v):
    s = unicodedata.normalize('NFD', str(v or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return s.strip().upper()


def _es_no_realizado(v):
    s = _status_normalizado(v)
    return 'NO SE REALIZO' in s or 'NO ASISTIO' in s


def _es_venta_registrada(v):
    """Replica esVentaRegistrada() del frontend: cuenta como venta un status
    REALIZO/COMPRO/COMPLETA/SESION/DEJO PAGADO, salvo que sea NO SE REALIZO
    o NO ASISTIO."""
    s = _status_normalizado(v)
    if _es_no_realizado(v):
        return False
    return any(x in s for x in am.COMPRA_POR_TEXTO)


def leer_dashboard(anio=None, mes=None):
    """Resumen del pipeline/ventas/agendados. Sin anio/mes, es histórico
    (todo lo que hay). Con ambos, sólo cuenta lo de ese mes: agendados por
    su fecha de agendado (no de cita) y ventas por su fecha de venta."""
    mes = str(mes).strip().upper() if mes else None
    anio = int(anio) if anio else None
    filtra_mes = bool(anio and mes)

    tar = leer_tarjetas()
    ag = cd.leer_agendados()['filas']
    ve = cd.leer_venta()['hojas']
    if filtra_mes:
        ag = [f for f in ag if str(f.get('D') or '').strip().upper() == mes
              and am.num(f.get('E')) == anio]

    funnel = {e: 0 for e in ETAPAS}
    for t in tar:
        funnel[t['etapa']] = funnel.get(t['etapa'], 0) + 1

    ventas_doctor = {}
    ventas_mes = {}
    total_ventas = 0
    ventas_registradas = 0
    for v in ve.values():
        for f in v['filas']:
            if not _es_venta_registrada(f.get('N')):
                continue
            if filtra_mes and (str(f.get('C') or '').strip().upper() != mes
                               or am.num(f.get('D')) != anio):
                continue
            monto = am.num(f.get('O')) or 0
            ventas_registradas += 1
            if monto:
                total_ventas += monto
                doc = str(f.get('M') or '').strip() or 'Sin doctor'
                ventas_doctor[doc] = ventas_doctor.get(doc, 0) + monto
                mes = str(f.get('C') or '').strip().upper()
                anio = am.num(f.get('D')) or datetime.now(TZ).year
                if mes:
                    k = f'{mes} {anio}'
                    ventas_mes[k] = ventas_mes.get(k, 0) + monto

    ag_por_crm = {}
    ag_por_campana = {}
    for f in ag:
        crm = str(f.get('B') or '').strip() or 'Sin CRM'
        camp = str(f.get('O') or '').strip() or 'Sin campaña'
        ag_por_crm[crm] = ag_por_crm.get(crm, 0) + 1
        ag_por_campana[camp] = ag_por_campana.get(camp, 0) + 1

    leads = sum(funnel.get(e, 0) for e in ('NUEVO', 'AGENDADO', 'CONFIRMADO',
                                            'ATENDIDO'))
    ganados = funnel.get('GANADO', 0)
    conversion = round(100 * ganados / leads, 1) if leads else 0

    return {
        'funnel': funnel,
        'conversion': conversion,
        'ganados': ganados,
        'perdidos': funnel.get('PERDIDO', 0),
        'total_ventas': round(total_ventas, 2),
        'ventas_registradas': ventas_registradas,
        'ventas_doctor': dict(sorted(ventas_doctor.items(),
                                     key=lambda kv: kv[1], reverse=True)),
        'ventas_mes': dict(sorted(ventas_mes.items(),
                                  key=lambda kv: kv[1], reverse=True)),
        'ag_por_crm': dict(sorted(ag_por_crm.items(),
                                  key=lambda kv: kv[1], reverse=True)),
        'ag_por_campana': dict(sorted(ag_por_campana.items(),
                                      key=lambda kv: kv[1], reverse=True)),
    }


# ============================================================
# ACTIVIDADES DE HOY
# ============================================================
# Ventana (en días) alrededor de hoy en la que se busca una venta escrita con
# la fecha equivocada. Más de esto ya no es un dedazo al escribir el día, es
# otra visita del paciente.
DIAS_DESFASE_VENTA = 3


def _cruzar_ventas_con_otra_fecha(citas, ag_filas, hojas_venta, hoy_iso):
    """Marca las citas de hoy cuyo paciente tiene una venta con OTRA fecha.

    La fila de VENTA DIARIA se escribe a mano en el Drive y ahí es fácil
    equivocarse de día (típico: se arrastra la fecha de la fila de arriba).
    Cuando pasa, la venta existe pero queda invisible para el resto del
    sistema: la caja del día sólo suma las filas con la fecha de hoy, y la
    cita se queda para siempre en "asistió, falta saber si compró". Nada
    cruzaba AGENDADOS con VENTA DIARIA salvo por fecha exacta, así que el
    error no se notaba hasta cuadrar caja.

    Sólo se marca la venta cuando el paciente NO tuvo cita ese día: si la
    tuvo, esa fecha es legítima (es otra visita suya) y no hay nada que
    corregir.
    """
    try:
        hoy_fecha = datetime.strptime(hoy_iso, '%Y-%m-%d').date()
    except ValueError:
        return

    citas_persona = set()
    for f in ag_filas:
        iso = _fecha_iso(f.get('L'), f.get('M'), f.get('N'))
        if not iso:
            continue
        for k in _claves_persona(f.get('I'), f.get('G')):
            citas_persona.add((k, iso))

    sueltas = []
    for hoja, v in hojas_venta.items():
        for f in v['filas']:
            iso = _fecha_iso(f.get('B'), f.get('C'), f.get('D'))
            if not iso or iso == hoy_iso:
                continue
            try:
                dias = (datetime.strptime(iso, '%Y-%m-%d').date() - hoy_fecha).days
            except ValueError:
                continue
            if abs(dias) > DIAS_DESFASE_VENTA:
                continue
            claves = set(_claves_persona(f.get('F'), f.get('G')))
            if not claves or any((k, iso) in citas_persona for k in claves):
                continue
            sueltas.append([claves, hoja, f])

    if not sueltas:
        return
    usadas = set()
    for c in citas:
        if 'CANCEL' in _status_normalizado(c.get('estado')):
            continue
        mias = set(_claves_persona(c.get('telefono'), c.get('nombre')))
        for i, (claves, hoja, f) in enumerate(sueltas):
            if i in usadas or not (mias & claves):
                continue
            usadas.add(i)
            c['venta_desfasada'] = {
                'hoja': hoja, 'fila': f.get('_fila'),
                'fecha': f'{f.get("B")}/{f.get("C")}/{f.get("D")}',
                'tratamiento': f.get('L'), 'status': f.get('N'),
                'monto': am.num(f.get('O')), 'pago': f.get('P'),
            }
            break


def leer_hoy():
    dia, mes, anio = _hoy()
    clave_hoy = f'{dia}/{mes}/{anio}'
    hoy_iso = datetime.now(TZ).strftime('%Y-%m-%d')

    ag_filas = cd.leer_agendados()['filas']

    riesgo = {}
    for f in ag_filas:
        st = str(f.get('Q') or '').upper().strip()
        if not st:
            continue
        tel = re.sub(r'\D', '', str(f.get('I') or ''))
        k = _clave(tel, f.get('G'))
        r = riesgo.setdefault(k, {'total': 0, 'no_show': 0})
        r['total'] += 1
        if 'NO ASISTIO' in st or 'NO CONTEST' in st:
            r['no_show'] += 1

    citas = []
    for f in ag_filas:
        if not (f.get('L') and f.get('M') and f.get('N')):
            continue
        if f'{f["L"]}/{f["M"]}/{f["N"]}' != clave_hoy:
            continue
        st = str(f.get('Q') or '').upper()
        # Cancelada se queda en el pipeline de hoy (con opción de reprogramar);
        # no asistió y el default "sin tocar" sí salen de la vista activa.
        if 'NO ASISTIO' in st or 'NO CONFIRM' in st:
            continue
        tel = re.sub(r'\D', '', str(f.get('I') or ''))
        r = riesgo.get(_clave(tel, f.get('G')), {'total': 0, 'no_show': 0})
        citas.append({'nombre': f.get('G'), 'telefono': f.get('I'),
                      'hora': f.get('P'), 'crm': f.get('B'),
                      'campana': f.get('O'), 'estado': f.get('Q'),
                      'dia': f.get('L'), 'mes': f.get('M'),
                      'fila': f.get('_fila'),
                      'riesgo_no_show': r['no_show'], 'riesgo_citas': r['total'],
                      'riesgo_alto': r['no_show'] >= 2})
    citas.sort(key=lambda c: c['hora'] or '99')

    hojas_venta = cd.leer_venta()['hojas']
    ventas = []
    for v in hojas_venta.values():
        for f in v['filas']:
            if not (f.get('B') and f.get('C') and f.get('D')):
                continue
            if f'{f["B"]}/{f["C"]}/{f["D"]}' != clave_hoy:
                continue
            ventas.append({'nombre': f.get('G'), 'telefono': f.get('F'),
                           'doctor': f.get('M'), 'status': f.get('N'),
                           'monto': am.num(f.get('O'))})
    ventas.sort(key=lambda v: v['monto'] or 0, reverse=True)

    _cruzar_ventas_con_otra_fecha(citas, ag_filas, hojas_venta, hoy_iso)

    tareas_hoy = []
    tareas_vencidas = []
    for t in leer_tareas():
        if t['estado'] != 'PENDIENTE':
            continue
        if t['fecha'] == hoy_iso:
            tareas_hoy.append(t)
        elif t['fecha'] and t['fecha'] < hoy_iso:
            tareas_vencidas.append(t)

    return {
        'fecha': clave_hoy,
        'citas': citas,
        'ventas': ventas,
        'tareas_hoy': tareas_hoy,
        'tareas_vencidas': tareas_vencidas,
    }


# ============================================================
# CRM Plus: Cuotas / pagos a plazos
# ============================================================
def _cuota_named(f):
    return {
        'id': am.num(f.get('A')),
        'paciente': f.get('B'),
        'telefono': f.get('C'),
        'tratamiento': f.get('D'),
        'monto_total': am.num(f.get('E')),
        'n_cuotas': am.num(f.get('F')),
        'pagadas': am.num(f.get('G')),
        'monto_cuota': am.num(f.get('H')),
        'prox_fecha': f.get('I'),
        'estado': f.get('J') or 'PENDIENTE',
        'nota': f.get('K'),
    }


def _aplicar_cuota(f, d):
    for campo, letra in (('paciente', 'B'), ('telefono', 'C'),
                         ('tratamiento', 'D'), ('prox_fecha', 'I'),
                         ('nota', 'K')):
        if campo in d and d[campo] not in (None, ''):
            f[letra] = str(d[campo]).strip()
    # monto_total debe ser positivo y n_cuotas al menos 1: si no, el saldo
    # (monto_total - monto_cuota*pagadas) puede dar cifras absurdas, o con
    # n_cuotas=0 la cuota queda "PAGADO" de entrada sin haberse cobrado nada
    # (en _estado_cuota, "pagadas >= n_cuotas" es cierto con n_cuotas=0). El
    # endpoint PATCH ya valida esto también, pero se repite aquí porque
    # crear_cuota pasa por esta misma función.
    if 'monto_total' in d and d['monto_total'] not in (None, ''):
        try:
            v = float(d['monto_total'])
            if v > 0:
                f['E'] = v
        except (TypeError, ValueError):
            pass
    if 'n_cuotas' in d and d['n_cuotas'] not in (None, ''):
        try:
            v = float(d['n_cuotas'])
            if v >= 1:
                f['F'] = v
        except (TypeError, ValueError):
            pass
    if 'pagadas' in d and d['pagadas'] not in (None, ''):
        try:
            v = float(d['pagadas'])
            if v >= 0:
                # No se permite dejar más cuotas "pagadas" que n_cuotas: si
                # no, el saldo (monto_total - monto_cuota*pagadas) se va
                # negativo, como si se le debiera dinero al paciente.
                n_actual = am.num(f.get('F')) or 0
                f['G'] = min(v, n_actual) if n_actual else v
        except (TypeError, ValueError):
            pass
    if 'monto_cuota' in d and d['monto_cuota'] not in (None, ''):
        try:
            v = float(d['monto_cuota'])
            if v > 0:
                f['H'] = v
        except (TypeError, ValueError):
            pass
    elif 'monto_total' in d or 'n_cuotas' in d:
        # Si se cambia el total o el número de cuotas sin dar un monto por
        # cuota explícito, se recalcula: si no, el saldo queda calculado
        # con un monto_cuota que ya no corresponde a los valores nuevos
        # (mismo recálculo que ya hacía crear_cuota, ahora también al editar).
        total, n = am.num(f.get('E')), am.num(f.get('F'))
        if total and n:
            f['H'] = total / n
    if 'estado' in d:
        e = str(d.get('estado') or '').upper().strip()
        if e in ('PENDIENTE', 'PAGADO', 'ATRASADO'):
            f['J'] = e


def _estado_cuota(cuota):
    """Recalcula estado según cuotas pagadas y fecha de pago."""
    n = cuota['n_cuotas'] or 0
    pag = cuota['pagadas'] or 0
    if n and pag >= n:
        return 'PAGADO'
    try:
        prox = datetime.strptime(cuota['prox_fecha'], '%d/%m/%Y').date()
    except (TypeError, ValueError):
        prox = None
    if prox is not None and prox < datetime.now(TZ).date():
        return 'ATRASADO'
    return 'PENDIENTE'


def leer_cuotas(estado=None, telefono=None):
    filas = _leer_hoja('CUOTAS')
    out = []
    for f in filas:
        c = _cuota_named(f)
        if not c['pagadas']:
            c['pagadas'] = 0
        if not c['n_cuotas']:
            c['n_cuotas'] = 0
        c['saldo'] = (c['monto_total'] or 0) - (c['monto_cuota'] or 0) * c['pagadas']
        c['estado'] = _estado_cuota(c)
        out.append(c)
    if estado:
        e = str(estado).upper().strip()
        out = [c for c in out if c['estado'] == e]
    if telefono:
        tel = re.sub(r'\D', '', str(telefono))
        out = [c for c in out if c['telefono'] and tel in str(c['telefono'])]
    out.sort(key=lambda c: c['estado'] != 'ATRASADO')  # atrasados primero
    return out


def crear_cuota(datos):
    with _lock:
        filas = _leer_hoja('CUOTAS')
        f = {'A': _siguiente_id(filas)}
        _aplicar_cuota(f, datos)
        f['F'] = f.get('F') or 1
        f['G'] = f.get('G') or 0
        if not f.get('H') and f.get('E') and f['F']:
            f['H'] = float(f['E']) / float(f['F'])
        f['J'] = f.get('J') or 'PENDIENTE'
        filas.append(f)
        _reescribir('CUOTAS', filas)
        c = _cuota_named(f)
        c['saldo'] = (c['monto_total'] or 0) - (c['monto_cuota'] or 0) * (c['pagadas'] or 0)
        c['estado'] = _estado_cuota(c)
        return c


def registrar_pago_cuota(cid):
    """Marca una cuota adicional como pagada."""
    with _lock:
        filas = _leer_hoja('CUOTAS')
        for f in filas:
            if am.num(f.get('A')) == cid:
                pagadas = am.num(f.get('G')) or 0
                n = am.num(f.get('F')) or 0
                if pagadas >= n:
                    return {'error': 'La cuota ya está pagada'}
                f['G'] = pagadas + 1
                c = _cuota_named(f)
                c['saldo'] = (c['monto_total'] or 0) - (c['monto_cuota'] or 0) * (pagadas + 1)
                c['estado'] = _estado_cuota(c)
                _reescribir('CUOTAS', filas)
                return c
    return None


def actualizar_cuota(cid, cambios):
    with _lock:
        filas = _leer_hoja('CUOTAS')
        for f in filas:
            if am.num(f.get('A')) == cid:
                _aplicar_cuota(f, cambios)
                _reescribir('CUOTAS', filas)
                return _cuota_named(f)
    return None


def borrar_cuota(cid):
    with _lock:
        filas = _leer_hoja('CUOTAS')
        nuevas = [f for f in filas if am.num(f.get('A')) != cid]
        if len(nuevas) == len(filas):
            return False
        _reescribir('CUOTAS', nuevas)
        return True


# ============================================================
# CRM Plus: Caja (apertura / movimientos / cierre del día)
# ============================================================
def _json_col(f, letra):
    crudo = f.get(letra)
    if not crudo:
        return None
    try:
        return json.loads(crudo)
    except (TypeError, ValueError):
        return None


def _caja_named(f):
    return {
        'id': am.num(f.get('A')),
        'fecha': f.get('B'),
        'hora': f.get('C'),
        'tipo': f.get('D'),
        'monto': am.num(f.get('E')) or 0,
        'concepto': f.get('F'),
        'responsable': f.get('G'),
        'desglose': _json_col(f, 'H'),
        'monto_usd': am.num(f.get('I')) or 0,
        'desglose_usd': _json_col(f, 'J'),
    }


def leer_caja(fecha=None):
    filas = [_caja_named(f) for f in _leer_hoja('CAJA')]
    if fecha:
        filas = [f for f in filas if f['fecha'] == fecha]
    return sorted(filas, key=lambda f: (f['fecha'] or '', f['hora'] or ''))


def _total_desglose(desglose, denominaciones):
    """Suma cantidad x valor para cada denominación."""
    total = 0.0
    for valor in denominaciones:
        cant = am.num((desglose or {}).get(str(valor))) or 0
        total += cant * valor
    return round(total, 2)


def _monto_desde_desglose(desglose):
    return _total_desglose(desglose, DENOMINACIONES_CAJA)


def _monto_usd_desde_desglose(desglose):
    return _total_desglose(desglose, DENOMINACIONES_USD)


def _crear_movimiento_caja(tipo, datos):
    with _lock:
        filas = _leer_hoja('CAJA')
        ahora = datetime.now(TZ)
        desglose = datos.get('desglose') or None
        desglose_usd = datos.get('desglose_usd') or None
        monto = (_monto_desde_desglose(desglose) if desglose
                 else (am.num(datos.get('monto')) or 0))
        monto_usd = (_monto_usd_desde_desglose(desglose_usd) if desglose_usd
                     else (am.num(datos.get('monto_usd')) or 0))
        f = {
            'A': _siguiente_id(filas),
            'B': datos.get('fecha') or _fecha_hoy_str(),
            'C': ahora.strftime('%H:%M'),
            'D': tipo,
            'E': monto,
            'F': str(datos.get('concepto') or '').strip(),
            'G': str(datos.get('responsable') or '').strip(),
            'H': json.dumps(desglose, ensure_ascii=False) if desglose else '',
            'I': monto_usd,
            'J': json.dumps(desglose_usd, ensure_ascii=False) if desglose_usd else '',
        }
        filas.append(f)
        _reescribir('CAJA', filas)
        return _caja_named(f)


def abrir_caja(datos):
    return _crear_movimiento_caja('APERTURA', datos)


def cerrar_caja(datos):
    return _crear_movimiento_caja('CIERRE', datos)


def registrar_movimiento_caja(datos):
    tipo = str(datos.get('tipo') or '').upper().strip()
    if tipo not in ('INGRESO', 'EGRESO'):
        raise ValueError("El tipo de movimiento debe ser 'INGRESO' o 'EGRESO'")
    return _crear_movimiento_caja(tipo, datos)


def borrar_movimiento_caja(cid):
    with _lock:
        filas = _leer_hoja('CAJA')
        nuevas = [f for f in filas if am.num(f.get('A')) != cid]
        if len(nuevas) == len(filas):
            return False
        _reescribir('CAJA', nuevas)
        return True


def _metodo_pago(valor):
    """Clasifica el texto de la columna PAGO en las categorías del arqueo."""
    p = str(valor or '').strip().upper()
    if not p:
        return 'sin_dato'
    if 'EFECTIVO' in p:
        return 'efectivo'
    if 'YAPE' in p or 'PLIN' in p:
        return 'yape_plin'
    if 'TARJETA' in p or p.startswith('T.') or 'DEBITO' in p or 'CREDITO' in p or 'POS' in p:
        return 'tarjeta'
    if 'TRANSF' in p or 'DEPOSITO' in p or 'DEPÓSITO' in p or 'LINK' in p:
        return 'deposito'
    return 'otros'


def ventas_del_dia(fecha):
    """Ventas de VENTA DIARIA para la fecha 'D/MES/AAAA', separadas por forma de pago."""
    res = {'efectivo': 0.0, 'tarjeta': 0.0, 'yape_plin': 0.0,
           'deposito': 0.0, 'otros': 0.0, 'sin_dato': 0.0}
    for v in cd.leer_venta()['hojas'].values():
        for f in v['filas']:
            if not (f.get('B') and f.get('C') and f.get('D')):
                continue
            if f'{f["B"]}/{f["C"]}/{f["D"]}' != fecha:
                continue
            monto = am.num(f.get('O')) or 0
            if not monto:
                continue
            res[_metodo_pago(f.get('P'))] += monto
    res = {k: round(v, 2) for k, v in res.items()}
    res['total'] = round(sum(res.values()), 2)
    return res


def estado_caja(fecha=None):
    """Arqueo de caja del día, con las secciones del formato de recepción.

    1. Inicio de día (soles + dólares)
    2. Efectivo contado al cierre (desglose por denominación)
    3. Ventas totales por forma de pago (desde VENTA DIARIA)
    4. Gastos de caja chica / ingresos extra
    5. Ajuste: sobrante o faltante
    """
    fecha = fecha or _fecha_hoy_str()
    filas = leer_caja(fecha)
    # `filas` viene ordenado ascendente por hora: si por lo que sea hay más
    # de una APERTURA o CIERRE el mismo día (una corrección, un reintento),
    # se toma la MÁS RECIENTE, no la primera — si no, una corrección de
    # arqueo queda registrada pero invisible para siempre en este cálculo.
    apertura = next((f for f in reversed(filas) if f['tipo'] == 'APERTURA'), None)
    cierre = next((f for f in reversed(filas) if f['tipo'] == 'CIERRE'), None)
    movimientos = [f for f in filas if f['tipo'] in ('INGRESO', 'EGRESO')]
    ingresos = round(sum(f['monto'] for f in movimientos if f['tipo'] == 'INGRESO'), 2)
    egresos = round(sum(f['monto'] for f in movimientos if f['tipo'] == 'EGRESO'), 2)
    egresos_usd = round(sum(f['monto_usd'] for f in movimientos if f['tipo'] == 'EGRESO'), 2)
    ingresos_usd = round(sum(f['monto_usd'] for f in movimientos if f['tipo'] == 'INGRESO'), 2)

    ventas = ventas_del_dia(fecha)
    inicio = apertura['monto'] if apertura else 0
    inicio_usd = apertura['monto_usd'] if apertura else 0

    esperado = round(inicio + ventas['efectivo'] + ingresos - egresos, 2)
    esperado_usd = round(inicio_usd + ingresos_usd - egresos_usd, 2)
    contado = cierre['monto'] if cierre else None
    contado_usd = cierre['monto_usd'] if cierre else None
    diferencia = round(contado - esperado, 2) if cierre else None
    diferencia_usd = round(contado_usd - esperado_usd, 2) if cierre else None

    return {
        'fecha': fecha,
        'abierta': apertura is not None,
        'cerrada': cierre is not None,
        'apertura': apertura,
        'cierre': cierre,
        'movimientos': movimientos,
        'ventas': ventas,
        'ingresos': ingresos,
        'egresos': egresos,
        'ingresos_usd': ingresos_usd,
        'egresos_usd': egresos_usd,
        'inicio': inicio,
        'inicio_usd': inicio_usd,
        'esperado': esperado,
        'esperado_usd': esperado_usd,
        'contado': contado,
        'contado_usd': contado_usd,
        'diferencia': diferencia,
        'diferencia_usd': diferencia_usd,
        # alias de compatibilidad
        'monto_inicial': inicio,
        'ventas_efectivo': ventas['efectivo'],
    }


def historial_caja(limite=30):
    filas = leer_caja()
    fechas = sorted({f['fecha'] for f in filas if f['tipo'] == 'APERTURA'},
                    key=_fecha_orden, reverse=True)[:limite]
    return [estado_caja(f) for f in fechas]


# ============================================================
# CRM Plus: Inventario (catálogo de productos + kardex de stock)
# ============================================================
def _producto_named(f):
    p = {
        'id': am.num(f.get('A')),
        'codigo': f.get('B'),
        'producto': f.get('C'),
        'costo_bruto': am.num(f.get('D')) or 0,
        'stock': am.num(f.get('E')) or 0,
        'stock_minimo': am.num(f.get('F')) or 0,
        'unidad': f.get('G') or 'UND',
        'fecha_alta': f.get('H'),
        'actualizado': f.get('I'),
    }
    p['stock_bajo'] = p['stock'] <= p['stock_minimo']
    return p


def leer_productos():
    out = [_producto_named(f) for f in _leer_hoja('PRODUCTOS')]
    out.sort(key=lambda p: str(p['producto'] or '').upper())
    return out


def producto_por_codigo(codigo):
    codigo = str(codigo or '').strip().upper()
    if not codigo:
        return None
    for f in _leer_hoja('PRODUCTOS'):
        if str(f.get('B') or '').strip().upper() == codigo:
            return _producto_named(f)
    return None


def crear_producto(datos):
    with _lock:
        filas = _leer_hoja('PRODUCTOS')
        codigo = str(datos.get('codigo') or '').strip().upper()
        if not codigo:
            raise ValueError('Indica el código del producto')
        if any(str(f.get('B') or '').strip().upper() == codigo for f in filas):
            raise ValueError(f'Ya existe un producto con el código "{codigo}"')
        producto = str(datos.get('producto') or '').strip()
        if not producto:
            raise ValueError('Indica el nombre del producto')
        f = {
            'A': _siguiente_id(filas), 'B': codigo, 'C': producto,
            'D': am.num(datos.get('costo_bruto')) or 0,
            'E': am.num(datos.get('stock')) or 0,
            'F': am.num(datos.get('stock_minimo')) or 0,
            'G': str(datos.get('unidad') or 'UND').strip().upper(),
            'H': _ahora(), 'I': _ahora(),
        }
        filas.append(f)
        _reescribir('PRODUCTOS', filas)
        return _producto_named(f)


def actualizar_producto(pid, cambios):
    with _lock:
        filas = _leer_hoja('PRODUCTOS')
        for f in filas:
            if am.num(f.get('A')) == pid:
                if cambios.get('codigo'):
                    nuevo = str(cambios['codigo']).strip().upper()
                    if any(am.num(g.get('A')) != pid
                           and str(g.get('B') or '').strip().upper() == nuevo
                           for g in filas):
                        raise ValueError(f'Ya existe un producto con el código "{nuevo}"')
                    f['B'] = nuevo
                for campo, letra in (('producto', 'C'), ('unidad', 'G')):
                    if cambios.get(campo):
                        f[letra] = str(cambios[campo]).strip()
                for campo, letra in (('costo_bruto', 'D'), ('stock_minimo', 'F')):
                    if campo in cambios and cambios[campo] not in (None, ''):
                        f[letra] = am.num(cambios[campo]) or 0
                f['I'] = _ahora()
                _reescribir('PRODUCTOS', filas)
                return _producto_named(f)
    return None


def borrar_producto(pid):
    with _lock:
        filas = _leer_hoja('PRODUCTOS')
        nuevas = [f for f in filas if am.num(f.get('A')) != pid]
        if len(nuevas) == len(filas):
            return False
        _reescribir('PRODUCTOS', nuevas)
        return True


def _movimiento_named(f):
    return {
        'id': am.num(f.get('A')), 'fecha': f.get('B'), 'codigo': f.get('C'),
        'producto': f.get('D'), 'tipo': f.get('E'), 'cantidad': am.num(f.get('F')) or 0,
        'stock_resultante': am.num(f.get('G')) or 0, 'referencia': f.get('H'),
        'nota': f.get('I'),
    }


def leer_movimientos_stock(codigo=None, limite=200):
    filas = [_movimiento_named(f) for f in _leer_hoja('MOVIMIENTOS_STOCK')]
    if codigo:
        c = str(codigo).strip().upper()
        filas = [f for f in filas if str(f['codigo'] or '').strip().upper() == c]
    filas.sort(key=lambda f: am.num(f['id']) or 0, reverse=True)
    return filas[:limite]


def registrar_movimiento_stock(codigo, tipo, cantidad, referencia='', nota=''):
    """Registra un movimiento de stock y actualiza el producto.

    ENTRADA suma, SALIDA resta, AJUSTE fija el stock al valor de
    ``cantidad`` directamente (para corregir conteos).
    """
    tipo = str(tipo or '').upper().strip()
    if tipo not in TIPO_MOVIMIENTO:
        raise ValueError(f"El tipo de movimiento debe ser {', '.join(TIPO_MOVIMIENTO)}")
    cantidad = am.num(cantidad)
    if cantidad is None or cantidad < 0:
        raise ValueError('Indica una cantidad válida')
    with _lock:
        todo = {n: _leer_hoja(n) for n in HOJAS}
        codigo_u = str(codigo or '').strip().upper()
        prod_f = next((f for f in todo['PRODUCTOS']
                       if str(f.get('B') or '').strip().upper() == codigo_u), None)
        if prod_f is None:
            raise ValueError(f'No existe un producto con el código "{codigo}"')
        stock_actual = am.num(prod_f.get('E')) or 0
        if tipo == 'ENTRADA':
            nuevo_stock = stock_actual + cantidad
        elif tipo == 'SALIDA':
            nuevo_stock = stock_actual - cantidad
            # No se puede "vender" stock que no existe: antes esto quedaba
            # en negativo sin ningún aviso (una venta con cantidad mal
            # tipeada, o dos ventas casi simultáneas del último producto,
            # dejaban el kardex corrupto en silencio). Los llamadores
            # (venta directa, "Compró" desde Agendados, salida manual) ya
            # atrapan esta excepción y la muestran como aviso sin bloquear
            # la venta en sí.
            if nuevo_stock < 0:
                raise ValueError(
                    f'Stock insuficiente de "{prod_f.get("C") or codigo}": '
                    f'quedan {stock_actual}, se pidieron {cantidad}')
        else:
            nuevo_stock = cantidad
        prod_f['E'] = nuevo_stock
        prod_f['I'] = _ahora()

        m = {'A': _siguiente_id(todo['MOVIMIENTOS_STOCK']), 'B': _ahora(),
             'C': prod_f.get('B'), 'D': prod_f.get('C'), 'E': tipo, 'F': cantidad,
             'G': nuevo_stock, 'H': str(referencia or '').strip(),
             'I': str(nota or '').strip()}
        todo['MOVIMIENTOS_STOCK'].append(m)
        _guardar({'PRODUCTOS': todo['PRODUCTOS'], 'MOVIMIENTOS_STOCK': todo['MOVIMIENTOS_STOCK']})
        return {'producto': _producto_named(prod_f), 'movimiento': _movimiento_named(m)}


def registrar_entrada_stock(codigo, cantidad, referencia='', nota=''):
    return registrar_movimiento_stock(codigo, 'ENTRADA', cantidad, referencia, nota)


def registrar_salida_stock(codigo, cantidad, referencia='', nota=''):
    return registrar_movimiento_stock(codigo, 'SALIDA', cantidad, referencia, nota)


def ajustar_stock(codigo, stock_nuevo, nota=''):
    return registrar_movimiento_stock(codigo, 'AJUSTE', stock_nuevo, 'Ajuste manual', nota)


def _prefijo_venta(hoja, fila):
    return f'VENTA:{hoja}:{fila}:'


def reversar_stock_de_venta(hoja, fila):
    """Si al crear la venta se descontó stock de un producto (venta con
    producto de inventario), al borrar esa venta hay que devolver el stock:
    si no, el inventario queda perdido para siempre aunque la venta ya no
    exista. Busca el movimiento por la referencia estructurada que deja
    crm_venta_nuevo ("VENTA:hoja:fila: nombre") y agrega una ENTRADA que lo
    compensa — no se borra el movimiento original, para no perder el
    historial del kardex.

    Cuenta candidatos vs reversas ya hechas (en vez de sólo "existe una
    reversa") para seguir funcionando bien si Drive reutiliza el mismo
    número de fila para una venta distinta más adelante.
    """
    prefijo = _prefijo_venta(hoja, fila)
    prefijo_reversa = f'Reversa venta borrada {prefijo}'
    movs = _leer_hoja('MOVIMIENTOS_STOCK')
    candidatos = [m for m in movs if str(m.get('H') or '').startswith(prefijo)]
    ya_revertidos = sum(1 for m in movs if str(m.get('H') or '').startswith(prefijo_reversa))
    if len(candidatos) <= ya_revertidos:
        return None
    mov = candidatos[-1]
    codigo = mov.get('C')
    cantidad = am.num(mov.get('F')) or 0
    if not codigo or cantidad <= 0:
        return None
    registrar_movimiento_stock(codigo, 'ENTRADA', cantidad,
                               referencia=f'{prefijo_reversa}{mov.get("D") or ""}',
                               nota='Automático al borrar la venta')
    return f'Se devolvieron {cantidad} unidad(es) de "{mov.get("D") or codigo}" al stock.'


# ============================================================
# CRM Plus: Planilla (trabajadores + pagos quincenales)
# ============================================================
def _trabajador_named(f):
    return {
        'id': am.num(f.get('A')),
        'nombre': f.get('B'),
        'cargo': f.get('C'),
        'sueldo_quincena': am.num(f.get('D')) or 0,
        'porcentaje_comision': am.num(f.get('E')) or 0,
        'metodo_pago': f.get('F'),
        'cuenta': f.get('G'),
        'estado': f.get('H') or 'ACTIVO',
        'fecha_alta': f.get('I'),
        'nombre_ventas': f.get('J'),
    }


# Títulos que se escriben en la columna DOCTOR de VENTA DIARIA pero no son
# parte del nombre ("DR. BORIS" vs "Boris Ramirez").
_TRATAMIENTOS_NOMBRE = {'DR', 'DRA', 'DOC', 'DOCTOR', 'DOCTORA', 'SR', 'SRA', 'SRTA', 'LIC'}


def _tokens_nombre(v):
    s = _status_normalizado(v).replace('.', ' ')
    return {t for t in s.split() if t and t not in _TRATAMIENTOS_NOMBRE}


def _coincide_doctor(valor, nombre, alias=''):
    """¿La celda DOCTOR de una venta corresponde a este trabajador?

    Si el trabajador tiene alias se exige coincidencia exacta con él. Si no,
    basta con que los nombres compartan todas las palabras del más corto: en la
    práctica el consultorio escribe "DR. BORIS" y el catálogo "Boris Ramirez"."""
    celda = _tokens_nombre(valor)
    if not celda:
        return False
    if str(alias or '').strip():
        return celda == _tokens_nombre(alias)
    propio = _tokens_nombre(nombre)
    if not propio:
        return False
    return celda <= propio or propio <= celda


def leer_trabajadores(solo_activos=False):
    out = [_trabajador_named(f) for f in _leer_hoja('TRABAJADORES')]
    if solo_activos:
        out = [t for t in out if t['estado'] == 'ACTIVO']
    out.sort(key=lambda t: str(t['nombre'] or '').upper())
    return out


def crear_trabajador(datos):
    with _lock:
        filas = _leer_hoja('TRABAJADORES')
        nombre = str(datos.get('nombre') or '').strip()
        if not nombre:
            raise ValueError('Indica el nombre del trabajador')
        f = {
            'A': _siguiente_id(filas), 'B': nombre,
            'C': str(datos.get('cargo') or '').strip(),
            'D': am.num(datos.get('sueldo_quincena')) or 0,
            'E': am.num(datos.get('porcentaje_comision')) or 0,
            'F': str(datos.get('metodo_pago') or '').strip(),
            'G': str(datos.get('cuenta') or '').strip(),
            'H': 'ACTIVO', 'I': _ahora(),
            'J': str(datos.get('nombre_ventas') or '').strip(),
        }
        filas.append(f)
        _reescribir('TRABAJADORES', filas)
        return _trabajador_named(f)


def actualizar_trabajador(tid, cambios):
    with _lock:
        filas = _leer_hoja('TRABAJADORES')
        for f in filas:
            if am.num(f.get('A')) == tid:
                if cambios.get('nombre'):
                    f['B'] = str(cambios['nombre']).strip()
                if 'cargo' in cambios:
                    f['C'] = str(cambios.get('cargo') or '').strip()
                if 'sueldo_quincena' in cambios and cambios['sueldo_quincena'] not in (None, ''):
                    f['D'] = am.num(cambios['sueldo_quincena']) or 0
                if 'porcentaje_comision' in cambios and cambios['porcentaje_comision'] not in (None, ''):
                    f['E'] = am.num(cambios['porcentaje_comision']) or 0
                if 'metodo_pago' in cambios:
                    f['F'] = str(cambios.get('metodo_pago') or '').strip()
                if 'cuenta' in cambios:
                    f['G'] = str(cambios.get('cuenta') or '').strip()
                if 'nombre_ventas' in cambios:
                    f['J'] = str(cambios.get('nombre_ventas') or '').strip()
                if cambios.get('estado') in ESTADO_TRABAJADOR:
                    f['H'] = cambios['estado']
                _reescribir('TRABAJADORES', filas)
                return _trabajador_named(f)
    return None


def borrar_trabajador(tid):
    with _lock:
        filas = _leer_hoja('TRABAJADORES')
        nuevas = [f for f in filas if am.num(f.get('A')) != tid]
        if len(nuevas) == len(filas):
            return False
        _reescribir('TRABAJADORES', nuevas)
        return True


def calcular_comision(nombre, anio, mes, quincena, porcentaje, alias=''):
    """Comisión = % del monto de VENTA DIARIA (columna VENTA) de las ventas
    realizadas (excluye NO SE REALIZO) atribuidas a ``nombre`` por la columna
    DOCTOR. Requiere que el nombre del trabajador coincida (sin distinguir
    mayúsculas) con lo escrito en esa columna.

    La comisión se paga una sola vez al mes: en la primera quincena es 0 y en
    la segunda se calcula sobre las ventas del mes completo."""
    porcentaje = am.num(porcentaje) or 0
    if not porcentaje or int(quincena) != QUINCENA_COMISION:
        return 0.0
    mes = str(mes or '').strip().upper()
    mes_num = _MM_IDX.get(mes)
    if not mes_num or not _tokens_nombre(nombre):
        return 0.0
    d1, d2 = 1, calendar.monthrange(int(anio), mes_num)[1]
    total = 0.0
    for v in cd.leer_venta()['hojas'].values():
        for f in v['filas']:
            if not _coincide_doctor(f.get('M'), nombre, alias):
                continue
            if str(f.get('C') or '').strip().upper() != mes:
                continue
            if am.num(f.get('D')) != int(anio):
                continue
            dia = am.num(f.get('B'))
            if dia is None or not (d1 <= dia <= d2):
                continue
            if not _es_venta_registrada(f.get('N')):
                continue
            total += am.num(f.get('O')) or 0
    return round(total * porcentaje / 100.0, 2)


def _pago_named(f):
    return {
        'id': am.num(f.get('A')),
        'trabajador_id': am.num(f.get('B')),
        'nombre': f.get('C'),
        'anio': am.num(f.get('D')),
        'mes': f.get('E'),
        'quincena': am.num(f.get('F')),
        'sueldo_base': am.num(f.get('G')) or 0,
        'comision': am.num(f.get('H')) or 0,
        'monto_total': am.num(f.get('I')) or 0,
        'estado': f.get('J') or 'PENDIENTE',
        'fecha_pago': f.get('K'),
        'metodo_pago': f.get('L'),
        'nota': f.get('M'),
        'extra': am.num(f.get('N')) or 0,
        'motivo_extra': f.get('O'),
        'estimado': False,
    }


def leer_planilla(anio=None, mes=None, quincena=None):
    out = [_pago_named(f) for f in _leer_hoja('PLANILLA')]
    if anio:
        out = [p for p in out if p['anio'] == int(anio)]
    if mes:
        out = [p for p in out if str(p['mes'] or '').strip().upper() == str(mes).strip().upper()]
    if quincena:
        out = [p for p in out if p['quincena'] == int(quincena)]
    out.sort(key=lambda p: str(p['nombre'] or '').upper())
    return out


def previsualizar_planilla(anio, mes, quincena):
    """Estimado de lo que cobraría cada trabajador ACTIVO en la quincena, sin
    escribir nada. Se usa para que la pantalla muestre el monto antes de
    generar la planilla; omite a quien ya tiene su pago generado."""
    anio = int(anio)
    mes = str(mes).strip().upper()
    quincena = int(quincena)
    if quincena not in (1, 2) or mes not in _MM_IDX:
        return []
    ya = {(am.num(f.get('B')), am.num(f.get('D')),
           str(f.get('E') or '').strip().upper(), am.num(f.get('F')))
          for f in _leer_hoja('PLANILLA')}
    out = []
    for t in leer_trabajadores(solo_activos=True):
        if (t['id'], anio, mes, quincena) in ya:
            continue
        comision = calcular_comision(t['nombre'], anio, mes, quincena,
                                     t['porcentaje_comision'], t['nombre_ventas'])
        out.append({
            'id': None, 'trabajador_id': t['id'], 'nombre': t['nombre'],
            'anio': anio, 'mes': mes, 'quincena': quincena,
            'sueldo_base': t['sueldo_quincena'], 'comision': comision,
            'extra': 0, 'motivo_extra': None,
            'monto_total': round(t['sueldo_quincena'] + comision, 2),
            'estado': 'ESTIMADO', 'fecha_pago': None,
            'metodo_pago': t['metodo_pago'], 'nota': None, 'estimado': True,
        })
    out.sort(key=lambda p: str(p['nombre'] or '').upper())
    return out


def generar_planilla_quincena(anio, mes, quincena):
    """Crea (o refresca) el pago de la quincena para cada trabajador ACTIVO.
    No toca los pagos que ya están marcados como PAGADO."""
    anio = int(anio)
    mes = str(mes).strip().upper()
    quincena = int(quincena)
    if quincena not in (1, 2):
        raise ValueError('La quincena debe ser 1 o 2')
    if mes not in _MM_IDX:
        raise ValueError(f'Mes inválido: {mes}')
    with _lock:
        todo = {n: _leer_hoja(n) for n in HOJAS}
        trabajadores = [_trabajador_named(f) for f in todo['TRABAJADORES']
                        if (f.get('H') or 'ACTIVO') == 'ACTIVO']
        if not trabajadores:
            raise ValueError('No hay trabajadores activos en el catálogo')
        existentes = {}
        for f in todo['PLANILLA']:
            clave = (am.num(f.get('B')), am.num(f.get('D')),
                     str(f.get('E') or '').strip().upper(), am.num(f.get('F')))
            existentes[clave] = f
        creados = actualizados = sin_tocar = 0
        for t in trabajadores:
            comision = calcular_comision(t['nombre'], anio, mes, quincena,
                                         t['porcentaje_comision'], t['nombre_ventas'])
            sueldo = t['sueldo_quincena']
            clave = (t['id'], anio, mes, quincena)
            f = existentes.get(clave)
            if f is not None:
                if (f.get('J') or 'PENDIENTE') == 'PAGADO':
                    sin_tocar += 1
                    continue
                # El extra se escribe a mano (horas extra, feriados): al
                # regenerar se respeta y sólo se recalcula el total.
                extra = am.num(f.get('N')) or 0
                f['G'] = sueldo; f['H'] = comision
                f['I'] = round(sueldo + comision + extra, 2)
                actualizados += 1
            else:
                f = {'A': _siguiente_id(todo['PLANILLA']), 'B': t['id'], 'C': t['nombre'],
                     'D': anio, 'E': mes, 'F': quincena, 'G': sueldo, 'H': comision,
                     'I': round(sueldo + comision, 2), 'J': 'PENDIENTE'}
                todo['PLANILLA'].append(f)
                creados += 1
        _guardar({'PLANILLA': todo['PLANILLA']})
        return {'creados': creados, 'actualizados': actualizados, 'sin_tocar': sin_tocar}


def registrar_extra(trabajador_id, anio, mes, quincena, extra, motivo=''):
    """Anota un monto extra (horas extra, feriados, bonos) a un trabajador.

    Si esa quincena todavía no estaba generada crea sólo la fila de ese
    trabajador, para no tener que generar la planilla entera nada más que para
    apuntar un extra. No toca un pago ya marcado como PAGADO."""
    anio, quincena = int(anio), int(quincena)
    mes = str(mes).strip().upper()
    if quincena not in (1, 2):
        raise ValueError('La quincena debe ser 1 o 2')
    if mes not in _MM_IDX:
        raise ValueError(f'Mes inválido: {mes}')
    extra = am.num(extra) or 0
    with _lock:
        todo = {n: _leer_hoja(n) for n in HOJAS}
        trabajador = next((_trabajador_named(f) for f in todo['TRABAJADORES']
                           if am.num(f.get('A')) == int(trabajador_id)), None)
        if not trabajador:
            raise ValueError('No se encontró ese trabajador')
        fila = next((f for f in todo['PLANILLA']
                     if am.num(f.get('B')) == trabajador['id']
                     and am.num(f.get('D')) == anio
                     and str(f.get('E') or '').strip().upper() == mes
                     and am.num(f.get('F')) == quincena), None)
        if fila is not None and (fila.get('J') or 'PENDIENTE') == 'PAGADO':
            raise ValueError('Ese pago ya está marcado como pagado: reábrelo antes de cambiarlo')
        if fila is None:
            comision = calcular_comision(trabajador['nombre'], anio, mes, quincena,
                                         trabajador['porcentaje_comision'],
                                         trabajador['nombre_ventas'])
            fila = {'A': _siguiente_id(todo['PLANILLA']), 'B': trabajador['id'],
                    'C': trabajador['nombre'], 'D': anio, 'E': mes, 'F': quincena,
                    'G': trabajador['sueldo_quincena'], 'H': comision,
                    'J': 'PENDIENTE'}
            todo['PLANILLA'].append(fila)
        fila['N'] = extra
        fila['O'] = str(motivo or '').strip()
        fila['I'] = round((am.num(fila.get('G')) or 0) + (am.num(fila.get('H')) or 0) + extra, 2)
        _guardar({'PLANILLA': todo['PLANILLA']})
        return _pago_named(fila)


def actualizar_pago_planilla(pid, cambios):
    with _lock:
        filas = _leer_hoja('PLANILLA')
        for f in filas:
            if am.num(f.get('A')) == pid:
                if 'sueldo_base' in cambios and cambios['sueldo_base'] not in (None, ''):
                    f['G'] = am.num(cambios['sueldo_base']) or 0
                if 'comision' in cambios and cambios['comision'] not in (None, ''):
                    f['H'] = am.num(cambios['comision']) or 0
                if 'extra' in cambios and cambios['extra'] not in (None, ''):
                    f['N'] = am.num(cambios['extra']) or 0
                if 'motivo_extra' in cambios:
                    f['O'] = str(cambios.get('motivo_extra') or '').strip()
                if {'sueldo_base', 'comision', 'extra'} & set(cambios):
                    f['I'] = round((am.num(f.get('G')) or 0) + (am.num(f.get('H')) or 0)
                                   + (am.num(f.get('N')) or 0), 2)
                if cambios.get('estado') in ESTADO_PAGO:
                    f['J'] = cambios['estado']
                    if cambios['estado'] == 'PAGADO':
                        f['K'] = f.get('K') or _fecha_hoy_str()
                    else:
                        f['K'] = None
                if 'metodo_pago' in cambios:
                    f['L'] = str(cambios.get('metodo_pago') or '').strip()
                if 'nota' in cambios:
                    f['M'] = str(cambios.get('nota') or '').strip()
                _reescribir('PLANILLA', filas)
                return _pago_named(f)
    return None


def borrar_pago_planilla(pid):
    with _lock:
        filas = _leer_hoja('PLANILLA')
        nuevas = [f for f in filas if am.num(f.get('A')) != pid]
        if len(nuevas) == len(filas):
            return False
        _reescribir('PLANILLA', nuevas)
        return True


# ============================================================
# CRM Plus: Historia clínica
# ============================================================
# Campos de texto de la historia: letra en la hoja -> nombre lógico.
HISTORIA_TEXTOS = {
    'I': 'motivo', 'J': 'antecedentes', 'K': 'alergias', 'L': 'diagnostico',
    'M': 'tratamiento', 'N': 'indicaciones', 'O': 'proximo_control',
    'P': 'observacion',
}

# Estructura de la ficha en papel "HISTORIA MEDICA DEL PACIENTE". Vive aquí y
# se le envía al navegador para que el formulario y lo que se guarda no puedan
# quedar desalineados. Cada bloque se guarda como JSON en una sola celda.
HISTORIA_PATOLOGICO = [
    ('alergias', 'Alergias'),
    ('int_quirurgica', 'Int. quirúrgica'),
    ('trata_anteriores', 'Tratamientos anteriores'),
    ('antec_cancerigenos', 'Antec. cancerígenos'),
    ('articulaciones', 'Articulaciones'),
    ('protesis', 'Prótesis'),
    ('varices', 'Várices'),
    ('enfermedad_renal', 'Enfermedad renal'),
    ('columna', 'Columna'),
    ('glandular', 'Glandular'),
    ('medicamentos', 'Medicamentos'),
    ('tiroides', 'Tiroides'),
    ('enfermedad_cardiaca', 'Enfer. cardiaca'),
    ('hipo', 'Hipo'),
    ('digestion', 'Digestión'),
    ('hiper', 'Hiper'),
    ('menopausia', 'Menopausia'),
    ('menstruaciones_regulares', 'Menstruaciones regulares'),
    ('actividad_fisica', 'Actividad física'),
    ('fuma', 'Fuma'),
    ('aumento_peso_menstrual', '¿Advierte aumento de peso en periodo menstrual?'),
    ('anticonceptivos', '¿Toma anticonceptivos o medicamento hormonal?'),
]
HISTORIA_PIEL = [
    ('piel_seca', 'Piel seca'),
    ('piel_grasa', 'Piel grasa'),
    ('levemente_seca', 'Levemente seca'),
    ('levemente_grasa', 'Levemente grasa'),
    ('medianamente_seca', 'Medianamente seca'),
    ('medianamente_grasa', 'Medianamente grasa'),
    ('muy_seca', 'Piel muy seca'),
    ('muy_grasa', 'Piel muy grasa'),
]
HISTORIA_ANATOMIA = [
    ('tox_rostro', 'Tox. botulínica rostro'),
    ('tox_masetero', 'Tox. botulínica masetero'),
    ('ah_labios', 'A.H. en labios'),
    ('ah_nasogenianos', 'A.H. surcos nasogenianos'),
    ('ah_anclaje', 'A.H. puntos de anclaje'),
    ('ah_ojeras', 'A.H. ojeras'),
    ('ah_maxilar', 'A.H. maxilar'),
    ('prp_facial', 'PRP facial'),
    ('prp_capilar', 'PRP capilar'),
    ('bioestimulador', 'Bioestimulador'),
    ('otros', 'Otros (especificar)'),
]
HISTORIA_CITA_COLS = ['fecha', 'tratamiento', 'cita', 'monto', 'saldo']
# Bloque estructurado -> letra de la celda donde se guarda su JSON.
HISTORIA_BLOQUES = {'patologico': 'S', 'piel': 'T', 'anatomia': 'U', 'citas': 'V'}

CATALOGO_HISTORIA = {
    'patologico': [{'clave': c, 'etiqueta': e} for c, e in HISTORIA_PATOLOGICO],
    'piel': [{'clave': c, 'etiqueta': e} for c, e in HISTORIA_PIEL],
    'anatomia': [{'clave': c, 'etiqueta': e} for c, e in HISTORIA_ANATOMIA],
    'cita_columnas': HISTORIA_CITA_COLS,
}


def _leer_json(valor, por_defecto):
    """Las celdas de bloques guardan JSON; una celda vacía o corrupta no debe
    tumbar la lectura de toda la historia."""
    if not valor:
        return por_defecto
    try:
        d = json.loads(valor)
    except (json.JSONDecodeError, TypeError):
        return por_defecto
    return d if isinstance(d, type(por_defecto)) else por_defecto


def _limpiar_sino(datos, catalogo):
    """Sólo se guardan las claves del catálogo y sólo SI/NO: nada de texto
    libre entrando por la puerta de atrás del JSON."""
    validas = {c for c, _ in catalogo}
    out = {}
    for k, v in (datos or {}).items():
        if k in validas:
            s = str(v or '').strip().upper()
            if s in ('SI', 'NO'):
                out[k] = s
    return out


def _limpiar_anatomia(datos):
    validas = {c for c, _ in HISTORIA_ANATOMIA}
    return {k: str(v or '').strip()[:300] for k, v in (datos or {}).items()
            if k in validas and str(v or '').strip()}


def _limpiar_citas(filas):
    out = []
    for f in (filas or [])[:40]:
        if not isinstance(f, dict):
            continue
        fila = {c: str(f.get(c) or '').strip()[:120] for c in HISTORIA_CITA_COLS}
        if any(fila.values()):
            out.append(fila)
    return out


def _historia_named(f):
    d = {
        'id': am.num(f.get('A')),
        'fecha': f.get('B'), 'hora': f.get('C'),
        'paciente': f.get('D'), 'telefono': f.get('E'),
        'dni': f.get('F'), 'edad': am.num(f.get('G')),
        'doctor': f.get('H'),
        'agendado_fila': am.num(f.get('Q')),
        'direccion': f.get('R'),
        'patologico': _leer_json(f.get('S'), {}),
        'piel': _leer_json(f.get('T'), {}),
        'anatomia': _leer_json(f.get('U'), {}),
        'citas': _leer_json(f.get('V'), []),
    }
    for letra, campo in HISTORIA_TEXTOS.items():
        d[campo] = f.get(letra)
    return d


def _tel_digitos(v):
    return re.sub(r'\D', '', str(v if v is not None else ''))


def leer_historias(telefono=None, paciente=None, desde=None, hasta=None):
    """Historias clínicas, de la más reciente a la más antigua.

    ``telefono`` y ``paciente`` filtran al paciente; si se pasan ambos, basta
    con que coincida uno (el teléfono es el criterio fuerte).
    """
    out = [_historia_named(f) for f in _leer_hoja('HISTORIAS')]
    tel = _tel_digitos(telefono)
    nom = _status_normalizado(paciente)
    if tel or nom:
        out = [h for h in out
               if (tel and _tel_digitos(h['telefono']) == tel)
               or (nom and _status_normalizado(h['paciente']) == nom)]
    if desde:
        out = [h for h in out if (h['fecha'] or '') >= desde]
    if hasta:
        out = [h for h in out if (h['fecha'] or '') <= hasta]
    out.sort(key=lambda h: (h['fecha'] or '', h['hora'] or ''), reverse=True)
    return out


def _fila_historia(datos, base=None):
    """Construye la fila de la hoja a partir de los datos del formulario.

    Con ``base`` (edición) sólo toca los campos presentes en ``datos``, así un
    PATCH parcial no borra lo que no envía; un campo enviado vacío sí se limpia.
    """
    f = dict(base or {})
    parcial = base is not None

    def poner(letra, campo, valor):
        if not parcial or campo in datos:
            f[letra] = valor

    poner('B', 'fecha', str(datos.get('fecha') or '').strip())
    poner('C', 'hora', str(datos.get('hora') or '').strip())
    poner('D', 'paciente', str(datos.get('paciente') or '').strip())
    poner('E', 'telefono', _tel_digitos(datos.get('telefono')))
    poner('F', 'dni', str(datos.get('dni') or '').strip())
    poner('G', 'edad', am.num(datos.get('edad')))
    poner('H', 'doctor', str(datos.get('doctor') or '').strip())
    for letra, campo in HISTORIA_TEXTOS.items():
        poner(letra, campo, str(datos.get(campo) or '').strip())
    poner('R', 'direccion', str(datos.get('direccion') or '').strip())
    # Los bloques van como JSON en una celda: se guardan enteros o no se
    # tocan si el campo ni siquiera viene en el PATCH. patologico/piel/
    # anatomia son diccionarios {clave: valor} — si SÍ vienen pero con sólo
    # algunas claves (p. ej. {"fuma": "SI"} de un catálogo con 10 ítems), se
    # mezclan sobre lo que ya había, en vez de reemplazar el bloque entero y
    # perder las demás respuestas ya guardadas. "citas" es una tabla (lista
    # de filas), no un diccionario: ahí sí se reemplaza completa, que es el
    # comportamiento correcto para una lista que el formulario reenvía entera.
    limpiadores = {
        'patologico': lambda v: _limpiar_sino(v, HISTORIA_PATOLOGICO),
        'piel': lambda v: _limpiar_sino(v, HISTORIA_PIEL),
        'anatomia': _limpiar_anatomia,
        'citas': _limpiar_citas,
    }
    for campo, letra in HISTORIA_BLOQUES.items():
        if not parcial or campo in datos:
            entrante = datos.get(campo)
            if parcial and campo != 'citas' and isinstance(entrante, dict):
                previo = _leer_json(f.get(letra), {})
                if isinstance(previo, dict):
                    entrante = {**previo, **entrante}
            valor = limpiadores[campo](entrante)
            f[letra] = json.dumps(valor, ensure_ascii=False) if valor else ''
    if am.num(datos.get('agendado_fila')):
        f['Q'] = am.num(datos['agendado_fila'])
    if not f.get('B'):
        f['B'] = _hoy_iso()
    return f


def _hoy_iso():
    return datetime.now(TZ).strftime('%Y-%m-%d')


def crear_historia(datos):
    if not str(datos.get('paciente') or '').strip():
        raise ValueError('Indica el nombre del paciente')
    with _lock:
        filas = _leer_hoja('HISTORIAS')
        f = _fila_historia(datos)
        f['A'] = _siguiente_id(filas)
        filas.append(f)
        _reescribir('HISTORIAS', filas)
        return _historia_named(f)


def actualizar_historia(hid, cambios):
    with _lock:
        filas = _leer_hoja('HISTORIAS')
        for f in filas:
            if am.num(f.get('A')) != hid:
                continue
            nueva = _fila_historia(cambios, base=f)
            nueva['A'] = f.get('A')
            f.clear()
            f.update(nueva)
            _reescribir('HISTORIAS', filas)
            return _historia_named(f)
        return None


def borrar_historia(hid):
    with _lock:
        filas = _leer_hoja('HISTORIAS')
        nuevas = [f for f in filas if am.num(f.get('A')) != hid]
        if len(nuevas) == len(filas):
            return False
        _reescribir('HISTORIAS', nuevas)
        return True
